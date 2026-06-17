from __future__ import annotations

from pathlib import Path
import warnings
import html
import json
import re
import base64

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import plotly.graph_objects as go

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=pd.errors.PerformanceWarning)

from module1_engine import (
    CompanyOverview,
    load_overview_from_csv,
    load_timeseries_from_csv,
    ensure_derived_metrics,
    append_ttm_row,
    latest_metric_cards,
    format_table_for_display,
)
from adapters.excel_financial_provider import ExcelFinancialProvider
from adapters.module2_web_research import WebEvidenceAgent
from adapters.vn_public_crawler import PublicFireAntCrawler, PublicVietstockCrawler, PublicSimplizeCrawler
from adapters.base import ProviderResult, MODULE1_OVERVIEW_COLUMNS, MODULE1_TIMESERIES_COLUMNS, normalize_columns
from module2_engine import (
    load_assumptions,
    classify_company,
    build_module2_valuation_table,
    build_valuation_range,
    build_porter_moat_scorecard,
    build_value_chain_table,
    build_risk_scenario_table,
    build_beneish_mscore_table,
    build_accrual_quality_table,
    build_modified_jones_kothari_table,
    build_real_earnings_management_table,
    build_module2_summary,
    export_module2_report_markdown,
)

from report_exporter import render_full_report_export_box

APP_DIR = Path(__file__).resolve().parent
LOGO_PATH = APP_DIR / "assets" / "trecapital_logo.png"
DEFAULT_OVERVIEW_CSV = APP_DIR / "sample_data" / "company_overview_sample.csv"
DEFAULT_YEAR_CSV = APP_DIR / "sample_data" / "financial_timeseries_year.csv"
DEFAULT_QUARTER_CSV = APP_DIR / "sample_data" / "financial_timeseries_quarter.csv"
BUNDLED_XLSM = APP_DIR / "data_sources" / "Financial-v1.3.0.xlsm"
DATA_CACHE_DIR = APP_DIR / "data_cache"
RAW_DIR = APP_DIR / "raw_data"
REPORT_DIR = APP_DIR / "reports"
ASSUMPTIONS_PATH = APP_DIR / "configs" / "valuation_assumptions.json"
APP_NAME = "Äá»‹nh giÃ¡ chuyÃªn sÃ¢u"
APP_VERSION = "V23.66-formula-source-audit-fix"

DATA_SOURCE_DISPLAY_TO_INTERNAL = {
    "Tá»± Ä‘á»™ng": "Tá»± Ä‘á»™ng tá»« dá»¯ liá»‡u tá»•ng quan",
    "Dá»¯ liá»‡u Æ°u tiÃªn 1": "FireAnt + Vietstock",
    "Dá»¯ liá»‡u Æ°u tiÃªn 2": "FireAnt",
    "Dá»¯ liá»‡u Æ°u tiÃªn 3": "Vietstock",
    "Dá»¯ liá»‡u tÃ­ch há»£p": "Financial tÃ­ch há»£p",
    "Dá»¯ liá»‡u máº«u": "CSV máº«u tÃ­ch há»£p",
}
DATA_SOURCE_INTERNAL_TO_DISPLAY = {v: k for k, v in DATA_SOURCE_DISPLAY_TO_INTERNAL.items()}
PEER_SOURCE_DISPLAY_TO_INTERNAL = {
    "CÃ¹ng cháº¿ Ä‘á»™ mÃ£ gá»‘c": "__same__",
    "Dá»¯ liá»‡u Æ°u tiÃªn": "FireAnt",
    "Dá»¯ liá»‡u tÃ­ch há»£p": "Financial tÃ­ch há»£p",
    "Dá»¯ liá»‡u máº«u": "CSV máº«u tÃ­ch há»£p",
}


def _to_internal_source(display_value: object) -> str:
    return DATA_SOURCE_DISPLAY_TO_INTERNAL.get(str(display_value), str(display_value))


def _to_internal_peer_source(display_value: object, current_source: str) -> str:
    val = PEER_SOURCE_DISPLAY_TO_INTERNAL.get(str(display_value), str(display_value))
    return current_source if val == "__same__" else val


def _public_text(value: object) -> str:
    # Pandas pd.NA cannot be evaluated in boolean context: bool(pd.NA) raises
    # "TypeError: boolean value of NA is ambiguous". Keep this helper safe for
    # all nullable/object columns before display.
    if value is None:
        text = ""
    else:
        try:
            if pd.isna(value):
                text = ""
            else:
                text = str(value)
        except Exception:
            text = str(value)
    for raw, public in DATA_SOURCE_INTERNAL_TO_DISPLAY.items():
        text = text.replace(raw, public)
    replacements = {
        "FireAnt": "Dá»¯ liá»‡u Æ°u tiÃªn",
        "Vietstock": "Dá»¯ liá»‡u Æ°u tiÃªn",
        "Simplize": "Danh sÃ¡ch cÃ¹ng ngÃ nh",
        "KBS": "nhÃ³m trá»±c tuyáº¿n",
        "VCI": "nhÃ³m trá»±c tuyáº¿n",
        "CafeF": "Tham kháº£o",
        "SSC": "CÃ´ng bá»‘ thÃ´ng tin",
        "HOSE": "CÃ´ng bá»‘ thÃ´ng tin",
        "HNX": "CÃ´ng bá»‘ thÃ´ng tin",
        "Financial tÃ­ch há»£p": "Dá»¯ liá»‡u tÃ­ch há»£p",
        "CSV máº«u tÃ­ch há»£p": "Dá»¯ liá»‡u máº«u",
        "raw_data": "nháº­t kÃ½ ná»™i bá»™",
        "data_cache": "bá»™ nhá»› dá»¯ liá»‡u",
    }
    for raw, public in replacements.items():
        text = text.replace(raw, public)
    # áº¨n cÃ¡c cá»¥m ká»¹ thuáº­t/nguá»“n ná»™i bá»™ trong má»i thÃ´ng bÃ¡o cÃ´ng khai.
    text = re.sub(r"(?:Dá»¯ liá»‡u Æ°u tiÃªn|Dá»¯ liá»‡u tÃ­ch há»£p|Dá»¯ liá»‡u máº«u|FireAnt|Vietstock|Simplize)?\s*(?:VBA\s+)?endpoints?\s*", "Dá»¯ liá»‡u cáº­p nháº­t ", text, flags=re.I)
    text = re.sub(r"\bCrawler\s+[^0-9]*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})", r"Dá»¯ liá»‡u cáº­p nháº­t \1", text, flags=re.I)
    text = text.replace("TCReport_", "")
    text = re.sub(r"https?://\S+", "liÃªn káº¿t ná»™i bá»™", text, flags=re.I)
    text = re.sub(r"/?[^\s<>]*\.(?:csv|json|html|txt|xlsm|xlsx|md)", "file ná»™i bá»™", text, flags=re.I)
    text = re.sub(r"[A-Za-z]:\\[^\s<>]+", "Ä‘Æ°á»ng dáº«n ná»™i bá»™", text)
    text = re.sub(r"/(?:mnt|home|Users|raw_data|data_cache)[^\s<>]+", "Ä‘Æ°á»ng dáº«n ná»™i bá»™", text)
    text = re.sub(r"\s{2,}", " ", text).strip()
    return text


def _hide_source_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    hidden_names = {
        "source", "Source", "Nguá»“n", "Nguá»“n/URL", "Nguá»“n dá»¯ liá»‡u", "URL", "Truy váº¥n",
        "updated_at", "note", "raw_path", "File", "File nÄƒm", "File quÃ½", "source_label"
    }
    out = df.drop(columns=[c for c in df.columns if str(c) in hidden_names or "url" in str(c).lower() or str(c).lower() == "source"], errors="ignore").copy()
    for c in out.columns:
        if out[c].dtype == object:
            out[c] = out[c].map(_public_text)
    return out

ICB_CODE_NAME_MAP = {
    "2353": "XÃ¢y dá»±ng & Váº­t liá»‡u",
    "2350": "XÃ¢y dá»±ng & Váº­t liá»‡u",
    "2357": "Váº­t liá»‡u xÃ¢y dá»±ng",
    "1353": "HÃ³a cháº¥t",
    "1357": "HÃ³a cháº¥t cÆ¡ báº£n/PhÃ¢n bÃ³n",
    "1753": "Dá»‹ch vá»¥ váº­n táº£i",
    "2777": "Háº¡ táº§ng giao thÃ´ng & dá»‹ch vá»¥ sÃ¢n bay",
    "8355": "NgÃ¢n hÃ ng",
    "8532": "Báº£o hiá»ƒm",
    "9533": "Báº¥t Ä‘á»™ng sáº£n",
}


def _display_industry_value(value: object) -> str:
    text = str(value or "").strip()
    if not text or text.lower() in {"nan", "none", "n/a"}:
        return "N/A"
    code_text = text[:-2] if re.fullmatch(r"\d{3,6}\.0", text) else text
    if re.fullmatch(r"\d{3,6}", code_text):
        return ICB_CODE_NAME_MAP.get(code_text, f"ChÆ°a nháº­n diá»‡n tÃªn ngÃ nh (mÃ£ {code_text})")
    return text


def _logo_data_uri() -> str:
    try:
        if LOGO_PATH.exists():
            return "data:image/png;base64," + base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
    except Exception:
        pass
    return ""


def _render_brand_page_header(title: str, subtitle: str) -> None:
    logo_uri = _logo_data_uri()
    logo_html = f"<img src='{logo_uri}' alt='Trecapital' class='page-logo-img'>" if logo_uri else ""
    st.markdown(
        f"""
        <div class="page-brand-shell">
          <div class="page-logo-wrap">{logo_html}</div>
          <div class="hero-card page-hero-card">
            <h1>{html.escape(title)}</h1>
            <p>{html.escape(subtitle)}</p>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# V23.20 helpers: render important conclusions without raw markdown markers and sync MOS globally.
MOS_OPTIONS_GLOBAL = [0, 10, 20, 25, 30, 35, 40, 45, 50, 60, 70]


def _normalize_mos_value(value, default: int = 50) -> int:
    try:
        v = int(float(value))
    except Exception:
        v = default
    if v not in MOS_OPTIONS_GLOBAL:
        # Choose nearest supported value.
        v = min(MOS_OPTIONS_GLOBAL, key=lambda x: abs(x - v))
    return v


def _prepare_mos_widget(widget_key: str) -> int:
    canonical = _normalize_mos_value(st.session_state.get("target_mos_pct", st.session_state.get("module2_target_mos_pct", 50)))
    # Náº¿u MOS vá»«a Ä‘á»•i á»Ÿ module khÃ¡c, Ã©p widget hiá»‡n táº¡i theo canonical chung.
    # Náº¿u MOS vá»«a Ä‘á»•i á»Ÿ chÃ­nh widget nÃ y, callback _commit_mos_widget Ä‘Ã£ cáº­p nháº­t canonical trÆ°á»›c khi rerun.
    if widget_key not in st.session_state or _normalize_mos_value(st.session_state.get(widget_key)) != canonical:
        st.session_state[widget_key] = canonical
    st.session_state["target_mos_pct"] = canonical
    st.session_state["module1_target_mos_pct"] = canonical
    st.session_state["module2_target_mos_pct"] = canonical
    return canonical


def _commit_mos_widget(widget_key: str) -> None:
    canonical = _normalize_mos_value(st.session_state.get(widget_key, st.session_state.get("target_mos_pct", 50)))
    st.session_state["target_mos_pct"] = canonical
    st.session_state["module1_target_mos_pct"] = canonical
    st.session_state["module2_target_mos_pct"] = canonical
    st.session_state["last_mos_source_widget"] = widget_key
    st.session_state["mos_sync_status"] = f"ÄÃ£ Ä‘á»“ng bá»™ MOS yÃªu cáº§u {canonical}% cho Tá»•ng quan doanh nghiá»‡p vÃ  Äá»‹nh giÃ¡ chuyÃªn sÃ¢u."


def _markdownish_to_html(text: object) -> str:
    raw = "" if text is None else str(text)
    # Normalize common markdown fragments from engine summaries into HTML so ** does not appear on screen.
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    escaped = html.escape(raw)
    escaped = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", escaped)
    escaped = re.sub(r"__(.+?)__", r"<b>\1</b>", escaped)
    escaped = escaped.replace("\n", "<br>")
    # Keep bullet-like output readable.
    escaped = escaped.replace("â€¢ ", "&bull; ").replace("- ", "&ndash; ")
    return escaped


def _render_important_red(title: str, body: object) -> None:
    st.markdown(
        f"""
        <div class="important-red" style="color:#B91C1C !important;font-size:1.08rem !important;line-height:1.65 !important;font-weight:850 !important;background:linear-gradient(180deg,#FFF1F2 0%,#FEF2F2 100%) !important;border:2px solid rgba(220,38,38,.34) !important;border-left:10px solid #DC2626 !important;padding:16px 18px !important;border-radius:16px !important;margin:12px 0 16px 0 !important;box-shadow:0 8px 22px rgba(185,28,28,.10) !important;">
            <div class="important-red-title" style="color:#991B1B !important;font-size:1.22rem !important;font-weight:950 !important;margin-bottom:8px !important;">{html.escape(str(title))}</div>
            <div>{_markdownish_to_html(body)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )




def _render_warning_card(title: str, body: object) -> None:
    st.markdown(
        f"""
        <div class="big-warning-card" style="border:2px solid #F5B21B;border-left:10px solid #F5B21B;border-radius:16px;padding:14px 16px;background:linear-gradient(135deg,#FFF7E6 0%,#FEF3C7 100%);margin:12px 0 16px 0;box-shadow:0 8px 22px rgba(245,178,27,.13);">
            <div class="big-warning-title" style="font-size:1.08rem;font-weight:950;color:#8A5A00;margin-bottom:7px;">{html.escape(str(title))}</div>
            <div class="big-warning-text" style="font-size:.96rem;font-weight:850;color:#5F3B00;line-height:1.55;">{_markdownish_to_html(body)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )




# V23.20: CSS pháº£i Ä‘Æ°á»£c inject trong render_dashboard(), khÃ´ng Ä‘á»ƒ á»Ÿ top-level import.
# LÃ½ do: Streamlit cache module Python; khi widget MOS lÃ m rerun thÃ¬ import khÃ´ng cháº¡y láº¡i,
# dáº«n Ä‘áº¿n máº¥t style tab vÃ  style nháº­n xÃ©t. HÃ m nÃ y cháº¡y á»Ÿ má»—i láº§n render Ä‘á»ƒ style luÃ´n cÃ²n.
def _inject_runtime_ui_css() -> None:
    st.markdown(
        """
        <style>
        :root {
            --tre-teal:#0B7F75;
            --tre-teal-dark:#064E47;
            --tre-teal-soft:#EAF7F1;
            --tre-yellow:#F5B21B;
            --tre-red:#B91C1C;
        }
        .main .block-container {padding-top: 1rem !important; padding-bottom: 2rem !important; max-width: none !important; width: 100% !important;}
        .stApp {background: radial-gradient(circle at 10% 0%, rgba(11,127,117,.08), transparent 28%), linear-gradient(180deg, #F7FBF8 0%, #FFFFFF 60%) !important;}
        section[data-testid="stSidebar"] {background: linear-gradient(180deg, #EAF7F1 0%, #FFFFFF 72%) !important; border-right: 1px solid rgba(11,127,117,.14) !important;}
    /* V23.25: hide Streamlit built-in multipage nav; app provides branded page links manually. */
    section[data-testid="stSidebar"] [data-testid="stSidebarNav"] {display:none !important;}
    div[data-testid="stPageLink"] a {
        border: 1.7px solid rgba(11,127,117,.28) !important;
        border-radius: 17px !important;
        margin: 6px 0 !important;
        padding: 11px 14px !important;
        background: linear-gradient(135deg, rgba(255,255,255,.95), rgba(248,255,251,.88)) !important;
        color: #064E47 !important;
        font-weight: 900 !important;
        box-shadow: 0 7px 17px rgba(11,127,117,.08) !important;
        text-decoration: none !important;
    }
    div[data-testid="stPageLink"] a:hover {
        border-color:#F5B21B !important;
        background: linear-gradient(135deg, #F8FFFB 0%, #FFF7E6 100%) !important;
        color:#0B5F58 !important;
        box-shadow: 0 10px 22px rgba(11,127,117,.16) !important;
    }


        /* V23.25: sidebar module navigation buttons - Trecapital brand identity */
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] {padding-top: .35rem !important;}
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] ul {
            display: flex !important;
            flex-direction: column !important;
            gap: 10px !important;
            padding: 0 8px !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] li {margin: 0 !important;}
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a {
            border: 1.7px solid rgba(11,127,117,.26) !important;
            border-radius: 17px !important;
            margin: 3px 0 !important;
            padding: 12px 14px !important;
            background: linear-gradient(135deg, rgba(255,255,255,.92), rgba(248,255,251,.86)) !important;
            color: #064E47 !important;
            font-weight: 900 !important;
            box-shadow: 0 7px 17px rgba(11,127,117,.08) !important;
            transition: all .16s ease-in-out !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a:hover {
            border-color:#F5B21B !important;
            background: linear-gradient(135deg, #F8FFFB 0%, #FFF7E6 100%) !important;
            color:#0B5F58 !important;
            box-shadow: 0 10px 22px rgba(11,127,117,.16) !important;
            transform: translateX(2px) !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[aria-current="page"],
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[data-selected="true"] {
            background: linear-gradient(135deg, #0B7F75 0%, #128C7E 78%, #F5B21B 132%) !important;
            color: #FFFFFF !important;
            border-color: #F5B21B !important;
            box-shadow: 0 12px 26px rgba(11,127,117,.24) !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a span,
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a p {
            color: inherit !important;
            font-size: .94rem !important;
            font-weight: 900 !important;
            line-height: 1.25 !important;
        }


        /* Streamlit Tabs - dÃ¹ng selector rá»™ng cho nhiá»u version Streamlit/BaseWeb */
        div[data-testid="stTabs"] {margin-top: 10px !important; margin-bottom: 8px !important;}
        div[data-testid="stTabs"] div[data-baseweb="tab-list"],
        div[data-testid="stTabs"] div[role="tablist"],
        div[data-baseweb="tab-list"],
        div[role="tablist"] {
            display: flex !important;
            flex-wrap: wrap !important;
            align-items: center !important;
            gap: 14px !important;
            min-height: 76px !important;
            padding: 14px 16px !important;
            margin: 12px 0 20px 0 !important;
            background: linear-gradient(180deg, #EAF7F1 0%, #F8FFFB 100%) !important;
            border: 2.5px solid rgba(11,127,117,.42) !important;
            border-radius: 28px !important;
            box-shadow: 0 12px 30px rgba(11,127,117,.16) !important;
            overflow: visible !important;
            visibility: visible !important;
            opacity: 1 !important;
        }
        div[data-testid="stTabs"] button[data-baseweb="tab"],
        div[data-testid="stTabs"] button[role="tab"],
        button[data-baseweb="tab"],
        button[role="tab"],
        div[data-testid="stTabs"] div[role="tab"] {
            min-height: 58px !important;
            height: 58px !important;
            padding: 0 28px !important;
            border-radius: 999px !important;
            border: 2.5px solid rgba(11,127,117,.45) !important;
            background: #FFFFFF !important;
            color: #0B5F58 !important;
            font-size: 17px !important;
            font-weight: 950 !important;
            letter-spacing: .01em !important;
            box-shadow: 0 8px 18px rgba(11,127,117,.14) !important;
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            visibility: visible !important;
            opacity: 1 !important;
        }
        div[data-testid="stTabs"] button[data-baseweb="tab"] p,
        div[data-testid="stTabs"] button[role="tab"] p,
        button[data-baseweb="tab"] p,
        button[role="tab"] p,
        div[data-testid="stTabs"] button[data-baseweb="tab"] span,
        div[data-testid="stTabs"] button[role="tab"] span,
        button[data-baseweb="tab"] span,
        button[role="tab"] span {
            font-size: 17px !important;
            font-weight: 950 !important;
            line-height: 1.2 !important;
            color: inherit !important;
        }
        div[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"],
        div[data-testid="stTabs"] button[role="tab"][aria-selected="true"],
        button[data-baseweb="tab"][aria-selected="true"],
        button[role="tab"][aria-selected="true"],
        div[data-testid="stTabs"] div[role="tab"][aria-selected="true"] {
            background: linear-gradient(135deg, #0B7F75 0%, #128C7E 100%) !important;
            color: #FFFFFF !important;
            border-color: #F5B21B !important;
            box-shadow: 0 12px 26px rgba(11,127,117,.32) !important;
            transform: translateY(-1px) !important;
        }
        div[data-baseweb="tab-highlight"], div[data-baseweb="tab-border"] {display:none !important;}
        button[data-baseweb="tab"][aria-selected="true"] p,
        button[role="tab"][aria-selected="true"] p,
        button[data-baseweb="tab"][aria-selected="true"] span,
        button[role="tab"][aria-selected="true"] span {color:#FFFFFF !important;}


        /* V23.54: tab buttons cÃ³ mÃ u rÃµ hÆ¡n, trÃ¡nh ná»n tráº¯ng khÃ³ nhÃ¬n. */
        div[data-testid="stTabs"] button[role="tab"],
        div[data-testid="stTabs"] button[data-baseweb="tab"],
        div[data-testid="stTabs"] div[role="tab"] {
            background: linear-gradient(135deg, #FFF7E6 0%, #EAF7F1 100%) !important;
            border: 2.6px solid #0B7F75 !important;
            color: #064E47 !important;
            box-shadow: 0 9px 20px rgba(11,127,117,.16) !important;
        }
        div[data-testid="stTabs"] button[role="tab"]:hover,
        div[data-testid="stTabs"] button[data-baseweb="tab"]:hover,
        div[data-testid="stTabs"] div[role="tab"]:hover {
            background: linear-gradient(135deg, #FEF3C7 0%, #CCFBF1 100%) !important;
            border-color: #F5B21B !important;
            transform: translateY(-1px) !important;
        }
        div[data-testid="stTabs"] button[role="tab"][aria-selected="true"],
        div[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"],
        div[data-testid="stTabs"] div[role="tab"][aria-selected="true"] {
            background: linear-gradient(135deg, #0B7F75 0%, #128C7E 70%, #F5B21B 132%) !important;
            color: #FFFFFF !important;
            border-color: #F5B21B !important;
            box-shadow: 0 12px 28px rgba(11,127,117,.30) !important;
        }

        /* Nháº­n xÃ©t/Ä‘Ã¡nh giÃ¡/káº¿t luáº­n quan trá»ng */
        .important-red, .important-red * {
            color: #B91C1C !important;
            font-size: 1.10rem !important;
            line-height: 1.68 !important;
            font-weight: 850 !important;
        }
        .important-red-title {
            color: #991B1B !important;
            font-size: 1.25rem !important;
            font-weight: 950 !important;
            margin-bottom: 9px !important;
        }
        .important-red {
            background: linear-gradient(180deg, #FFF1F2 0%, #FEF2F2 100%) !important;
            border: 2px solid rgba(220,38,38,.36) !important;
            border-left: 11px solid #DC2626 !important;
            padding: 17px 19px !important;
            border-radius: 17px !important;
            margin: 12px 0 17px 0 !important;
            box-shadow: 0 9px 24px rgba(185,28,28,.11) !important;
        }
        .hero-card {padding: 23px 28px; border-radius: 22px; background: linear-gradient(135deg, #0B7F75 0%, #128C7E 56%, #F5B21B 135%); color: white; margin-bottom: 18px; box-shadow: 0 14px 34px rgba(11,127,117,.20); border:1px solid rgba(255,255,255,.28);}
        .hero-card h1 {font-size: 2.16rem; margin: 0 0 6px 0; color: white; letter-spacing: -.02em;}
        .hero-card p {font-size: 1.06rem; margin: 0; opacity: .95;}
        .logo-card {display:flex; align-items:center; justify-content:center; padding: 12px 7px; border-radius:23px; background: transparent; border:1px solid rgba(11,127,117,.10); box-shadow: 0 10px 24px rgba(11,127,117,.10); margin-bottom: 16px;}
        .page-brand-shell {display:grid; grid-template-columns: 158px minmax(0,1fr); gap:18px; align-items:center; margin:8px 0 20px 0;}
        .page-logo-wrap {height:156px; display:flex; align-items:center; justify-content:center; border-radius:28px; background:linear-gradient(180deg,#FFFFFF 0%,#F8FFFB 100%); border:1.8px solid rgba(11,127,117,.18); box-shadow:0 12px 30px rgba(11,127,117,.10);}
        .page-logo-img {max-height:138px; max-width:138px; object-fit:contain;}
        .page-hero-card {margin-bottom:0 !important; min-height:118px; display:flex; flex-direction:column; justify-content:center;}
        div[data-testid="stDataFrame"] [role="columnheader"], div[data-testid="stDataEditor"] [role="columnheader"] {font-weight:950 !important; color:#064E47 !important;}
        .workflow-card {border: 1px solid rgba(11,127,117,.28); border-radius: 18px; padding: 13px 14px; background: linear-gradient(180deg, rgba(234,247,241,.96), rgba(255,255,255,.88)); margin-bottom: 14px;}
        .source-card {border: 1px solid rgba(11,127,117,.28); border-radius: 16px; padding: 11px 13px; background: rgba(234,247,241,.72); margin-top: 8px; margin-bottom: 12px;}
        .note-card {border: 1px solid rgba(11,127,117,.22); border-radius: 16px; padding: 11px 13px; background: rgba(255,251,235,.75); margin-bottom: 10px;}
        .reason-yellow-card {border:2px solid rgba(245,178,27,.62); border-left:10px solid #F5B21B; border-radius:18px; padding:15px 18px; background:linear-gradient(135deg,#FFF7E6 0%,#FFF3C4 100%); color:#5F3B00; margin:12px 0 15px 0; box-shadow:0 9px 24px rgba(245,178,27,.16);}
        .reason-yellow-card .reason-title {font-size:1.03rem; font-weight:1000; color:#8A5A00; margin-bottom:8px;}
        .reason-yellow-card ul {margin:6px 0 0 20px; padding:0;}
        .reason-yellow-card li {margin:6px 0; font-weight:800; line-height:1.45;}
        .valuation-tab-compact div[data-testid="stVerticalBlock"] {gap:.35rem !important;}
        .ok-card {border: 1px solid rgba(11,127,117,.32); border-radius: 16px; padding: 13px 14px; background: rgba(234,247,241,.76); margin-bottom: 12px;}
        .warn-card {border: 1px solid rgba(245,178,27,.48); border-radius: 16px; padding: 13px 14px; background: rgba(255,247,230,.88); margin-bottom: 12px;}
        .big-warning-card {border: 2px solid #0B7F75; border-left: 8px solid #F5B21B; border-radius: 16px; padding: 13px 14px; background: linear-gradient(135deg, #FFF4C7 0%, #FFE68A 100%); margin: 10px 0 12px 0; box-shadow: 0 8px 20px rgba(245,178,27,.18);}
        .big-warning-title {font-size: .92rem; font-weight: 950; color: #0B7F75; margin-bottom: 4px; letter-spacing:-.01em;}
        .big-warning-text {font-size: .86rem; font-weight: 900; color: #5F3B00; line-height: 1.38;}
        .ticker-title-card {
            border: 2.8px solid rgba(11,127,117,.30);
            border-left: 14px solid #F5B21B;
            border-radius: 26px;
            padding: 20px 26px;
            background: linear-gradient(135deg, rgba(234,247,241,.95) 0%, rgba(255,255,255,.96) 100%);
            margin: 12px 0 16px 0;
            box-shadow: 0 12px 31px rgba(11,127,117,.12);
        }
        .ticker-title-main {font-size: 2.10rem; font-weight: 980; color:#064E47; letter-spacing:-.02em;}
        .ticker-title-code {font-size: 2.40rem; font-weight: 1000; color:#0B7F75;}
        .ticker-title-name {font-size: 1.68rem; font-weight: 930; color:#8A5A00;}
        .ticker-title-meta {font-size: .95rem; color:#0B5F58; margin-top:7px;}
        .current-price-inline-card {display:inline-flex;align-items:center;gap:21px;margin-top:17px;margin-bottom:11px;padding:17px 23px;border-radius:23px;border:2.1px solid #F5B21B;border-left:10px solid #F5B21B;background:#FFFFFF;box-shadow:0 10px 24px rgba(245,178,27,.16);min-width:389px;}
        .current-price-inline-card .price-label {font-size:.93rem;line-height:1.08;color:#50646B;font-weight:950;text-transform:uppercase;letter-spacing:.03em;}
        .current-price-inline-card .price-value {font-size:1.56rem;line-height:1.12;color:#064E47;font-weight:1000;}
        .current-price-inline-card .price-note {font-size:1.02rem;color:#64748B;font-weight:780;}
        div[data-testid="stMetric"] {background: rgba(255,255,255,.92); border: 1.7px solid rgba(11,127,117,.22); border-radius: 28px; padding: 22px 24px; min-height: 127px; box-shadow: 0 7px 23px rgba(11,127,117,.09);}
        div.stButton > button {border-radius: 999px !important; border: 1px solid rgba(11,127,117,.35) !important; background: linear-gradient(135deg, #0B7F75, #139486) !important; color: white !important; font-weight: 800 !important;}
        div.stButton > button:hover {border-color:#F5B21B !important; color:white !important; box-shadow: 0 0 0 3px rgba(245,178,27,.16) !important;}
        .small-muted {font-size: .88rem; color: #64748b;}
        .glossary-fit-table {width:100%; border-collapse:collapse; table-layout:auto; font-size:.94rem; line-height:1.42;}
        .glossary-fit-table th {background:#EAF7F1; color:#064E47; border:1px solid rgba(11,127,117,.22); padding:8px 10px; text-align:left; white-space:nowrap; font-weight:950;}
        .glossary-fit-table td {border:1px solid rgba(11,127,117,.14); padding:8px 10px; vertical-align:top; background:#FFFFFF; color:#12343B;}
        .glossary-fit-table .stt {width:52px; text-align:center; white-space:nowrap;}
        .glossary-fit-table .term {width:220px; min-width:150px; white-space:nowrap; font-weight:850; color:#064E47;}
        .glossary-fit-table .desc {width:auto; white-space:normal;}
        /* V23.33 extra CSS */

        section[data-testid="stSidebar"] img {display:none !important;}
        div[data-testid="stDataFrame"] [role="columnheader"], div[data-testid="stDataEditor"] [role="columnheader"],
        div[data-testid="stDataFrame"] [data-testid="stHeader"], div[data-testid="stDataEditor"] [data-testid="stHeader"],
        .stDataFrame th, .stDataEditor th {font-weight:950 !important; color:#064E47 !important;}
        .page-brand-shell {display:grid; grid-template-columns: 176px minmax(0,1fr); gap:22px; align-items:center; margin:6px 0 22px 0;}
        .page-logo-wrap {height:166px; display:flex; align-items:center; justify-content:center; border-radius:30px; background:linear-gradient(180deg,#FFFFFF 0%,#F8FFFB 100%); border:2px solid rgba(11,127,117,.20); box-shadow:0 14px 34px rgba(11,127,117,.12);}
        .page-logo-img {max-height:146px; max-width:146px; object-fit:contain; display:block !important;}
        .page-hero-card {margin-bottom:0 !important; min-height:118px; display:flex; flex-direction:column; justify-content:center;}

        /* V23.54 final override: má»i tab Ä‘á»u cÃ³ mÃ u ná»n dá»… nháº­n diá»‡n. */
        div[data-testid="stTabs"] button[role="tab"],
        div[data-testid="stTabs"] button[data-baseweb="tab"],
        div[data-testid="stTabs"] div[role="tab"] {
            background: linear-gradient(135deg, #FFF7E6 0%, #EAF7F1 100%) !important;
            border: 2.6px solid #0B7F75 !important;
            color: #064E47 !important;
        }
        div[data-testid="stTabs"] button[role="tab"][aria-selected="true"],
        div[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"],
        div[data-testid="stTabs"] div[role="tab"][aria-selected="true"] {
            background: linear-gradient(135deg, #0B7F75 0%, #128C7E 70%, #F5B21B 132%) !important;
            color: #FFFFFF !important;
            border-color: #F5B21B !important;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )
    try:
        from ui_oaktree_theme import inject_oaktree_theme

        inject_oaktree_theme()
    except Exception:
        pass
    # V23.56: final override after theme injection so sidebar nav remains colored like tabs.
    st.markdown("""
    <style>
    section[data-testid="stSidebar"] div[data-testid="stPageLink"] a,
    section[data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"],
    section[data-testid="stSidebar"] nav a {
        border-radius:999px !important; border:2.4px solid #0B7F75 !important; border-bottom:4px solid rgba(6,78,71,.38) !important;
        background:linear-gradient(135deg,#FFF7E6 0%,#EAF7F1 100%) !important; color:#064E47 !important; -webkit-text-fill-color:#064E47 !important;
        font-weight:950 !important; box-shadow:0 8px 18px rgba(11,127,117,.14) !important; margin:7px 2px !important; padding:10px 16px !important;
    }
    section[data-testid="stSidebar"] div[data-testid="stPageLink"] a:hover,
    section[data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"]:hover,
    section[data-testid="stSidebar"] nav a:hover {background:linear-gradient(135deg,#FFE8A3 0%,#D8F3E4 100%) !important; border-color:#F5B21B !important;}
    section[data-testid="stSidebar"] div[data-testid="stPageLink"] a[aria-current="page"],
    section[data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"],
    section[data-testid="stSidebar"] nav a[aria-current="page"] {background:linear-gradient(135deg,#064E47 0%,#0B7F75 70%,#F5B21B 132%) !important; color:#FFFFFF !important; -webkit-text-fill-color:#FFFFFF !important; border-color:#F5B21B !important;}
    div[data-testid="stDataFrame"] [role="columnheader"], div[data-testid="stDataEditor"] [role="columnheader"] {position:sticky !important; top:0 !important; z-index:50 !important; background:#EAF7F1 !important; color:#064E47 !important; font-weight:950 !important;}

    /* V23.59: logo Trecapital card styled like colored tabs. */
    .page-logo-wrap, .logo-card {
        background:linear-gradient(135deg,#FFF7E6 0%,#EAF7F1 100%) !important;
        border:2.6px solid #0B7F75 !important;
        border-bottom:4px solid rgba(6,78,71,.38) !important;
        box-shadow:0 12px 30px rgba(11,127,117,.16) !important;
        transition:all .16s ease-in-out !important;
    }
    .page-logo-wrap:hover, .logo-card:hover {
        background:linear-gradient(135deg,#FFE8A3 0%,#D8F3E4 100%) !important;
        border-color:#F5B21B !important;
        box-shadow:0 16px 36px rgba(245,178,27,.20), 0 10px 22px rgba(11,127,117,.16) !important;
        transform:translateY(-1px) !important;
    }

    </style>
    """, unsafe_allow_html=True)

@st.cache_data(show_spinner=False)
def _available_financial_tickers_cached(xlsm_path: str) -> list[str]:
    """Return only tickers with full financial blocks in the integrated workbook.

    Previous build scanned column B in all financial sheets and accidentally treated labels such
    as ROE/ROIC/BasicEPS as tickers. This function reads the configured stock list in the workbook.
    """
    try:
        from openpyxl import load_workbook
        import re
        wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_vba=False)
        tickers: list[str] = []
        if "BÃO CÃO TÃ€I CHÃNH" in wb.sheetnames:
            ws = wb["BÃO CÃO TÃ€I CHÃNH"]
            for row in range(15, ws.max_row + 1):
                code = ws.cell(row=row, column=3).value
                if isinstance(code, str) and re.fullmatch(r"[A-Z0-9]{2,8}", code.strip().upper()):
                    tickers.append(code.strip().upper())
        return sorted(dict.fromkeys(tickers))
    except Exception:
        return []


@st.cache_data(show_spinner=False)
def _listed_ticker_info_cached(xlsm_path: str, ticker: str) -> dict:
    """Find company name/exchange in DANH SÃCH MÃƒ even if BCTC is not bundled."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_vba=False)
        ws = wb["DANH SÃCH MÃƒ"] if "DANH SÃCH MÃƒ" in wb.sheetnames else None
        if ws is None:
            return {}
        ticker = ticker.upper().strip()
        for row in range(5, ws.max_row + 1):
            code = ws.cell(row=row, column=3).value
            if isinstance(code, str) and code.strip().upper() == ticker:
                return {
                    "ticker": ticker,
                    "company_name": str(ws.cell(row=row, column=4).value or ticker),
                    "exchange": str(ws.cell(row=row, column=5).value or ""),
                    "sub_industry": str(ws.cell(row=row, column=7).value or ""),
                }
    except Exception:
        pass
    return {}


st.set_page_config(
    page_title=APP_NAME,
    page_icon="ðŸ§ ",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    .main .block-container {padding-top: 1rem; padding-bottom: 2rem; max-width: none; width: 100%;}
    .stApp {background: radial-gradient(circle at 10% 0%, rgba(11,127,117,.08), transparent 28%), linear-gradient(180deg, #F7FBF8 0%, #FFFFFF 60%);} 
    section[data-testid="stSidebar"] {background: linear-gradient(180deg, #EAF7F1 0%, #FFFFFF 72%); border-right: 1px solid rgba(11,127,117,.14);}
    /* V23.25: hide Streamlit built-in multipage nav; app provides branded page links manually. */
    section[data-testid="stSidebar"] [data-testid="stSidebarNav"] {display:none !important;}
    div[data-testid="stPageLink"] a {
        border: 1.7px solid rgba(11,127,117,.28) !important;
        border-radius: 17px !important;
        margin: 6px 0 !important;
        padding: 11px 14px !important;
        background: linear-gradient(135deg, rgba(255,255,255,.95), rgba(248,255,251,.88)) !important;
        color: #064E47 !important;
        font-weight: 900 !important;
        box-shadow: 0 7px 17px rgba(11,127,117,.08) !important;
        text-decoration: none !important;
    }
    div[data-testid="stPageLink"] a:hover {
        border-color:#F5B21B !important;
        background: linear-gradient(135deg, #F8FFFB 0%, #FFF7E6 100%) !important;
        color:#0B5F58 !important;
        box-shadow: 0 10px 22px rgba(11,127,117,.16) !important;
    }
 

        /* V23.25: sidebar module navigation buttons - Trecapital brand identity */
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] {padding-top: .35rem ;}
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] ul {
            display: flex ;
            flex-direction: column ;
            gap: 10px ;
            padding: 0 8px ;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] li {margin: 0 ;}
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a {
            border: 1.7px solid rgba(11,127,117,.26) ;
            border-radius: 17px ;
            margin: 3px 0 ;
            padding: 12px 14px ;
            background: linear-gradient(135deg, rgba(255,255,255,.92), rgba(248,255,251,.86)) ;
            color: #064E47 ;
            font-weight: 900 ;
            box-shadow: 0 7px 17px rgba(11,127,117,.08) ;
            transition: all .16s ease-in-out ;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a:hover {
            border-color:#F5B21B ;
            background: linear-gradient(135deg, #F8FFFB 0%, #FFF7E6 100%) ;
            color:#0B5F58 ;
            box-shadow: 0 10px 22px rgba(11,127,117,.16) ;
            transform: translateX(2px) ;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[aria-current="page"],
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[data-selected="true"] {
            background: linear-gradient(135deg, #0B7F75 0%, #128C7E 78%, #F5B21B 132%) ;
            color: #FFFFFF ;
            border-color: #F5B21B ;
            box-shadow: 0 12px 26px rgba(11,127,117,.24) ;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a span,
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a p {
            color: inherit ;
            font-size: .94rem ;
            font-weight: 900 ;
            line-height: 1.25 ;
        }

    .hero-card {padding: 23px 28px; border-radius: 22px; background: linear-gradient(135deg, #0B7F75 0%, #128C7E 56%, #F5B21B 135%); color: white; margin-bottom: 18px; box-shadow: 0 14px 34px rgba(11,127,117,.20); border:1px solid rgba(255,255,255,.28);} 
    .hero-card h1 {font-size: 2.16rem; margin: 0 0 6px 0; color: white; letter-spacing: -.02em;}
    .hero-card p {font-size: 1.06rem; margin: 0; opacity: .95;}
    .logo-card {display:flex; align-items:center; justify-content:center; padding: 12px 7px; border-radius:23px; background: transparent; border:1px solid rgba(11,127,117,.10); box-shadow: 0 10px 24px rgba(11,127,117,.10); margin-bottom: 16px;}
    .note-card {border: 1px solid rgba(11,127,117,.22); border-radius: 16px; padding: 11px 13px; background: rgba(255,251,235,.75); margin-bottom: 10px;}
    .ok-card {border: 1px solid rgba(11,127,117,.32); border-radius: 16px; padding: 13px 14px; background: rgba(234,247,241,.76); margin-bottom: 12px;}
    .warn-card {border: 1px solid rgba(245,178,27,.48); border-radius: 16px; padding: 13px 14px; background: rgba(255,247,230,.88); margin-bottom: 12px;}
    .big-warning-card {border: 2px solid #F5B21B; border-left: 8px solid #0B7F75; border-radius: 16px; padding: 11px 13px; background: linear-gradient(135deg, #FFF7E6 0%, #FEF3C7 100%); margin: 9px 0 11px 0; box-shadow: 0 7px 17px rgba(245,178,27,.16);}
    .big-warning-title {font-size: .88rem; font-weight: 900; color: #0B7F75; margin-bottom: 4px; letter-spacing:-.01em;}
    .big-warning-text {font-size: .84rem; font-weight: 850; color: #5f3b00; line-height: 1.36;}
    div[data-testid="stAlert"] p {font-size: 1.02rem;}
    div[data-testid="stMetric"] {background: rgba(255,255,255,.90); border: 1.5px solid rgba(11,127,117,.20); border-radius: 21px; padding: 15px 18px; min-height: 77px; box-shadow: 0 5px 19px rgba(11,127,117,.07);} 
    div.stButton > button {border-radius: 999px; border: 1px solid rgba(11,127,117,.35); background: linear-gradient(135deg, #0B7F75, #139486); color: white; font-weight: 700;}
    div.stButton > button:hover {border-color:#F5B21B; color:white; box-shadow: 0 0 0 3px rgba(245,178,27,.16);}
    .stTabs [data-baseweb="tab-list"], div[data-testid="stTabs"] [role="tablist"] {gap: 14px !important; background: rgba(234,247,241,.96) !important; padding: 14px 16px !important; border-radius: 26px !important; border:2px solid rgba(11,127,117,.30) !important; box-shadow:0 10px 26px rgba(11,127,117,.12) !important;}
    .stTabs [data-baseweb="tab"], div[data-testid="stTabs"] button[role="tab"] {min-height: 58px !important; height: 58px !important; border-radius: 999px !important; padding: 0 28px !important; border: 2.5px solid rgba(11,127,117,.40) !important; background: #FFFFFF !important; color:#0B5F58 !important; font-size: 1.08rem !important; font-weight: 900 !important; box-shadow:0 6px 16px rgba(11,127,117,.10) !important;}
    .stTabs [aria-selected="true"], div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {background: linear-gradient(135deg, #0B7F75, #128C7E) !important; color: #FFFFFF !important; border-color:#F5B21B !important; box-shadow:0 10px 24px rgba(11,127,117,.28) !important; transform: translateY(-1px);}
    .stTabs [data-baseweb="tab"] p, div[data-testid="stTabs"] button[role="tab"] p {font-size: 1.08rem !important; font-weight: 900 !important;}
    .important-red {color:#B91C1C !important; font-size:1.26rem !important; line-height:1.66 !important; font-weight:900 !important; background:rgba(254,242,242,.92) !important; border:2px solid rgba(239,68,68,.28) !important; border-left:10px solid #DC2626 !important; padding:16px 18px !important; border-radius:16px !important; margin:12px 0 16px 0 !important; box-shadow:0 8px 22px rgba(185,28,28,.08) !important;}
    .small-muted {font-size: .88rem; color: #64748b;}
    
    /* V23.20: robust global Streamlit tab styling for Tá»•ng quan doanh nghiá»‡p + Äá»‹nh giÃ¡ chuyÃªn sÃ¢u */
    div[data-testid="stTabs"] {margin-top: 12px !important;}
    div[data-testid="stTabs"] div[role="tablist"],
    div[role="tablist"] {
        display: flex !important;
        flex-wrap: wrap !important;
        gap: 14px !important;
        min-height: 72px !important;
        padding: 14px 16px !important;
        margin: 10px 0 18px 0 !important;
        background: linear-gradient(180deg, #EAF7F1 0%, #F8FFFB 100%) !important;
        border: 2px solid rgba(11,127,117,.36) !important;
        border-radius: 26px !important;
        box-shadow: 0 10px 26px rgba(11,127,117,.14) !important;
        visibility: visible !important;
        opacity: 1 !important;
        overflow: visible !important;
    }
    div[data-testid="stTabs"] button[role="tab"],
    div[data-testid="stTabs"] div[role="tab"],
    div[role="tablist"] button,
    button[role="tab"],
    div[role="tab"] {
        min-height: 56px !important;
        height: 56px !important;
        padding: 0 26px !important;
        border-radius: 999px !important;
        border: 2px solid rgba(11,127,117,.42) !important;
        background: #FFFFFF !important;
        color: #0B5F58 !important;
        font-size: 16px !important;
        font-weight: 900 !important;
        box-shadow: 0 6px 16px rgba(11,127,117,.12) !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        visibility: visible !important;
        opacity: 1 !important;
    }
    button[role="tab"] p, div[role="tab"] p,
    button[role="tab"] span, div[role="tab"] span,
    div[data-testid="stTabs"] p, div[data-testid="stTabs"] span {
        font-size: 16px !important;
        font-weight: 900 !important;
        color: inherit !important;
        line-height: 1.2 !important;
    }
    button[role="tab"][aria-selected="true"],
    div[role="tab"][aria-selected="true"] {
        background: linear-gradient(135deg, #0B7F75 0%, #128C7E 100%) !important;
        color: #FFFFFF !important;
        border-color: #F5B21B !important;
        box-shadow: 0 10px 24px rgba(11,127,117,.28) !important;
        transform: translateY(-1px) !important;
    }
    button[role="tab"][aria-selected="true"] p, div[role="tab"][aria-selected="true"] p,
    button[role="tab"][aria-selected="true"] span, div[role="tab"][aria-selected="true"] span {
        color: #FFFFFF !important;
    }

    /* V23.54 runtime override: tab thÆ°á»ng cÃ³ mÃ u vÃ ng/xanh, tab active ná»•i báº­t. */
    div[data-testid="stTabs"] button[role="tab"],
    div[data-testid="stTabs"] div[role="tab"],
    div[role="tablist"] button,
    button[role="tab"],
    div[role="tab"] {
        background: linear-gradient(135deg, #FFF7E6 0%, #EAF7F1 100%) !important;
        border: 2.6px solid #0B7F75 !important;
        color: #064E47 !important;
        box-shadow: 0 9px 20px rgba(11,127,117,.16) !important;
    }
    div[data-testid="stTabs"] button[role="tab"]:hover,
    div[data-testid="stTabs"] div[role="tab"]:hover,
    div[role="tablist"] button:hover,
    button[role="tab"]:hover,
    div[role="tab"]:hover {
        background: linear-gradient(135deg, #FEF3C7 0%, #CCFBF1 100%) !important;
        border-color: #F5B21B !important;
    }
    button[role="tab"][aria-selected="true"],
    div[role="tab"][aria-selected="true"] {
        background: linear-gradient(135deg, #0B7F75 0%, #128C7E 70%, #F5B21B 132%) !important;
        color: #FFFFFF !important;
        border-color: #F5B21B !important;
        box-shadow: 0 12px 28px rgba(11,127,117,.30) !important;
    }
    button[role="tab"][aria-selected="true"] p, div[role="tab"][aria-selected="true"] p,
    button[role="tab"][aria-selected="true"] span, div[role="tab"][aria-selected="true"] span {
        color: #FFFFFF !important;
    }
    .important-red, .important-red * {
        color:#B91C1C !important;
        font-size:1.08rem !important;
        line-height:1.65 !important;
        font-weight:850 !important;
    }
    .important-red-title {
        color:#991B1B !important;
        font-size:1.22rem !important;
        font-weight:950 !important;
        margin-bottom:8px !important;
    }
    .important-red {
        background:linear-gradient(180deg, #FFF1F2 0%, #FEF2F2 100%) !important;
        border:2px solid rgba(220,38,38,.34) !important;
        border-left:10px solid #DC2626 !important;
        padding:16px 18px !important;
        border-radius:16px !important;
        margin:12px 0 16px 0 !important;
        box-shadow:0 8px 22px rgba(185,28,28,.10) !important;
    }
    
    </style>
    """,
    unsafe_allow_html=True,
)



def _safe_ticker(ticker: str) -> str:
    return "".join(ch for ch in ticker.upper().strip() if ch.isalnum())[:12]


@st.cache_data(show_spinner=False)
def _export_bundled_financial_cached(xlsm_path: str, ticker: str, cache_dir: str) -> tuple[str, str, str, str]:
    out_dir = Path(cache_dir) / "financial_xlsm" / ticker.upper()
    result = ExcelFinancialProvider(xlsm_path).export_csv(ticker, out_dir)
    return (
        str(out_dir / "company_overview_sample.csv"),
        str(out_dir / "financial_timeseries_year.csv"),
        str(out_dir / "financial_timeseries_quarter.csv"),
        f"Dá»¯ liá»‡u tÃ­ch há»£p | NÄƒm: {len(result.annual)} dÃ²ng | QuÃ½: {len(result.quarterly)} dÃ²ng",
    )


def _provider_result_score(result: ProviderResult) -> int:
    score = 0
    if isinstance(result.overview, pd.DataFrame) and not result.overview.empty:
        score += 5
    if isinstance(result.annual, pd.DataFrame) and not result.annual.empty:
        score += min(len(result.annual), 12) * 10
    if isinstance(result.quarterly, pd.DataFrame) and not result.quarterly.empty:
        score += min(len(result.quarterly), 24) * 5
    return score


def _minimal_overview_df(ticker: str, source_name: str) -> pd.DataFrame:
    return normalize_columns(pd.DataFrame([{
        "ticker": ticker.upper(),
        "company_name": f"{ticker.upper()} - Ä‘ang cáº­p nháº­t há»“ sÆ¡ doanh nghiá»‡p",
        "exchange": "",
        "industry": "",
        "sub_industry": "",
        "updated_at": f"Crawler {source_name} {pd.Timestamp.now():%Y-%m-%d %H:%M:%S}",
    }]), MODULE1_OVERVIEW_COLUMNS)


def _write_provider_df(df: pd.DataFrame, path: Path, fallback_columns: list[str]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not isinstance(df, pd.DataFrame) or df.empty:
        pd.DataFrame(columns=fallback_columns).to_csv(path, index=False, encoding="utf-8-sig")
        return 0
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return len(df)


def _export_provider_result_to_module2_cache(result: ProviderResult, ticker: str, source_name: str, cache_dir: str) -> tuple[str, str, str, str]:
    """Ghi káº¿t quáº£ crawler chuáº©n Tá»•ng quan doanh nghiá»‡p ra cache Ä‘á»ƒ Äá»‹nh giÃ¡ chuyÃªn sÃ¢u dÃ¹ng trá»±c tiáº¿p.

    ÄÃ¢y lÃ  Ä‘iá»ƒm tÃ­ch há»£p quan trá»ng: Äá»‹nh giÃ¡ chuyÃªn sÃ¢u khÃ´ng tá»± bá»‹a dá»¯ liá»‡u vÃ  khÃ´ng chá»‰ tÃ¬m báº±ng chá»©ng Ä‘á»‹nh tÃ­nh;
    nÃ³ dÃ¹ng Ä‘Ãºng pipeline crawler/normalizer cá»§a Tá»•ng quan doanh nghiá»‡p rá»“i Ä‘Æ°a vÃ o engine Ä‘á»‹nh giÃ¡.
    """
    ticker = _safe_ticker(ticker)
    source_key = source_name.lower().replace(" ", "_").replace("+", "plus").replace("/", "_")
    out_dir = Path(cache_dir) / "module1_crawler" / source_key / ticker
    overview_path = out_dir / "company_overview_sample.csv"
    year_path = out_dir / "financial_timeseries_year.csv"
    quarter_path = out_dir / "financial_timeseries_quarter.csv"

    overview_df = result.overview if isinstance(result.overview, pd.DataFrame) else pd.DataFrame()
    annual_df = result.annual if isinstance(result.annual, pd.DataFrame) else pd.DataFrame()
    quarterly_df = result.quarterly if isinstance(result.quarterly, pd.DataFrame) else pd.DataFrame()
    if overview_df.empty:
        overview_df = _minimal_overview_df(ticker, source_name)

    counts = {
        "overview": _write_provider_df(normalize_columns(overview_df, MODULE1_OVERVIEW_COLUMNS), overview_path, MODULE1_OVERVIEW_COLUMNS),
        "annual": _write_provider_df(normalize_columns(annual_df, MODULE1_TIMESERIES_COLUMNS), year_path, MODULE1_TIMESERIES_COLUMNS),
        "quarterly": _write_provider_df(normalize_columns(quarterly_df, MODULE1_TIMESERIES_COLUMNS), quarter_path, MODULE1_TIMESERIES_COLUMNS),
    }
    raw_note = ""
    label = f"Dá»¯ liá»‡u Ä‘Ã£ chuáº©n hÃ³a tá»« Tá»•ng quan doanh nghiá»‡p | Tá»•ng quan: {counts['overview']} dÃ²ng | NÄƒm: {counts['annual']} dÃ²ng | QuÃ½: {counts['quarterly']} dÃ²ng"
    return str(overview_path), str(year_path), str(quarter_path), label


def _fetch_module1_crawler_result(ticker: str, source: str, raw_dir: str) -> ProviderResult:
    ticker = _safe_ticker(ticker)
    if source == "FireAnt":
        return PublicFireAntCrawler(raw_dir).fetch(ticker)
    if source == "Vietstock":
        return PublicVietstockCrawler(raw_dir).fetch(ticker)
    if source in {"FireAnt + Vietstock", "Tá»± Ä‘á»™ng tá»« dá»¯ liá»‡u tá»•ng quan"}:
        candidates: list[tuple[str, ProviderResult]] = []
        errors: list[str] = []
        for name, crawler in [("FireAnt", PublicFireAntCrawler(raw_dir)), ("Vietstock", PublicVietstockCrawler(raw_dir))]:
            try:
                candidates.append((name, crawler.fetch(ticker)))
            except Exception as exc:
                errors.append(f"{name}: {exc}")
        if not candidates:
            raise RuntimeError("KhÃ´ng gá»i Ä‘Æ°á»£c crawler Tá»•ng quan doanh nghiá»‡p: " + " | ".join(errors))
        # chá»n nguá»“n nÃ o tráº£ Ä‘Æ°á»£c báº£ng nhiá»u ká»³ tá»‘t nháº¥t
        best_name, best_result = max(candidates, key=lambda item: _provider_result_score(item[1]))
        notes = [f"{n}: overview={len(r.overview)}, nÄƒm={len(r.annual)}, quÃ½={len(r.quarterly)}, note={r.note}" for n, r in candidates]
        if errors:
            notes.extend(errors)
        best_result.note = f"Chá»n {best_name}. " + " | ".join(notes)
        return best_result
    raise ValueError(f"Cháº¿ Ä‘á»™ dá»¯ liá»‡u Tá»•ng quan doanh nghiá»‡p khÃ´ng há»£p lá»‡: {_public_text(source)}")


@st.cache_data(show_spinner=False)
def _export_module1_crawler_cached(ticker: str, source: str, cache_dir: str, raw_dir: str) -> tuple[str, str, str, str]:
    result = _fetch_module1_crawler_result(ticker, source, raw_dir)
    return _export_provider_result_to_module2_cache(result, ticker, source, cache_dir)


@st.cache_data(show_spinner=False)
def _load_overview_cached(path: str, ticker: str):
    return load_overview_from_csv(path, ticker)


@st.cache_data(show_spinner=False)
def _load_timeseries_cached(path: str, ticker: str, period_type: str, limit: int) -> pd.DataFrame:
    return ensure_derived_metrics(load_timeseries_from_csv(path, ticker, period_type, limit))



def _set_active_module1_bundle(overview_csv: Path, year_csv: Path, quarter_csv: Path, label: str, ticker: str) -> None:
    """DÃ¹ng chung má»™t bá»™ dá»¯ liá»‡u hoáº¡t Ä‘á»™ng cho Tá»•ng quan doanh nghiá»‡p vÃ  Äá»‹nh giÃ¡ chuyÃªn sÃ¢u trong multipage app."""
    ticker = _safe_ticker(ticker)
    st.session_state["active_ticker"] = ticker
    st.session_state["module1_ticker"] = ticker
    st.session_state["module2_ticker"] = ticker
    st.session_state["shared_ticker"] = ticker
    st.session_state["active_overview_csv"] = str(overview_csv)
    st.session_state["active_year_csv"] = str(year_csv)
    st.session_state["active_quarter_csv"] = str(quarter_csv)
    st.session_state["active_source_label"] = label
    st.session_state["module_sync_status"] = f"ÄÃ£ Ä‘á»“ng bá»™ {ticker} vÃ o Tá»•ng quan doanh nghiá»‡p + Äá»‹nh giÃ¡ chuyÃªn sÃ¢u lÃºc {pd.Timestamp.now():%Y-%m-%d %H:%M:%S}"


def _active_module1_bundle_for_ticker(ticker: str) -> tuple[Path, Path, Path, str] | None:
    """Æ¯u tiÃªn dá»¯ liá»‡u Ä‘Ã£ Ä‘Æ°á»£c Tá»•ng quan doanh nghiá»‡p kÃ­ch hoáº¡t trong cÃ¹ng app."""
    ticker = _safe_ticker(ticker)
    active_ticker = _safe_ticker(str(st.session_state.get("active_ticker", "")))
    paths = [st.session_state.get("active_overview_csv"), st.session_state.get("active_year_csv"), st.session_state.get("active_quarter_csv")]
    if active_ticker == ticker and all(p and Path(str(p)).exists() for p in paths):
        return Path(str(paths[0])), Path(str(paths[1])), Path(str(paths[2])), str(st.session_state.get("active_source_label", "Dá»¯ liá»‡u Ä‘Ã£ Ä‘á»“ng bá»™ tá»« Tá»•ng quan doanh nghiá»‡p"))
    return None


def _existing_cache_bundle_for_ticker(ticker: str) -> tuple[Path, Path, Path, str] | None:
    """TÃ¬m cache chuáº©n mÃ  Tá»•ng quan doanh nghiá»‡p/Äá»‹nh giÃ¡ chuyÃªn sÃ¢u Ä‘Ã£ táº¡o trÆ°á»›c Ä‘Ã³ cho cÃ¹ng mÃ£."""
    ticker = _safe_ticker(ticker)
    roots = [
        DATA_CACHE_DIR / "fireant_vietstock" / ticker,
        DATA_CACHE_DIR / "fireant" / ticker,
        DATA_CACHE_DIR / "vietstock" / ticker,
        DATA_CACHE_DIR / "module1_crawler" / "fireant_plus_vietstock" / ticker,
        DATA_CACHE_DIR / "module1_crawler" / "fireant" / ticker,
        DATA_CACHE_DIR / "module1_crawler" / "vietstock" / ticker,
        DATA_CACHE_DIR / "financial_xlsm" / ticker,
    ]
    for root in roots:
        overview = root / "company_overview_sample.csv"
        year = root / "financial_timeseries_year.csv"
        quarter = root / "financial_timeseries_quarter.csv"
        if overview.exists() and year.exists() and quarter.exists():
            try:
                company, annual, quarterly = _load_csv_bundle(overview, year, quarter, ticker)
                if _has_real_financial_data(annual):
                    return overview, year, quarter, f"Cache Tá»•ng quan doanh nghiá»‡p Ä‘Ã£ Ä‘á»“ng bá»™: {root}"
            except Exception:
                continue
    return None


def _format_note_value(value: object) -> str:
    if value is None:
        return "N/A"
    try:
        if pd.isna(value):
            return "N/A"
    except Exception:
        pass
    if isinstance(value, float):
        return f"{value:,.1f}" if abs(value) < 1000 else f"{value:,.0f}"
    if isinstance(value, int):
        return f"{value:,.0f}"
    return str(value)


# ===== V23.20: giá»¯ láº¡i thuáº­t ngá»¯ tÃ i chÃ­nh tiáº¿ng Anh nhÆ° V23.15 =====
# CÃ¡c thuáº­t ngá»¯ nhÆ° Owner Earnings, FCF, Moat score, Low/Base/High/Weighted Ä‘Æ°á»£c giá»¯ nguyÃªn
# Ä‘á»ƒ Ä‘á»c Ä‘Ãºng báº£n cháº¥t phÃ¢n tÃ­ch. HÃ m dÆ°á»›i Ä‘Ã¢y chá»‰ lÃ  compatibility no-op cho code báº£ng.
def _vi_text(value: object) -> object:
    return value

def _vi_dataframe_for_display(df: pd.DataFrame) -> pd.DataFrame:
    return df

# V23.33: bá» block _build_row_note/_render_explainable_table cÅ©; chá»‰ dÃ¹ng block má»›i á»Ÿ dÆ°á»›i Ä‘á»ƒ note hoáº¡t Ä‘á»™ng á»•n Ä‘á»‹nh.

def _load_csv_bundle(overview_csv: Path, year_csv: Path, quarter_csv: Path, ticker: str):
    company = _load_overview_cached(str(overview_csv), ticker)
    annual_raw = _load_timeseries_cached(str(year_csv), ticker, "Y", 12)
    quarterly = _load_timeseries_cached(str(quarter_csv), ticker, "Q", 24)
    annual = append_ttm_row(annual_raw, quarterly)
    return company, annual, quarterly


def _load_data(ticker: str, source: str) -> tuple[object, pd.DataFrame, pd.DataFrame, str, tuple[Path, Path, Path]]:
    ticker = _safe_ticker(ticker) or "DCM"

    if source == "Tá»± Ä‘á»™ng tá»« dá»¯ liá»‡u tá»•ng quan":
        # 0) Æ¯u tiÃªn bá»™ dá»¯ liá»‡u Ä‘ang hoáº¡t Ä‘á»™ng cá»§a Tá»•ng quan doanh nghiá»‡p trong cÃ¹ng app.
        active = _active_module1_bundle_for_ticker(ticker)
        if active:
            overview_csv, year_csv, quarter_csv, label = active
            company, annual, quarterly = _load_csv_bundle(overview_csv, year_csv, quarter_csv, ticker)
            if _has_real_financial_data(annual):
                _set_active_module1_bundle(overview_csv, year_csv, quarter_csv, label + " | Äá»‹nh giÃ¡ chuyÃªn sÃ¢u dÃ¹ng trá»±c tiáº¿p", ticker)
                return company, annual, quarterly, label + " | Äá»‹nh giÃ¡ chuyÃªn sÃ¢u dÃ¹ng trá»±c tiáº¿p", (overview_csv, year_csv, quarter_csv)

        # 1) Náº¿u Ä‘Ã£ cÃ³ cache Tá»•ng quan doanh nghiá»‡p/Äá»‹nh giÃ¡ chuyÃªn sÃ¢u cho Ä‘Ãºng mÃ£ thÃ¬ dÃ¹ng ngay, khÃ´ng crawl láº¡i.
        cached = _existing_cache_bundle_for_ticker(ticker)
        if cached:
            overview_csv, year_csv, quarter_csv, label = cached
            company, annual, quarterly = _load_csv_bundle(overview_csv, year_csv, quarter_csv, ticker)
            _set_active_module1_bundle(overview_csv, year_csv, quarter_csv, label, ticker)
            return company, annual, quarterly, label, (overview_csv, year_csv, quarter_csv)

        # 2) KhÃ´ng cÃ³ cache thÃ¬ tá»± gá»i pipeline crawler Tá»•ng quan doanh nghiá»‡p.
        overview, year, quarter, label = _export_module1_crawler_cached(ticker, "FireAnt + Vietstock", str(DATA_CACHE_DIR), str(RAW_DIR))
        overview_csv, year_csv, quarter_csv = Path(overview), Path(year), Path(quarter)
        company, annual, quarterly = _load_csv_bundle(overview_csv, year_csv, quarter_csv, ticker)
        if _has_real_financial_data(annual):
            _set_active_module1_bundle(overview_csv, year_csv, quarter_csv, label + " | Auto-sync tá»« Äá»‹nh giÃ¡ chuyÃªn sÃ¢u", ticker)
            return company, annual, quarterly, label + " | Auto-sync tá»« Äá»‹nh giÃ¡ chuyÃªn sÃ¢u", (overview_csv, year_csv, quarter_csv)

        # 3) Fallback cuá»‘i: dá»¯ liá»‡u tÃ­ch há»£p, chá»‰ dÃ¹ng náº¿u mÃ£ cÃ³ dá»¯ liá»‡u tháº­t.
        if BUNDLED_XLSM.exists():
            overview, year, quarter, label2 = _export_bundled_financial_cached(str(BUNDLED_XLSM), ticker, str(DATA_CACHE_DIR))
            overview_csv, year_csv, quarter_csv = Path(overview), Path(year), Path(quarter)
            company, annual, quarterly = _load_csv_bundle(overview_csv, year_csv, quarter_csv, ticker)
            if _has_real_financial_data(annual):
                _set_active_module1_bundle(overview_csv, year_csv, quarter_csv, label2 + " | fallback Financial", ticker)
                return company, annual, quarterly, label2 + " | fallback Financial", (overview_csv, year_csv, quarter_csv)
        return company, annual, quarterly, label, (Path(overview), Path(year), Path(quarter))

    elif source in {"FireAnt", "Vietstock", "FireAnt + Vietstock"}:
        overview, year, quarter, label = _export_module1_crawler_cached(ticker, source, str(DATA_CACHE_DIR), str(RAW_DIR))
        overview_csv, year_csv, quarter_csv = Path(overview), Path(year), Path(quarter)
        company, annual, quarterly = _load_csv_bundle(overview_csv, year_csv, quarter_csv, ticker)
        if _has_real_financial_data(annual):
            _set_active_module1_bundle(overview_csv, year_csv, quarter_csv, label + " | Ä‘á»“ng bá»™ chá»§ Ä‘á»™ng", ticker)
        return company, annual, quarterly, label, (overview_csv, year_csv, quarter_csv)
    elif source == "Financial tÃ­ch há»£p" and BUNDLED_XLSM.exists():
        overview, year, quarter, label = _export_bundled_financial_cached(str(BUNDLED_XLSM), ticker, str(DATA_CACHE_DIR))
        overview_csv, year_csv, quarter_csv = Path(overview), Path(year), Path(quarter)
        company, annual, quarterly = _load_csv_bundle(overview_csv, year_csv, quarter_csv, ticker)
        if _has_real_financial_data(annual):
            _set_active_module1_bundle(overview_csv, year_csv, quarter_csv, label + " | Ä‘á»“ng bá»™ Financial", ticker)
        return company, annual, quarterly, label, (overview_csv, year_csv, quarter_csv)
    else:
        overview_csv, year_csv, quarter_csv = DEFAULT_OVERVIEW_CSV, DEFAULT_YEAR_CSV, DEFAULT_QUARTER_CSV
        label = "CSV máº«u tÃ­ch há»£p"
    company, annual, quarterly = _load_csv_bundle(overview_csv, year_csv, quarter_csv, ticker)
    return company, annual, quarterly, label, (overview_csv, year_csv, quarter_csv)

def _parse_num(value: object) -> float | None:
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).replace(",", "").replace("%", "").strip()
    if not text or text in {"-", "--", "N/A", "None"}:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _style_table(df: pd.DataFrame):
    if df is None or df.empty:
        return df
    def styles(data: pd.DataFrame) -> pd.DataFrame:
        out = pd.DataFrame("", index=data.index, columns=data.columns)
        numeric_cols = []
        for col in data.columns:
            if any(k in col.lower() for k in ["giÃ¡", "mos", "Ä‘iá»ƒm", "trá»ng sá»‘", "%", "vnd", "cp", "tá»·", "láº§n", "value", "score"]):
                numeric_cols.append(col)
        vals = []
        for col in numeric_cols:
            vals.extend([v for v in data[col].map(_parse_num).dropna().tolist()])
        max_pos = max([v for v in vals if v > 0], default=1.0)
        max_neg = abs(min([v for v in vals if v < 0], default=-1.0))
        for col in numeric_cols:
            for idx, raw in data[col].items():
                v = _parse_num(raw)
                if v is None:
                    continue
                if v < 0:
                    alpha = min(0.65, 0.12 + 0.53 * min(abs(v) / max_neg, 1))
                    out.loc[idx, col] = f"background-color: rgba(239,68,68,{alpha:.2f}); color:#7f1d1d;"
                elif v > 0:
                    alpha = min(0.55, 0.10 + 0.45 * min(v / max_pos, 1))
                    out.loc[idx, col] = f"background-color: rgba(16,185,129,{alpha:.2f}); color:#064e3b;"
        # Heatmap riÃªng cho cÃ¡c má»©c Moat level dáº¡ng chá»¯ Ä‘á»ƒ khÃ´ng bá»‹ bá» sÃ³t nhÆ° 'Lá»£i tháº¿ khÃ¡'.
        for col in data.columns:
            cl = str(col).strip().lower()
            if cl in {"moat level", "má»©c moat", "moat_level"} or "moat level" in cl:
                for idx, raw in data[col].items():
                    txt = str(raw).strip().lower()
                    if not txt:
                        continue
                    if "ráº¥t máº¡nh" in txt or "very strong" in txt:
                        out.loc[idx, col] = "background-color:#0B7F75; color:#FFFFFF; font-weight:950;"
                    elif "máº¡nh" in txt or "strong" in txt:
                        out.loc[idx, col] = "background-color:#D1FAE5; color:#065F46; font-weight:900;"
                    elif "khÃ¡" in txt or "good" in txt:
                        out.loc[idx, col] = "background-color:#FFF4C7; color:#7A4B00; font-weight:900;"
                    elif "bÃ¬nh" in txt or "trung" in txt or "normal" in txt or "average" in txt:
                        out.loc[idx, col] = "background-color:#FEF3C7; color:#92400E; font-weight:850;"
                    elif "yáº¿u" in txt or "khÃ´ng" in txt or "weak" in txt or "no moat" in txt:
                        out.loc[idx, col] = "background-color:#FEE2E2; color:#991B1B; font-weight:850;"
        return out
    fmt = {}
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            low = str(col).lower()
            if str(col).strip().upper() in {"STT", "#"}:
                fmt[col] = "{:,.0f}"
            elif "%" in str(col) or "mos" in low or "Ä‘iá»ƒm" in low or "trá»ng sá»‘" in low:
                fmt[col] = "{:,.1f}"
            elif "giÃ¡" in low or "cp" in low:
                fmt[col] = "{:,.0f}"
            else:
                fmt[col] = "{:,.1f}"
    return df.style.format(fmt, na_rep="").apply(styles, axis=None)


def _show_table(df: pd.DataFrame, height: int | None = 520) -> None:
    if df is None or df.empty:
        st.info("ChÆ°a cÃ³ dá»¯ liá»‡u.")
    else:
        safe_df = _hide_source_columns(df)
        st.dataframe(_style_table(safe_df), use_container_width=True, height=(height or 520), hide_index=True)


def _render_static_html_table(df: pd.DataFrame, table_kind: str = "", height: int = 520) -> None:
    """Render a non-clickable HTML table using the same visual language as explainable tables.

    Used for small summary tables where the user needs the standard table format but not row notes.
    """
    if df is None or df.empty:
        st.info("ChÆ°a cÃ³ dá»¯ liá»‡u.")
        return
    display_df = _vi_dataframe_for_display(df.copy())
    if "Note" in display_df.columns:
        display_df = display_df.drop(columns=["Note"], errors="ignore")
    display_df = display_df.drop(columns=[c for c in ["Nguá»“n/logic", "Nguá»“n / logic"] if c in display_df.columns], errors="ignore")

    table_id = "static_tbl_" + str(abs(hash((table_kind, tuple(display_df.columns), len(display_df), APP_VERSION))))[0:10]
    header_cells = []
    for c in display_df.columns:
        hcls = "summary-layer-header" if table_kind == "financial_manipulation_summary" and str(c).strip() == "Lá»›p" else ""
        header_cells.append(f"<th class='{hcls}'>{html.escape(str(c))}</th>")
    headers = "".join(header_cells)
    rows_html = []
    for _, row in display_df.iterrows():
        tds = []
        for c in display_df.columns:
            val = row.get(c)
            text = _format_note_value(val)
            cls = _signal_class(val) if c in {"TÃ­n hiá»‡u", "Má»©c Ä‘á»™", "Má»©c cáº£nh bÃ¡o", "TÃ¬nh tráº¡ng", "Khuyáº¿n nghá»‹", "Káº¿t luáº­n", "Káº¿t luáº­n theo mÃ£", "Moat level", "Má»©c moat", "Äá»™ tin cáº­y", "ÄÃ¡nh giÃ¡ sÆ¡ bá»™", "Loáº¡i lá»£i tháº¿", "Vai trÃ²"} else ""
            num = _parse_num(val)
            if not cls and num is not None and any(k in str(c).lower() for k in ["giÃ¡", "mos", "Ä‘iá»ƒm", "Ä‘iá»ƒm nhiá»‡t", "trá»ng", "%", "value", "score"]):
                cls = "pos" if num > 0 else "neg" if num < 0 else ""
            cell_classes = [cls] if cls else []
            if table_kind == "financial_manipulation_summary" and str(c).strip() == "Lá»›p":
                cell_classes.append("summary-layer")
            tds.append(f"<td class='{' '.join(cell_classes)}'>{html.escape(text)}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    # V23.61: báº£ng tá»•ng há»£p thao tÃºng tÃ i chÃ­nh chá»‰ cÃ³ 4 lá»›p nÃªn tá»± tÃ­nh chiá»u cao gá»n,
    # khÃ´ng Ä‘á»ƒ dÆ° khoáº£ng tráº¯ng lá»›n lÃ m pháº§n phÃ¢n tÃ­ch tá»«ng lá»›p bá»‹ Ä‘áº©y xuá»‘ng xa.
    if table_kind == "financial_manipulation_summary":
        component_height = max(300, min(int(height), 92 + 78 * max(len(display_df), 1)))
    else:
        component_height = max(320, int(height))
    html_doc = f"""
    <div class='wrap'>
      <table id='{table_id}'>
        <thead><tr>{headers}</tr></thead>
        <tbody>{''.join(rows_html)}</tbody>
      </table>
    </div>
    <style>
      .wrap {{border:1px solid #e2e8f0; border-radius:12px; overflow:visible; background:#FFFFFF;}}
      table {{border-collapse:collapse; width:100%; font-family: system-ui, -apple-system, Segoe UI, sans-serif; font-size:13px; table-layout:auto;}}
      th {{background:#EAF7F1; color:#123D3A; text-align:left; border-bottom:1px solid #e2e8f0; padding:8px; font-weight:950; box-shadow:0 2px 0 rgba(11,127,117,.18); white-space:nowrap;}}
      td {{border-bottom:1px solid #edf2f7; padding:7px 8px; vertical-align:top; color:#123D3A; line-height:1.38;}}
      th.summary-layer-header, td.summary-layer {{font-weight:1000 !important; color:#075E54 !important; background:linear-gradient(135deg,#ECFDF5 0%,#FFF8D6 100%) !important; white-space:nowrap;}}
      tbody tr:nth-child(even) td {{background:#FBFDFB;}}
      tr:hover td {{background:#F7FBF8;}}
      td.pos {{background:rgba(16,185,129,.11); color:#064e3b; font-weight:600;}}
      td.neg {{background:rgba(239,68,68,.11); color:#7f1d1d; font-weight:600;}}
      td.sig-red-strong {{background:#FECACA !important; color:#7F1D1D !important; font-weight:900; border-left:5px solid #DC2626;}}
      td.sig-red {{background:#FEE2E2 !important; color:#991B1B !important; font-weight:800; border-left:4px solid #EF4444;}}
      td.sig-purple-strong {{background:#E9D5FF !important; color:#581C87 !important; font-weight:900; border-left:5px solid #7E22CE;}}
      td.sig-purple {{background:#F3E8FF !important; color:#6B21A8 !important; font-weight:800; border-left:4px solid #A855F7;}}
      td.sig-yellow {{background:#FEF3C7 !important; color:#92400E !important; font-weight:800; border-left:4px solid #F59E0B;}}
      td.heat-green-strong {{background:#047857 !important; color:#FFFFFF !important; font-weight:950 !important;}}
      td.heat-green {{background:#A7F3D0 !important; color:#064E3B !important; font-weight:900 !important;}}
      td.heat-yellow {{background:#FEF3C7 !important; color:#92400E !important; font-weight:900 !important;}}
      td.heat-orange {{background:#FED7AA !important; color:#9A3412 !important; font-weight:900 !important;}}
      td.heat-red {{background:#FECACA !important; color:#7F1D1D !important; font-weight:900 !important;}}
    </style>
    """
    components.html(html_doc, height=component_height, scrolling=False)


def _has_real_financial_data(annual_df: pd.DataFrame) -> bool:
    if annual_df is None or annual_df.empty:
        return False
    core_cols = ["revenue_bil", "net_profit_bil", "equity_bil", "total_assets_bil", "cfo_bil", "owner_earnings_bil"]
    existing = [c for c in core_cols if c in annual_df.columns]
    if not existing:
        return False
    return pd.to_numeric(annual_df[existing].stack(), errors="coerce").notna().sum() >= 5



# ===== V23.9: note engine theo tá»«ng doanh nghiá»‡p, sá»‘ liá»‡u vÃ  tÃ i liá»‡u nguá»“n =====
def _ctx() -> dict:
    obj = st.session_state.get("module2_note_context", {})
    return obj if isinstance(obj, dict) else {}


def _series(df: pd.DataFrame, col: str) -> pd.Series:
    if df is None or df.empty or col not in df.columns:
        return pd.Series(dtype="float64")
    return pd.to_numeric(df[col], errors="coerce")


def _latest_dict(df: pd.DataFrame) -> dict:
    if df is None or df.empty:
        return {}
    if "period" in df.columns:
        ttm = df[df["period"].astype(str).str.upper().str.contains("TTM|T12M", regex=True, na=False)]
        if not ttm.empty:
            return ttm.iloc[-1].to_dict()
    return df.iloc[-1].to_dict()


def _recent_median2(df: pd.DataFrame, col: str, n: int = 5) -> float | None:
    s = _series(df, col).dropna().tail(n)
    return float(s.median()) if not s.empty else None


def _recent_mean2(df: pd.DataFrame, col: str, n: int = 5) -> float | None:
    s = _series(df, col).dropna().tail(n)
    return float(s.mean()) if not s.empty else None


def _recent_positive_ratio2(df: pd.DataFrame, col: str, n: int = 5) -> float | None:
    s = _series(df, col).dropna().tail(n)
    return float((s > 0).mean()) if not s.empty else None


def _cagr2(df: pd.DataFrame, col: str, years: int = 5) -> float | None:
    s = _series(df, col).dropna().tail(years + 1)
    if len(s) < 2:
        return None
    start, end = float(s.iloc[0]), float(s.iloc[-1])
    periods = len(s) - 1
    if start <= 0 or end <= 0 or periods <= 0:
        return None
    return (end / start) ** (1 / periods) - 1


def _cv2(df: pd.DataFrame, col: str, n: int = 7) -> float | None:
    s = _series(df, col).dropna().tail(n)
    s = s[s.abs() > 1e-9]
    if len(s) < 3 or abs(float(s.mean())) < 1e-9:
        return None
    return float(s.std(ddof=0) / abs(s.mean()))


def _share_count_from_context(company: object, annual_df: pd.DataFrame) -> float | None:
    latest = _latest_dict(annual_df)
    direct = _parse_num(latest.get("shares_outstanding_mil"))
    overview = _parse_num(getattr(company, "shares_outstanding_mil", None))
    np_bil = _parse_num(latest.get("net_profit_bil"))
    eps_vnd = _parse_num(latest.get("eps_vnd"))
    inferred = None
    if np_bil is not None and eps_vnd and eps_vnd > 0:
        inferred = np_bil * 1000 / eps_vnd
    if direct and direct > 0:
        return direct
    if inferred and inferred > 0:
        if overview and overview > 0 and abs(inferred - overview) / max(overview, 1e-9) <= 0.30:
            return overview
        return inferred
    return overview if overview and overview > 0 else None


def _per_share_bil(value_bil: float | None, shares_mil: float | None) -> float | None:
    if value_bil is None or shares_mil is None or shares_mil <= 0:
        return None
    return value_bil * 1000 / shares_mil



def _signal_class(value: object) -> str:
    s = str(value or "").strip().lower()
    if not s or s in {"nan", "none", "n/a", "na", "-"}:
        return ""
    # Heatmap Ä‘áº§y Ä‘á»§ cho Moat level / tÃ­n hiá»‡u chá»¯.
    if "lá»£i tháº¿" in s or "moat" in s:
        if "ráº¥t máº¡nh" in s or "very strong" in s or "xuáº¥t sáº¯c" in s:
            return "sig-purple-strong"
        if "máº¡nh" in s or "strong" in s:
            return "sig-purple"
        if "khÃ¡" in s or "good" in s:
            return "sig-yellow"
        if "bÃ¬nh" in s or "trung" in s or "normal" in s or "average" in s:
            return "sig-yellow"
        if "yáº¿u" in s or "khÃ´ng" in s or "weak" in s or "no moat" in s:
            return "sig-red"
    # Äá»™ tin cáº­y.
    if s in {"cao", "high"}:
        return "sig-purple-strong"
    if s in {"trung bÃ¬nh", "medium", "moderate"}:
        return "sig-yellow"
    if s in {"tháº¥p", "low", "khÃ´ng cÃ³ dá»¯ liá»‡u", "no data"}:
        return "sig-red"
    # Rá»§i ro/cáº£nh bÃ¡o.
    red_terms = ["cáº£nh bÃ¡o", "rá»§i ro", "rá»§i ro chu ká»³", "yáº¿u", "Ã¢m", "suy giáº£m", "khÃ´ng Ä‘áº¡t", "chÆ°a Ä‘áº¡t", "khÃ´ng phÃ¹ há»£p", "thiáº¿u dá»¯ liá»‡u", "khÃ´ng cÃ³ dá»¯ liá»‡u", "chÆ°a Ä‘á»§", "chÆ°a cÃ³", "lá»—i", "Ä‘Ã²n báº©y cao", "xáº¥u"]
    if any(k in s for k in red_terms):
        if any(k in s for k in ["nghiÃªm trá»ng", "ráº¥t", "khÃ´ng Ä‘áº¡t", "chÆ°a Ä‘áº¡t", "rá»§i ro", "yáº¿u", "xáº¥u"]):
            return "sig-red-strong"
        return "sig-red"
    # TÃ­n hiá»‡u tÃ­ch cá»±c.
    purple_terms = ["tá»‘t", "Ä‘áº¡t", "máº¡nh", "an toÃ n", "hiá»‡u quáº£", "tÃ­ch cá»±c", "vÆ°á»£t", "cao", "bá»n", "á»•n Ä‘á»‹nh", "cÃ³ runway", "runway", "pricing power", "cÃ³ báº±ng chá»©ng", "quality", "cash tá»‘t", "táº¡o giÃ¡ trá»‹"]
    if any(k in s for k in purple_terms):
        if any(k in s for k in ["ráº¥t", "máº¡nh", "vÆ°á»£t", "tá»‘t", "cao", "bá»n"]):
            return "sig-purple-strong"
        return "sig-purple"
    # Theo dÃµi/cáº§n kiá»ƒm chá»©ng.
    yellow_terms = ["theo dÃµi", "cáº§n kiá»ƒm", "cáº§n soi", "cáº§n xÃ¡c minh", "cáº§n kiá»ƒm chá»©ng", "cáº§n bá»• sung", "cáº§n tÃ¬m", "cáº©n trá»ng", "trung bÃ¬nh", "bÃ¬nh thÆ°á»ng", "khÃ¡", "chÆ°a rÃµ", "háº¡n cháº¿", "gáº§n vÃ¹ng", "chá»", "kiá»ƒm chá»©ng", "chÆ°a káº¿t luáº­n"]
    if any(k in s for k in yellow_terms):
        return "sig-yellow"
    return ""


def _money(value: object, suffix: str = "") -> str:
    n = _parse_num(value)
    if n is None:
        return "N/A"
    return f"{n:,.0f}{suffix}"


def _pct(value: object) -> str:
    n = _parse_num(value)
    if n is None:
        return "N/A"
    return f"{n:,.1f}%"


def _ratio(value: object) -> str:
    n = _parse_num(value)
    if n is None:
        return "N/A"
    return f"{n:,.1f} láº§n"


def _bil(value: object) -> str:
    n = _parse_num(value)
    if n is None:
        return "N/A"
    return f"{n:,.0f} tá»· Ä‘á»“ng"


def _period_label(annual_df: pd.DataFrame) -> str:
    latest = _latest_dict(annual_df)
    return str(latest.get("period") or latest.get("year") or "ká»³ má»›i nháº¥t/TTM")


def _source_principle_text(kind: str, cls_name: str) -> str:
    base = {
        "owner": "CÆ¡ sá»Ÿ tÆ° duy: Buffett xem Owner Earnings lÃ  chá»‰ tiÃªu phÃ¹ há»£p cho Ä‘á»‹nh giÃ¡ chá»§ sá»Ÿ há»¯u: lá»£i nhuáº­n bÃ¡o cÃ¡o + kháº¥u hao/phi tiá»n máº·t - capex duy trÃ¬ cáº§n thiáº¿t Ä‘á»ƒ giá»¯ vá»‹ tháº¿ cáº¡nh tranh vÃ  sáº£n lÆ°á»£ng. VÃ¬ váº­y vá»›i doanh nghiá»‡p vá»‘n nháº¹/dÃ²ng tiá»n tháº­t, OE Ä‘Æ°á»£c Æ°u tiÃªn; vá»›i doanh nghiá»‡p Ä‘ang má»Ÿ rá»™ng capex máº¡nh, cáº§n tÃ¡ch capex duy trÃ¬ vÃ  capex tÄƒng trÆ°á»Ÿng.",
        "fcf": "CÆ¡ sá»Ÿ tÆ° duy: FCF lÃ  kiá»ƒm tra dÃ²ng tiá»n, nhÆ°ng khÃ´ng tá»± Ä‘á»™ng thay tháº¿ Owner Earnings. Náº¿u FCF Ã¢m do Ä‘áº§u tÆ° má»Ÿ rá»™ng táº¡o ROIC cao thÃ¬ khÃ´ng káº¿t luáº­n xáº¥u ngay; náº¿u FCF Ã¢m kÃ©o dÃ i trong khi ROIC/biÃªn lá»£i nhuáº­n giáº£m thÃ¬ pháº£i cáº£nh bÃ¡o cháº¥t lÆ°á»£ng lá»£i nhuáº­n.",
        "epv": "CÆ¡ sá»Ÿ tÆ° duy: Graham/Dodd Ä‘áº·t trá»ng tÃ¢m vÃ o earning power vÃ  sá»± khÃ¡c biá»‡t giá»¯a giÃ¡ vÃ  giÃ¡ trá»‹. EPS/LNST pháº£i Ä‘Æ°á»£c chuáº©n hÃ³a qua nhiá»u ká»³, Ä‘áº·c biá»‡t vá»›i doanh nghiá»‡p chu ká»³ hoáº·c cÃ³ lá»£i nhuáº­n báº¥t thÆ°á»ng.",
        "asset": "CÆ¡ sá»Ÿ tÆ° duy: Graham Æ°u tiÃªn biÃªn an toÃ n vÃ  giÃ¡ trá»‹ tÃ i sáº£n khi doanh nghiá»‡p rÆ¡i vÃ o nhÃ³m asset play/deep value. TÃ i sáº£n thanh khoáº£n Ä‘Æ°á»£c haircut báº£o thá»§ Ä‘á»ƒ báº£o vá»‡ downside, khÃ´ng dÃ¹ng book value thÃ´ khi cháº¥t lÆ°á»£ng tÃ i sáº£n chÆ°a Ä‘Æ°á»£c kiá»ƒm tra.",
        "porter": "CÆ¡ sá»Ÿ tÆ° duy: Porter khÃ´ng xem lá»£i tháº¿ cáº¡nh tranh lÃ  kháº©u hiá»‡u chung. Moat pháº£i truy vá» cÃ¡c hoáº¡t Ä‘á»™ng cá»¥ thá»ƒ trong chuá»—i giÃ¡ trá»‹ táº¡o giÃ¡ trá»‹ cho khÃ¡ch hÃ ng hoáº·c giáº£m chi phÃ­, vÃ  pháº£i cÃ³ báº±ng chá»©ng Ä‘á»‹nh lÆ°á»£ng/Ä‘á»‹nh tÃ­nh cá»§a chÃ­nh doanh nghiá»‡p.",
        "risk": "CÆ¡ sá»Ÿ tÆ° duy: Howard Marks nháº¥n máº¡nh kiá»ƒm soÃ¡t rá»§i ro vÃ  dáº£i káº¿t quáº£ cÃ³ thá»ƒ xáº£y ra. VÃ¬ váº­y note dÃ¹ng Bear/Base/Bull vÃ  sá»‘ liá»‡u thá»±c táº¿ cá»§a doanh nghiá»‡p thay vÃ¬ má»™t káº¿t luáº­n cá»‘ Ä‘á»‹nh.",
        "mos": "CÆ¡ sá»Ÿ tÆ° duy: Graham/Li Lu coi margin of safety lÃ  trung tÃ¢m vÃ¬ giÃ¡ trá»‹ chá»‰ lÃ  Æ°á»›c tÃ­nh. MOS pháº£i phá»¥ thuá»™c loáº¡i doanh nghiá»‡p, Ä‘á»™ tin cáº­y dÃ²ng tiá»n, tÃ i sáº£n báº£o vá»‡ downside vÃ  rá»§i ro chu ká»³ cá»§a chÃ­nh mÃ£ Ä‘ang phÃ¢n tÃ­ch.",
    }
    txt = base.get(kind, base["mos"])
    if "Cyclical" in cls_name:
        txt += " Vá»›i doanh nghiá»‡p cÃ³ tÃ­nh chu ká»³, há»‡ thá»‘ng giáº£m Ã½ nghÄ©a cá»§a lá»£i nhuáº­n 1 nÄƒm vÃ  Æ°u tiÃªn trung vá»‹ nhiá»u ká»³/kiá»ƒm tra biÃªn an toÃ n cao hÆ¡n."
    elif "Quality" in cls_name:
        txt += " Vá»›i doanh nghiá»‡p cháº¥t lÆ°á»£ng/compounder, há»‡ thá»‘ng Æ°u tiÃªn ROIC, Owner Earnings vÃ  kháº£ nÄƒng tÃ¡i Ä‘áº§u tÆ°; tuy nhiÃªn váº«n kiá»ƒm tra CFO/LNST vÃ  FCF Ä‘á»ƒ trÃ¡nh lá»£i nhuáº­n káº¿ toÃ¡n."
    elif "Asset" in cls_name:
        txt += " Vá»›i doanh nghiá»‡p asset play/deep value, há»‡ thá»‘ng Æ°u tiÃªn tÃ i sáº£n thanh khoáº£n rÃ²ng vÃ  downside protection trÆ°á»›c khi tin vÃ o tÄƒng trÆ°á»Ÿng."
    elif "Financial" in cls_name:
        txt += " Vá»›i ngÃ¢n hÃ ng/báº£o hiá»ƒm/tÃ i chÃ­nh, há»‡ thá»‘ng khÃ´ng dÃ¹ng FCF/VLÄ tá»•ng quÃ¡t lÃ m lÃµi, mÃ  Æ°u tiÃªn P/B, ROE vÃ  cháº¥t lÆ°á»£ng tÃ i sáº£n/vá»‘n."
    return txt


def _company_snapshot() -> str:
    c = _ctx().get("company")
    annual = _ctx().get("annual_df", pd.DataFrame())
    cls = _ctx().get("classification")
    latest = _latest_dict(annual)
    period = _period_label(annual)
    roic = _recent_median2(annual, "roic_standard_pct") or _recent_median2(annual, "roic_pct")
    roe = _recent_median2(annual, "roe_actual_pct") or _recent_median2(annual, "roe_pct")
    cfo_np = _recent_median2(annual, "cfo_to_net_profit")
    fcf_np = _recent_median2(annual, "fcf_to_net_profit")
    rev_cagr = _cagr2(annual, "revenue_bil")
    profit_cv = _cv2(annual, "net_profit_bil")
    price = _parse_num(getattr(c, "current_price", None)) if c is not None else None
    parts = [
        f"Doanh nghiá»‡p: {getattr(c, 'ticker', '')} - {getattr(c, 'company_name', '')}",
        f"Ká»³ dá»¯ liá»‡u dÃ¹ng chÃ­nh: {period}; bá»™ dá»¯ liá»‡u: {_public_text(_ctx().get('source_label', 'N/A'))}",
        f"PhÃ¢n loáº¡i hiá»‡n táº¡i: {getattr(cls, 'company_type', 'N/A')} | GiÃ¡ hiá»‡n táº¡i: {_money(price, ' Ä‘/cp')}",
        f"TÃ­n hiá»‡u chÃ­nh cá»§a chÃ­nh doanh nghiá»‡p: ROIC trung vá»‹ {_pct(roic)}, ROE trung vá»‹ {_pct(roe)}, CFO/LNST {_ratio(cfo_np)}, FCF/LNST {_ratio(fcf_np)}, CAGR doanh thu {_pct(None if rev_cagr is None else rev_cagr*100)}, Ä‘á»™ biáº¿n Ä‘á»™ng LNST {_pct(None if profit_cv is None else profit_cv*100)}.",
    ]
    return "\n".join(parts)


def _module2_numeric_evidence_for_note(topic: str = "") -> str:
    """Return concrete metrics used in click-notes so explanation is not generic."""
    annual = _ctx().get("annual_df", pd.DataFrame())
    latest = _latest_dict(annual)
    topic_l = str(topic or "").lower()
    period = _period_label(annual)
    roic = _recent_median2(annual, "roic_standard_pct") or _recent_median2(annual, "roic_pct")
    roe = _recent_median2(annual, "roe_actual_pct") or _recent_median2(annual, "roe_pct")
    gross = _recent_median2(annual, "gross_margin_pct")
    net = _recent_median2(annual, "net_margin_pct")
    ebit = _recent_median2(annual, "ebit_margin_pct")
    cfo_np = _recent_median2(annual, "cfo_to_net_profit")
    fcf_np = _recent_median2(annual, "fcf_to_net_profit")
    fcf_pos = _recent_positive_ratio2(annual, "free_cash_flow_bil")
    rev_cagr = _cagr2(annual, "revenue_bil")
    np_cagr = _cagr2(annual, "net_profit_bil")
    oe_cagr = _cagr2(annual, "owner_earnings_bil")
    profit_cv = _cv2(annual, "net_profit_bil")
    margin_cv = _cv2(annual, "gross_margin_pct")
    ccc = _recent_median2(annual, "cash_conversion_cycle_days")
    inv_turn = _recent_median2(annual, "inventory_turnover")
    debt_ebitda = _parse_num(latest.get("net_debt_to_ebitda"))
    net_debt_equity = _parse_num(latest.get("net_debt_to_equity"))
    wacc = _recent_median2(annual, "wacc_pct")
    capex = _recent_median2(annual, "capex_bil")
    fcf = _recent_median2(annual, "free_cash_flow_bil")
    oe = _recent_median2(annual, "owner_earnings_bil")
    rev = _recent_median2(annual, "revenue_bil")
    np = _recent_median2(annual, "net_profit_bil")
    total_inv = _recent_median2(annual, "total_investment_bil")
    lines = [f"- Ká»³/chuá»—i dá»¯ liá»‡u dÃ¹ng: {period}."]
    if any(k in topic_l for k in ["tÃ¡i Ä‘áº§u tÆ°", "runway", "compounder"]):
        lines += [
            f"- TÄƒng trÆ°á»Ÿng: CAGR doanh thu {_pct(None if rev_cagr is None else rev_cagr*100)}, CAGR LNST {_pct(None if np_cagr is None else np_cagr*100)}, CAGR Owner Earnings {_pct(None if oe_cagr is None else oe_cagr*100)}.",
            f"- Hiá»‡u quáº£ vá»‘n: ROIC trung vá»‹ {_pct(roic)}, WACC trung vá»‹ {_pct(wacc)}, ROE trung vá»‹ {_pct(roe)}.",
            f"- DÃ²ng tiá»n/tÃ¡i Ä‘áº§u tÆ°: CFO/LNST {_ratio(cfo_np)}, FCF/LNST {_ratio(fcf_np)}, FCF trung vá»‹ {_bil(fcf)}, Owner Earnings trung vá»‹ {_bil(oe)}, Capex trung vá»‹ {_bil(capex)}, tá»•ng Ä‘áº§u tÆ° trung vá»‹ {_bil(total_inv)}.",
            f"- Diá»…n giáº£i cá»¥ thá»ƒ: náº¿u doanh thu/LNST tÄƒng nhÆ°ng ROIC khÃ´ng cao hÆ¡n WACC hoáº·c FCF/LNST tháº¥p, tÄƒng trÆ°á»Ÿng cÃ³ thá»ƒ Ä‘ang tiÃªu tiá»n/Ä‘Ã²n báº©y chá»© chÆ°a cháº¯c táº¡o giÃ¡ trá»‹.",
        ]
    elif any(k in topic_l for k in ["dÃ²ng tiá»n", "cash", "fcf", "owner"]):
        lines += [
            f"- Cháº¥t lÆ°á»£ng tiá»n: CFO/LNST {_ratio(cfo_np)}, FCF/LNST {_ratio(fcf_np)}, tá»· lá»‡ ká»³ FCF dÆ°Æ¡ng {_pct(None if fcf_pos is None else fcf_pos*100)}.",
            f"- Quy mÃ´ tiá»n: CFO ká»³ má»›i nháº¥t {_bil(latest.get('cfo_bil'))}, FCF ká»³ má»›i nháº¥t {_bil(latest.get('free_cash_flow_bil'))}, Owner Earnings ká»³ má»›i nháº¥t {_bil(latest.get('owner_earnings_bil'))}, LNST ká»³ má»›i nháº¥t {_bil(latest.get('net_profit_bil'))}.",
        ]
    elif any(k in topic_l for k in ["cost", "chi phÃ­", "váº­n hÃ nh", "logistics", "chuá»—i giÃ¡ trá»‹"]):
        lines += [
            f"- Hiá»‡u quáº£ hoáº¡t Ä‘á»™ng: biÃªn gá»™p trung vá»‹ {_pct(gross)}, biÃªn EBIT {_pct(ebit)}, biÃªn rÃ²ng {_pct(net)}, biáº¿n Ä‘á»™ng biÃªn gá»™p {_pct(None if margin_cv is None else margin_cv*100)}.",
            f"- Vá»‘n lÆ°u Ä‘á»™ng/váº­n hÃ nh: CCC {_money(ccc, ' ngÃ y')}, vÃ²ng quay HTK {_ratio(inv_turn)}, CFO/LNST {_ratio(cfo_np)}.",
        ]
    elif any(k in topic_l for k in ["cáº¥u trÃºc", "chu ká»³", "ngÃ nh"]):
        lines += [
            f"- Chu ká»³: Ä‘á»™ biáº¿n Ä‘á»™ng LNST {_pct(None if profit_cv is None else profit_cv*100)}, CAGR doanh thu {_pct(None if rev_cagr is None else rev_cagr*100)}, CAGR LNST {_pct(None if np_cagr is None else np_cagr*100)}.",
            f"- BiÃªn lá»£i nhuáº­n: biÃªn gá»™p {_pct(gross)}, biÃªn rÃ²ng {_pct(net)}; ná»£/EBITDA {_ratio(debt_ebitda)}.",
        ]
    elif any(k in topic_l for k in ["quáº£n trá»‹", "phÃ¢n bá»• vá»‘n", "an toÃ n"]):
        lines += [
            f"- PhÃ¢n bá»• vá»‘n: ROIC {_pct(roic)}, WACC {_pct(wacc)}, ROE {_pct(roe)}, Capex {_bil(capex)}, tá»•ng Ä‘áº§u tÆ° {_bil(total_inv)}.",
            f"- An toÃ n tÃ i chÃ­nh: ná»£ rÃ²ng/EBITDA {_ratio(debt_ebitda)}, ná»£ rÃ²ng/VCSH {_ratio(net_debt_equity)}, CFO/LNST {_ratio(cfo_np)}.",
        ]
    else:
        lines += [
            f"- Sinh lá»i: ROIC {_pct(roic)}, ROE {_pct(roe)}, biÃªn gá»™p {_pct(gross)}, biÃªn rÃ²ng {_pct(net)}.",
            f"- TÄƒng trÆ°á»Ÿng & dÃ²ng tiá»n: doanh thu trung vá»‹ {_bil(rev)}, LNST trung vá»‹ {_bil(np)}, CAGR doanh thu {_pct(None if rev_cagr is None else rev_cagr*100)}, CFO/LNST {_ratio(cfo_np)}, FCF/LNST {_ratio(fcf_np)}.",
        ]
    return "\n".join(lines)


def _valuation_method_note(rowd: dict) -> str:
    c = _ctx().get("company")
    annual = _ctx().get("annual_df", pd.DataFrame())
    assumptions = _ctx().get("assumptions", {})
    cls = _ctx().get("classification")
    cls_name = getattr(cls, "company_type", "N/A")
    latest = _latest_dict(annual)
    method = str(rowd.get("PhÆ°Æ¡ng phÃ¡p", ""))
    shares = _share_count_from_context(c, annual)
    current_price = _parse_num(rowd.get("GiÃ¡ hiá»‡n táº¡i")) or (_parse_num(getattr(c, "current_price", None)) if c is not None else None)
    intrinsic = _parse_num(rowd.get("GiÃ¡ trá»‹ ná»™i táº¡i/cp"))
    mos = _parse_num(rowd.get("MOS hiá»‡n táº¡i %"))
    lines = [_company_snapshot(), "", f"CHá»ˆ TIÃŠU/PHÆ¯Æ NG PHÃP ÄANG CHá»ŒN: {method}"]
    if "Earnings Power" in method:
        net_profit_norm = _recent_median2(annual, "net_profit_bil") or _parse_num(latest.get("net_profit_bil"))
        eps_norm = _recent_median2(annual, "eps_vnd") or _per_share_bil(net_profit_norm, shares) or (_parse_num(getattr(c, "eps", None)) if c is not None else None)
        target_pe = assumptions.get("target_pe_quality", 14.0) if cls_name == "Quality Compounder" else assumptions.get("target_pe_normal", assumptions.get("target_pe_default", 10.0))
        calc = eps_norm * target_pe if eps_norm is not None else None
        lines += [
            "CÃ¡ch tÃ­nh theo dá»¯ liá»‡u doanh nghiá»‡p:",
            f"- LNST chuáº©n hÃ³a = trung vá»‹ cÃ¡c ká»³ gáº§n Ä‘Ã¢y = {_bil(net_profit_norm)}.",
            f"- Sá»‘ cá»• phiáº¿u pha loÃ£ng/Æ°á»›c tÃ­nh = {_ratio(shares)} triá»‡u cp. Náº¿u file khÃ´ng cÃ³ cá»• phiáº¿u, há»‡ thá»‘ng suy ra tá»« LNST vÃ  EPS.",
            f"- EPS chuáº©n hÃ³a = LNST chuáº©n hÃ³a / sá»‘ cá»• phiáº¿u = {_money(eps_norm, ' Ä‘/cp')}.",
            f"- P/E má»¥c tiÃªu Ã¡p dá»¥ng cho {cls_name} = {_ratio(target_pe)}.",
            f"- GiÃ¡ trá»‹ tÃ­nh láº¡i = EPS chuáº©n hÃ³a x P/E má»¥c tiÃªu = {_money(calc, ' Ä‘/cp')}; giÃ¡ trá»‹ Ä‘ang dÃ¹ng trong báº£ng = {_money(intrinsic, ' Ä‘/cp')}.",
            f"- MOS hiá»‡n táº¡i = (giÃ¡ trá»‹ - giÃ¡ thá»‹ trÆ°á»ng) / giÃ¡ trá»‹ = {_pct(mos)} vá»›i giÃ¡ hiá»‡n táº¡i {_money(current_price, ' Ä‘/cp')}.",
            _source_principle_text("epv", cls_name),
        ]
    elif "Owner Earnings" in method:
        oe_norm = _recent_median2(annual, "owner_earnings_bil") or _parse_num(latest.get("owner_earnings_bil"))
        oeps = _recent_median2(annual, "oeps_vnd") or _per_share_bil(oe_norm, shares)
        rev_cagr = _cagr2(annual, "revenue_bil") or 0.0
        oe_cagr = _cagr2(annual, "owner_earnings_bil")
        growth_base = min(max(oe_cagr if oe_cagr is not None else rev_cagr, 0.0), assumptions.get("base_growth_cap_pct", 8.0) / 100)
        discount = assumptions.get("required_return", assumptions.get("required_return_pct", 13.0) / 100)
        terminal = assumptions.get("terminal_growth", assumptions.get("terminal_growth_pct", 3.0) / 100)
        calc = oeps * (1 + growth_base) / max(discount - terminal, 0.04) if oeps is not None else None
        lines += [
            "CÃ¡ch tÃ­nh theo dá»¯ liá»‡u doanh nghiá»‡p:",
            f"- Owner Earnings chuáº©n hÃ³a = trung vá»‹ cÃ¡c ká»³ gáº§n Ä‘Ã¢y = {_bil(oe_norm)}.",
            f"- OEPS chuáº©n hÃ³a = OE / sá»‘ cá»• phiáº¿u = {_money(oeps, ' Ä‘/cp')}.",
            f"- TÄƒng trÆ°á»Ÿng cÆ¡ sá»Ÿ láº¥y theo OE CAGR náº¿u cÃ³, náº¿u khÃ´ng dÃ¹ng CAGR doanh thu, sau Ä‘Ã³ cháº·n tráº§n = {_pct(growth_base*100)}.",
            f"- Required return = {_pct(discount*100)}, terminal growth = {_pct(terminal*100)}.",
            f"- GiÃ¡ trá»‹ tÃ­nh láº¡i = OEPS x (1+g) / (r-g) = {_money(calc, ' Ä‘/cp')}; báº£ng Ä‘ang dÃ¹ng = {_money(intrinsic, ' Ä‘/cp')}.",
            f"- Kiá»ƒm tra phÃ¹ há»£p: CFO/LNST {_ratio(_recent_median2(annual, 'cfo_to_net_profit'))}, FCF/LNST {_ratio(_recent_median2(annual, 'fcf_to_net_profit'))}. Náº¿u 2 chá»‰ tiÃªu nÃ y tháº¥p, OE pháº£i bá»‹ giáº£m Ä‘á»™ tin cáº­y.",
            _source_principle_text("owner", cls_name),
        ]
    elif "FCF" in method:
        cfo = _recent_median2(annual, "cfo_bil")
        capex = _recent_median2(annual, "capex_bil")
        fcf_norm = _recent_median2(annual, "free_cash_flow_bil") or _parse_num(latest.get("free_cash_flow_bil"))
        fcf_ps = _per_share_bil(fcf_norm, shares)
        discount = assumptions.get("required_return", assumptions.get("required_return_pct", 13.0) / 100)
        conservative_growth = assumptions.get("conservative_growth_pct", 0.0) / 100
        calc = fcf_ps / max(discount - conservative_growth, 0.06) if fcf_ps is not None and fcf_ps > 0 else None
        lines += [
            "CÃ¡ch tÃ­nh theo dá»¯ liá»‡u doanh nghiá»‡p:",
            f"- CFO chuáº©n hÃ³a = {_bil(cfo)}; Capex chuáº©n hÃ³a = {_bil(capex)}.",
            f"- FCF chuáº©n hÃ³a = CFO - Capex = {_bil(fcf_norm)}.",
            f"- FCF/cp = {_money(fcf_ps, ' Ä‘/cp')}.",
            f"- Suáº¥t vá»‘n hÃ³a báº£o thá»§ = required return - tÄƒng trÆ°á»Ÿng báº£o thá»§ = {_pct((discount-conservative_growth)*100)}.",
            f"- GiÃ¡ trá»‹ tÃ­nh láº¡i = FCF/cp / suáº¥t vá»‘n hÃ³a = {_money(calc, ' Ä‘/cp')}; báº£ng Ä‘ang dÃ¹ng = {_money(intrinsic, ' Ä‘/cp')}.",
            f"- TrÆ°á»ng há»£p doanh nghiá»‡p Ä‘ang má»Ÿ rá»™ng: náº¿u capex lá»›n nhÆ°ng ROIC váº«n cao ({_pct(_recent_median2(annual, 'roic_standard_pct') or _recent_median2(annual, 'roic_pct'))}), FCF Ã¢m khÃ´ng nháº¥t thiáº¿t xáº¥u; náº¿u FCF Ã¢m vÃ  ROIC giáº£m thÃ¬ cáº£nh bÃ¡o cháº¥t lÆ°á»£ng tÄƒng trÆ°á»Ÿng.",
            _source_principle_text("fcf", cls_name),
        ]
    elif "Book Value" in method or "P-B" in method:
        equity = _parse_num(latest.get("equity_bil"))
        bvps = _per_share_bil(equity, shares)
        target_pb = assumptions.get("target_pb_bank", 1.2) if cls_name == "Financial / Bank / Insurance" else 1.0
        calc = bvps * target_pb if bvps is not None else None
        lines += [
            "CÃ¡ch tÃ­nh theo dá»¯ liá»‡u doanh nghiá»‡p:",
            f"- Vá»‘n chá»§ sá»Ÿ há»¯u ká»³ má»›i nháº¥t = {_bil(equity)}.",
            f"- Sá»‘ cá»• phiáº¿u = {_ratio(shares)} triá»‡u cp.",
            f"- BVPS = vá»‘n chá»§ sá»Ÿ há»¯u / cá»• phiáº¿u = {_money(bvps, ' Ä‘/cp')}.",
            f"- P/B má»¥c tiÃªu Ã¡p dá»¥ng cho {cls_name} = {_ratio(target_pb)}.",
            f"- GiÃ¡ trá»‹ tÃ­nh láº¡i = BVPS x P/B má»¥c tiÃªu = {_money(calc, ' Ä‘/cp')}; báº£ng Ä‘ang dÃ¹ng = {_money(intrinsic, ' Ä‘/cp')}.",
            f"- Cáº§n Ä‘á»c cÃ¹ng cháº¥t lÆ°á»£ng tÃ i sáº£n: tá»•ng tÃ i sáº£n {_bil(latest.get('total_assets_bil'))}, ná»£ pháº£i tráº£ {_bil(latest.get('liabilities_bil'))}, ROE trung vá»‹ {_pct(_recent_median2(annual, 'roe_actual_pct') or _recent_median2(annual, 'roe_pct'))}.",
            _source_principle_text("asset", cls_name),
        ]
    elif "Net Liquid" in method or "NCAV" in method:
        cash = _parse_num(latest.get("cash_equivalents_bil")) or 0.0
        sti = _parse_num(latest.get("short_term_investments_bil")) or 0.0
        recv = _parse_num(latest.get("accounts_receivable_bil")) or 0.0
        inv = _parse_num(latest.get("inventory_bil")) or 0.0
        liab = _parse_num(latest.get("liabilities_bil")) or _parse_num(latest.get("current_liabilities_bil")) or 0.0
        h_cash = assumptions.get("asset_haircut_cash_pct", 0.0)
        h_recv = assumptions.get("asset_haircut_receivables_pct", 25.0)
        h_inv = assumptions.get("asset_haircut_inventory_pct", 50.0)
        nlav = cash*(1-h_cash/100) + sti*(1-h_cash/100) + recv*(1-h_recv/100) + inv*(1-h_inv/100) - liab
        calc = _per_share_bil(nlav, shares) if nlav > 0 else None
        lines += [
            "CÃ¡ch tÃ­nh theo dá»¯ liá»‡u doanh nghiá»‡p:",
            f"- Tiá»n = {_bil(cash)}, Ä‘áº§u tÆ° ngáº¯n háº¡n = {_bil(sti)}, pháº£i thu = {_bil(recv)}, tá»“n kho = {_bil(inv)}, ná»£ pháº£i tráº£ = {_bil(liab)}.",
            f"- Haircut Ã¡p dá»¥ng: tiá»n/ÄT ngáº¯n háº¡n {h_cash:.0f}%, pháº£i thu {h_recv:.0f}%, tá»“n kho {h_inv:.0f}%.",
            f"- NLA/NCAV báº£o thá»§ = tiá»n + ÄT ngáº¯n háº¡n + pháº£i thu sau haircut + tá»“n kho sau haircut - ná»£ = {_bil(nlav)}.",
            f"- NLA/NCAV/cp = {_money(calc, ' Ä‘/cp')}; báº£ng Ä‘ang dÃ¹ng = {_money(intrinsic, ' Ä‘/cp')}.",
            f"- Náº¿u NLA Ã¢m, phÆ°Æ¡ng phÃ¡p nÃ y chá»‰ Ä‘Ã³ng vai trÃ² kiá»ƒm tra downside, khÃ´ng dÃ¹ng lÃ m fair value chÃ­nh.",
            _source_principle_text("asset", cls_name),
        ]
    else:
        lines += [
            f"Vai trÃ²: {_format_note_value(rowd.get('Vai trÃ²'))}; trá»ng sá»‘ {_pct(rowd.get('Trá»ng sá»‘ %'))}; Ä‘á»™ tin cáº­y {_format_note_value(rowd.get('Äá»™ tin cáº­y'))}.",
            f"CÆ¡ sá»Ÿ tÃ­nh trong báº£ng: {_format_note_value(rowd.get('CÆ¡ sá»Ÿ tÃ­nh'))}.",
            f"Cáº£nh bÃ¡o: {_format_note_value(rowd.get('Cáº£nh bÃ¡o'))}.",
        ]
    return "\n".join(lines)


def _valuation_range_note(rowd: dict) -> str:
    valuation = _ctx().get("valuation_df", pd.DataFrame())
    rng = _ctx().get("value_range")
    c = _ctx().get("company")
    cls = _ctx().get("classification")
    method_lines = []
    if isinstance(valuation, pd.DataFrame) and not valuation.empty:
        valid = valuation[pd.to_numeric(valuation.get("GiÃ¡ trá»‹ ná»™i táº¡i/cp"), errors="coerce").notna()].copy()
        valid = valid[pd.to_numeric(valid.get("GiÃ¡ trá»‹ ná»™i táº¡i/cp"), errors="coerce") > 0]
        for _, r in valid.iterrows():
            method_lines.append(f"- {r.get('PhÆ°Æ¡ng phÃ¡p')}: giÃ¡ trá»‹ {_money(r.get('GiÃ¡ trá»‹ ná»™i táº¡i/cp'), ' Ä‘/cp')}, trá»ng sá»‘ {_pct(r.get('Trá»ng sá»‘ %'))}, vai trÃ² {r.get('Vai trÃ²')}")
    current_price = _parse_num(getattr(c, "current_price", None)) if c is not None else None
    chosen = str(rowd.get("Chá»‰ tiÃªu", ""))
    return "\n".join([
        _company_snapshot(),
        "",
        f"Dáº¢I GIÃ TRá»Š ÄANG CHá»ŒN: {chosen} = {_money(rowd.get('GiÃ¡ trá»‹/cp'), ' Ä‘/cp')}",
        f"- Low = phÃ¢n vá»‹ 25% cá»§a cÃ¡c phÆ°Æ¡ng phÃ¡p há»£p lá»‡: {_money(getattr(rng, 'low_vnd', None), ' Ä‘/cp')}.",
        f"- Base/Median = trung vá»‹ cÃ¡c phÆ°Æ¡ng phÃ¡p há»£p lá»‡: {_money(getattr(rng, 'base_vnd', None), ' Ä‘/cp')}.",
        f"- High = phÃ¢n vá»‹ 75% cá»§a cÃ¡c phÆ°Æ¡ng phÃ¡p há»£p lá»‡: {_money(getattr(rng, 'high_vnd', None), ' Ä‘/cp')}.",
        f"- Weighted = trung bÃ¬nh trá»ng sá»‘ theo vai trÃ²/phÃ¹ há»£p vá»›i {getattr(cls, 'company_type', 'N/A')}: {_money(getattr(rng, 'weighted_vnd', None), ' Ä‘/cp')}.",
        f"- MOS hiá»‡n táº¡i = (Weighted - giÃ¡ thá»‹ trÆ°á»ng) / Weighted = {_pct(getattr(rng, 'mos_to_weighted_pct', None))}; giÃ¡ thá»‹ trÆ°á»ng {_money(current_price, ' Ä‘/cp')}.",
        f"- MOS yÃªu cáº§u Ä‘ang chá»n = {_pct(_ctx().get('target_mos_pct'))}; giÃ¡ mua tá»‘i Ä‘a theo giÃ¡ trá»‹ trá»ng sá»‘ = {_money(getattr(rng, 'weighted_vnd', None) * (1 - float(_ctx().get('target_mos_pct', 30)) / 100) if getattr(rng, 'weighted_vnd', None) else None, ' Ä‘/cp')}.",
        "CÃ¡c phÆ°Æ¡ng phÃ¡p Ä‘ang tham gia dáº£i giÃ¡ trá»‹:",
        "\n".join(method_lines) if method_lines else "- ChÆ°a cÃ³ phÆ°Æ¡ng phÃ¡p há»£p lá»‡.",
        _source_principle_text("mos", getattr(cls, 'company_type', 'N/A')),
    ])


def _moat_note(rowd: dict) -> str:
    annual = _ctx().get("annual_df", pd.DataFrame())
    cls = _ctx().get("classification")
    group = str(rowd.get("NhÃ³m Porter/Moat", ""))
    roic = _recent_median2(annual, "roic_standard_pct") or _recent_median2(annual, "roic_pct")
    roe = _recent_median2(annual, "roe_actual_pct") or _recent_median2(annual, "roe_pct")
    gross = _recent_median2(annual, "gross_margin_pct")
    ebit = _recent_median2(annual, "ebit_margin_pct")
    cfo_np = _recent_median2(annual, "cfo_to_net_profit")
    fcf_np = _recent_median2(annual, "fcf_to_net_profit")
    fcf_pos = _recent_positive_ratio2(annual, "free_cash_flow_bil")
    rev_cagr = _cagr2(annual, "revenue_bil")
    oe_cagr = _cagr2(annual, "owner_earnings_bil")
    ccc = _recent_median2(annual, "cash_conversion_cycle_days")
    inv_turn = _recent_median2(annual, "inventory_turnover")
    sga_ratio = None
    rev = _recent_median2(annual, "revenue_bil")
    if rev and rev > 0:
        sga_ratio = ((abs(_recent_median2(annual, "selling_expense_bil") or 0) + abs(_recent_median2(annual, "admin_expense_bil") or 0)) / rev * 100)
    mapping = {
        "Hiá»‡u quáº£ vá»‘n": f"ROIC trung vá»‹ {_pct(roic)} vÃ  ROE trung vá»‹ {_pct(roe)}. Äiá»ƒm cao chá»‰ há»£p lÃ½ khi lá»£i suáº¥t trÃªn vá»‘n cao láº·p láº¡i nhiá»u ká»³, khÃ´ng pháº£i má»™t nÄƒm Ä‘á»™t biáº¿n.",
        "Cost advantage": f"BiÃªn gá»™p trung vá»‹ {_pct(gross)}, biÃªn EBIT {_pct(ebit)}, CCC {_money(ccc, ' ngÃ y')}, vÃ²ng quay HTK {_ratio(inv_turn)}, SG&A/DT {_pct(sga_ratio)}. ÄÃ¢y lÃ  dáº¥u hiá»‡u cá»¥ thá»ƒ xem cÃ´ng ty cÃ³ cost advantage khÃ´ng.",
        "Differentiation": f"BiÃªn gá»™p {_pct(gross)} vÃ  má»©c duy trÃ¬ biÃªn qua nhiá»u ká»³ lÃ  tÃ­n hiá»‡u Ä‘á»‹nh lÆ°á»£ng cá»§a pricing power. Cáº§n bá»• sung BCTN/tin IR vá» thÆ°Æ¡ng hiá»‡u, khÃ¡ch hÃ ng, sáº£n pháº©m, kÃªnh phÃ¢n phá»‘i cá»§a chÃ­nh doanh nghiá»‡p.",
        "Cáº¥u trÃºc ngÃ nh": f"Há»‡ thá»‘ng hiá»‡n dÃ¹ng biáº¿n Ä‘á»™ng lá»£i nhuáº­n {_pct(None if _cv2(annual, 'net_profit_bil') is None else _cv2(annual, 'net_profit_bil')*100)} vÃ  ngÃ nh cá»§a mÃ£ Ä‘á»ƒ nháº­n diá»‡n chu ká»³. Vá»›i ngÃ nh cáº¡nh tranh/chu ká»³, Ä‘iá»ƒm cáº¥u trÃºc ngÃ nh sáº½ tháº­n trá»ng hÆ¡n.",
        "Cháº¥t lÆ°á»£ng dÃ²ng tiá»n": f"CFO/LNST {_ratio(cfo_np)}, FCF/LNST {_ratio(fcf_np)}, tá»· lá»‡ ká»³ FCF dÆ°Æ¡ng {_pct(None if fcf_pos is None else fcf_pos*100)}. Náº¿u LNST cao nhÆ°ng tiá»n khÃ´ng vá», moat/valuation pháº£i giáº£m Ä‘á»™ tin cáº­y.",
        "TÃ¡i Ä‘áº§u tÆ°": f"CAGR doanh thu {_pct(None if rev_cagr is None else rev_cagr*100)}, CAGR Owner Earnings {_pct(None if oe_cagr is None else oe_cagr*100)}, ROIC {_pct(roic)}. Compounder tháº­t cáº§n vá»«a tÄƒng trÆ°á»Ÿng vá»«a duy trÃ¬ ROIC.",
        "Quáº£n trá»‹ & phÃ¢n bá»• vá»‘n": f"Kiá»ƒm tra ROIC {_pct(roic)}, ná»£/EBITDA {_ratio(_parse_num(_latest_dict(annual).get('net_debt_to_ebitda')))}, cá»• tá»©c/capex/ná»£ vay. ÄÃ¢y lÃ  dáº¥u hiá»‡u xem quáº£n trá»‹ cÃ³ phÃ¢n bá»• vá»‘n táº¡o giÃ¡ trá»‹ hay khÃ´ng.",
        "Chuá»—i giÃ¡ trá»‹ Porter": f"DÃ¹ng dá»¯ liá»‡u váº­n hÃ nh: biÃªn gá»™p {_pct(gross)}, SG&A/DT {_pct(sga_ratio)}, CCC {_money(ccc, ' ngÃ y')}, vÃ²ng quay HTK {_ratio(inv_turn)} Ä‘á»ƒ truy nguá»“n lá»£i tháº¿ tá»« hoáº¡t Ä‘á»™ng cá»¥ thá»ƒ.",
    }
    detail = next((v for k, v in mapping.items() if k in group), rowd.get("Diá»…n giáº£i", ""))
    return "\n".join([
        _company_snapshot(),
        "",
        f"NHÃ“M ÄÃNH GIÃ: {group}",
        f"- Äiá»ƒm Ä‘áº¡t: {_format_note_value(rowd.get('Äiá»ƒm Ä‘áº¡t'))}/{_format_note_value(rowd.get('Trá»ng sá»‘ %'))}.",
        f"- TÃ­n hiá»‡u trong báº£ng: {_format_note_value(rowd.get('TÃ­n hiá»‡u'))}.",
        f"- Diá»…n giáº£i theo dá»¯ liá»‡u cá»§a mÃ£ nÃ y: {detail}",
        "Sá»‘ liá»‡u cá»¥ thá»ƒ dáº«n Ä‘áº¿n diá»…n giáº£i:",
        _module2_numeric_evidence_for_note(group),
        f"- Báº±ng chá»©ng cáº§n xem thÃªm: {_format_note_value(rowd.get('Báº±ng chá»©ng Ä‘á»‹nh lÆ°á»£ng cáº§n xem'))}.",
        _source_principle_text("porter", getattr(cls, 'company_type', 'N/A')),
    ])


def _value_chain_note(rowd: dict) -> str:
    annual = _ctx().get("annual_df", pd.DataFrame())
    cls = _ctx().get("classification")
    activity = str(rowd.get("Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹", ""))
    gross = _recent_median2(annual, "gross_margin_pct")
    ebit = _recent_median2(annual, "ebit_margin_pct")
    ccc = _recent_median2(annual, "cash_conversion_cycle_days")
    inv_turn = _recent_median2(annual, "inventory_turnover")
    roic = _recent_median2(annual, "roic_standard_pct") or _recent_median2(annual, "roic_pct")
    cfo_np = _recent_median2(annual, "cfo_to_net_profit")
    rev = _recent_median2(annual, "revenue_bil")
    sga_ratio = None
    if rev and rev > 0:
        sga_ratio = ((abs(_recent_median2(annual, "selling_expense_bil") or 0) + abs(_recent_median2(annual, "admin_expense_bil") or 0)) / rev * 100)
    activity_detail = {
        "Logistics Ä‘áº§u vÃ o": f"VÃ²ng quay tá»“n kho {_ratio(inv_turn)}. Náº¿u vÃ²ng quay cao vÃ  tá»“n kho khÃ´ng phÃ¬nh ra khi doanh thu tÄƒng, cÃ³ thá»ƒ cÃ³ lá»£i tháº¿ mua hÃ ng/quáº£n trá»‹ tá»“n kho. Náº¿u vÃ²ng quay tháº¥p, pháº£i soi tá»“n kho cháº­m luÃ¢n chuyá»ƒn vÃ  giÃ¡ nguyÃªn liá»‡u.",
        "Váº­n hÃ nh/sáº£n xuáº¥t": f"BiÃªn gá»™p {_pct(gross)}, biÃªn EBIT {_pct(ebit)}, ROIC {_pct(roic)}. Náº¿u biÃªn cao Ä‘i kÃ¨m ROIC cao nhiá»u ká»³, lá»£i tháº¿ cÃ³ thá»ƒ Ä‘áº¿n tá»« quy mÃ´, cÃ´ng nghá»‡ hoáº·c hiá»‡u suáº¥t váº­n hÃ nh.",
        "Logistics Ä‘áº§u ra": f"CCC {_money(ccc, ' ngÃ y')}, CFO/LNST {_ratio(cfo_np)}. CCC tháº¥p hoáº·c Ã¢m cho tháº¥y doanh nghiá»‡p thu tiá»n nhanh/chiáº¿m dá»¥ng vá»‘n tá»‘t; CCC cao lÃ m giáº£m cháº¥t lÆ°á»£ng dÃ²ng tiá»n.",
        "Marketing & bÃ¡n hÃ ng": f"SG&A/DT {_pct(sga_ratio)} vÃ  biÃªn gá»™p {_pct(gross)}. Náº¿u chi phÃ­ bÃ¡n hÃ ng tháº¥p nhÆ°ng biÃªn gá»™p cao, cÃ³ thá»ƒ cÃ³ thÆ°Æ¡ng hiá»‡u/kÃªnh phÃ¢n phá»‘i máº¡nh; náº¿u SG&A cao mÃ  biÃªn khÃ´ng tÄƒng, cáº§n cáº£nh bÃ¡o.",
        "Dá»‹ch vá»¥ sau bÃ¡n hÃ ng": "BCTC thÆ°á»ng khÃ´ng Ä‘á»§ dá»¯ liá»‡u Ä‘á»‹nh lÆ°á»£ng. Cáº§n BCTN/IR: báº£o hÃ nh, tá»· lá»‡ khÃ¡ch hÃ ng láº·p láº¡i, há»£p Ä‘á»“ng dÃ i háº¡n, khiáº¿u náº¡i, churn. Note nÃ y khÃ´ng káº¿t luáº­n moat náº¿u thiáº¿u báº±ng chá»©ng doanh nghiá»‡p.",
        "CÃ´ng nghá»‡/R&D": "Cáº§n dá»¯ liá»‡u BCTN/IR vá» R&D, báº±ng sÃ¡ng cháº¿, chá»©ng chá»‰, tá»± Ä‘á»™ng hÃ³a, chi phÃ­ cÃ´ng nghá»‡. Náº¿u ROIC/biÃªn gá»™p cao nhÆ°ng khÃ´ng cÃ³ báº±ng chá»©ng hoáº¡t Ä‘á»™ng, chá»‰ cháº¥m moat tháº­n trá»ng.",
        "NhÃ¢n sá»±": "Cáº§n dá»¯ liá»‡u nhÃ¢n sá»±/nÄƒng suáº¥t/Ä‘Ã o táº¡o. Náº¿u doanh thu/nhÃ¢n viÃªn, nÄƒng suáº¥t hoáº·c tá»· lá»‡ nghá»‰ viá»‡c khÃ´ng cÃ³, há»‡ thá»‘ng chá»‰ gá»£i Ã½ kiá»ƒm tra, khÃ´ng káº¿t luáº­n moat.",
        "Háº¡ táº§ng quáº£n trá»‹": f"ROIC {_pct(roic)}, CFO/LNST {_ratio(cfo_np)}, ná»£/EBITDA {_ratio(_parse_num(_latest_dict(annual).get('net_debt_to_ebitda')))}. Quáº£n trá»‹ tá»‘t pháº£i thá»ƒ hiá»‡n á»Ÿ phÃ¢n bá»• vá»‘n, kiá»ƒm soÃ¡t ná»£, minh báº¡ch vÃ  khÃ´ng lÃ m loÃ£ng cá»• Ä‘Ã´ng.",
    }
    detail = next((v for k, v in activity_detail.items() if k in activity), "Cáº§n Ä‘á»c BCTN/IR vÃ  Ä‘á»‘i chiáº¿u sá»‘ liá»‡u nhiá»u ká»³.")
    return "\n".join([
        _company_snapshot(),
        "",
        f"HOáº T Äá»˜NG CHUá»–I GIÃ TRá»Š: {activity}",
        f"- ÄÃ¡nh giÃ¡ sÆ¡ bá»™ trong báº£ng: {_format_note_value(rowd.get('ÄÃ¡nh giÃ¡ sÆ¡ bá»™'))}; loáº¡i lá»£i tháº¿: {_format_note_value(rowd.get('Loáº¡i lá»£i tháº¿'))}.",
        f"- CÃ¡ch Ä‘á»c theo dá»¯ liá»‡u doanh nghiá»‡p: {detail}",
        "Sá»‘ liá»‡u cá»¥ thá»ƒ dáº«n Ä‘áº¿n diá»…n giáº£i:",
        _module2_numeric_evidence_for_note(activity),
        f"- Báº±ng chá»©ng hiá»‡n cÃ³/cáº§n tÃ¬m: {_format_note_value(rowd.get('Báº±ng chá»©ng hiá»‡n cÃ³/cáº§n tÃ¬m'))}.",
        _source_principle_text("porter", getattr(cls, 'company_type', 'N/A')),
    ])


def _scenario_note(rowd: dict) -> str:
    annual = _ctx().get("annual_df", pd.DataFrame())
    cls = _ctx().get("classification")
    current = _parse_num(getattr(_ctx().get("company"), "current_price", None))
    value = _parse_num(rowd.get("GiÃ¡ trá»‹/cp"))
    mos = ((value - current) / value * 100) if value and current else _parse_num(rowd.get("MOS so vá»›i giÃ¡ hiá»‡n táº¡i %"))
    return "\n".join([
        _company_snapshot(),
        "",
        f"Ká»ŠCH Báº¢N: {_format_note_value(rowd.get('Ká»‹ch báº£n'))}",
        f"- GiÃ¡ trá»‹/cp cá»§a ká»‹ch báº£n = {_money(value, ' Ä‘/cp')}; giÃ¡ hiá»‡n táº¡i = {_money(current, ' Ä‘/cp')}; MOS = {_pct(mos)}.",
        f"- Giáº£ Ä‘á»‹nh chÃ­nh: {_format_note_value(rowd.get('Giáº£ Ä‘á»‹nh chÃ­nh'))}.",
        f"- Rá»§i ro Ä‘Æ°á»£c kÃ­ch hoáº¡t bá»Ÿi dá»¯ liá»‡u: {_format_note_value(rowd.get('Rá»§i ro cáº§n kiá»ƒm tra'))}.",
        f"- Sá»‘ liá»‡u rá»§i ro cá»¥ thá»ƒ: biáº¿n Ä‘á»™ng LNST {_pct(None if _cv2(annual, 'net_profit_bil') is None else _cv2(annual, 'net_profit_bil')*100)}, ná»£/EBITDA {_ratio(_parse_num(_latest_dict(annual).get('net_debt_to_ebitda')))}, FCF/LNST {_ratio(_recent_median2(annual, 'fcf_to_net_profit'))}.",
        _source_principle_text("risk", getattr(cls, 'company_type', 'N/A')),
    ])


def _latest_card_note(rowd: dict) -> str:
    return "\n".join([
        _company_snapshot(),
        "",
        f"CHá»ˆ TIÃŠU Ká»² Gáº¦N NHáº¤T: {_format_note_value(rowd.get('Chá»‰ tiÃªu'))}",
        f"- GiÃ¡ trá»‹ Ä‘ang hiá»ƒn thá»‹: {_format_note_value(rowd.get('GiÃ¡ trá»‹'))}.",
        "- Sá»‘ liá»‡u láº¥y tá»« báº£ng BCTC Ä‘Ã£ chuáº©n hÃ³a cá»§a Tá»•ng quan doanh nghiá»‡p, Æ°u tiÃªn TTM náº¿u Ä‘á»§ 4 quÃ½ gáº§n nháº¥t; náº¿u khÃ´ng cÃ³ TTM thÃ¬ dÃ¹ng nÄƒm/ká»³ má»›i nháº¥t.",
        "- CÃ¡ch Ä‘á»c khÃ´ng dÃ¹ng chung: so sÃ¡nh chá»‰ tiÃªu nÃ y vá»›i phÃ¢n loáº¡i doanh nghiá»‡p, ROIC, CFO/LNST, FCF/LNST vÃ  chu ká»³ ngÃ nh cá»§a mÃ£ Ä‘ang nháº­p. VÃ­ dá»¥ cÃ¹ng má»™t FCF Ã¢m: vá»›i doanh nghiá»‡p má»Ÿ rá»™ng ROIC cao thÃ¬ cáº§n tÃ¡ch capex tÄƒng trÆ°á»Ÿng; vá»›i doanh nghiá»‡p suy giáº£m ROIC thÃ¬ lÃ  cáº£nh bÃ¡o.",
    ])



def _median_pct_for_note(df: pd.DataFrame, cols: list[str], n: int = 5) -> float | None:
    for col in cols:
        v = _recent_median2(df, col, n=n)
        if v is not None:
            return v
    return None


def _profit_sustainability_assessment(annual_df: pd.DataFrame) -> tuple[str, str, str]:
    roic = _median_pct_for_note(annual_df, ["roic_standard_pct", "roic_pct"])
    roe = _median_pct_for_note(annual_df, ["roe_actual_pct", "roe_pct"])
    cfo_np = _recent_median2(annual_df, "cfo_to_net_profit")
    fcf_np = _recent_median2(annual_df, "fcf_to_net_profit")
    profit_pos = _recent_positive_ratio2(annual_df, "net_profit_bil")
    fcf_pos = _recent_positive_ratio2(annual_df, "free_cash_flow_bil")
    rev_cagr = _cagr2(annual_df, "revenue_bil")
    profit_cagr = _cagr2(annual_df, "net_profit_bil")
    profit_cv = _cv2(annual_df, "net_profit_bil")
    margin_cv = _cv2(annual_df, "net_margin_pct")
    net_margin = _recent_median2(annual_df, "net_margin_pct")
    facts = (
        f"ROIC trung vá»‹ {_pct(roic)}, ROE trung vá»‹ {_pct(roe)}, CFO/LNST {_ratio(cfo_np)}, FCF/LNST {_ratio(fcf_np)}, "
        f"tá»· lá»‡ LNST dÆ°Æ¡ng {_pct(None if profit_pos is None else profit_pos*100)}, tá»· lá»‡ FCF dÆ°Æ¡ng {_pct(None if fcf_pos is None else fcf_pos*100)}, "
        f"CAGR doanh thu {_pct(None if rev_cagr is None else rev_cagr*100)}, CAGR LNST {_pct(None if profit_cagr is None else profit_cagr*100)}, "
        f"Ä‘á»™ biáº¿n Ä‘á»™ng LNST {_pct(None if profit_cv is None else profit_cv*100)}, biÃªn rÃ²ng trung vá»‹ {_pct(net_margin)}."
    )
    if profit_pos is not None and profit_pos < 0.6:
        conclusion = "ChÆ°a bá»n vá»¯ng: doanh nghiá»‡p cÃ³ nhiá»u ká»³ khÃ´ng táº¡o lá»£i nhuáº­n dÆ°Æ¡ng."
    elif profit_cv is not None and profit_cv > 0.65:
        conclusion = "CÃ³ tÃ­nh chu ká»³/biáº¿n Ä‘á»™ng cao: lá»£i nhuáº­n hiá»‡n táº¡i cáº§n chuáº©n hÃ³a qua nhiá»u ká»³."
    elif cfo_np is not None and cfo_np < 0.7:
        conclusion = "Cáº§n kiá»ƒm tra: lá»£i nhuáº­n káº¿ toÃ¡n chÆ°a chuyá»ƒn hÃ³a tá»‘t thÃ nh dÃ²ng tiá»n."
    elif roic is not None and roic >= 15 and cfo_np is not None and cfo_np >= 0.8 and (profit_cv is None or profit_cv <= 0.45):
        conclusion = "KhÃ¡ bá»n vá»¯ng: ROIC cao, dÃ²ng tiá»n há»— trá»£ lá»£i nhuáº­n vÃ  biáº¿n Ä‘á»™ng lá»£i nhuáº­n khÃ´ng quÃ¡ lá»›n."
    elif profit_cv is not None and profit_cv <= 0.45 and profit_pos is not None and profit_pos >= 0.8:
        conclusion = "TÆ°Æ¡ng Ä‘á»‘i bá»n vá»¯ng nhÆ°ng cáº§n theo dÃµi thÃªm dÃ²ng tiá»n vÃ  kháº£ nÄƒng duy trÃ¬ biÃªn lá»£i nhuáº­n."
    else:
        conclusion = "ChÆ°a Ä‘á»§ cháº¯c cháº¯n: cáº§n káº¿t há»£p thÃªm BCTN, cÆ¡ cáº¥u sáº£n pháº©m, chu ká»³ ngÃ nh vÃ  báº±ng chá»©ng internet."
    principle = (
        "KhÃ´ng dÃ¹ng chung má»™t nguyÃªn táº¯c cho má»i mÃ£: náº¿u lá»£i nhuáº­n á»•n Ä‘á»‹nh + CFO/LNST tá»‘t thÃ¬ Æ°u tiÃªn earning power/Owner Earnings; "
        "náº¿u lá»£i nhuáº­n biáº¿n Ä‘á»™ng máº¡nh thÃ¬ dÃ¹ng lá»£i nhuáº­n chuáº©n hÃ³a qua chu ká»³; náº¿u lá»£i nhuáº­n khÃ´ng Ä‘i kÃ¨m tiá»n tháº­t thÃ¬ giáº£m trá»ng sá»‘ Ä‘á»‹nh giÃ¡ theo earnings."
    )
    return conclusion, facts, principle


def _infer_advantage_sources(company: object, annual_df: pd.DataFrame, moat_df: pd.DataFrame | None = None) -> tuple[str, str, str]:
    industry = f"{getattr(company, 'industry', '')} {getattr(company, 'sub_industry', '')}".lower()
    roic = _median_pct_for_note(annual_df, ["roic_standard_pct", "roic_pct"])
    roe = _median_pct_for_note(annual_df, ["roe_actual_pct", "roe_pct"])
    gross_margin = _recent_median2(annual_df, "gross_margin_pct")
    net_margin = _recent_median2(annual_df, "net_margin_pct")
    asset_turnover = _recent_median2(annual_df, "asset_turnover")
    cfo_np = _recent_median2(annual_df, "cfo_to_net_profit")
    fcf_np = _recent_median2(annual_df, "fcf_to_net_profit")
    rev_cagr = _cagr2(annual_df, "revenue_bil")
    margin_cv = _cv2(annual_df, "gross_margin_pct")
    latest = _latest_dict(annual_df)
    cash = _parse_num(latest.get("cash_bil")) or _parse_num(latest.get("cash_and_equivalents_bil"))
    debt = _parse_num(latest.get("interest_bearing_debt_bil")) or _parse_num(latest.get("debt_bil"))
    equity = _parse_num(latest.get("equity_bil"))

    sources = []
    evidence = []
    if gross_margin is not None and gross_margin >= 25 and (margin_cv is None or margin_cv <= 0.25):
        sources.append("khÃ¡c biá»‡t hÃ³a/pricing power")
        evidence.append(f"biÃªn gá»™p trung vá»‹ {_pct(gross_margin)} vÃ  biáº¿n Ä‘á»™ng biÃªn gá»™p {_pct(None if margin_cv is None else margin_cv*100)}")
    elif gross_margin is not None and gross_margin < 15 and asset_turnover is not None and asset_turnover >= 1.0:
        sources.append("hiá»‡u quáº£ chi phÃ­/vÃ²ng quay")
        evidence.append(f"biÃªn gá»™p tháº¥p {_pct(gross_margin)} nhÆ°ng vÃ²ng quay tÃ i sáº£n {_ratio(asset_turnover)}")
    if roic is not None and roic >= 15:
        sources.append("hiá»‡u quáº£ vá»‘n/kháº£ nÄƒng tÃ¡i Ä‘áº§u tÆ°")
        evidence.append(f"ROIC trung vá»‹ {_pct(roic)}")
    if cfo_np is not None and cfo_np >= 0.8 and fcf_np is not None and fcf_np >= 0:
        sources.append("mÃ´ hÃ¬nh táº¡o tiá»n tá»‘t")
        evidence.append(f"CFO/LNST {_ratio(cfo_np)}, FCF/LNST {_ratio(fcf_np)}")
    if rev_cagr is not None and rev_cagr > 0.05 and roic is not None and roic >= 12:
        sources.append("quy mÃ´/phÃ¢n phá»‘i hoáº·c nhu cáº§u thá»‹ trÆ°á»ng thuáº­n lá»£i")
        evidence.append(f"CAGR doanh thu {_pct(rev_cagr*100)} Ä‘i cÃ¹ng ROIC {_pct(roic)}")
    if debt is not None and equity is not None and equity > 0 and debt / equity < 0.5:
        sources.append("báº£ng cÃ¢n Ä‘á»‘i tháº­n trá»ng")
        evidence.append(f"ná»£ vay/vá»‘n chá»§ khoáº£ng {_ratio(debt/equity)}")
    if any(k in industry for k in ["dÆ°á»£c", "pharma", "Ä‘iá»‡n", "power", "nÆ°á»›c", "water", "cáº£ng", "port"]):
        sources.append("giáº¥y phÃ©p/tÃ i sáº£n Ä‘áº·c thÃ¹ ngÃ nh")
        evidence.append(f"ngÃ nh/phÃ¢n ngÃ nh: {getattr(company, 'industry', '')} {getattr(company, 'sub_industry', '')}")

    # KhÃ´ng kháº³ng Ä‘á»‹nh switching cost/thÆ°Æ¡ng hiá»‡u náº¿u chá»‰ cÃ³ sá»‘ tÃ i chÃ­nh.
    direct_lack = "ThÆ°Æ¡ng hiá»‡u, switching cost, giáº¥y phÃ©p Ä‘á»™c quyá»n vÃ  kÃªnh phÃ¢n phá»‘i cáº§n Ä‘á»‘i chiáº¿u thÃªm vá»›i BCTN/tin IR; há»‡ thá»‘ng khÃ´ng tá»± kháº³ng Ä‘á»‹nh náº¿u chÆ°a cÃ³ báº±ng chá»©ng Ä‘á»‹nh tÃ­nh."
    if not sources:
        sources = ["chÆ°a xÃ¡c Ä‘á»‹nh rÃµ nguá»“n moat"]
        evidence.append("chá»‰ sá»‘ hiá»‡n táº¡i chÆ°a Ä‘á»§ máº¡nh hoáº·c thiáº¿u dá»¯ liá»‡u so sÃ¡nh ngang ngÃ nh")
    # loáº¡i trÃ¹ng nhÆ°ng giá»¯ thá»© tá»±
    uniq_sources = list(dict.fromkeys(sources))
    conclusion = "; ".join(uniq_sources)
    facts = "; ".join(evidence) + ". " + direct_lack
    principle = (
        "Theo Porter, lá»£i tháº¿ cáº¡nh tranh pháº£i truy vá» hoáº¡t Ä‘á»™ng cá»¥ thá»ƒ trong chuá»—i giÃ¡ trá»‹: táº¡o chi phÃ­ tháº¥p hÆ¡n, táº¡o khÃ¡c biá»‡t hÃ³a Ä‘Æ°á»£c khÃ¡ch hÃ ng tráº£ tiá»n, hoáº·c táº¡o rÃ o cáº£n khÃ³ báº¯t chÆ°á»›c. "
        "VÃ¬ váº­y káº¿t luáº­n nÃ y thay Ä‘á»•i theo sá»‘ liá»‡u tá»«ng doanh nghiá»‡p vÃ  khÃ´ng dÃ¹ng chung má»™t checklist cá»‘ Ä‘á»‹nh."
    )
    return conclusion, facts, principle


def _roic_moat_vs_cycle_assessment(company: object, annual_df: pd.DataFrame, cls_name: str) -> tuple[str, str, str]:
    industry = f"{getattr(company, 'industry', '')} {getattr(company, 'sub_industry', '')}".lower()
    roic = _median_pct_for_note(annual_df, ["roic_standard_pct", "roic_pct"])
    # ROCE fallback: EBIT/capital employed náº¿u cÃ³; náº¿u khÃ´ng dÃ¹ng ROIC proxy nhÆ°ng nÃ³i rÃµ.
    latest = _latest_dict(annual_df)
    ebit = _parse_num(latest.get("ebit_bil")) or _parse_num(latest.get("operating_profit_bil")) or _parse_num(latest.get("pretax_profit_bil"))
    total_assets = _parse_num(latest.get("total_assets_bil"))
    current_liabilities = _parse_num(latest.get("current_liabilities_bil"))
    capital_employed = None
    roce = None
    if total_assets is not None and current_liabilities is not None:
        capital_employed = total_assets - current_liabilities
        if ebit is not None and capital_employed and capital_employed > 0:
            roce = ebit / capital_employed * 100
    cfo_np = _recent_median2(annual_df, "cfo_to_net_profit")
    fcf_np = _recent_median2(annual_df, "fcf_to_net_profit")
    profit_cv = _cv2(annual_df, "net_profit_bil")
    gross_cv = _cv2(annual_df, "gross_margin_pct")
    rev_cagr = _cagr2(annual_df, "revenue_bil")
    cyc_keywords = ["thÃ©p", "steel", "phÃ¢n bÃ³n", "fertil", "hÃ³a cháº¥t", "chemical", "cao su", "rubber", "dáº§u", "oil", "báº¥t Ä‘á»™ng sáº£n", "real estate", "commodity", "than", "coal"]
    cyclic_flag = any(k in industry for k in cyc_keywords) or (profit_cv is not None and profit_cv > 0.65)
    facts = (
        f"ROIC trung vá»‹ {_pct(roic)}, ROCE ká»³ má»›i nháº¥t {_pct(roce)}" + (f" = EBIT/Pretax {_bil(ebit)} / capital employed {_bil(capital_employed)}" if roce is not None else " (chÆ°a Ä‘á»§ dá»¯ liá»‡u EBIT/current liabilities Ä‘á»ƒ tÃ­nh ROCE riÃªng, dÃ¹ng ROIC lÃ m proxy)") +
        f"; CFO/LNST {_ratio(cfo_np)}, FCF/LNST {_ratio(fcf_np)}, biáº¿n Ä‘á»™ng LNST {_pct(None if profit_cv is None else profit_cv*100)}, biáº¿n Ä‘á»™ng biÃªn gá»™p {_pct(None if gross_cv is None else gross_cv*100)}, CAGR doanh thu {_pct(None if rev_cagr is None else rev_cagr*100)}."
    )
    if roic is None:
        conclusion = "ChÆ°a káº¿t luáº­n: thiáº¿u ROIC/ROCE nhiá»u ká»³."
    elif roic >= 15 and not cyclic_flag and (profit_cv is None or profit_cv <= 0.45) and (cfo_np is None or cfo_np >= 0.8):
        conclusion = "ROIC cao nghiÃªng vá» moat tháº­t/hiá»‡u quáº£ hoáº¡t Ä‘á»™ng bá»n vá»¯ng hÆ¡n lÃ  chu ká»³ ngáº¯n háº¡n."
    elif roic >= 15 and cyclic_flag:
        conclusion = "ROIC cao nhÆ°ng cÃ³ rá»§i ro Ä‘áº¿n tá»« chu ká»³/ngÃ nh Ä‘ang thuáº­n lá»£i; cáº§n chuáº©n hÃ³a qua chu ká»³ trÆ°á»›c khi tráº£ premium."
    elif roic >= 15 and cfo_np is not None and cfo_np < 0.7:
        conclusion = "ROIC cao nhÆ°ng dÃ²ng tiá»n chÆ°a há»— trá»£ Ä‘á»§; cáº§n kiá»ƒm tra vá»‘n lÆ°u Ä‘á»™ng, pháº£i thu, tá»“n kho vÃ  capex."
    elif roic >= 10:
        conclusion = "ROIC khÃ¡, nhÆ°ng chÆ°a Ä‘á»§ báº±ng chá»©ng Ä‘á»ƒ gá»i lÃ  moat máº¡nh."
    else:
        conclusion = "ROIC/ROCE chÆ°a cao; lá»£i tháº¿ cáº¡nh tranh náº¿u cÃ³ cáº§n chá»©ng minh báº±ng tÃ i sáº£n, giáº¥y phÃ©p hoáº·c phá»¥c há»“i chu ká»³."
    principle = (
        "Há»‡ thá»‘ng khÃ´ng máº·c Ä‘á»‹nh ROIC cao lÃ  moat. ROIC chá»‰ Ä‘Æ°á»£c xem lÃ  moat tháº­t khi duy trÃ¬ nhiá»u ká»³, Ã­t biáº¿n Ä‘á»™ng, cÃ³ dÃ²ng tiá»n há»— trá»£ vÃ  khÃ´ng chá»‰ xuáº¥t hiá»‡n Ä‘Ãºng lÃºc ngÃ nh thuáº­n lá»£i. "
        "Vá»›i doanh nghiá»‡p chu ká»³, dÃ¹ng trung vá»‹ nhiá»u ká»³ vÃ  Ä‘Ã¡nh giÃ¡ downside trÆ°á»›c."
    )
    return conclusion, facts, principle


def _mos_assessment(value_range, current_price: float | None, cls_name: str, cls_conf: float, target_mos_pct: float = 30.0) -> tuple[str, str, str]:
    mos = value_range.mos_to_weighted_pct
    target_mos_pct = 30.0 if target_mos_pct is None else float(target_mos_pct)
    buy_price = value_range.weighted_vnd * (1 - target_mos_pct / 100) if getattr(value_range, 'weighted_vnd', None) else None
    facts = f"GiÃ¡ hiá»‡n táº¡i {_money(current_price, ' Ä‘/cp')}; giÃ¡ trá»‹ tháº¥p {_money(value_range.low_vnd, ' Ä‘/cp')}; giÃ¡ trá»‹ cÆ¡ sá»Ÿ {_money(value_range.base_vnd, ' Ä‘/cp')}; giÃ¡ trá»‹ cao {_money(value_range.high_vnd, ' Ä‘/cp')}; giÃ¡ trá»‹ trá»ng sá»‘ {_money(value_range.weighted_vnd, ' Ä‘/cp')}; MOS hiá»‡n táº¡i {_pct(mos)}; MOS yÃªu cáº§u {target_mos_pct:.0f}%; giÃ¡ mua tá»‘i Ä‘a theo MOS chá»n {_money(buy_price, ' Ä‘/cp')}; Ä‘á»™ tin cáº­y phÃ¢n loáº¡i {cls_conf:,.0f}/100."
    if mos is None:
        conclusion = "ChÆ°a Ä‘á»§ dá»¯ liá»‡u Ä‘á»ƒ káº¿t luáº­n biÃªn an toÃ n."
    elif mos >= target_mos_pct:
        conclusion = f"Äáº¡t MOS yÃªu cáº§u {target_mos_pct:.0f}% so vá»›i giÃ¡ trá»‹ weighted."
    elif mos >= 30:
        conclusion = f"CÃ³ biÃªn an toÃ n Ä‘Ã¡ng chÃº Ã½ nhÆ°ng chÆ°a Ä‘áº¡t MOS yÃªu cáº§u {target_mos_pct:.0f}%."
    elif mos >= 15:
        conclusion = "CÃ³ biÃªn an toÃ n vá»«a pháº£i; phÃ¹ há»£p theo dÃµi thÃªm, chÆ°a quÃ¡ ráº» náº¿u dá»¯ liá»‡u khÃ´ng tháº­t máº¡nh."
    elif mos >= 0:
        conclusion = "BiÃªn an toÃ n má»ng; giÃ¡ khÃ´ng cÃ²n ráº» rÃµ rÃ ng so vá»›i giÃ¡ trá»‹ ná»™i táº¡i weighted."
    else:
        conclusion = "GiÃ¡ thá»‹ trÆ°á»ng cao hÆ¡n giÃ¡ trá»‹ weighted; chÆ°a cÃ³ biÃªn an toÃ n theo mÃ´ hÃ¬nh hiá»‡n táº¡i."
    if "Cyclical" in cls_name and mos is not None and mos < 50:
        conclusion += " Do cÃ³ tÃ­nh chu ká»³, nÃªn yÃªu cáº§u MOS cao hÆ¡n doanh nghiá»‡p á»•n Ä‘á»‹nh."
    if "Quality" in cls_name and mos is not None and mos >= 20:
        conclusion += " Vá»›i compounder cháº¥t lÆ°á»£ng, MOS khÃ´ng nháº¥t thiáº¿t pháº£i cá»±c sÃ¢u nhÆ°ng pháº£i xÃ¡c nháº­n Ä‘Æ°á»£c kháº£ nÄƒng tÃ¡i Ä‘áº§u tÆ° vÃ  moat."
    principle = f"Theo Graham/Li Lu, biÃªn an toÃ n lÃ  lá»›p báº£o vá»‡ khi Æ°á»›c tÃ­nh giÃ¡ trá»‹ cÃ³ thá»ƒ sai. Trong láº§n cháº¡y nÃ y, app dÃ¹ng MOS yÃªu cáº§u {target_mos_pct:.0f}% do ngÆ°á»i dÃ¹ng chá»n; cÃ¡c giÃ¡ mua MOS vÃ  káº¿t luáº­n Ä‘á»§/chÆ°a Ä‘á»§ MOS Ä‘Æ°á»£c tÃ­nh láº¡i theo má»©c nÃ y."
    return conclusion, facts, principle


def _build_strategic_assessment_table(company: object, annual_df: pd.DataFrame, cls, value_range, moat_df: pd.DataFrame, target_mos_pct: float = 30.0) -> pd.DataFrame:
    cls_name = getattr(cls, "company_type", "N/A")
    cls_conf = float(getattr(cls, "confidence", 0) or 0)
    current_price = _parse_num(getattr(company, "current_price", None))
    profit_c, profit_f, profit_p = _profit_sustainability_assessment(annual_df)
    adv_c, adv_f, adv_p = _infer_advantage_sources(company, annual_df, moat_df)
    roic_c, roic_f, roic_p = _roic_moat_vs_cycle_assessment(company, annual_df, cls_name)
    mos_c, mos_f, mos_p = _mos_assessment(value_range, current_price, cls_name, cls_conf, target_mos_pct)
    cls_facts = "; ".join(getattr(cls, "reasons", [])[:5]) or _company_snapshot()
    cls_principle = "PhÃ¢n loáº¡i dÃ¹ng dá»¯ liá»‡u chÃ­nh cá»§a mÃ£ Ä‘ang phÃ¢n tÃ­ch: ngÃ nh/phÃ¢n ngÃ nh, ROIC/ROE, CFO/LNST, FCF, CAGR doanh thu, Ä‘á»™ biáº¿n Ä‘á»™ng lá»£i nhuáº­n, P/B vÃ  tÃ i sáº£n ngáº¯n háº¡n rÃ²ng. Do Ä‘Ã³ má»—i doanh nghiá»‡p sáº½ ra káº¿t luáº­n khÃ¡c nhau."
    return pd.DataFrame([
        {"CÃ¢u há»i Ä‘Ã¡nh giÃ¡": "1. Doanh nghiá»‡p thuá»™c loáº¡i nÃ o?", "Káº¿t luáº­n theo mÃ£": f"{cls_name} (Ä‘á»™ tin cáº­y {cls_conf:.0f}/100)", "Sá»‘ liá»‡u/chá»©ng cá»© chÃ­nh": cls_facts, "NguyÃªn táº¯c Ã¡p dá»¥ng riÃªng": cls_principle},
        {"CÃ¢u há»i Ä‘Ã¡nh giÃ¡": "2. Lá»£i nhuáº­n hiá»‡n táº¡i cÃ³ bá»n vá»¯ng khÃ´ng?", "Káº¿t luáº­n theo mÃ£": profit_c, "Sá»‘ liá»‡u/chá»©ng cá»© chÃ­nh": profit_f, "NguyÃªn táº¯c Ã¡p dá»¥ng riÃªng": profit_p},
        {"CÃ¢u há»i Ä‘Ã¡nh giÃ¡": "3. Lá»£i tháº¿ cáº¡nh tranh Ä‘áº¿n tá»« Ä‘Ã¢u?", "Káº¿t luáº­n theo mÃ£": adv_c, "Sá»‘ liá»‡u/chá»©ng cá»© chÃ­nh": adv_f, "NguyÃªn táº¯c Ã¡p dá»¥ng riÃªng": adv_p},
        {"CÃ¢u há»i Ä‘Ã¡nh giÃ¡": "4. ROIC/ROCE cao do moat tháº­t hay chu ká»³?", "Káº¿t luáº­n theo mÃ£": roic_c, "Sá»‘ liá»‡u/chá»©ng cá»© chÃ­nh": roic_f, "NguyÃªn táº¯c Ã¡p dá»¥ng riÃªng": roic_p},
        {"CÃ¢u há»i Ä‘Ã¡nh giÃ¡": "5. GiÃ¡ hiá»‡n táº¡i cÃ³ Ä‘á»§ biÃªn an toÃ n khÃ´ng?", "Káº¿t luáº­n theo mÃ£": mos_c, "Sá»‘ liá»‡u/chá»©ng cá»© chÃ­nh": mos_f, "NguyÃªn táº¯c Ã¡p dá»¥ng riÃªng": mos_p},
    ])


def _canonical_company_type_key(cls_text: object) -> str:
    """Map engine classification names/aliases to the guidance keys used by the UI."""
    raw = str(cls_text or "").strip()
    low = raw.lower()
    if not low:
        return "Normal Business"
    if "chÆ°a cÃ³ dá»¯ liá»‡u" in low or "khÃ´ng cÃ³ dá»¯ liá»‡u" in low:
        return "ChÆ°a cÃ³ dá»¯ liá»‡u tÃ i chÃ­nh"
    if "financial" in low or "bank" in low or "insurance" in low or "ngÃ¢n hÃ ng" in low or "báº£o hiá»ƒm" in low:
        return "Bank/Insurance"
    if "asset" in low or "deep value" in low or "net-net" in low or "ncav" in low or "nla" in low:
        return "Asset Play"
    if "quality compounder" in low:
        return "Quality Compounder"
    if "compounder" in low:
        return "Compounder"
    if "cyclical" in low or "chu ká»³" in low:
        return "Cyclical"
    if "turnaround" in low or "phá»¥c há»“i" in low:
        return "Turnaround"
    key = next((k for k in COMPANY_TYPE_GUIDANCE.keys() if raw.lower().startswith(k.lower())), None)
    if key is None:
        key = next((k for k in COMPANY_TYPE_GUIDANCE.keys() if k.lower() in low), None)
    return key or "Normal Business"


def _company_type_info(cls_text: object) -> tuple[str, dict]:
    key = _canonical_company_type_key(cls_text)
    return key, COMPANY_TYPE_GUIDANCE.get(key, COMPANY_TYPE_GUIDANCE["Normal Business"])


def _company_type_guidance_for_note(cls_text: str) -> str:
    key, info = _company_type_info(cls_text)
    return "\n".join([
        f"DIá»„N GIáº¢I THEO LOáº I DOANH NGHIá»†P: {key}",
        f"- CÆ¡ sá»Ÿ tÆ° duy: {info.get('CÆ¡ sá»Ÿ tÆ° duy', 'N/A')}",
        f"- Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra: {info.get('Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra', 'N/A')}",
        f"- Cáº§n phÃ¢n tÃ­ch thÃªm: {info.get('Cáº§n phÃ¢n tÃ­ch thÃªm', 'N/A')}",
        f"- Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn: {info.get('Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn', 'N/A')}",
    ])


def _strategic_assessment_note(rowd: dict) -> str:
    question = str(rowd.get("CÃ¢u há»i Ä‘Ã¡nh giÃ¡", "ÄÃ¡nh giÃ¡ trá»ng yáº¿u"))
    conclusion = str(rowd.get('Káº¿t luáº­n theo mÃ£', 'N/A'))
    extra_type_guidance = ""
    if "Doanh nghiá»‡p thuá»™c loáº¡i nÃ o" in question:
        extra_type_guidance = "\n\n" + _company_type_guidance_for_note(conclusion)
    return "\n".join([
        _company_snapshot(),
        "",
        question,
        f"Káº¿t luáº­n theo mÃ£: {conclusion}",
        "",
        "Sá»‘ liá»‡u/chá»©ng cá»© chÃ­nh:",
        str(rowd.get("Sá»‘ liá»‡u/chá»©ng cá»© chÃ­nh", "N/A")),
        "",
        "NguyÃªn táº¯c Ã¡p dá»¥ng riÃªng cho doanh nghiá»‡p nÃ y:",
        str(rowd.get("NguyÃªn táº¯c Ã¡p dá»¥ng riÃªng", "N/A")) + extra_type_guidance,
        "",
        "Sá»‘ liá»‡u cá»¥ thá»ƒ bá»• sung tá»« chuá»—i BCTC:",
        _module2_numeric_evidence_for_note(question),
        "",
        "CÃ¡ch ra káº¿t luáº­n: app láº¥y dá»¯ liá»‡u Tá»•ng quan doanh nghiá»‡p Ä‘Ã£ chuáº©n hÃ³a, Ä‘á»c nhÃ³m chá»‰ tiÃªu liÃªn quan Ä‘áº¿n cÃ¢u há»i nÃ y, so vá»›i ngÆ°á»¡ng trong engine, sau Ä‘Ã³ káº¿t há»£p phÃ¢n loáº¡i doanh nghiá»‡p vÃ  MOS yÃªu cáº§u hiá»‡n táº¡i. Káº¿t luáº­n khÃ´ng dÃ¹ng má»™t nguyÃªn táº¯c chung mÃ  phá»¥ thuá»™c trá»±c tiáº¿p vÃ o ROIC/ROE, CFO/LNST, FCF/LNST, CAGR doanh thu, Ä‘á»™ biáº¿n Ä‘á»™ng LNST, ná»£ vay, WACC vÃ  MOS cá»§a chÃ­nh mÃ£ Ä‘ang xem.",
        "LÆ°u Ã½: Ä‘Ã¢y lÃ  Ä‘Ã¡nh giÃ¡ tá»± Ä‘á»™ng dá»±a trÃªn dá»¯ liá»‡u Ä‘ang cÃ³ trong tá»«ng pháº§n 1 + Äá»‹nh giÃ¡ chuyÃªn sÃ¢u. Khi evidence internet/BCTN Ä‘Æ°á»£c cáº­p nháº­t, pháº§n moat/nguá»“n lá»£i tháº¿ cáº§n Ä‘Æ°á»£c Ä‘á»‘i chiáº¿u láº¡i vá»›i báº±ng chá»©ng Ä‘á»‹nh tÃ­nh."
    ])



def _beneish_note(rowd: dict) -> str:
    c = _ctx().get("company")
    annual = _ctx().get("annual_df", pd.DataFrame())
    latest = _latest_dict(annual)
    period = str(rowd.get("Ká»³", "N/A"))
    mscore = _parse_num(rowd.get("M-Score"))
    dsri = _parse_num(rowd.get("DSRI"))
    gmi = _parse_num(rowd.get("GMI"))
    aqi = _parse_num(rowd.get("AQI"))
    sgi = _parse_num(rowd.get("SGI"))
    depi = _parse_num(rowd.get("DEPI"))
    sgai = _parse_num(rowd.get("SGAI"))
    tata = _parse_num(rowd.get("TATA"))
    lvgi = _parse_num(rowd.get("LVGI"))
    return "\n".join([
        _company_snapshot(),
        "",
        f"BENEISH M-SCORE - Ká»² ÄANG CHá»ŒN: {period}",
        f"M-Score: {_format_note_value(mscore)} | NgÆ°á»¡ng cáº£nh bÃ¡o: -2.22 | Má»©c cáº£nh bÃ¡o: {rowd.get('Má»©c cáº£nh bÃ¡o', 'N/A')}",
        "",
        "CÃ´ng thá»©c sá»­ dá»¥ng:",
        "M = -4.84 + 0.920Ã—DSRI + 0.528Ã—GMI + 0.404Ã—AQI + 0.892Ã—SGI + 0.115Ã—DEPI - 0.172Ã—SGAI + 4.679Ã—TATA - 0.327Ã—LVGI.",
        "Náº¿u M-Score > -2.22, mÃ´ hÃ¬nh gáº¯n cá» rá»§i ro thao tÃºng lá»£i nhuáº­n. ÄÃ¢y lÃ  cáº£nh bÃ¡o Ä‘á»‹nh lÆ°á»£ng, khÃ´ng pháº£i káº¿t luáº­n phÃ¡p lÃ½ vá» gian láº­n.",
        "",
        "8 biáº¿n Ä‘áº§u vÃ o vÃ  Ã½ nghÄ©a theo dá»¯ liá»‡u doanh nghiá»‡p:",
        f"- DSRI {_format_note_value(dsri)}: pháº£i thu/doanh thu ká»³ hiá»‡n táº¡i so vá»›i ká»³ trÆ°á»›c. DSRI cao cÃ³ thá»ƒ bÃ¡o hiá»‡u doanh thu ghi nháº­n lá»ng hoáº·c thu tiá»n cháº­m.",
        f"- GMI {_format_note_value(gmi)}: biÃªn gá»™p ká»³ trÆ°á»›c / biÃªn gá»™p ká»³ hiá»‡n táº¡i. GMI > 1 nghÄ©a lÃ  biÃªn gá»™p suy giáº£m, tÄƒng Ã¡p lá»±c lÃ m Ä‘áº¹p lá»£i nhuáº­n.",
        f"- AQI {_format_note_value(aqi)}: tá»· trá»ng tÃ i sáº£n cháº¥t lÆ°á»£ng tháº¥p/chi phÃ­ hoÃ£n láº¡i tÄƒng. AQI > 1 cáº§n soi tÃ i sáº£n dÃ i háº¡n, chi phÃ­ vá»‘n hÃ³a, khoáº£n pháº£i thu/tá»“n kho.",
        f"- SGI {_format_note_value(sgi)}: tÄƒng trÆ°á»Ÿng doanh thu. TÄƒng trÆ°á»Ÿng cao cÃ³ thá»ƒ táº¡o Ã¡p lá»±c duy trÃ¬ ká»³ vá»ng.",
        f"- DEPI {_format_note_value(depi)}: tá»· lá»‡ kháº¥u hao giáº£m hay khÃ´ng. DEPI > 1 cáº§n kiá»ƒm tra thay Ä‘á»•i thá»i gian há»¯u dá»¥ng/phÆ°Æ¡ng phÃ¡p kháº¥u hao.",
        f"- SGAI {_format_note_value(sgai)}: chi phÃ­ bÃ¡n hÃ ng & quáº£n lÃ½/doanh thu. SGAI > 1 pháº£n Ã¡nh chi phÃ­ váº­n hÃ nh tÄƒng nhanh hÆ¡n doanh thu.",
        f"- TATA {_format_note_value(tata)}: tá»•ng accruals/tá»•ng tÃ i sáº£n. TATA dÆ°Æ¡ng cao nghÄ©a lÃ  lá»£i nhuáº­n phá»¥ thuá»™c accruals nhiá»u hÆ¡n tiá»n tháº­t.",
        f"- LVGI {_format_note_value(lvgi)}: Ä‘Ã²n báº©y ká»³ hiá»‡n táº¡i so vá»›i ká»³ trÆ°á»›c. LVGI > 1 cÃ³ thá»ƒ tÄƒng Ä‘á»™ng cÆ¡ Ä‘Ã¡p á»©ng covenant hoáº·c má»¥c tiÃªu ná»£.",
        "",
        "Biáº¿n ná»•i báº­t/cáº§n kiá»ƒm tra:",
        str(rowd.get("Biáº¿n ná»•i báº­t/cáº§n kiá»ƒm tra", "N/A")),
        "",
        "Biáº¿n thiáº¿u/cáº§n kiá»ƒm tra:",
        str(rowd.get("Biáº¿n thiáº¿u/cáº§n kiá»ƒm tra", "N/A")),
        "",
        "Sá»‘ liá»‡u ká»³ má»›i nháº¥t trong app Ä‘á»ƒ Ä‘á»‘i chiáº¿u cháº¥t lÆ°á»£ng lá»£i nhuáº­n:",
        f"- Doanh thu: {_bil(latest.get('revenue_bil'))}; LNST: {_bil(latest.get('net_profit_bil'))}; CFO: {_bil(latest.get('cfo_bil'))}; CFO/LNST: {_ratio(latest.get('cfo_to_net_profit'))}.",
        f"- Pháº£i thu: {_bil(latest.get('accounts_receivable_bil'))}; Tá»“n kho: {_bil(latest.get('inventory_bil'))}; Tá»•ng tÃ i sáº£n: {_bil(latest.get('total_assets_bil'))}; Ná»£ pháº£i tráº£: {_bil(latest.get('liabilities_bil'))}.",
        "",
        "CÃ¡ch dÃ¹ng trong Ä‘áº§u tÆ° giÃ¡ trá»‹: náº¿u Beneish cáº£nh bÃ¡o cao, app khÃ´ng káº¿t luáº­n doanh nghiá»‡p gian láº­n; thay vÃ o Ä‘Ã³ giáº£m Ä‘á»™ tin cáº­y cá»§a lá»£i nhuáº­n/Ä‘á»‹nh giÃ¡, yÃªu cáº§u Ä‘á»c BCTC kiá»ƒm toÃ¡n, thuyáº¿t minh doanh thu, pháº£i thu, tá»“n kho, kháº¥u hao, giao dá»‹ch bÃªn liÃªn quan vÃ  so sÃ¡nh vá»›i dÃ²ng tiá»n."
    ])



def _financial_manipulation_layer_note(rowd: dict, layer_name: str) -> str:
    """Detailed notes for financial manipulation layers 2-4."""
    lines = [
        _company_snapshot(),
        "",
        f"{layer_name.upper()} - Ká»² ÄANG CHá»ŒN: {rowd.get('Ká»³', 'N/A')}",
        f"Má»©c cáº£nh bÃ¡o: {rowd.get('Má»©c cáº£nh bÃ¡o', 'N/A')} | Äiá»ƒm nhiá»‡t: {_format_note_value(rowd.get('Äiá»ƒm nhiá»‡t'))}",
        "",
        "CÃ´ng thá»©c/logic app Ä‘ang dÃ¹ng:",
        str(rowd.get("CÃ´ng thá»©c/logic", "N/A")),
        "",
        "Sá»‘ liá»‡u/cÃ¡ch tÃ­nh trÃªn dÃ²ng Ä‘ang chá»n:",
    ]
    for key, val in rowd.items():
        if key in {"Nguá»“n/logic", "CÃ´ng thá»©c/logic"}:
            continue
        lines.append(f"- {key}: {_format_note_value(val)}")
    lines += [
        "",
        "Diá»…n giáº£i káº¿t quáº£:",
        str(rowd.get("TÃ­n hiá»‡u", "N/A")),
        "",
        "Cáº§n kiá»ƒm tra sÃ¢u:",
        str(rowd.get("Cáº§n kiá»ƒm tra", "N/A")),
        "",
        "LÆ°u Ã½ sá»­ dá»¥ng:",
        "CÃ¡c mÃ´ hÃ¬nh thao tÃºng tÃ i chÃ­nh chá»‰ lÃ  cá» Ä‘á» Ä‘á»‹nh lÆ°á»£ng. App khÃ´ng káº¿t luáº­n doanh nghiá»‡p gian láº­n; káº¿t quáº£ Ä‘Æ°á»£c dÃ¹ng Ä‘á»ƒ giáº£m Ä‘á»™ tin cáº­y cá»§a lá»£i nhuáº­n káº¿ toÃ¡n, yÃªu cáº§u Ä‘á»c thuyáº¿t minh, BCTC kiá»ƒm toÃ¡n, biáº¿n Ä‘á»™ng vá»‘n lÆ°u Ä‘á»™ng, giao dá»‹ch bÃªn liÃªn quan vÃ  so sÃ¡nh vá»›i dÃ²ng tiá»n tháº­t.",
    ]
    if "Accrual" in layer_name or "Sloan" in layer_name:
        lines += [
            "",
            "NgÆ°á»¡ng tham chiáº¿u ná»™i bá»™:",
            "- Sloan accrual ratio > 7% tÃ i sáº£n bÃ¬nh quÃ¢n: cáº§n theo dÃµi; > 12%: rá»§i ro cao.",
            "- CFO/LNST < 0.8: lá»£i nhuáº­n chÆ°a chuyá»ƒn hÃ³a tá»‘t thÃ nh tiá»n; < 0.5 hoáº·c CFO Ã¢m: rá»§i ro cao hÆ¡n.",
        ]
    elif "Jones" in layer_name or "Kothari" in layer_name:
        lines += [
            "",
            "NgÆ°á»¡ng tham chiáº¿u ná»™i bá»™:",
            "- |DA| > 7% tá»•ng tÃ i sáº£n Ä‘áº§u ká»³: cáº§n theo dÃµi; |DA| > 12%: rá»§i ro cao.",
            "- DA dÆ°Æ¡ng thÆ°á»ng lÃ  accruals lÃ m tÄƒng lá»£i nhuáº­n; DA Ã¢m sÃ¢u cÃ³ thá»ƒ lÃ  big-bath/ghi nháº­n chi phÃ­ trÆ°á»›c.",
        ]
    elif "Real" in layer_name or "REM" in layer_name:
        lines += [
            "",
            "NgÆ°á»¡ng tham chiáº¿u ná»™i bá»™:",
            "- Abnormal CFO Ã¢m: nghi ngá» kÃ©o doanh thu báº±ng giáº£m giÃ¡/ná»›i tÃ­n dá»¥ng.",
            "- Abnormal PROD dÆ°Æ¡ng: nghi ngá» sáº£n xuáº¥t dÆ°/lÃ m giáº£m giÃ¡ vá»‘n Ä‘Æ¡n vá»‹/tá»“n kho tÄƒng.",
            "- Abnormal DISEXP Ã¢m: nghi ngá» cáº¯t chi phÃ­ tÃ¹y Ã½ nhÆ° quáº£ng cÃ¡o, R&D, báº£o trÃ¬ Ä‘á»ƒ nÃ¢ng lá»£i nhuáº­n ngáº¯n háº¡n.",
        ]
    return "\n".join(lines)

def _render_big_recommendation(text: str) -> None:
    """Render a highly visible recommendation block using inline CSS so it survives Streamlit CSS isolation."""
    message = html.escape(str(text or "ChÆ°a cÃ³ khuyáº¿n nghá»‹"))
    st.markdown(
        f"""
        <div style="border:3px solid #0B7F75; border-left:11px solid #F5B21B; border-radius:18px;
                    padding:16px 20px; margin:14px 0 16px 0;
                    background:linear-gradient(135deg,#FFF176 0%,#FFD54F 46%,#FFF3B0 100%);
                    box-shadow:0 10px 24px rgba(11,127,117,.18), 0 0 0 3px rgba(245,178,27,.12);">
          <div style="font-size:18px; font-weight:1000; color:#0B7F75; margin-bottom:6px; letter-spacing:-.01em;">ðŸ“Œ Cáº¢NH BÃO / KHUYáº¾N NGHá»Š Ná»”I Báº¬T</div>
          <div style="font-size:17px; font-weight:950; color:#3B2600; line-height:1.35;">{message}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _latest_financial_manipulation_layer_summary(layer_name: str, df: pd.DataFrame, metric_candidates: list[str]) -> dict:
    """One-row latest summary for a financial manipulation warning layer."""
    src = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    if src.empty:
        return {
            "Lá»›p": layer_name,
            "Ká»³": "N/A",
            "Chá»‰ tiÃªu chÃ­nh": metric_candidates[0] if metric_candidates else "N/A",
            "GiÃ¡ trá»‹": "N/A",
            "Má»©c cáº£nh bÃ¡o": "ChÆ°a Ä‘á»§ dá»¯ liá»‡u",
            "Äiá»ƒm nhiá»‡t": "N/A",
            "TÃ­n hiá»‡u": "ChÆ°a Ä‘á»§ dá»¯ liá»‡u Ä‘á»ƒ tÃ­nh lá»›p cáº£nh bÃ¡o nÃ y.",
            "Cáº§n kiá»ƒm tra": "Bá»• sung dá»¯ liá»‡u BCTC theo nÄƒm, Ä‘áº·c biá»‡t doanh thu, pháº£i thu, tÃ i sáº£n, CFO, tá»“n kho, chi phÃ­ vÃ  kháº¥u hao.",
        }
    latest = src.iloc[-1].to_dict()
    metric_name = next((m for m in metric_candidates if m in latest), metric_candidates[0] if metric_candidates else "Chá»‰ tiÃªu")
    return {
        "Lá»›p": layer_name,
        "Ká»³": latest.get("Ká»³", latest.get("period", "N/A")),
        "Chá»‰ tiÃªu chÃ­nh": metric_name,
        "GiÃ¡ trá»‹": latest.get(metric_name, "N/A"),
        "Má»©c cáº£nh bÃ¡o": latest.get("Má»©c cáº£nh bÃ¡o", "N/A"),
        "Äiá»ƒm nhiá»‡t": latest.get("Äiá»ƒm nhiá»‡t", "N/A"),
        "TÃ­n hiá»‡u": latest.get("TÃ­n hiá»‡u", latest.get("Nháº­n xÃ©t", "N/A")),
        "Cáº§n kiá»ƒm tra": latest.get("Cáº§n kiá»ƒm tra", latest.get("Biáº¿n ná»•i báº­t/cáº§n kiá»ƒm tra", latest.get("Biáº¿n thiáº¿u/cáº§n kiá»ƒm tra", "N/A"))),
    }


def _build_financial_manipulation_summary_df(
    beneish_df: pd.DataFrame,
    accrual_quality_df: pd.DataFrame,
    modified_jones_df: pd.DataFrame,
    rem_df: pd.DataFrame,
) -> pd.DataFrame:
    """Build the 4-layer summary table shown above the financial manipulation layer tabs."""
    rows = [
        _latest_financial_manipulation_layer_summary("1. Beneish M-Score", beneish_df, ["M-Score"]),
        _latest_financial_manipulation_layer_summary("2. Accrual Quality/Sloan", accrual_quality_df, ["Sloan accrual ratio", "CFO/LNST", "FCF/LNST"]),
        _latest_financial_manipulation_layer_summary("3. Modified Jones/Kothari", modified_jones_df, ["DA Modified Jones", "DA Kothari"]),
        _latest_financial_manipulation_layer_summary("4. REM - hoáº¡t Ä‘á»™ng tháº­t", rem_df, ["REM Score", "Abnormal CFO", "Abnormal PROD", "Abnormal DISEXP"]),
    ]
    return pd.DataFrame(rows)


def _financial_manipulation_summary_note(rowd: dict) -> str:
    return "\n".join([
        _company_snapshot(),
        "",
        "Tá»”NG Há»¢P THAO TÃšNG TÃ€I CHÃNH 4 Lá»šP",
        f"Lá»›p: {rowd.get('Lá»›p', 'N/A')}",
        f"Ká»³ má»›i nháº¥t: {rowd.get('Ká»³', 'N/A')}",
        f"Chá»‰ tiÃªu chÃ­nh: {rowd.get('Chá»‰ tiÃªu chÃ­nh', 'N/A')} = {_format_note_value(rowd.get('GiÃ¡ trá»‹', 'N/A'))}",
        f"Má»©c cáº£nh bÃ¡o: {rowd.get('Má»©c cáº£nh bÃ¡o', 'N/A')} | Äiá»ƒm nhiá»‡t: {_format_note_value(rowd.get('Äiá»ƒm nhiá»‡t', 'N/A'))}",
        "",
        "TÃ­n hiá»‡u:",
        str(rowd.get("TÃ­n hiá»‡u", "N/A")),
        "",
        "Cáº§n kiá»ƒm tra:",
        str(rowd.get("Cáº§n kiá»ƒm tra", "N/A")),
        "",
        "CÃ¡ch Ä‘á»c: báº£ng nÃ y gom ká»³ má»›i nháº¥t cá»§a 4 lá»›p Ä‘á»ƒ xem nhanh lá»›p nÃ o Ä‘ang phÃ¡t tÃ­n hiá»‡u máº¡nh nháº¥t. Khi má»™t hoáº·c nhiá»u lá»›p cáº£nh bÃ¡o cao, cáº§n giáº£m Ä‘á»™ tin cáº­y cá»§a lá»£i nhuáº­n káº¿ toÃ¡n vÃ  Ä‘á»‘i chiáº¿u ká»¹ CFO/LNST/FCF, thuyáº¿t minh doanh thu, pháº£i thu, tá»“n kho, kháº¥u hao, vá»‘n hÃ³a chi phÃ­, giao dá»‹ch bÃªn liÃªn quan vÃ  Ã½ kiáº¿n kiá»ƒm toÃ¡n.",
    ])

def _build_row_note(row: pd.Series, table_kind: str) -> str:
    rowd = row.to_dict()
    if "CÃ¢u há»i Ä‘Ã¡nh giÃ¡" in rowd:
        return _strategic_assessment_note(rowd)
    if "PhÆ°Æ¡ng phÃ¡p" in rowd:
        return _valuation_method_note(rowd)
    if "NhÃ³m Porter/Moat" in rowd:
        return _moat_note(rowd)
    if "Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹" in rowd:
        return _value_chain_note(rowd)
    if "Ká»‹ch báº£n" in rowd:
        return _scenario_note(rowd)
    if "M-Score" in rowd and "DSRI" in rowd:
        return _beneish_note(rowd)
    if table_kind == "financial_manipulation_summary":
        return _financial_manipulation_summary_note(rowd)
    if table_kind == "accrual_quality" or "Sloan accrual ratio" in rowd:
        return _financial_manipulation_layer_note(rowd, "Lá»›p 2 - Accrual Quality/Sloan")
    if table_kind == "modified_jones" or "DA Modified Jones" in rowd:
        return _financial_manipulation_layer_note(rowd, "Lá»›p 3 - Modified Jones/Kothari")
    if table_kind == "real_earnings_management" or "Abnormal CFO" in rowd:
        return _financial_manipulation_layer_note(rowd, "Lá»›p 4 - Real Earnings Management")
    if "Chá»‰ tiÃªu" in rowd and "GiÃ¡ trá»‹/cp" in rowd:
        return _valuation_range_note(rowd)
    if "Chá»‰ tiÃªu" in rowd and "GiÃ¡ trá»‹" in rowd:
        return _latest_card_note(rowd)
    if "MÃ£" in rowd and "Äiá»ƒm tá»•ng há»£p" in rowd:
        return _build_peer_row_note(rowd)
    return "\n".join([_company_snapshot(), "", "Dá»® LIá»†U DÃ’NG ÄANG CHá»ŒN:", "\n".join([f"- {k}: {_format_note_value(v)}" for k, v in rowd.items()])])


def _render_explainable_table(df: pd.DataFrame, table_kind: str = "", height: int = 420) -> None:
    """Báº£ng HTML cÃ³ báº¯t sá»± kiá»‡n click má»™t láº§n Ä‘á»ƒ hiá»‡n note theo dá»¯ liá»‡u doanh nghiá»‡p."""
    if df is None or df.empty:
        st.info("ChÆ°a cÃ³ dá»¯ liá»‡u.")
        return
    raw_df = df.copy()
    notes = [_build_row_note(row, table_kind) for _, row in raw_df.iterrows()]
    display_df = _vi_dataframe_for_display(raw_df)
    if "Note" in display_df.columns:
        display_df = display_df.drop(columns=["Note"])
    if table_kind in {"beneish_mscore", "accrual_quality", "modified_jones", "real_earnings_management"}:
        # V23.58: hide source/logic and redundant layer columns in the financial manipulation detail tables.
        # The layer is already visible in the sub-tab title; source/logic remains available in row notes.
        display_df = display_df.drop(columns=[c for c in ["Nguá»“n/logic", "Nguá»“n / logic", "Lá»›p"] if c in display_df.columns], errors="ignore")
    # Internal marker for highlighting the company currently under analysis in peer comparison.
    if table_kind == "peer_compare":
        drop_internal = [c for c in ["MÃ£ Ä‘ang phÃ¢n tÃ­ch", "Nguá»“n dá»¯ liá»‡u", "source", "Source", "NgÃ nh", "PhÃ¢n ngÃ nh"] if c in display_df.columns]
        if drop_internal:
            display_df = display_df.drop(columns=drop_internal)
    table_id = "tbl_" + str(abs(hash((table_kind, tuple(display_df.columns), len(display_df), APP_VERSION))))[0:10]
    full_table = table_kind == "strategic_assessment"
    if full_table:
        # V23.36: báº£ng Ä‘Ã¡nh giÃ¡ trá»ng yáº¿u chá»‰ cÃ³ vÃ i dÃ²ng nhÆ°ng pháº§n note ráº¥t dÃ i.
        # Giá»¯ báº£ng gá»n Ä‘á»ƒ khÃ´ng táº¡o khoáº£ng trá»‘ng lá»›n, Ä‘á»“ng thá»i cho note cÃ³ vÃ¹ng cuá»™n riÃªng.
        table_max_height = min(max(230, 140 + len(display_df) * 34), 300)
        wrap_css = f"max-height:{table_max_height}px; overflow:auto;"
        note_css_extra = "min-height:220px; max-height:360px; overflow-y:auto;"
        component_height = 720
    elif table_kind in {"valuation_range", "valuation_methods"}:
        # V23.38: tÄƒng 20% vÃ¹ng báº£ng vÃ  note cho Dáº£i giÃ¡ trá»‹ ná»™i táº¡i / Báº£ng Ä‘á»‹nh giÃ¡ theo phÆ°Æ¡ng phÃ¡p.
        wrap_css = f"max-height:{height}px; overflow:auto;"
        note_css_extra = "min-height:240px; max-height:432px; overflow-y:auto;"
        component_height = min(max(height + 312, 516), 1128)
    elif table_kind in {"beneish_mscore", "accrual_quality", "modified_jones", "real_earnings_management"}:
        # V23.55: thao tÃºng tÃ i chÃ­nh cÃ³ note dÃ i vÃ¬ pháº£i diá»…n giáº£i cÃ´ng thá»©c, biáº¿n Ä‘áº§u vÃ o,
        # ngÆ°á»¡ng cáº£nh bÃ¡o vÃ  cÃ¡ch dÃ¹ng. TÄƒng vÃ¹ng Ä‘á»c cho cáº£ 4 lá»›p.
        wrap_css = f"max-height:{height}px; overflow:auto;"
        note_css_extra = "min-height:430px; max-height:860px; overflow-y:auto; font-size:14px; line-height:1.62;"
        component_height = min(max(height + 660, 1000), 1660)
    else:
        wrap_css = f"max-height:{height}px; overflow:auto;"
        note_css_extra = "max-height:360px; overflow-y:auto;"
        component_height = min(max(height + 260, 430), 940)
    header_cells = []
    for c in display_df.columns:
        hcls = "summary-layer-header" if table_kind == "financial_manipulation_summary" and str(c).strip() == "Lá»›p" else ""
        header_cells.append(f"<th class='{hcls}'>{html.escape(str(c))}</th>")
    headers = "".join(header_cells)
    rows_html = []
    current_ticker_for_highlight = _safe_ticker(str(st.session_state.get("module3_base_ticker") or st.session_state.get("active_ticker") or st.session_state.get("module1_ticker") or ""))
    for i, (_, row) in enumerate(display_df.iterrows()):
        raw_rowd = raw_df.iloc[i].to_dict() if i < len(raw_df) else {}
        row_is_current = False
        if table_kind == "peer_compare":
            marker_val = raw_rowd.get("MÃ£ Ä‘ang phÃ¢n tÃ­ch")
            row_is_current = bool(marker_val is True or str(marker_val).strip().lower() in {"true", "1", "yes", "mÃ£ gá»‘c", "ma goc"})
            row_is_current = row_is_current or (_safe_ticker(str(raw_rowd.get("MÃ£", ""))) == current_ticker_for_highlight and current_ticker_for_highlight != "")
        row_class = "base-peer-row" if row_is_current else ""
        tds = []
        for c in display_df.columns:
            val = row.get(c)
            text = _format_note_value(val)
            cls = _signal_class(val) if c in {"TÃ­n hiá»‡u", "Má»©c Ä‘á»™", "Má»©c cáº£nh bÃ¡o", "TÃ¬nh tráº¡ng", "Khuyáº¿n nghá»‹", "Káº¿t luáº­n", "Káº¿t luáº­n theo mÃ£", "Moat level", "Má»©c moat", "Äá»™ tin cáº­y", "ÄÃ¡nh giÃ¡ sÆ¡ bá»™", "Loáº¡i lá»£i tháº¿", "Vai trÃ²"} else ""
            num = _parse_num(val)
            if table_kind == "peer_compare" and str(c).strip() == "Äiá»ƒm tá»•ng há»£p" and num is not None:
                if num >= 80:
                    cls = "heat-green-strong"
                elif num >= 65:
                    cls = "heat-green"
                elif num >= 50:
                    cls = "heat-yellow"
                elif num >= 35:
                    cls = "heat-orange"
                else:
                    cls = "heat-red"
            elif not cls and num is not None and any(k in str(c).lower() for k in ["giÃ¡", "mos", "Ä‘iá»ƒm", "Ä‘iá»ƒm nhiá»‡t", "trá»ng", "%", "value", "score"]):
                cls = "pos" if num > 0 else "neg" if num < 0 else ""
            tds.append(f"<td class='{cls}'>{html.escape(text)}</td>")
        rows_html.append(f"<tr class='{row_class}' data-note='{html.escape(json.dumps(notes[i], ensure_ascii=False), quote=True)}'>{''.join(tds)}</tr>")
    html_doc = f"""
    <div class='hint'>ðŸ’¡ Nháº¥p má»™t láº§n vÃ o má»™t dÃ²ng/chá»‰ tiÃªu Ä‘á»ƒ xem note giáº£i thÃ­ch theo chÃ­nh dá»¯ liá»‡u cá»§a doanh nghiá»‡p.</div>
    <div class='wrap'>
      <table id='{table_id}'>
        <thead><tr>{headers}</tr></thead>
        <tbody>{''.join(rows_html)}</tbody>
      </table>
    </div>
    <div id='{table_id}_note' class='note'>ChÆ°a chá»n chá»‰ tiÃªu. HÃ£y nháº¥p má»™t láº§n vÃ o má»™t dÃ²ng trong báº£ng.</div>
    <script>
      const table = document.getElementById('{table_id}');
      const note = document.getElementById('{table_id}_note');
      if (table && note) {{
        table.querySelectorAll('tbody tr').forEach(function(row) {{
          row.addEventListener('click', function() {{
            table.querySelectorAll('tbody tr').forEach(function(r) {{ r.classList.remove('selected'); }});
            row.classList.add('selected');
            let raw = row.getAttribute('data-note') || '""';
            let msg = '';
            try {{ msg = JSON.parse(raw); }} catch(e) {{ msg = raw; }}
            note.innerText = msg;
          }});
        }});
      }}
    </script>
    <style>
      .hint {{font-size:13px; color:#0B7F75; margin: 2px 0 8px 0; font-weight:600;}}
      .wrap {{{wrap_css} border:1px solid #e2e8f0; border-radius:12px;}}
      table {{border-collapse:collapse; width:100%; font-family: system-ui, -apple-system, Segoe UI, sans-serif; font-size:13px;}}
      th {{position: sticky; top:0; background:#EAF7F1; color:#123D3A; text-align:left; border-bottom:1px solid #e2e8f0; padding:8px; z-index:8; font-weight:950; box-shadow:0 2px 0 rgba(11,127,117,.18);}}
      td {{border-bottom:1px solid #edf2f7; padding:7px 8px; vertical-align:top; color:#123D3A;}}
      tr:hover td {{background:#F7FBF8; cursor:pointer;}}
      tr.selected td {{background:#FEF3C7 !important;}}
      tr.base-peer-row td {{background:#FFF3C4 !important; color:#064E47 !important; font-weight:900 !important; border-top:2px solid #F5B21B; border-bottom:2px solid #F5B21B;}}
      tr.base-peer-row td:first-child {{border-left:7px solid #F5B21B;}}
      td.pos {{background:rgba(16,185,129,.11); color:#064e3b; font-weight:600;}}
      td.neg {{background:rgba(239,68,68,.11); color:#7f1d1d; font-weight:600;}}
      td.sig-red-strong {{background:#FECACA !important; color:#7F1D1D !important; font-weight:900; border-left:5px solid #DC2626;}}
      td.sig-red {{background:#FEE2E2 !important; color:#991B1B !important; font-weight:800; border-left:4px solid #EF4444;}}
      td.sig-purple-strong {{background:#E9D5FF !important; color:#581C87 !important; font-weight:900; border-left:5px solid #7E22CE;}}
      td.sig-purple {{background:#F3E8FF !important; color:#6B21A8 !important; font-weight:800; border-left:4px solid #A855F7;}}
      td.sig-yellow {{background:#FEF3C7 !important; color:#92400E !important; font-weight:800; border-left:4px solid #F59E0B;}}
      td.heat-green-strong {{background:#047857 !important; color:#FFFFFF !important; font-weight:950 !important;}}
      td.heat-green {{background:#A7F3D0 !important; color:#064E3B !important; font-weight:900 !important;}}
      td.heat-yellow {{background:#FEF3C7 !important; color:#92400E !important; font-weight:900 !important;}}
      td.heat-orange {{background:#FED7AA !important; color:#9A3412 !important; font-weight:900 !important;}}
      td.heat-red {{background:#FECACA !important; color:#7F1D1D !important; font-weight:900 !important;}}
      .critical-note-label {{color:#B91C1C !important; font-weight:950 !important;}}
      .note {{white-space:pre-wrap; margin-top:10px; padding:13px 15px; border-radius:14px; background:#FFF7E6; border:1px solid rgba(245,178,27,.52); color:#5f3b00; font-size:13px; line-height:1.48; {note_css_extra}}}
    </style>
    """
    components.html(html_doc, height=component_height, scrolling=True)



GLOSSARY_TERMS = {
    "Asset Play": "Doanh nghiá»‡p cÃ³ giÃ¡ thá»‹ trÆ°á»ng tháº¥p so vá»›i tÃ i sáº£n cÃ³ thá»ƒ Ä‘á»‹nh giÃ¡/thu há»“i. Trá»ng tÃ¢m lÃ  cháº¥t lÆ°á»£ng tÃ i sáº£n, kháº£ nÄƒng thu há»“i tiá»n, tÃ i sáº£n áº©n, ná»£ tiá»m tÃ ng vÃ  downside protection.",
    "Bank/Insurance": "NgÃ¢n hÃ ng/báº£o hiá»ƒm cÃ³ BCTC Ä‘áº·c thÃ¹. KhÃ´ng nÃªn dÃ¹ng FCF/working capital nhÆ° doanh nghiá»‡p sáº£n xuáº¥t; Æ°u tiÃªn ROE, P/B, cháº¥t lÆ°á»£ng tÃ i sáº£n, NIM, CASA, NPL, dá»± phÃ²ng vÃ  biÃªn an toÃ n vá»‘n.",
    "Base/Median": "GiÃ¡ trá»‹ cÆ¡ sá»Ÿ/trung vá»‹ cá»§a cÃ¡c phÆ°Æ¡ng phÃ¡p Ä‘á»‹nh giÃ¡ há»£p lá»‡. DÃ¹ng Ä‘á»ƒ trÃ¡nh phá»¥ thuá»™c vÃ o má»™t mÃ´ hÃ¬nh duy nháº¥t.",
    "Beta": "Má»©c nháº¡y cá»§a cá»• phiáº¿u/doanh nghiá»‡p vá»›i thá»‹ trÆ°á»ng hoáº·c proxy rá»§i ro. Náº¿u thiáº¿u dá»¯ liá»‡u giÃ¡ lá»‹ch sá»­, app chá»‰ dÃ¹ng beta proxy vÃ  pháº£i ghi rÃµ cháº¥t lÆ°á»£ng tháº¥p hÆ¡n beta thá»‹ trÆ°á»ng.",
    "BVPS": "Book Value per Share - giÃ¡ trá»‹ sá»• sÃ¡ch trÃªn má»—i cá»• phiáº¿u = vá»‘n chá»§ sá»Ÿ há»¯u / sá»‘ cá»• phiáº¿u.",
    "Capex": "Capital Expenditure - chi Ä‘áº§u tÆ° tÃ i sáº£n cá»‘ Ä‘á»‹nh/capex. Capex cÃ³ thá»ƒ lÃ  duy trÃ¬ hoáº·c má»Ÿ rá»™ng; Ä‘á»‹nh giÃ¡ Owner Earnings cáº§n Æ°á»›c tÃ­nh capex duy trÃ¬.",
    "Capital Employed": "Vá»‘n sá»­ dá»¥ng cho hoáº¡t Ä‘á»™ng. CÃ³ thá»ƒ tÃ­nh báº±ng tá»•ng tÃ i sáº£n - ná»£ ngáº¯n háº¡n, hoáº·c tÃ i sáº£n cá»‘ Ä‘á»‹nh + vá»‘n lÆ°u Ä‘á»™ng; pháº£i dÃ¹ng nháº¥t quÃ¡n.",
    "CCC": "Cash Conversion Cycle - chu ká»³ chuyá»ƒn Ä‘á»•i tiá»n = DSO + DIO - DPO. CCC tháº¥p thÆ°á»ng tá»‘t hÆ¡n, nhÆ°ng pháº£i Ä‘á»c theo mÃ´ hÃ¬nh kinh doanh.",
    "CFO": "Cash Flow from Operations - dÃ²ng tiá»n thuáº§n tá»« hoáº¡t Ä‘á»™ng kinh doanh. CFO/LNST cao thÆ°á»ng cho tháº¥y lá»£i nhuáº­n káº¿ toÃ¡n chuyá»ƒn hÃ³a tá»‘t thÃ nh tiá»n.",
    "CFO/LNST": "Tá»· lá»‡ CFO trÃªn lá»£i nhuáº­n sau thuáº¿. Náº¿u tháº¥p kÃ©o dÃ i, lá»£i nhuáº­n káº¿ toÃ¡n cÃ³ thá»ƒ chÆ°a chuyá»ƒn hÃ³a thÃ nh tiá»n.",
    "Beneish M-Score": "MÃ´ hÃ¬nh 8 biáº¿n dÃ¹ng Ä‘á»ƒ cáº£nh bÃ¡o kháº£ nÄƒng thao tÃºng lá»£i nhuáº­n. M-Score > -2.22 lÃ  vÃ¹ng cáº£nh bÃ¡o; Ä‘Ã¢y lÃ  tÃ­n hiá»‡u Ä‘á»‹nh lÆ°á»£ng, khÃ´ng pháº£i káº¿t luáº­n phÃ¡p lÃ½ vá» gian láº­n.",
    "Earnings Management": "Quáº£n trá»‹/thao tÃºng lá»£i nhuáº­n - viá»‡c ban Ä‘iá»u hÃ nh sá»­ dá»¥ng xÃ©t Ä‘oÃ¡n káº¿ toÃ¡n, Æ°á»›c tÃ­nh hoáº·c cáº¥u trÃºc giao dá»‹ch Ä‘á»ƒ lÃ m lá»£i nhuáº­n trÃ¬nh bÃ y khÃ¡c vá»›i cháº¥t lÆ°á»£ng kinh táº¿ thá»±c. Trong app chá»‰ gá»i lÃ  cá» Ä‘á»/cáº§n kiá»ƒm tra, khÃ´ng káº¿t luáº­n gian láº­n náº¿u chÆ°a cÃ³ báº±ng chá»©ng phÃ¡p lÃ½ hoáº·c kiá»ƒm toÃ¡n.",
    "Thao tÃºng tÃ i chÃ­nh": "NhÃ³m dáº¥u hiá»‡u cho tháº¥y sá»‘ liá»‡u tÃ i chÃ­nh cÃ³ thá»ƒ bá»‹ lÃ m Ä‘áº¹p hoáº·c trÃ¬nh bÃ y chÆ°a pháº£n Ã¡nh Ä‘Ãºng báº£n cháº¥t kinh táº¿, vÃ­ dá»¥ doanh thu ghi nháº­n sá»›m, pháº£i thu tÄƒng nhanh, vá»‘n hÃ³a chi phÃ­, thay Ä‘á»•i kháº¥u hao, hoÃ n nháº­p dá»± phÃ²ng hoáº·c lá»£i nhuáº­n khÃ´ng Ä‘i kÃ¨m dÃ²ng tiá»n.",
    "Thao tÃºng lá»£i nhuáº­n": "Má»™t dáº¡ng thao tÃºng tÃ i chÃ­nh táº­p trung vÃ o chá»‰ tiÃªu lá»£i nhuáº­n. Dáº¥u hiá»‡u thÆ°á»ng gáº·p lÃ  LNST tÄƒng nhÆ°ng CFO/FCF yáº¿u, accruals cao, biÃªn lá»£i nhuáº­n báº¥t thÆ°á»ng, thay Ä‘á»•i chÃ­nh sÃ¡ch káº¿ toÃ¡n hoáº·c nhiá»u khoáº£n má»¥c Æ°á»›c tÃ­nh.",
    "Accruals": "CÃ¡c khoáº£n dá»“n tÃ­ch káº¿ toÃ¡n lÃ m lá»£i nhuáº­n khÃ¡c dÃ²ng tiá»n. Accruals cao khÃ´ng máº·c nhiÃªn xáº¥u, nhÆ°ng náº¿u kÃ©o dÃ i hoáº·c tÄƒng Ä‘á»™t biáº¿n thÃ¬ cáº§n kiá»ƒm tra cháº¥t lÆ°á»£ng lá»£i nhuáº­n.",
    "Accrual-based Earnings Management (AEM)": "Quáº£n trá»‹ lá»£i nhuáº­n qua accruals: dÃ¹ng xÃ©t Ä‘oÃ¡n káº¿ toÃ¡n, Æ°á»›c tÃ­nh, dá»± phÃ²ng, ghi nháº­n doanh thu/chi phÃ­ Ä‘á»ƒ Ä‘iá»u chá»‰nh lá»£i nhuáº­n káº¿ toÃ¡n mÃ  chÆ°a nháº¥t thiáº¿t lÃ m thay Ä‘á»•i dÃ²ng tiá»n ngay.",
    "Real Earnings Management (REM)": "Quáº£n trá»‹ lá»£i nhuáº­n qua hoáº¡t Ä‘á»™ng tháº­t: thay Ä‘á»•i quyáº¿t Ä‘á»‹nh kinh doanh nhÆ° giáº£m giÃ¡/ná»›i tÃ­n dá»¥ng Ä‘á»ƒ kÃ©o doanh thu, sáº£n xuáº¥t dÆ° Ä‘á»ƒ giáº£m giÃ¡ vá»‘n Ä‘Æ¡n vá»‹, hoáº·c cáº¯t chi phÃ­ quáº£ng cÃ¡o/R&D/báº£o trÃ¬ Ä‘á»ƒ lÃ m Ä‘áº¹p lá»£i nhuáº­n ngáº¯n háº¡n.",
    "Sloan accrual ratio": "Chá»‰ tiÃªu cháº¥t lÆ°á»£ng lá»£i nhuáº­n = (LNST - CFO) / Tá»•ng tÃ i sáº£n bÃ¬nh quÃ¢n. Tá»· lá»‡ dÆ°Æ¡ng cao cho tháº¥y lá»£i nhuáº­n phá»¥ thuá»™c nhiá»u vÃ o accruals hÆ¡n dÃ²ng tiá»n tháº­t.",
    "Discretionary Accruals (DA)": "Pháº§n accruals báº¥t thÆ°á»ng do mÃ´ hÃ¬nh Æ°á»›c lÆ°á»£ng khÃ´ng giáº£i thÃ­ch Ä‘Æ°á»£c bá»Ÿi tÄƒng trÆ°á»Ÿng doanh thu, thay Ä‘á»•i pháº£i thu, PPE/TSCÄ vÃ  hiá»‡u quáº£ hoáº¡t Ä‘á»™ng. DA dÆ°Æ¡ng cao thÆ°á»ng lÃ  cá» Ä‘á» lÃ m tÄƒng lá»£i nhuáº­n.",
    "Modified Jones Model": "MÃ´ hÃ¬nh Æ°á»›c lÆ°á»£ng discretionary accruals: TA/A(t-1)=Î±0+Î±1(1/A(t-1))+Î±2((Î”REV-Î”REC)/A(t-1))+Î±3(PPE/A(t-1))+Îµ. Residual Îµ lÃ  DA.",
    "Kothari Model": "Báº£n má»Ÿ rá»™ng cá»§a Modified Jones, thÃªm ROA Ä‘á»ƒ kiá»ƒm soÃ¡t hiá»‡u quáº£ hoáº¡t Ä‘á»™ng. Má»¥c tiÃªu lÃ  trÃ¡nh gáº¯n cá» sai cÃ¡c doanh nghiá»‡p tÄƒng trÆ°á»Ÿng/hiá»‡u quáº£ cao nhÆ°ng accruals tÄƒng do kinh doanh tháº­t.",
    "Abnormal CFO": "Pháº§n CFO báº¥t thÆ°á»ng sau khi kiá»ƒm soÃ¡t doanh thu vÃ  tÄƒng trÆ°á»Ÿng doanh thu. Abnormal CFO Ã¢m cÃ³ thá»ƒ bÃ¡o hiá»‡u kÃ©o doanh thu báº±ng giáº£m giÃ¡ hoáº·c ná»›i Ä‘iá»u kiá»‡n tÃ­n dá»¥ng.",
    "Abnormal PROD": "Chi phÃ­ sáº£n xuáº¥t báº¥t thÆ°á»ng. Trong app, PROD = GiÃ¡ vá»‘n hÃ ng bÃ¡n + Î”Tá»“n kho. Abnormal PROD dÆ°Æ¡ng cÃ³ thá»ƒ bÃ¡o hiá»‡u sáº£n xuáº¥t dÆ°/tá»“n kho tÄƒng Ä‘á»ƒ giáº£m giÃ¡ vá»‘n Ä‘Æ¡n vá»‹.",
    "Abnormal DISEXP": "Chi phÃ­ tÃ¹y Ã½ báº¥t thÆ°á»ng. DISEXP thÆ°á»ng gá»“m R&D, quáº£ng cÃ¡o, SG&A; app dÃ¹ng SG&A/chi phÃ­ bÃ¡n hÃ ng + quáº£n lÃ½ lÃ m proxy náº¿u thiáº¿u chi tiáº¿t. Abnormal DISEXP Ã¢m cÃ³ thá»ƒ lÃ  cáº¯t chi phÃ­ Ä‘á»ƒ nÃ¢ng lá»£i nhuáº­n ngáº¯n háº¡n.",
    "AQI proxy": "CÃ¡ch tÃ­nh thay tháº¿ khi nguá»“n dá»¯ liá»‡u khÃ´ng tÃ¡ch Ä‘Æ°á»£c TSCÄ/PPE thuáº§n. App dÃ¹ng tá»· trá»ng tÃ i sáº£n dÃ i háº¡n/tá»•ng tÃ i sáº£n lÃ m Ä‘áº¡i diá»‡n vÃ  ghi rÃµ proxy Ä‘á»ƒ trÃ¡nh nháº§m vá»›i AQI chuáº©n cá»§a Beneish.",
    "PP&E/TSCÄ thuáº§n": "Property, Plant and Equipment - tÃ i sáº£n cá»‘ Ä‘á»‹nh há»¯u hÃ¬nh thuáº§n sau kháº¥u hao. ÄÃ¢y lÃ  biáº¿n Ä‘áº§u vÃ o tá»‘t hÆ¡n tÃ i sáº£n dÃ i háº¡n khi tÃ­nh AQI chuáº©n.",
    "Doanh thu ghi nháº­n sá»›m": "Rá»§i ro doanh nghiá»‡p ghi nháº­n doanh thu trÆ°á»›c khi hoÃ n táº¥t nghÄ©a vá»¥ hoáº·c trÆ°á»›c khi kháº£ nÄƒng thu tiá»n Ä‘á»§ cháº¯c cháº¯n. Trong app thÆ°á»ng thá»ƒ hiá»‡n qua DSRI/DSO tÄƒng, pháº£i thu tÄƒng nhanh hÆ¡n doanh thu vÃ  CFO yáº¿u.",
    "Cháº¥t lÆ°á»£ng lá»£i nhuáº­n": "Má»©c Ä‘á»™ lá»£i nhuáº­n káº¿ toÃ¡n Ä‘Æ°á»£c há»— trá»£ bá»Ÿi dÃ²ng tiá»n tháº­t, biÃªn lá»£i nhuáº­n bá»n vá»¯ng, chÃ­nh sÃ¡ch káº¿ toÃ¡n tháº­n trá»ng vÃ  Ã­t phá»¥ thuá»™c khoáº£n má»¥c má»™t láº§n/Æ°á»›c tÃ­nh chá»§ quan.",
    "BÃºt toÃ¡n cuá»‘i ká»³": "CÃ¡c bÃºt toÃ¡n Ä‘iá»u chá»‰nh gáº§n cuá»‘i ká»³ hoáº·c sau ngÃ y khÃ³a sá»•. Náº¿u cÃ³ giÃ¡ trá»‹ lá»›n, láº·p láº¡i báº¥t thÆ°á»ng hoáº·c thiáº¿u giáº£i thÃ­ch, cáº§n xem xÃ©t rá»§i ro lÃ m Ä‘áº¹p lá»£i nhuáº­n.",
    "Management Override": "Rá»§i ro ban Ä‘iá»u hÃ nh vÆ°á»£t qua kiá»ƒm soÃ¡t ná»™i bá»™ Ä‘á»ƒ Ä‘iá»u chá»‰nh sá»‘ liá»‡u káº¿ toÃ¡n, Æ°á»›c tÃ­nh hoáº·c bÃºt toÃ¡n. ÄÃ¢y lÃ  nhÃ³m rá»§i ro kiá»ƒm toÃ¡n trá»ng yáº¿u vÃ  cáº§n Ä‘á»‘i chiáº¿u vá»›i kiá»ƒm soÃ¡t ná»™i bá»™/Ã½ kiáº¿n kiá»ƒm toÃ¡n.",
    "Revenue Recognition": "Ghi nháº­n doanh thu. ÄÃ¢y lÃ  khu vá»±c dá»… phÃ¡t sinh rá»§i ro thao tÃºng lá»£i nhuáº­n, cáº§n kiá»ƒm tra Ä‘iá»u kiá»‡n ghi nháº­n, cutoff, pháº£i thu, hoÃ n tráº£, chiáº¿t kháº¥u, bÃªn liÃªn quan vÃ  dÃ²ng tiá»n thu tá»« khÃ¡ch hÃ ng.",
    "Cutoff": "Kiá»ƒm tra doanh thu/chi phÃ­ cÃ³ Ä‘Æ°á»£c ghi nháº­n Ä‘Ãºng ká»³ hay khÃ´ng. Sai cutoff cÃ³ thá»ƒ lÃ m doanh thu/lá»£i nhuáº­n ká»³ hiá»‡n táº¡i bá»‹ Ä‘áº©y lÃªn hoáº·c Ä‘áº©y xuá»‘ng khÃ´ng Ä‘Ãºng báº£n cháº¥t.",
    "Capitalized Expense": "Chi phÃ­ Ä‘Æ°á»£c vá»‘n hÃ³a thÃ nh tÃ i sáº£n thay vÃ¬ ghi nháº­n vÃ o chi phÃ­ ká»³ hiá»‡n táº¡i. Náº¿u vá»‘n hÃ³a quÃ¡ má»©c, lá»£i nhuáº­n hiá»‡n táº¡i cÃ³ thá»ƒ bá»‹ thá»•i phá»“ng vÃ  AQI/TATA thÆ°á»ng cáº§n Ä‘Æ°á»£c soi ká»¹.",
    "One-off Income": "Thu nháº­p báº¥t thÆ°á»ng/khÃ´ng láº·p láº¡i nhÆ° thanh lÃ½ tÃ i sáº£n, hoÃ n nháº­p lá»›n, lÃ£i Ä‘Ã¡nh giÃ¡ láº¡i. Cáº§n tÃ¡ch khá»i earnings power khi Ä‘á»‹nh giÃ¡.",
    "Restatement": "Viá»‡c Ä‘iá»u chá»‰nh láº¡i BCTC Ä‘Ã£ cÃ´ng bá»‘. ÄÃ¢y lÃ  cá» Ä‘á» cáº§n kiá»ƒm tra nguyÃªn nhÃ¢n, quy mÃ´ Ä‘iá»u chá»‰nh vÃ  áº£nh hÆ°á»Ÿng Ä‘áº¿n lá»£i nhuáº­n, vá»‘n chá»§, dÃ²ng tiá»n.",
    "DSRI": "Days Sales in Receivables Index - chá»‰ sá»‘ pháº£i thu/doanh thu ká»³ hiá»‡n táº¡i so vá»›i ká»³ trÆ°á»›c. DSRI cao cÃ³ thá»ƒ bÃ¡o hiá»‡u doanh thu ghi nháº­n lá»ng hoáº·c thu tiá»n cháº­m.",
    "GMI": "Gross Margin Index - biÃªn gá»™p ká»³ trÆ°á»›c chia biÃªn gá»™p ká»³ hiá»‡n táº¡i. GMI > 1 nghÄ©a lÃ  biÃªn gá»™p suy giáº£m, tÄƒng Ä‘á»™ng cÆ¡ lÃ m Ä‘áº¹p lá»£i nhuáº­n.",
    "AQI": "Asset Quality Index - chá»‰ sá»‘ cháº¥t lÆ°á»£ng tÃ i sáº£n. AQI > 1 cÃ³ thá»ƒ cho tháº¥y tÃ i sáº£n kÃ©m thanh khoáº£n/chi phÃ­ hoÃ£n láº¡i tÄƒng.",
    "SGI": "Sales Growth Index - chá»‰ sá»‘ tÄƒng trÆ°á»Ÿng doanh thu. TÄƒng trÆ°á»Ÿng cao cÃ³ thá»ƒ táº¡o Ã¡p lá»±c duy trÃ¬ má»¥c tiÃªu lá»£i nhuáº­n.",
    "DEPI": "Depreciation Index - chá»‰ sá»‘ kháº¥u hao. DEPI > 1 cÃ³ thá»ƒ cho tháº¥y tá»· lá»‡ kháº¥u hao giáº£m, cáº§n kiá»ƒm tra thá»i gian há»¯u dá»¥ng/phÆ°Æ¡ng phÃ¡p kháº¥u hao.",
    "SGAI": "SG&A Expense Index - chá»‰ sá»‘ chi phÃ­ bÃ¡n hÃ ng vÃ  quáº£n lÃ½/doanh thu. SGAI > 1 pháº£n Ã¡nh chi phÃ­ váº­n hÃ nh tÄƒng nhanh hÆ¡n doanh thu.",
    "TATA": "Total Accruals to Total Assets - tá»•ng accruals/tá»•ng tÃ i sáº£n. TATA dÆ°Æ¡ng cao cho tháº¥y lá»£i nhuáº­n phá»¥ thuá»™c accruals nhiá»u hÆ¡n dÃ²ng tiá»n.",
    "LVGI": "Leverage Index - chá»‰ sá»‘ Ä‘Ã²n báº©y. LVGI > 1 nghÄ©a lÃ  Ä‘Ã²n báº©y tÄƒng, cÃ³ thá»ƒ tÄƒng Ä‘á»™ng cÆ¡ lÃ m Ä‘áº¹p BCTC Ä‘á»ƒ Ä‘Ã¡p á»©ng covenant/ná»£.",
    "Compounder": "Doanh nghiá»‡p cháº¥t lÆ°á»£ng cÃ³ thá»ƒ tÃ¡i Ä‘áº§u tÆ° lá»£i nhuáº­n vá»›i ROIC cao trong thá»i gian dÃ i, tá»« Ä‘Ã³ lÃ m giÃ¡ trá»‹ ná»™i táº¡i tÄƒng kÃ©p.",
    "Cost Advantage": "Lá»£i tháº¿ chi phÃ­ - kháº£ nÄƒng sáº£n xuáº¥t/phÃ¢n phá»‘i/váº­n hÃ nh vá»›i chi phÃ­ tháº¥p hÆ¡n Ä‘á»‘i thá»§ má»™t cÃ¡ch bá»n vá»¯ng.",
    "Cyclical": "Doanh nghiá»‡p chu ká»³, lá»£i nhuáº­n phá»¥ thuá»™c máº¡nh vÃ o giÃ¡ hÃ ng hÃ³a, cung cáº§u ngÃ nh, cÃ´ng suáº¥t, tá»“n kho hoáº·c chu ká»³ kinh táº¿.",
    "Deployed Capital": "Vá»‘n triá»ƒn khai vÃ o hoáº¡t Ä‘á»™ng kinh doanh, thÆ°á»ng loáº¡i bá»›t tiá»n/Ä‘áº§u tÆ° tÃ i chÃ­nh dÆ° thá»«a Ä‘á»ƒ nhÃ¬n hiá»‡u quáº£ vá»‘n váº­n hÃ nh.",
    "Differentiation": "KhÃ¡c biá»‡t hÃ³a - kháº£ nÄƒng táº¡o giÃ¡ trá»‹ cho khÃ¡ch hÃ ng Ä‘á»ƒ duy trÃ¬ biÃªn lá»£i nhuáº­n, thÆ°Æ¡ng hiá»‡u, giÃ¡ bÃ¡n hoáº·c lÃ²ng trung thÃ nh.",
    "DIO": "Days Inventory Outstanding - sá»‘ ngÃ y tá»“n kho bÃ¬nh quÃ¢n. DIO cao/tÄƒng cÃ³ thá»ƒ bÃ¡o hiá»‡u hÃ ng cháº­m luÃ¢n chuyá»ƒn.",
    "DPO": "Days Payable Outstanding - sá»‘ ngÃ y pháº£i tráº£ bÃ¬nh quÃ¢n. DPO cao cÃ³ thá»ƒ há»— trá»£ dÃ²ng tiá»n nhÆ°ng cÅ©ng cÃ³ thá»ƒ pháº£n Ã¡nh Ã¡p lá»±c thanh toÃ¡n.",
    "DSO": "Days Sales Outstanding - sá»‘ ngÃ y pháº£i thu bÃ¬nh quÃ¢n. DSO tÄƒng máº¡nh cÃ³ thá»ƒ bÃ¡o hiá»‡u rá»§i ro thu tiá»n.",
    "EBIT": "Earnings Before Interest and Taxes - lá»£i nhuáº­n trÆ°á»›c lÃ£i vay vÃ  thuáº¿. DÃ¹ng Ä‘á»ƒ Ä‘o lá»£i nhuáº­n hoáº¡t Ä‘á»™ng trÆ°á»›c tÃ¡c Ä‘á»™ng cáº¥u trÃºc vá»‘n.",
    "EBITDA": "EBIT + kháº¥u hao vÃ  phÃ¢n bá»•. Há»¯u Ã­ch Ä‘á»ƒ tham kháº£o kháº£ nÄƒng táº¡o lá»£i nhuáº­n trÆ°á»›c capex, nhÆ°ng khÃ´ng thay tháº¿ dÃ²ng tiá»n tháº­t.",
    "EPS": "Earnings per Share - lá»£i nhuáº­n trÃªn má»—i cá»• phiáº¿u = LNST thuá»™c cá»• Ä‘Ã´ng cÃ´ng ty máº¹ / sá»‘ cá»• phiáº¿u.",
    "FCF": "Free Cash Flow - dÃ²ng tiá»n tá»± do. Trong app thÆ°á»ng tÃ­nh FCF = CFO - Capex. Cáº§n phÃ¢n biá»‡t FCF Ã¢m do má»Ÿ rá»™ng hiá»‡u quáº£ vá»›i FCF Ã¢m do mÃ´ hÃ¬nh kinh doanh hÃºt tiá»n.",
    "FCF/LNST": "Tá»· lá»‡ FCF trÃªn lá»£i nhuáº­n sau thuáº¿. DÃ¹ng Ä‘á»ƒ kiá»ƒm tra lá»£i nhuáº­n cÃ³ Ä‘i kÃ¨m dÃ²ng tiá»n tá»± do hay khÃ´ng.",
    "GiÃ¡ trá»‹ ná»™i táº¡i": "Æ¯á»›c tÃ­nh giÃ¡ trá»‹ kinh táº¿ cá»§a má»™t cá»• phiáº¿u dá»±a trÃªn lá»£i nhuáº­n, dÃ²ng tiá»n, tÃ i sáº£n vÃ  cháº¥t lÆ°á»£ng doanh nghiá»‡p; khÃ´ng pháº£i má»™t con sá»‘ tuyá»‡t Ä‘á»‘i.",
    "High": "Ká»‹ch báº£n giÃ¡ trá»‹ cao trong dáº£i Ä‘á»‹nh giÃ¡. Chá»‰ nÃªn dÃ¹ng khi cháº¥t lÆ°á»£ng dá»¯ liá»‡u, moat vÃ  triá»ƒn vá»ng tÄƒng trÆ°á»Ÿng Ä‘á»§ thuyáº¿t phá»¥c.",
    "Kd": "Cost of Debt - chi phÃ­ ná»£ vay trÆ°á»›c thuáº¿ = chi phÃ­ lÃ£i vay / ná»£ vay chá»‹u lÃ£i bÃ¬nh quÃ¢n.",
    "Ke": "Cost of Equity - chi phÃ­ vá»‘n chá»§ sá»Ÿ há»¯u. App dÃ¹ng mÃ´ hÃ¬nh: lÃ£i suáº¥t phi rá»§i ro + beta x pháº§n bÃ¹ rá»§i ro thá»‹ trÆ°á»ng, hoáº·c proxy khi thiáº¿u dá»¯ liá»‡u thá»‹ trÆ°á»ng.",
    "Low": "Ká»‹ch báº£n giÃ¡ trá»‹ tháº¥p trong dáº£i Ä‘á»‹nh giÃ¡. DÃ¹ng Ä‘á»ƒ kiá»ƒm tra downside vÃ  má»©c chá»‹u Ä‘á»±ng khi giáº£ Ä‘á»‹nh xáº¥u hÆ¡n xáº£y ra.",
    "Maintenance Capex": "Capex duy trÃ¬ - pháº§n vá»‘n Ä‘áº§u tÆ° cáº§n thiáº¿t Ä‘á»ƒ giá»¯ nÄƒng lá»±c cáº¡nh tranh vÃ  sáº£n lÆ°á»£ng dÃ i háº¡n hiá»‡n táº¡i.",
    "Moat": "Lá»£i tháº¿ cáº¡nh tranh bá»n vá»¯ng giÃºp doanh nghiá»‡p duy trÃ¬ lá»£i nhuáº­n cao trÃªn vá»‘n trong thá»i gian dÃ i.",
    "MOS": "Margin of Safety - biÃªn an toÃ n. MOS hiá»‡n táº¡i = (giÃ¡ trá»‹ ná»™i táº¡i - giÃ¡ thá»‹ trÆ°á»ng) / giÃ¡ trá»‹ ná»™i táº¡i. MOS yÃªu cáº§u lÃ  má»©c chiáº¿t kháº¥u ngÆ°á»i dÃ¹ng chá»n trÆ°á»›c khi xem xÃ©t mua.",
    "NCAV": "Net Current Asset Value - giÃ¡ trá»‹ tÃ i sáº£n ngáº¯n háº¡n rÃ²ng, thÆ°á»ng dÃ¹ng trong Ä‘á»‹nh giÃ¡ tÃ i sáº£n theo Graham.",
    "NLA": "Net Liquid Assets - tÃ i sáº£n thanh khoáº£n rÃ²ng = tiá»n + chá»©ng khoÃ¡n thanh khoáº£n + pháº£i thu cÃ³ thá»ƒ thu há»“i - ná»£ ngáº¯n háº¡n/ná»£ pháº£i tráº£ liÃªn quan.",
    "NOPAT": "Net Operating Profit After Tax - lá»£i nhuáº­n hoáº¡t Ä‘á»™ng sau thuáº¿ = EBIT x (1 - thuáº¿ suáº¥t). DÃ¹ng lÃ m tá»­ sá»‘ phá»• biáº¿n khi tÃ­nh ROIC.",
    "OEPS": "Owner Earnings per Share - Owner Earnings trÃªn má»—i cá»• phiáº¿u = Owner Earnings / sá»‘ cá»• phiáº¿u lÆ°u hÃ nh.",
    "Owner Earnings": "Lá»£i nhuáº­n chá»§ sá»Ÿ há»¯u theo Buffett: lá»£i nhuáº­n bÃ¡o cÃ¡o + kháº¥u hao vÃ  chi phÃ­ phi tiá»n máº·t - capex duy trÃ¬ cáº§n thiáº¿t Â± thay Ä‘á»•i vá»‘n lÆ°u Ä‘á»™ng váº­n hÃ nh cáº§n thiáº¿t.",
    "P/B": "Price to Book - giÃ¡ / giÃ¡ trá»‹ sá»• sÃ¡ch. Há»¯u Ã­ch hÆ¡n vá»›i ngÃ¢n hÃ ng, báº£o hiá»ƒm, asset play hoáº·c doanh nghiá»‡p tÃ i sáº£n lá»›n.",
    "P/E": "Price to Earnings - giÃ¡ / EPS. DÃ¹ng tá»‘t hÆ¡n vá»›i doanh nghiá»‡p cÃ³ EPS á»•n Ä‘á»‹nh, Ã­t chu ká»³, lá»£i nhuáº­n káº¿ toÃ¡n Ä‘Ã¡ng tin.",
    "Porter Value Chain": "Chuá»—i giÃ¡ trá»‹ Porter - phÃ¢n rÃ£ doanh nghiá»‡p thÃ nh cÃ¡c hoáº¡t Ä‘á»™ng cá»¥ thá»ƒ Ä‘á»ƒ tÃ¬m nguá»“n gá»‘c lá»£i tháº¿ chi phÃ­/khÃ¡c biá»‡t hÃ³a.",
    "QoQ": "Quarter over Quarter - tÄƒng/giáº£m so vá»›i quÃ½ liá»n trÆ°á»›c.",
    "ROCE": "Return on Capital Employed - lá»£i nhuáº­n trÃªn capital employed. ThÆ°á»ng tÃ­nh EBIT / capital employed.",
    "ROIC": "Return on Invested Capital - lá»£i nhuáº­n trÃªn vá»‘n Ä‘áº§u tÆ°. App Æ°u tiÃªn ROIC Operating Profit = NOPAT / vá»‘n Ä‘áº§u tÆ° bÃ¬nh quÃ¢n.",
    "ROIC Operating Profit": "ROIC dá»±a trÃªn lá»£i nhuáº­n hoáº¡t Ä‘á»™ng: NOPAT chia cho vá»‘n Ä‘áº§u tÆ°/capital employed bÃ¬nh quÃ¢n. DÃ¹ng Ä‘á»ƒ so sÃ¡nh vá»›i WACC.",
    "Switching Cost": "Chi phÃ­ chuyá»ƒn Ä‘á»•i - má»©c khÃ³ khÄƒn/chi phÃ­ khi khÃ¡ch hÃ ng chuyá»ƒn sang nhÃ  cung cáº¥p khÃ¡c.",
    "TTM": "Trailing Twelve Months - sá»‘ liá»‡u 12 thÃ¡ng gáº§n nháº¥t, thÆ°á»ng cá»™ng 4 quÃ½ gáº§n nháº¥t Ä‘á»ƒ cÃ³ cÃ¡i nhÃ¬n cáº­p nháº­t hÆ¡n nÄƒm tÃ i chÃ­nh cÅ©.",
    "Turnaround": "Doanh nghiá»‡p Ä‘ang trong quÃ¡ trÃ¬nh phá»¥c há»“i; Ä‘á»‹nh giÃ¡ cáº§n ká»‹ch báº£n vÃ  biÃªn an toÃ n lá»›n hÆ¡n do rá»§i ro thá»±c thi.",
    "WACC": "Weighted Average Cost of Capital - chi phÃ­ vá»‘n bÃ¬nh quÃ¢n gia quyá»n = We x Ke + Wd x Kd x (1 - thuáº¿ suáº¥t). DÃ¹ng so sÃ¡nh vá»›i ROIC, nhÆ°ng cáº§n dá»¯ liá»‡u ná»£ vay, chi phÃ­ lÃ£i vay, vá»‘n hÃ³a vÃ  chi phÃ­ vá»‘n chá»§.",
    "Weighted": "GiÃ¡ trá»‹ ná»™i táº¡i trung bÃ¬nh trá»ng sá»‘ tá»« cÃ¡c phÆ°Æ¡ng phÃ¡p Ä‘á»‹nh giÃ¡ há»£p lá»‡. Trá»ng sá»‘ cao hÆ¡n Ä‘Æ°á»£c trao cho phÆ°Æ¡ng phÃ¡p phÃ¹ há»£p hÆ¡n vá»›i loáº¡i doanh nghiá»‡p vÃ  cháº¥t lÆ°á»£ng dá»¯ liá»‡u.",
    "YoY": "Year over Year - tÄƒng/giáº£m so vá»›i cÃ¹ng ká»³ nÄƒm trÆ°á»›c.",
}

COMPANY_TYPE_GUIDANCE = {
    "Quality Compounder": {
        "CÆ¡ sá»Ÿ tÆ° duy": "Buffett/Munger + ROIC/Moat + Owner Earnings",
        "Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra": "ROIC/ROCE cao vÃ  á»•n Ä‘á»‹nh; biÃªn lá»£i nhuáº­n bá»n; CFO/LNST vÃ  FCF/LNST tá»‘t; tÃ¡i Ä‘áº§u tÆ° Ä‘Æ°á»£c vá»›i lá»£i suáº¥t cao; ná»£ vay kiá»ƒm soÃ¡t; moat Ä‘áº¿n tá»« thÆ°Æ¡ng hiá»‡u, phÃ¢n phá»‘i, quy mÃ´, switching cost hoáº·c chi phÃ­ tháº¥p.",
        "Cáº§n phÃ¢n tÃ­ch thÃªm": "Runway tÃ¡i Ä‘áº§u tÆ° cÃ²n dÃ i khÃ´ng; ROIC cao cÃ³ bá»‹ kÃ©o xuá»‘ng khi má»Ÿ rá»™ng khÃ´ng; Owner Earnings cÃ³ tháº­t khÃ´ng hay bá»‹ capex duy trÃ¬ lá»›n; ban lÃ£nh Ä‘áº¡o phÃ¢n bá»• vá»‘n cÃ³ ká»· luáº­t khÃ´ng.",
        "Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn": "Owner Earnings, Earnings Power, ROIC Reinvestment, FCF Yield; chá»‰ tráº£ premium khi moat vÃ  tÄƒng trÆ°á»Ÿng tháº­t sá»± bá»n.",
    },
    "Compounder": {
        "CÆ¡ sá»Ÿ tÆ° duy": "Buffett/Munger + ROIC/Moat + Owner Earnings",
        "Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra": "ROIC cao hÆ¡n chi phÃ­ vá»‘n; dÃ²ng tiá»n tá»‘t; doanh thu/lá»£i nhuáº­n tÄƒng Ä‘á»u; doanh nghiá»‡p cÃ³ kháº£ nÄƒng tÃ¡i Ä‘áº§u tÆ°; lá»£i tháº¿ cáº¡nh tranh khÃ´ng phá»¥ thuá»™c vÃ o má»™t chu ká»³ ngáº¯n.",
        "Cáº§n phÃ¢n tÃ­ch thÃªm": "TÄƒng trÆ°á»Ÿng Ä‘áº¿n tá»« sáº£n lÆ°á»£ng, giÃ¡ bÃ¡n, má»Ÿ rá»™ng kÃªnh hay M&A; cháº¥t lÆ°á»£ng lá»£i nhuáº­n; capex duy trÃ¬; rá»§i ro pha loÃ£ng; sá»©c máº¡nh Ä‘á»‹nh giÃ¡.",
        "Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn": "Owner Earnings, FCF, Earnings Power; kiá»ƒm tra MOS báº±ng dáº£i giÃ¡ trá»‹ thay vÃ¬ má»™t con sá»‘ duy nháº¥t.",
    },
    "Cyclical": {
        "CÆ¡ sá»Ÿ tÆ° duy": "Peter Lynch vá» cá»• phiáº¿u chu ká»³ + Howard Marks vá» chu ká»³/rá»§i ro + Graham vá» chuáº©n hÃ³a lá»£i nhuáº­n",
        "Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra": "Lá»£i nhuáº­n biáº¿n Ä‘á»™ng theo giÃ¡ hÃ ng hÃ³a, cung cáº§u ngÃ nh, cÃ´ng suáº¥t, tá»“n kho, lÃ£i suáº¥t hoáº·c chu ká»³ Ä‘áº§u tÆ°. P/E tháº¥p á»Ÿ Ä‘á»‰nh chu ká»³ cÃ³ thá»ƒ lÃ  báº«y; P/E cao á»Ÿ Ä‘Ã¡y chu ká»³ chÆ°a cháº¯c lÃ  Ä‘áº¯t.",
        "Cáº§n phÃ¢n tÃ­ch thÃªm": "Äang á»Ÿ giai Ä‘oáº¡n nÃ o cá»§a chu ká»³; biÃªn lá»£i nhuáº­n hiá»‡n táº¡i so vá»›i trung bÃ¬nh 5-10 nÄƒm; sáº£n lÆ°á»£ng/cÃ´ng suáº¥t; tá»“n kho; ná»£ vay; capex lá»›n cÃ³ rÆ¡i vÃ o cuá»‘i chu ká»³ khÃ´ng; lá»£i nhuáº­n chuáº©n hÃ³a lÃ  bao nhiÃªu.",
        "Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn": "Normalized earnings, P/B, ROCE qua chu ká»³, EV/EBITDA chuáº©n hÃ³a, asset value; MOS cáº§n rá»™ng hÆ¡n doanh nghiá»‡p á»•n Ä‘á»‹nh.",
    },
    "Asset Play": {
        "CÆ¡ sá»Ÿ tÆ° duy": "Graham/Dodd + Li Lu Timberland case + Net Liquid Assets",
        "Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra": "GiÃ¡ thá»‹ trÆ°á»ng tháº¥p so vá»›i tÃ i sáº£n há»¯u hÃ¬nh/thanh khoáº£n; downside Ä‘Æ°á»£c báº£o vá»‡ bá»Ÿi tiá»n, tÃ i sáº£n, báº¥t Ä‘á»™ng sáº£n, vá»‘n lÆ°u Ä‘á»™ng hoáº·c giÃ¡ trá»‹ thanh lÃ½.",
        "Cáº§n phÃ¢n tÃ­ch thÃªm": "Cháº¥t lÆ°á»£ng tÃ i sáº£n; khoáº£n pháº£i thu/tá»“n kho cÃ³ thu há»“i Ä‘Æ°á»£c khÃ´ng; tÃ i sáº£n áº©n; ná»£ tiá»m tÃ ng; giao dá»‹ch bÃªn liÃªn quan; kháº£ nÄƒng hiá»‡n thá»±c hÃ³a giÃ¡ trá»‹; quáº£n trá»‹ cÃ³ thÃ¢n thiá»‡n cá»• Ä‘Ã´ng khÃ´ng.",
        "Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn": "Book Value Ä‘iá»u chá»‰nh, NCAV, NLA, Liquidation Value, P/B; khÃ´ng nÃªn tráº£ cao chá»‰ vÃ¬ tÃ i sáº£n lá»›n náº¿u tÃ i sáº£n khÃ³ chuyá»ƒn thÃ nh tiá»n.",
    },
    "Turnaround": {
        "CÆ¡ sá»Ÿ tÆ° duy": "Graham vá» báº£o vá»‡ downside + Howard Marks vá» kiá»ƒm soÃ¡t rá»§i ro",
        "Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra": "Doanh nghiá»‡p Ä‘ang phá»¥c há»“i tá»« suy giáº£m, lá»—, tÃ¡i cáº¥u trÃºc hoáº·c thay Ä‘á»•i chiáº¿n lÆ°á»£c. Rá»§i ro thá»±c thi cao vÃ  dá»¯ liá»‡u quÃ¡ khá»© cÃ³ thá»ƒ chÆ°a Ä‘áº¡i diá»‡n tÆ°Æ¡ng lai.",
        "Cáº§n phÃ¢n tÃ­ch thÃªm": "NguyÃªn nhÃ¢n suy giáº£m Ä‘Ã£ xá»­ lÃ½ chÆ°a; dÃ²ng tiá»n cÃ³ Ä‘á»§ sá»‘ng sÃ³t khÃ´ng; ná»£ vay/Ä‘Ã¡o háº¡n; tÃ i sáº£n cÃ³ bÃ¡n Ä‘Æ°á»£c khÃ´ng; ban lÃ£nh Ä‘áº¡o má»›i; dáº¥u hiá»‡u cáº£i thiá»‡n biÃªn lá»£i nhuáº­n vÃ  vÃ²ng quay vá»‘n.",
        "Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn": "Downside asset value, bear/base/bull scenario, normalized earnings sau phá»¥c há»“i; yÃªu cáº§u MOS ráº¥t rá»™ng.",
    },
    "Bank/Insurance": {
        "CÆ¡ sá»Ÿ tÆ° duy": "Peter Lynch/regional banks + tiÃªu chÃ­ ngÃ¢n hÃ ng trong bá»™ nguá»“n",
        "Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra": "BCTC Ä‘áº·c thÃ¹; Ä‘Ã²n báº©y cao; ROE/P/B quan trá»ng hÆ¡n FCF; cháº¥t lÆ°á»£ng tÃ i sáº£n vÃ  quáº£n trá»‹ rá»§i ro quyáº¿t Ä‘á»‹nh giÃ¡ trá»‹.",
        "Cáº§n phÃ¢n tÃ­ch thÃªm": "NIM, CASA, tÄƒng trÆ°á»Ÿng tÃ­n dá»¥ng, NPL, ná»£ nhÃ³m 2, bao phá»§ ná»£ xáº¥u, chi phÃ­ tÃ­n dá»¥ng, CAR, cháº¥t lÆ°á»£ng trÃ¡i phiáº¿u/tÃ i sáº£n Ä‘áº§u tÆ°, governance.",
        "Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn": "P/B so vá»›i ROE bá»n vá»¯ng, residual income/earning power, cháº¥t lÆ°á»£ng tÃ i sáº£n; khÃ´ng dÃ¹ng FCF cÃ´ng nghiá»‡p lÃ m phÆ°Æ¡ng phÃ¡p lÃµi.",
    },
    "ChÆ°a cÃ³ dá»¯ liá»‡u tÃ i chÃ­nh": {
        "CÆ¡ sá»Ÿ tÆ° duy": "NguyÃªn táº¯c audit dá»¯ liá»‡u: khÃ´ng Ä‘á»‹nh giÃ¡ khi chÆ°a cÃ³ BCTC nhiá»u ká»³ Ä‘á»§ kiá»ƒm chá»©ng",
        "Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra": "ChÆ°a Ä‘á»§ chuá»—i BCTC nÄƒm/quÃ½ Ä‘á»ƒ phÃ¢n loáº¡i Ä‘Ã¡ng tin cáº­y; cáº§n kiá»ƒm tra nguá»“n dá»¯ liá»‡u, mÃ£ chá»©ng khoÃ¡n, ká»³ bÃ¡o cÃ¡o vÃ  file import/crawler.",
        "Cáº§n phÃ¢n tÃ­ch thÃªm": "Táº£i hoáº·c import BCTC nhiá»u ká»³; Ä‘á»‘i chiáº¿u doanh thu, LNST, CFO, capex, ROE/ROIC, ná»£ vay vÃ  vá»‘n chá»§ trÆ°á»›c khi cháº¡y Ä‘á»‹nh giÃ¡.",
        "Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn": "Táº¡m dá»«ng Ä‘á»‹nh giÃ¡ tá»± Ä‘á»™ng; chá»‰ má»Ÿ láº¡i P/E, P/B, FCF, Owner Earnings, MOS khi dá»¯ liá»‡u Ä‘Ã£ Ä‘á»§ vÃ  cÃ³ kiá»ƒm chá»©ng ná»™i bá»™.",
    },
    "Normal Business": {
        "CÆ¡ sá»Ÿ tÆ° duy": "Graham/Buffett/Peter Lynch káº¿t há»£p",
        "Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra": "Doanh nghiá»‡p chÆ°a Ä‘á»§ báº±ng chá»©ng Ä‘á»ƒ xáº¿p compounder, cyclical, asset play hay turnaround. Cáº§n Ä‘á»c cháº¥t lÆ°á»£ng lá»£i nhuáº­n vÃ  lá»£i tháº¿ cáº¡nh tranh trÆ°á»›c khi tráº£ premium.",
        "Cáº§n phÃ¢n tÃ­ch thÃªm": "ROIC so vá»›i WACC; CFO/LNST; FCF; tÄƒng trÆ°á»Ÿng doanh thu/lá»£i nhuáº­n; biáº¿n Ä‘á»™ng biÃªn lá»£i nhuáº­n; ná»£ vay; ngÃ nh cÃ³ cáº¡nh tranh gay gáº¯t khÃ´ng; cÃ³ moat tháº­t khÃ´ng.",
        "Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn": "Káº¿t há»£p P/E chuáº©n hÃ³a, FCF, Earnings Power, P/B; giáº£m trá»ng sá»‘ cÃ¡c phÆ°Æ¡ng phÃ¡p thiáº¿u dá»¯ liá»‡u vÃ  yÃªu cáº§u MOS phÃ¹ há»£p má»©c báº¥t Ä‘á»‹nh.",
    },
}


def _render_company_type_guidance(current_type: str | None = None) -> None:
    raw_type = str(current_type or "Normal Business")
    current_type, info = _company_type_info(raw_type)
    st.subheader("Diá»…n giáº£i loáº¡i hÃ¬nh doanh nghiá»‡p & Ä‘iá»ƒm cáº§n phÃ¢n tÃ­ch thÃªm")
    st.markdown(
        f"""
        <div class='note-card'>
        <b style='color:#0B7F75;font-size:1.05rem'>Loáº¡i hiá»‡n táº¡i: {html.escape(current_type)}</b><br>
        <b>CÆ¡ sá»Ÿ tÆ° duy:</b> {html.escape(info['CÆ¡ sá»Ÿ tÆ° duy'])}<br>
        <b>Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra:</b> {html.escape(info['Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra'])}<br>
        <b>Cáº§n phÃ¢n tÃ­ch thÃªm:</b> {html.escape(info['Cáº§n phÃ¢n tÃ­ch thÃªm'])}<br>
        <b>Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn:</b> {html.escape(info['Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn'])}
        </div>
        """,
        unsafe_allow_html=True,
    )
    type_df = pd.DataFrame([
        {"Loáº¡i doanh nghiá»‡p": k, **v} for k, v in COMPANY_TYPE_GUIDANCE.items()
    ]).sort_values("Loáº¡i doanh nghiá»‡p").reset_index(drop=True)
    type_df.insert(0, "STT", range(1, len(type_df) + 1))
    st.download_button("â¬‡ï¸ Táº£i báº£ng loáº¡i hÃ¬nh doanh nghiá»‡p", type_df.to_csv(index=False, encoding="utf-8-sig"), file_name="company_type_guidance.csv", mime="text/csv", use_container_width=True)
    rows_html = []
    for _, r in type_df.iterrows():
        rows_html.append(
            "<tr>"
            f"<td>{html.escape(str(r.get('STT','')))}</td>"
            f"<td class='type-name'>{html.escape(str(r.get('Loáº¡i doanh nghiá»‡p','')))}</td>"
            f"<td>{html.escape(str(r.get('CÆ¡ sá»Ÿ tÆ° duy','')))}</td>"
            f"<td>{html.escape(str(r.get('Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra','')))}</td>"
            f"<td>{html.escape(str(r.get('Cáº§n phÃ¢n tÃ­ch thÃªm','')))}</td>"
            f"<td>{html.escape(str(r.get('Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn','')))}</td>"
            "</tr>"
        )
    st.markdown(
        """
        <div style='overflow-x:auto; border:1px solid rgba(11,127,117,.22); border-radius:14px;'>
        <table class='type-fit-table'>
          <thead><tr><th>STT</th><th>Loáº¡i doanh nghiá»‡p</th><th>CÆ¡ sá»Ÿ tÆ° duy</th><th>Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra</th><th>Cáº§n phÃ¢n tÃ­ch thÃªm</th><th>Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn</th></tr></thead>
          <tbody>{rows}</tbody>
        </table>
        </div>
        <style>
        .type-fit-table {{width:100%; border-collapse:collapse; font-size:.92rem; line-height:1.42;}}
        .type-fit-table th {{position:sticky; top:0; background:#EAF7F1; color:#064E47; border:1px solid rgba(11,127,117,.20); padding:8px 10px; text-align:left; font-weight:950;}}
        .type-fit-table td {{border:1px solid rgba(11,127,117,.12); padding:8px 10px; vertical-align:top;}}
        .type-fit-table td.type-name {{font-weight:950; color:#0B5F58; background:#FFF4C7;}}
        </style>
        """.format(rows="".join(rows_html)),
        unsafe_allow_html=True,
    )


def _render_company_type_summary_callout(classification: object | None = None, current_type: str | None = None) -> None:
    """Render the yellow company-type card with ticker-specific classification reasons."""
    if classification is not None and hasattr(classification, "company_type"):
        raw_type = str(getattr(classification, "company_type", "Normal Business") or "Normal Business")
        confidence = getattr(classification, "confidence", None)
        reasons = [str(x).strip() for x in getattr(classification, "reasons", []) if str(x).strip()]
        preferred = [str(x).strip() for x in getattr(classification, "preferred_methods", []) if str(x).strip()]
    else:
        raw_type = str(current_type or classification or "Normal Business")
        confidence = None
        reasons = []
        preferred = []
    guide_type, info = _company_type_info(raw_type)
    confidence_text = f" Â· Ä‘á»™ tin cáº­y {float(confidence):.0f}/100" if confidence is not None else ""
    reason_items = "".join(f"<li>{html.escape(reason)}</li>" for reason in reasons)
    if not reason_items:
        reason_items = "<li>ChÆ°a cÃ³ Ä‘á»§ dá»¯ liá»‡u Ä‘á»‹nh lÆ°á»£ng Ä‘á»ƒ giáº£i thÃ­ch phÃ¢n loáº¡i; cáº§n cáº­p nháº­t BCTC/nguá»“n dá»¯ liá»‡u trÆ°á»›c khi káº¿t luáº­n.</li>"
    preferred_text = "; ".join(preferred) if preferred else info.get("Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn", "N/A")
    st.markdown(
        f"""
        <div style="border:2px solid rgba(245,178,27,.65); border-left:12px solid #F5B21B; border-radius:18px;
                    padding:15px 18px; margin:10px 0 16px 0; background:linear-gradient(135deg,#FFF9E8 0%,#FFF3C4 100%);
                    color:#5F3B00; font-weight:900; line-height:1.58; box-shadow:0 9px 22px rgba(245,178,27,.13);">
          <div style="font-size:1.08rem; color:#0B7F75; font-weight:1000; margin-bottom:6px;">ðŸ“Œ PhÃ¢n loáº¡i doanh nghiá»‡p: {html.escape(raw_type)}{html.escape(confidence_text)}</div>
          <div style="font-size:.90rem; color:#7A4B00; margin-bottom:8px;"><b>NhÃ³m hÆ°á»›ng dáº«n Ã¡p dá»¥ng:</b> {html.escape(guide_type)}</div>
          <div><b style="color:#9A6600;">Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra:</b> {html.escape(info.get('Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra','N/A'))}</div>
          <div><b style="color:#9A6600;">Cáº§n phÃ¢n tÃ­ch thÃªm:</b> {html.escape(info.get('Cáº§n phÃ¢n tÃ­ch thÃªm','N/A'))}</div>
          <div><b style="color:#9A6600;">Äá»‹nh giÃ¡ nÃªn Æ°u tiÃªn:</b> {html.escape(preferred_text)}</div>
          <div style="margin-top:10px; padding-top:9px; border-top:1px solid rgba(154,102,0,.22);">
            <div style="font-size:1.02rem; color:#8A5A00; font-weight:1000; margin-bottom:4px;">ðŸŸ¡ LÃ½ do phÃ¢n loáº¡i theo dá»¯ liá»‡u mÃ£ Ä‘ang phÃ¢n tÃ­ch</div>
            <ul style="margin:6px 0 0 20px; padding:0;">{reason_items}</ul>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _augment_auto_summary_with_checklist(summary: object, current_type: str | None = None) -> str:
    """KhÃ´ng chÃ¨n Äáº·c Ä‘iá»ƒm cáº§n kiá»ƒm tra vÃ o card Ä‘á» TÃ³m táº¯t tá»± Ä‘á»™ng."""
    return "" if summary is None else str(summary)


def _render_glossary_panel() -> None:
    st.subheader("Diá»…n giáº£i thuáº­t ngá»¯ vÃ  tá»« viáº¿t táº¯t")
    st.caption("Báº£ng thuáº­t ngá»¯ Ä‘Æ°á»£c sáº¯p xáº¿p theo thá»© tá»± chá»¯ cÃ¡i; cá»™t Thuáº­t ngá»¯ tá»± Ã´m sÃ¡t ná»™i dung, cá»™t Diá»…n giáº£i má»Ÿ rá»™ng vÃ  xuá»‘ng dÃ²ng Ä‘á»ƒ dá»… Ä‘á»c.")
    glossary_df = pd.DataFrame([{"Thuáº­t ngá»¯": k, "Diá»…n giáº£i": v} for k, v in GLOSSARY_TERMS.items()])
    glossary_df = glossary_df.sort_values("Thuáº­t ngá»¯", key=lambda s: s.str.lower()).reset_index(drop=True)
    glossary_df.insert(0, "STT", range(1, len(glossary_df) + 1))
    st.download_button("â¬‡ï¸ Táº£i báº£ng thuáº­t ngá»¯", glossary_df.to_csv(index=False, encoding="utf-8-sig"), file_name="glossary.csv", mime="text/csv", use_container_width=True)
    rows = []
    for _, r in glossary_df.iterrows():
        rows.append(
            "<tr>"
            f"<td class='stt'>{html.escape(str(r.get('STT', '')))}</td>"
            f"<td class='term'>{html.escape(str(r.get('Thuáº­t ngá»¯', '')))}</td>"
            f"<td class='desc'>{html.escape(str(r.get('Diá»…n giáº£i', '')))}</td>"
            "</tr>"
        )
    table_html = """
    <div style='max-height:560px; overflow:auto; border-radius:14px; box-shadow:0 8px 20px rgba(11,127,117,.07);'>
      <table class='glossary-fit-table'>
        <thead><tr><th class='stt'>STT</th><th class='term'>Thuáº­t ngá»¯</th><th class='desc'>Diá»…n giáº£i</th></tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>
    """.format(rows="".join(rows))
    st.markdown(table_html, unsafe_allow_html=True)

def _render_no_data(ticker: str, source: str, available_tickers: list[str], error: str | None = None) -> None:
    info = _listed_ticker_info_cached(str(BUNDLED_XLSM), ticker) if BUNDLED_XLSM.exists() else {}
    name = info.get("company_name", ticker)
    exchange = info.get("exchange", "")
    st.error(f"ChÆ°a cÃ³ dá»¯ liá»‡u BCTC nhiá»u ká»³ Ä‘á»ƒ Ä‘á»‹nh giÃ¡ mÃ£ {ticker} tá»« nguá»“n '{source}'.")
    if info:
        st.info(f"ÄÃ£ nháº­n diá»‡n mÃ£ {ticker}: {name} ({exchange}). Tuy nhiÃªn dá»¯ liá»‡u tÃ­ch há»£p hiá»‡n chÆ°a cÃ³ block BCTC nhiá»u ká»³ cho mÃ£ nÃ y.")
    if error:
        st.warning(error)
    st.markdown(
        f"""
        <div class='warn-card'>
        <b>Ã nghÄ©a mÃ n hÃ¬nh nÃ y:</b><br>
        App khÃ´ng bá»‹ treo. Äá»‹nh giÃ¡ chuyÃªn sÃ¢u Ä‘ang cháº·n viá»‡c Ä‘á»‹nh giÃ¡ khi chÆ°a cÃ³ Ä‘á»§ BCTC, Ä‘á»ƒ trÃ¡nh hiá»‡n N/A hoáº·c cháº¥m moat áº£o.<br><br>
        <b>CÃ¡ch xá»­ lÃ½:</b><br>
        1) Chá»n má»™t mÃ£ cÃ³ dá»¯ liá»‡u trong danh sÃ¡ch bÃªn trÃ¡i Ä‘á»ƒ test ngay.<br>
        2) Vá»›i mÃ£ <b>{ticker}</b>, chá»n cháº¿ Ä‘á»™ <b>Tá»± Ä‘á»™ng</b> hoáº·c <b>Dá»¯ liá»‡u Æ°u tiÃªn</b> Ä‘á»ƒ app dÃ¹ng láº¡i bá»™ BCTC Ä‘Ã£ chuáº©n hÃ³a cá»§a Tá»•ng quan doanh nghiá»‡p.<br>
        3) Náº¿u dá»¯ liá»‡u chÆ°a Ä‘á»§, cÃ³ thá»ƒ thá»­ cháº¿ Ä‘á»™ <b>Dá»¯ liá»‡u trá»±c tuyáº¿n</b> hoáº·c cáº­p nháº­t/import file dá»¯ liá»‡u tÃ­ch há»£p cÃ³ chá»©a block BCTC cá»§a mÃ£ nÃ y rá»“i cháº¡y láº¡i.
        </div>
        """,
        unsafe_allow_html=True,
    )
    if available_tickers:
        st.success("CÃ¡c mÃ£ cÃ³ Ä‘á»§ dá»¯ liá»‡u tÃ­ch há»£p: " + ", ".join(available_tickers))


def _update_module2_web_evidence(company) -> None:
    """TÃ¬m evidence internet vÃ  lÆ°u vÃ o session cho Ä‘Ãºng mÃ£ Ä‘ang phÃ¢n tÃ­ch."""
    try:
        ticker = _safe_ticker(getattr(company, "ticker", ""))
        company_name = str(getattr(company, "company_name", "") or "")
        result = WebEvidenceAgent(RAW_DIR).search(ticker, company_name)
        st.session_state["module2_web_table"] = result.table
        st.session_state["module2_web_note"] = result.note
        st.session_state["module2_web_ticker"] = ticker
        st.session_state["module2_web_updated_at"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
        st.session_state["module2_auto_update_status"] = (
            f"ÄÃ£ cáº­p nháº­t BCTC/cache Tá»•ng quan doanh nghiá»‡p â†’ Äá»‹nh giÃ¡ chuyÃªn sÃ¢u vÃ  evidence internet cho {ticker} lÃºc {st.session_state['module2_web_updated_at']}"
        )
    except Exception as exc:
        st.session_state["module2_auto_update_status"] = f"KhÃ´ng cáº­p nháº­t Ä‘Æ°á»£c evidence internet: {exc}"



PEER_UNIVERSE_CSV = DATA_CACHE_DIR / "peer_universe.csv"


SIMPLIZE_PEER_COLUMNS = [
    "ticker", "company_name", "current_price", "price_change_pct", "change_7d_pct", "change_1y_pct",
    "pe", "pb", "roe_pct", "forecast_profit_growth_3y_pct", "dividend_yield_pct", "exchange",
    "market_cap_bil", "chart_30d", "industry", "sub_industry", "peer_group", "source", "note", "updated_at"
]


PEER_BAD_TICKERS = {
    "ROE", "ROA", "ROIC", "EPS", "PE", "PB", "PS", "P/E", "P/B", "FCF", "CFO", "EBIT", "EBITDA",
    "LNST", "LNTT", "DT", "TTM", "YOY", "QOQ", "MOM", "DPS", "WACC", "MOS", "NIM", "CASA", "NPL",
    "HTML", "HTTP", "HTTPS", "JSON", "POST", "GET", "API", "CSS", "JS", "PDF", "XLS",
    "TRUE", "FALSE", "NULL", "NONE", "NAN", "STOCK", "STOCKS", "HOSTC", "HOST",
    "HOSE", "HNX", "UPCOM", "INDEX", "CODE", "NAME", "SYMBOL", "TICKER", "MARKET",
    "FLOOR", "HOME", "LOGIN", "LOGO", "DATA", "IR", "PR", "ETF", "GDP", "PMI",
}


def _is_probable_peer_ticker(code: object) -> bool:
    """Strict filter for peer tickers; prevents table headers like ROE from becoming rows."""
    text = _safe_ticker(str(code or ""))
    if not re.fullmatch(r"[A-Z][A-Z0-9]{1,5}", text):
        return False
    if text in PEER_BAD_TICKERS:
        return False
    if len(text) > 4 and not any(ch.isdigit() for ch in text):
        return False
    if re.search(r"(HTML|HTTP|JSON|STOCK|HOST|INDEX|LOGIN|DATA|ROE|ROA|EPS|FCF|CFO|MOS|WACC)", text):
        return False
    return True


def _empty_peer_universe() -> pd.DataFrame:
    return pd.DataFrame(columns=SIMPLIZE_PEER_COLUMNS)


def _normalize_peer_universe(df: pd.DataFrame | None) -> pd.DataFrame:
    base_cols = SIMPLIZE_PEER_COLUMNS
    if df is None or df.empty:
        return _empty_peer_universe()
    out = df.copy()
    rename = {}
    for c in out.columns:
        lc = str(c).strip().lower()
        if lc in {"mÃ£", "ma", "ma_cp", "mÃ£ cp", "mÃ£ cá»• phiáº¿u", "code", "symbol"}:
            rename[c] = "ticker"
        elif lc in {"tÃªn", "ten", "company", "company_name", "tÃªn doanh nghiá»‡p", "doanh nghiá»‡p"}:
            rename[c] = "company_name"
        elif lc in {"sÃ n", "san", "exchange"}:
            rename[c] = "exchange"
        elif lc in {"ngÃ nh", "nganh", "industry"}:
            rename[c] = "industry"
        elif lc in {"phÃ¢n ngÃ nh", "phan nganh", "sub_industry", "sub industry"}:
            rename[c] = "sub_industry"
        elif lc in {"nhÃ³m ngÃ nh", "nhom nganh", "peer_group", "group"}:
            rename[c] = "peer_group"
        elif lc in {"nguá»“n", "nguon", "source"}:
            rename[c] = "source"
        elif lc in {"ghi chÃº", "ghi chu", "note"}:
            rename[c] = "note"
        elif lc in {"giÃ¡ hiá»‡n táº¡i", "gia hien tai", "current_price", "last_price", "price"}:
            rename[c] = "current_price"
        elif lc in {"biáº¿n Ä‘á»™ng giÃ¡", "bien dong gia", "price_change_pct", "change_pct", "thay Ä‘á»•i giÃ¡ %"}:
            rename[c] = "price_change_pct"
        elif lc in {"7 ngÃ y", "7 ngay", "7d", "change_7d_pct"}:
            rename[c] = "change_7d_pct"
        elif lc in {"1 nÄƒm", "1 nam", "1y", "change_1y_pct"}:
            rename[c] = "change_1y_pct"
        elif lc in {"p/e", "pe"}:
            rename[c] = "pe"
        elif lc in {"p/b", "pb"}:
            rename[c] = "pb"
        elif lc in {"roe", "roe %", "roe_pct"}:
            rename[c] = "roe_pct"
        elif lc in {"t.trÆ°á»Ÿng lnst 3 nÄƒm dá»± phÃ³ng", "tang truong lnst 3 nam du phong", "forecast_profit_growth_3y_pct"}:
            rename[c] = "forecast_profit_growth_3y_pct"
        elif lc in {"tá»· suáº¥t cá»• tá»©c", "ty suat co tuc", "dividend_yield_pct"}:
            rename[c] = "dividend_yield_pct"
        elif lc in {"vá»‘n hÃ³a", "von hoa", "market_cap", "market_cap_bil", "vá»‘n hÃ³a (tá»· Ä‘á»“ng)"}:
            rename[c] = "market_cap_bil"
        elif lc in {"biá»ƒu Ä‘á»“ giÃ¡ 30d", "bieu do gia 30d", "chart_30d"}:
            rename[c] = "chart_30d"
    out = out.rename(columns=rename)
    for c in base_cols:
        if c not in out.columns:
            out[c] = ""
    out = out[base_cols].copy()
    out["ticker"] = out["ticker"].astype(str).map(_safe_ticker)
    out = out[out["ticker"].map(_is_probable_peer_ticker)]
    for num_col in ["current_price", "price_change_pct", "change_7d_pct", "change_1y_pct", "pe", "pb", "roe_pct", "forecast_profit_growth_3y_pct", "dividend_yield_pct", "market_cap_bil"]:
        if num_col in out.columns:
            out[num_col] = pd.to_numeric(out[num_col], errors="coerce")
    out["peer_group"] = out["peer_group"].replace({None: ""}).astype(str)
    fallback_group = out["industry"].fillna("").astype(str).where(out["industry"].fillna("").astype(str).str.len() > 0, "ChÆ°a phÃ¢n nhÃ³m")
    out["peer_group"] = out["peer_group"].where(out["peer_group"].str.strip().str.len() > 0, fallback_group)
    # Giá»¯ báº£n ghi Ä‘áº§y Ä‘á»§ nháº¥t khi trÃ¹ng mÃ£; trÃ¡nh dÃ²ng mÃ£ Ä‘ang phÃ¢n tÃ­ch bá»‹ trá»‘ng do cache cÅ©/row ká»¹ thuáº­t ghi Ä‘Ã¨ dá»¯ liá»‡u cÃ¹ng ngÃ nh.
    non_empty = out.replace("", pd.NA).notna().sum(axis=1)
    out = out.assign(_complete_score=non_empty).sort_values(["ticker", "_complete_score"]).drop_duplicates(subset=["ticker"], keep="last").drop(columns=["_complete_score"])
    if "market_cap_bil" in out.columns and pd.to_numeric(out["market_cap_bil"], errors="coerce").notna().any():
        out = out.assign(_sort_cap=pd.to_numeric(out["market_cap_bil"], errors="coerce")).sort_values(["peer_group", "_sort_cap", "ticker"], ascending=[True, False, True]).drop(columns=["_sort_cap"])
    else:
        out = out.sort_values(["peer_group", "ticker"])
    out = out.reset_index(drop=True)
    return out


def _load_peer_universe() -> pd.DataFrame:
    if PEER_UNIVERSE_CSV.exists():
        try:
            return _normalize_peer_universe(pd.read_csv(PEER_UNIVERSE_CSV))
        except Exception:
            return _empty_peer_universe()
    return _empty_peer_universe()


def _save_peer_universe(df: pd.DataFrame) -> None:
    PEER_UNIVERSE_CSV.parent.mkdir(parents=True, exist_ok=True)
    _normalize_peer_universe(df).to_csv(PEER_UNIVERSE_CSV, index=False, encoding="utf-8-sig")


@st.cache_data(show_spinner=False, ttl=60 * 60 * 12)
def _simplize_industry_peers_cached(ticker: str, raw_dir: str, industry_url: str = "") -> tuple[pd.DataFrame, str, str]:
    raw_path, peer_df, note = PublicSimplizeCrawler(raw_dir).fetch_industry_peers(ticker, industry_url or None)
    return _normalize_peer_universe(peer_df), note, str(raw_path)


def _load_simplize_peer_rows(ticker: str, industry_url: str = "") -> tuple[pd.DataFrame, str, str]:
    ticker = _safe_ticker(ticker)
    if not ticker:
        return _empty_peer_universe(), "ChÆ°a cÃ³ mÃ£ cá»• phiáº¿u Ä‘á»ƒ láº¥y danh sÃ¡ch ngÃ nh.", ""
    try:
        return _simplize_industry_peers_cached(ticker, str(RAW_DIR), industry_url or "")
    except Exception as exc:
        return _empty_peer_universe(), f"KhÃ´ng cáº­p nháº­t Ä‘Æ°á»£c danh sÃ¡ch cÃ¹ng ngÃ nh cho {ticker}: {_public_text(exc)}", ""


def _company_peer_group(company) -> str:
    sub = _display_industry_value(getattr(company, "sub_industry", ""))
    ind = _display_industry_value(getattr(company, "industry", ""))
    return (sub if sub != "N/A" else "") or (ind if ind != "N/A" else "") or "ChÆ°a phÃ¢n nhÃ³m"


def _company_to_peer_row(company, source_label: str = "MÃ£ Ä‘ang phÃ¢n tÃ­ch", note: str = "") -> dict:
    ticker = _safe_ticker(str(getattr(company, "ticker", "")))
    return {
        "ticker": ticker,
        "company_name": str(getattr(company, "company_name", "") or ticker),
        "exchange": str(getattr(company, "exchange", "") or ""),
        "industry": _display_industry_value(getattr(company, "industry", "")),
        "sub_industry": _display_industry_value(getattr(company, "sub_industry", "")),
        "peer_group": _company_peer_group(company),
        "current_price": _parse_num(getattr(company, "current_price", None)),
        "market_cap_bil": _parse_num(getattr(company, "market_cap_bil", None)),
        "pe": _parse_num(getattr(company, "pe", None)),
        "pb": _parse_num(getattr(company, "pb", None)),
        "roe_pct": _parse_num(getattr(company, "roe", None)),
        "source": source_label,
        "note": note,
        "updated_at": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _merge_peer_rows(base: pd.DataFrame, rows: list[dict]) -> pd.DataFrame:
    add = _normalize_peer_universe(pd.DataFrame(rows)) if rows else _empty_peer_universe()
    return _normalize_peer_universe(pd.concat([base, add], ignore_index=True))


def _ensure_current_peer_row(df: pd.DataFrame, company) -> pd.DataFrame:
    """Ensure the base ticker row is never blank and remains visible in the peer list."""
    out = _normalize_peer_universe(df)
    cur = _company_to_peer_row(company, "MÃ£ Ä‘ang phÃ¢n tÃ­ch", "DÃ²ng mÃ£ gá»‘c Ä‘Æ°á»£c ghim vÃ  tick máº·c Ä‘á»‹nh")
    ticker = _safe_ticker(cur.get("ticker"))
    if not ticker:
        return out
    if out.empty or "ticker" not in out.columns:
        return _normalize_peer_universe(pd.DataFrame([cur]))
    mask = out["ticker"].astype(str).map(_safe_ticker).eq(ticker)
    if not mask.any():
        out = pd.concat([pd.DataFrame([cur]), out], ignore_index=True)
        return _normalize_peer_universe(out)
    idx = out.index[mask][0]
    # Always keep identity fields from the active company; preserve richer peer metrics if present.
    out.at[idx, "ticker"] = ticker
    for col in ["company_name", "exchange", "industry", "sub_industry", "peer_group", "source", "note", "updated_at"]:
        cur_val = cur.get(col, "")
        old_val = out.at[idx, col] if col in out.columns else ""
        if str(old_val).strip() in {"", "nan", "None", "N/A"} and str(cur_val).strip():
            out.at[idx, col] = cur_val
    for col in ["current_price", "market_cap_bil", "pe", "pb", "roe_pct"]:
        old_num = _parse_num(out.at[idx, col]) if col in out.columns else None
        cur_num = _parse_num(cur.get(col))
        if old_num is None and cur_num is not None:
            out.at[idx, col] = cur_num
    return _normalize_peer_universe(out)


def _recent_median_local(df: pd.DataFrame, *cols: str, n: int = 5) -> float | None:
    if df is None or df.empty:
        return None
    for col in cols:
        if col in df.columns:
            s = pd.to_numeric(df[col], errors="coerce").dropna().tail(n)
            if not s.empty:
                return float(s.median())
    return None


def _latest_num_local(df: pd.DataFrame, *cols: str) -> float | None:
    if df is None or df.empty:
        return None
    row = df.iloc[-1].to_dict()
    for col in cols:
        val = _parse_num(row.get(col))
        if val is not None:
            return val
    return None


def _cagr_local(df: pd.DataFrame, col: str, years: int = 5) -> float | None:
    if df is None or df.empty or col not in df.columns:
        return None
    s = pd.to_numeric(df[col], errors="coerce").dropna().tail(years + 1)
    if len(s) < 2:
        return None
    start, end = float(s.iloc[0]), float(s.iloc[-1])
    periods = len(s) - 1
    if start <= 0 or end <= 0 or periods <= 0:
        return None
    return ((end / start) ** (1 / periods) - 1) * 100


def _score_high(value: float | None, good: float, ok: float, weak: float) -> float:
    if value is None:
        return 45.0
    if value >= good:
        return 100.0
    if value >= ok:
        return 75.0
    if value >= weak:
        return 50.0
    return 20.0


def _score_low(value: float | None, good: float, ok: float, bad: float) -> float:
    if value is None or value <= 0:
        return 45.0
    if value <= good:
        return 100.0
    if value <= ok:
        return 75.0
    if value <= bad:
        return 50.0
    return 20.0


def _mean_ignore_none(values: list[float | None]) -> float:
    vals = [float(v) for v in values if v is not None and not pd.isna(v)]
    return sum(vals) / len(vals) if vals else 45.0


def _peer_snapshot(ticker: str, source_for_peer: str, assumptions: dict, target_mos_pct: float) -> tuple[dict, dict | None]:
    ticker = _safe_ticker(ticker)
    try:
        c, annual, quarterly, label, _paths = _load_data(ticker, source_for_peer)
        if not _has_real_financial_data(annual):
            raise RuntimeError("KhÃ´ng cÃ³ dá»¯ liá»‡u tÃ i chÃ­nh nhiá»u ká»³ há»£p lá»‡")
        valuation = build_module2_valuation_table(c, annual, assumptions)
        current_price = _parse_num(getattr(c, "current_price", None))
        value_range = build_valuation_range(valuation, current_price, float(target_mos_pct))
        moat = build_porter_moat_scorecard(c, annual)
        cls = classify_company(c, annual)
        roe = _recent_median_local(annual, "roe_actual_pct", "roe_pct")
        roic = _recent_median_local(annual, "roic_standard_pct", "roic_pct")
        gross_margin = _recent_median_local(annual, "gross_margin_pct")
        net_margin = _recent_median_local(annual, "net_margin_pct")
        cfo_np = _recent_median_local(annual, "cfo_to_net_profit")
        fcf_np = _recent_median_local(annual, "fcf_to_net_profit")
        revenue_cagr = _cagr_local(annual, "revenue_bil")
        profit_cagr = _cagr_local(annual, "net_profit_bil")
        net_debt_equity = _latest_num_local(annual, "net_debt_to_equity")
        pe = _parse_num(getattr(c, "pe", None))
        pb = _parse_num(getattr(c, "pb", None))
        mos = value_range.mos_to_weighted_pct
        moat_score = float(moat.attrs.get("total_score", 0) or 0)

        quality_score = _mean_ignore_none([_score_high(roic, 15, 10, 5), _score_high(roe, 18, 12, 6), _score_high(gross_margin, 30, 18, 10)])
        cash_score = _mean_ignore_none([_score_high(cfo_np, 1.0, 0.7, 0.3), _score_high(fcf_np, 0.8, 0.3, 0.0)])
        valuation_score = _mean_ignore_none([_score_high(mos, float(target_mos_pct), 0, -20), _score_low(pe, 8, 12, 20), _score_low(pb, 1.0, 1.8, 3.0)])
        risk_penalty = 10 if net_debt_equity is not None and net_debt_equity > 1.5 else 0
        total_score = max(0.0, min(100.0, 0.30 * quality_score + 0.25 * cash_score + 0.25 * moat_score + 0.20 * valuation_score - risk_penalty))
        if total_score >= 80 and mos is not None and mos >= float(target_mos_pct):
            conclusion = "Æ¯u tiÃªn cao: cháº¥t lÆ°á»£ng/Ä‘á»‹nh giÃ¡ cÃ¹ng thuáº­n lá»£i, cáº§n xÃ¡c nháº­n thÃªm báº±ng BCTN vÃ  rá»§i ro ngÃ nh."
        elif total_score >= 72 and (mos or -999) >= 0:
            conclusion = "Theo dÃµi tá»‘t: cháº¥t lÆ°á»£ng tÆ°Æ¡ng Ä‘á»‘i cao nhÆ°ng cáº§n kiá»ƒm tra MOS/chu ká»³ trÆ°á»›c khi giáº£i ngÃ¢n."
        elif total_score >= 60:
            conclusion = "Trung bÃ¬nh: chá»‰ nÃªn xem lÃ  mÃ£ Ä‘á»‘i chiáº¿u hoáº·c chá» giÃ¡/triá»ƒn vá»ng rÃµ hÆ¡n."
        else:
            conclusion = "Tháº­n trá»ng: Ä‘iá»ƒm tá»•ng há»£p yáº¿u hoáº·c dá»¯ liá»‡u/rá»§i ro chÆ°a á»§ng há»™."
        row = {
            "MÃ£": ticker,
            "TÃªn doanh nghiá»‡p": getattr(c, "company_name", ""),
            "SÃ n": getattr(c, "exchange", ""),
            "NgÃ nh": getattr(c, "industry", ""),
            "PhÃ¢n ngÃ nh": getattr(c, "sub_industry", ""),
            "Loáº¡i DN": cls.company_type,
            "GiÃ¡ hiá»‡n táº¡i": current_price,
            "GiÃ¡ trá»‹ weighted": value_range.weighted_vnd,
            "MOS hiá»‡n táº¡i %": mos,
            "P/E": pe,
            "P/B": pb,
            "ROE %": roe,
            "ROIC %": roic,
            "BiÃªn gá»™p %": gross_margin,
            "BiÃªn rÃ²ng %": net_margin,
            "CAGR DT 5Y %": revenue_cagr,
            "CAGR LNST 5Y %": profit_cagr,
            "CFO/LNST": cfo_np,
            "FCF/LNST": fcf_np,
            "Ná»£ rÃ²ng/VCSH": net_debt_equity,
            "Vá»‘n hÃ³a (tá»· Ä‘á»“ng)": _parse_num(getattr(c, "market_cap_bil", None)),
            "Moat score": moat_score,
            "Moat level": moat.attrs.get("level", "N/A"),
            "Äiá»ƒm cháº¥t lÆ°á»£ng": quality_score,
            "Äiá»ƒm dÃ²ng tiá»n": cash_score,
            "Äiá»ƒm Ä‘á»‹nh giÃ¡": valuation_score,
            "Äiá»ƒm tá»•ng há»£p": total_score,
            "Xáº¿p háº¡ng": None,
            "Káº¿t luáº­n so sÃ¡nh": conclusion,
        }
        return row, _company_to_peer_row(c, label, "ÄÃ£ cáº­p nháº­t tá»« lá»‡nh so sÃ¡nh")
    except Exception as exc:
        return {
            "MÃ£": ticker,
            "TÃªn doanh nghiá»‡p": "",
            "SÃ n": "",
            "NgÃ nh": "",
            "PhÃ¢n ngÃ nh": "",
            "Loáº¡i DN": "KhÃ´ng Ä‘á»§ dá»¯ liá»‡u",
            "GiÃ¡ hiá»‡n táº¡i": None,
            "GiÃ¡ trá»‹ weighted": None,
            "MOS hiá»‡n táº¡i %": None,
            "P/E": None,
            "P/B": None,
            "ROE %": None,
            "ROIC %": None,
            "BiÃªn gá»™p %": None,
            "BiÃªn rÃ²ng %": None,
            "CAGR DT 5Y %": None,
            "CAGR LNST 5Y %": None,
            "CFO/LNST": None,
            "FCF/LNST": None,
            "Ná»£ rÃ²ng/VCSH": None,
            "Vá»‘n hÃ³a (tá»· Ä‘á»“ng)": None,
            "Moat score": None,
            "Moat level": "N/A",
            "Äiá»ƒm cháº¥t lÆ°á»£ng": None,
            "Äiá»ƒm dÃ²ng tiá»n": None,
            "Äiá»ƒm Ä‘á»‹nh giÃ¡": None,
            "Äiá»ƒm tá»•ng há»£p": 0,
            "Xáº¿p háº¡ng": None,
            "Káº¿t luáº­n so sÃ¡nh": f"KhÃ´ng so sÃ¡nh Ä‘Æ°á»£c: {exc}",
        }, None


def _rank_peer_comparison(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out["Äiá»ƒm tá»•ng há»£p"] = pd.to_numeric(out.get("Äiá»ƒm tá»•ng há»£p"), errors="coerce").fillna(0)
    out["_mos_sort"] = pd.to_numeric(out.get("MOS hiá»‡n táº¡i %"), errors="coerce").fillna(-999)
    out["_moat_sort"] = pd.to_numeric(out.get("Moat score"), errors="coerce").fillna(-999)
    out = out.sort_values(["Äiá»ƒm tá»•ng há»£p", "_mos_sort", "_moat_sort"], ascending=[False, False, False]).drop(columns=["_mos_sort", "_moat_sort"]).reset_index(drop=True)
    out["Xáº¿p háº¡ng"] = range(1, len(out) + 1)
    return out


def _peer_comparison_summary(df: pd.DataFrame, target_mos_pct: float) -> str:
    if df is None or df.empty:
        return "ChÆ°a cÃ³ káº¿t quáº£ so sÃ¡nh."
    ok = df[pd.to_numeric(df.get("Äiá»ƒm tá»•ng há»£p"), errors="coerce").fillna(0) > 0].copy()
    if ok.empty:
        return "CÃ¡c mÃ£ Ä‘Ã£ chá»n chÆ°a cÃ³ Ä‘á»§ dá»¯ liá»‡u Ä‘á»ƒ so sÃ¡nh."
    top = ok.sort_values("Äiá»ƒm tá»•ng há»£p", ascending=False).iloc[0]
    moat_leaders = ok.sort_values("Moat score", ascending=False).head(3)["MÃ£"].astype(str).tolist() if "Moat score" in ok else []
    mos_candidates = ok[pd.to_numeric(ok.get("MOS hiá»‡n táº¡i %"), errors="coerce").fillna(-999) >= float(target_mos_pct)]["MÃ£"].astype(str).tolist()
    text = (
        f"MÃ£ Ä‘á»©ng Ä‘áº§u theo Ä‘iá»ƒm tá»•ng há»£p lÃ  **{top.get('MÃ£')}** vá»›i {top.get('Äiá»ƒm tá»•ng há»£p', 0):,.1f}/100. "
        f"CÃ¡c mÃ£ cÃ³ moat score ná»•i báº­t: {', '.join(moat_leaders) if moat_leaders else 'chÆ°a Ä‘á»§ dá»¯ liá»‡u'}. "
        f"CÃ¡c mÃ£ Ä‘áº¡t MOS yÃªu cáº§u {float(target_mos_pct):.0f}%: {', '.join(mos_candidates) if mos_candidates else 'chÆ°a cÃ³ mÃ£ nÃ o Ä‘áº¡t'}. "
        "CÃ¡ch Ä‘á»c: báº£ng nÃ y chá»‰ lÃ  bá»™ lá»c tÆ°Æ¡ng Ä‘á»‘i trong cÃ¹ng ngÃ nh; quyáº¿t Ä‘á»‹nh cuá»‘i cÃ¹ng váº«n pháº£i kiá»ƒm tra BCTC gá»‘c, lá»£i tháº¿ cáº¡nh tranh, chu ká»³ ngÃ nh vÃ  sá»± kiá»‡n báº¥t thÆ°á»ng."
    )
    return text


def _build_peer_row_note(rowd: dict) -> str:
    return "\n".join([
        f"SO SÃNH CÃ™NG NGÃ€NH: {rowd.get('MÃ£', 'N/A')} - {rowd.get('TÃªn doanh nghiá»‡p', '')}",
        f"- Xáº¿p háº¡ng: {rowd.get('Xáº¿p háº¡ng', 'N/A')}; Ä‘iá»ƒm tá»•ng há»£p: {_format_note_value(rowd.get('Äiá»ƒm tá»•ng há»£p'))}/100.",
        f"- Cháº¥t lÆ°á»£ng vá»‘n: ROE {_format_note_value(rowd.get('ROE %'))}%; ROIC {_format_note_value(rowd.get('ROIC %'))}%; biÃªn gá»™p {_format_note_value(rowd.get('BiÃªn gá»™p %'))}%.",
        f"- Cháº¥t lÆ°á»£ng dÃ²ng tiá»n: CFO/LNST {_format_note_value(rowd.get('CFO/LNST'))}; FCF/LNST {_format_note_value(rowd.get('FCF/LNST'))}.",
        f"- Äá»‹nh giÃ¡: giÃ¡ hiá»‡n táº¡i {_format_note_value(rowd.get('GiÃ¡ hiá»‡n táº¡i'))}; giÃ¡ trá»‹ weighted {_format_note_value(rowd.get('GiÃ¡ trá»‹ weighted'))}; MOS {_format_note_value(rowd.get('MOS hiá»‡n táº¡i %'))}%; P/E {_format_note_value(rowd.get('P/E'))}; P/B {_format_note_value(rowd.get('P/B'))}.",
        f"- Porter/Moat: {rowd.get('Moat score', 'N/A')}/100 - {rowd.get('Moat level', 'N/A')}.",
        f"- Káº¿t luáº­n tá»± Ä‘á»™ng: {rowd.get('Káº¿t luáº­n so sÃ¡nh', 'N/A')}",
        "NguyÃªn táº¯c: Ä‘iá»ƒm tá»•ng há»£p = 30% cháº¥t lÆ°á»£ng sinh lá»i/vá»‘n + 25% cháº¥t lÆ°á»£ng dÃ²ng tiá»n + 25% Porter Moat + 20% Ä‘á»‹nh giÃ¡/MOS, cÃ³ pháº¡t rá»§i ro náº¿u Ä‘Ã²n báº©y cao. ÄÃ¢y lÃ  bá»™ lá»c tÆ°Æ¡ng Ä‘á»‘i, khÃ´ng thay tháº¿ phÃ¢n tÃ­ch riÃªng tá»«ng mÃ£.",
    ])


def _render_value_chain_spider_chart(value_chain_df: pd.DataFrame, company=None) -> None:
    """Render radar/spider chart from Porter value-chain heat scores."""
    if value_chain_df is None or value_chain_df.empty or "Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹" not in value_chain_df.columns or "Äiá»ƒm nhiá»‡t" not in value_chain_df.columns:
        st.info("ChÆ°a Ä‘á»§ dá»¯ liá»‡u Ä‘iá»ƒm nhiá»‡t Ä‘á»ƒ váº½ biá»ƒu Ä‘á»“ mÃ ng nhá»‡n chuá»—i giÃ¡ trá»‹.")
        return
    chart_df = value_chain_df[["Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹", "Äiá»ƒm nhiá»‡t", "ÄÃ¡nh giÃ¡ sÆ¡ bá»™", "Má»©c Ä‘á»™"]].copy()
    chart_df["Äiá»ƒm nhiá»‡t"] = pd.to_numeric(chart_df["Äiá»ƒm nhiá»‡t"], errors="coerce").fillna(0).clip(0, 100)
    theta = chart_df["Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹"].astype(str).tolist()
    r = chart_df["Äiá»ƒm nhiá»‡t"].astype(float).tolist()
    custom = [
        f"{act}<br>Äiá»ƒm nhiá»‡t: {score:.1f}/100<br>ÄÃ¡nh giÃ¡: {rating}<br>Má»©c Ä‘á»™: {level}"
        for act, score, rating, level in zip(chart_df["Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹"], chart_df["Äiá»ƒm nhiá»‡t"], chart_df["ÄÃ¡nh giÃ¡ sÆ¡ bá»™"], chart_df["Má»©c Ä‘á»™"])
    ]
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=r + r[:1],
        theta=theta + theta[:1],
        fill="toself",
        name=str(getattr(company, "ticker", "Chuá»—i giÃ¡ trá»‹") or "Chuá»—i giÃ¡ trá»‹"),
        text=custom + custom[:1],
        hovertemplate="%{text}<extra></extra>",
    ))
    fig.update_layout(
        title="Biá»ƒu Ä‘á»“ mÃ ng nhá»‡n Ä‘iá»ƒm nhiá»‡t Chuá»—i giÃ¡ trá»‹ Porter",
        polar=dict(radialaxis=dict(visible=True, range=[0, 100], tickfont=dict(size=10))),
        showlegend=False,
        height=520,
        margin=dict(l=50, r=50, t=70, b=45),
    )
    st.plotly_chart(fig, use_container_width=True)
    st.caption("Äiá»ƒm nhiá»‡t láº¥y trá»±c tiáº¿p tá»« báº£ng 'Báº£n Ä‘á»“ chuá»—i giÃ¡ trá»‹ theo Porter': Tá»‘t = 100, Trung bÃ¬nh = 55, Yáº¿u = 15, ChÆ°a Ä‘á»§ dá»¯ liá»‡u/Cáº§n bá»• sung = 35.")





def _render_value_chain_yellow_assessment_card(value_chain_df: pd.DataFrame) -> None:
    """Render a brand-yellow executive assessment card for the Porter value-chain map."""
    if value_chain_df is None or value_chain_df.empty:
        st.markdown(
            """
            <div style="border:2px solid rgba(245,178,27,.72); border-left:12px solid #F5B21B; border-radius:18px;
                        padding:15px 18px; margin:10px 0 14px 0; background:linear-gradient(135deg,#FFF9E8 0%,#FEF3C7 100%);
                        color:#7A4A00; font-weight:850; line-height:1.55; box-shadow:0 9px 22px rgba(245,178,27,.16);">
              <div style="font-size:1.12rem; color:#0B7F75; font-weight:1000; margin-bottom:6px;">ðŸŸ¡ ÄÃ¡nh giÃ¡ chuá»—i giÃ¡ trá»‹ Porter</div>
              <div><b>Tráº¡ng thÃ¡i:</b> ChÆ°a cÃ³ dá»¯ liá»‡u Ä‘á»ƒ tá»•ng há»£p.</div>
              <div><b>Cáº§n bá»• sung:</b> BCTC/BCTN/tin IR hoáº·c dá»¯ liá»‡u Ä‘á»‹nh lÆ°á»£ng tá»« Tá»•ng quan doanh nghiá»‡p Ä‘á»ƒ nháº­n diá»‡n hoáº¡t Ä‘á»™ng nÃ o táº¡o lá»£i tháº¿ chi phÃ­, khÃ¡c biá»‡t hÃ³a hoáº·c rÃ o cáº£n khÃ³ báº¯t chÆ°á»›c.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    df = value_chain_df.copy()
    score_series = pd.to_numeric(df.get("Äiá»ƒm nhiá»‡t", pd.Series(dtype=float)), errors="coerce").dropna()
    avg_score = float(score_series.mean()) if not score_series.empty else 0.0
    rating_series = df.get("ÄÃ¡nh giÃ¡ sÆ¡ bá»™", pd.Series(dtype=object)).fillna("").astype(str)

    good_mask = rating_series.str.contains("Tá»‘t", case=False, na=False)
    medium_mask = rating_series.str.contains("Trung bÃ¬nh", case=False, na=False)
    weak_mask = rating_series.str.contains("Yáº¿u", case=False, na=False)
    missing_mask = rating_series.str.contains("ChÆ°a Ä‘á»§|Cáº§n bá»• sung|Thiáº¿u", case=False, regex=True, na=False)

    good_count = int(good_mask.sum())
    medium_count = int(medium_mask.sum())
    weak_count = int(weak_mask.sum())
    missing_count = int(missing_mask.sum())

    def _items(mask, limit: int = 4) -> list[str]:
        if "Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹" not in df.columns:
            return []
        return df.loc[mask, "Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹"].dropna().astype(str).head(limit).tolist()

    strengths = _items(good_mask) or ["ChÆ°a cÃ³ hoáº¡t Ä‘á»™ng Ä‘Æ°á»£c cháº¥m Tá»‘t"]
    risk_items = (_items(weak_mask) + _items(missing_mask & ~weak_mask)) or ["ChÆ°a cÃ³ Ä‘iá»ƒm yáº¿u ná»•i báº­t tá»« báº£ng hiá»‡n táº¡i"]

    if avg_score >= 75 and weak_count == 0:
        level = "Chuá»—i giÃ¡ trá»‹ máº¡nh"
        conclusion = "Nhiá»u hoáº¡t Ä‘á»™ng cÃ³ tÃ­n hiá»‡u lá»£i tháº¿ rÃµ; cÃ³ thá»ƒ xem lÃ  Ä‘iá»ƒm cá»™ng cho moat náº¿u báº±ng chá»©ng Ä‘á»‹nh tÃ­nh trong BCTN/IR xÃ¡c nháº­n Ä‘Æ°á»£c tÃ­nh bá»n vá»¯ng."
    elif avg_score >= 55:
        level = "Chuá»—i giÃ¡ trá»‹ khÃ¡ / cáº§n kiá»ƒm chá»©ng"
        conclusion = "CÃ³ tÃ­n hiá»‡u lá»£i tháº¿ á»Ÿ má»™t sá»‘ hoáº¡t Ä‘á»™ng, nhÆ°ng chÆ°a nÃªn káº¿t luáº­n moat máº¡nh náº¿u cÃ²n nhiá»u khÃ¢u thiáº¿u báº±ng chá»©ng hoáº·c chá»‰ sá»‘ Ä‘á»‹nh lÆ°á»£ng chÆ°a Ä‘á»“ng thuáº­n."
    elif avg_score >= 40:
        level = "Chuá»—i giÃ¡ trá»‹ trung bÃ¬nh / chÆ°a rÃµ moat"
        conclusion = "Lá»£i tháº¿ cáº¡nh tranh chÆ°a Ä‘á»§ rÃµ; cáº§n Æ°u tiÃªn kiá»ƒm tra hoáº¡t Ä‘á»™ng táº¡o chi phÃ­ tháº¥p, khÃ¡c biá»‡t hÃ³a vÃ  kháº£ nÄƒng duy trÃ¬ ROIC/biÃªn lá»£i nhuáº­n qua chu ká»³."
    else:
        level = "Cáº£nh bÃ¡o chuá»—i giÃ¡ trá»‹ yáº¿u"
        conclusion = "CÃ¡c tÃ­n hiá»‡u hiá»‡n táº¡i nghiÃªng vá» yáº¿u hoáº·c thiáº¿u dá»¯ liá»‡u; khÃ´ng nÃªn gÃ¡n moat náº¿u chÆ°a cÃ³ báº±ng chá»©ng máº¡nh tá»« bÃ¡o cÃ¡o gá»‘c vÃ  so sÃ¡nh ngÃ nh."

    evidence_lines: list[str] = []
    evidence_col = "Báº±ng chá»©ng hiá»‡n cÃ³/cáº§n tÃ¬m"
    if evidence_col in df.columns:
        for _, row in df.head(8).iterrows():
            act = str(row.get("Hoáº¡t Ä‘á»™ng chuá»—i giÃ¡ trá»‹", "")).strip()
            ev = str(row.get(evidence_col, "")).strip()
            if act and ev:
                evidence_lines.append(f"<li><b>{html.escape(act)}:</b> {html.escape(ev)}</li>")
            if len(evidence_lines) >= 4:
                break

    evidence_html = ""
    if evidence_lines:
        evidence_html = "<ul style='margin:10px 0 5px 18px; padding:0; line-height:1.45;'>" + "".join(evidence_lines) + "</ul>"

    note = (
        "Theo Porter, lá»£i tháº¿ cáº¡nh tranh pháº£i truy vá» cÃ¡c hoáº¡t Ä‘á»™ng cá»¥ thá»ƒ trong chuá»—i giÃ¡ trá»‹: "
        "hoáº¡t Ä‘á»™ng nÃ o lÃ m chi phÃ­ tháº¥p hÆ¡n, hoáº¡t Ä‘á»™ng nÃ o táº¡o khÃ¡c biá»‡t hÃ³a Ä‘Æ°á»£c khÃ¡ch hÃ ng tráº£ tiá»n, "
        "vÃ  hoáº¡t Ä‘á»™ng nÃ o khÃ³ bá»‹ Ä‘á»‘i thá»§ sao chÃ©p. Card nÃ y lÃ  Ä‘Ã¡nh giÃ¡ tá»± Ä‘á»™ng; click/chá»n tá»«ng dÃ²ng trong báº£ng Ä‘á»ƒ xem note chi tiáº¿t."
    )

    st.markdown(
        f"""
        <div style="border:2px solid rgba(245,178,27,.78); border-left:14px solid #F5B21B; border-radius:20px;
                    padding:16px 19px; margin:10px 0 16px 0; background:linear-gradient(135deg,#FFF9E8 0%,#FEF3C7 72%,#FFE8A3 100%);
                    color:#7A4A00; font-weight:850; line-height:1.55; box-shadow:0 10px 25px rgba(245,178,27,.18);">
            <div style="display:flex; align-items:flex-start; justify-content:space-between; gap:16px; flex-wrap:wrap;">
                <div>
                    <div style="font-size:1.16rem; color:#0B7F75; font-weight:1000; margin-bottom:4px;">ðŸŸ¡ ÄÃ¡nh giÃ¡ tá»•ng há»£p chuá»—i giÃ¡ trá»‹ Porter</div>
                    <div style="font-size:.97rem; color:#7A4A00; font-weight:900;">Má»©c Ä‘Ã¡nh giÃ¡: <b>{html.escape(level)}</b></div>
                </div>
                <div style="min-width:142px; text-align:center; border:1.8px solid rgba(245,178,27,.75); border-radius:18px; padding:10px 13px; background:rgba(255,255,255,.84);">
                    <div style="font-size:.78rem; color:#64748B; font-weight:950; text-transform:uppercase;">Äiá»ƒm nhiá»‡t TB</div>
                    <div style="font-size:1.74rem; line-height:1.05; color:#064E47; font-weight:1000;">{avg_score:.1f}/100</div>
                </div>
            </div>
            <div style="display:grid; grid-template-columns:repeat(4,minmax(110px,1fr)); gap:10px; margin:12px 0 11px 0;">
                <div style="border-radius:14px; padding:9px 10px; background:rgba(255,255,255,.76); border:1px solid rgba(245,178,27,.42);"><b>{good_count}</b><br><span style="font-size:.84rem;">Hoáº¡t Ä‘á»™ng tá»‘t</span></div>
                <div style="border-radius:14px; padding:9px 10px; background:rgba(255,255,255,.76); border:1px solid rgba(245,178,27,.42);"><b>{medium_count}</b><br><span style="font-size:.84rem;">Trung bÃ¬nh</span></div>
                <div style="border-radius:14px; padding:9px 10px; background:rgba(255,255,255,.76); border:1px solid rgba(245,178,27,.42);"><b>{weak_count}</b><br><span style="font-size:.84rem;">Cáº£nh bÃ¡o yáº¿u</span></div>
                <div style="border-radius:14px; padding:9px 10px; background:rgba(255,255,255,.76); border:1px solid rgba(245,178,27,.42);"><b>{missing_count}</b><br><span style="font-size:.84rem;">Cáº§n bá»• sung</span></div>
            </div>
            <div>
                <b>Hoáº¡t Ä‘á»™ng ná»•i báº­t:</b> {html.escape(', '.join(strengths[:5]))}<br>
                <b>Äiá»ƒm yáº¿u/cáº§n kiá»ƒm tra:</b> {html.escape(', '.join(risk_items[:5]))}<br>
                <b>Káº¿t luáº­n tá»± Ä‘á»™ng:</b> {html.escape(conclusion)}
            </div>
            {evidence_html}
            <div style="font-size:.88rem; color:#475569; margin-top:9px;">{html.escape(note)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def _render_moat_spider_chart(moat_df: pd.DataFrame, company=None) -> None:
    """Render radar/spider chart from Porter moat score heat values.

    Äiá»ƒm nhiá»‡t cá»§a báº£ng moat Ä‘Æ°á»£c chuáº©n hÃ³a = Äiá»ƒm Ä‘áº¡t / Trá»ng sá»‘ * 100. NhÆ° váº­y má»—i trá»¥c
    Ä‘á»u náº±m trong thang 0-100, khÃ´ng bá»‹ mÃ©o vÃ¬ má»—i nhÃ³m Porter cÃ³ trá»ng sá»‘ khÃ¡c nhau.
    """
    if moat_df is None or moat_df.empty or "NhÃ³m Porter/Moat" not in moat_df.columns:
        st.info("ChÆ°a Ä‘á»§ dá»¯ liá»‡u Ä‘á»ƒ váº½ biá»ƒu Ä‘á»“ mÃ ng nhá»‡n Porter Moat Score.")
        return
    chart_df = moat_df.copy()
    if "Trá»ng sá»‘ %" not in chart_df.columns or "Äiá»ƒm Ä‘áº¡t" not in chart_df.columns:
        st.info("Báº£ng Porter Moat chÆ°a cÃ³ cá»™t Trá»ng sá»‘ %/Äiá»ƒm Ä‘áº¡t Ä‘á»ƒ váº½ Ä‘iá»ƒm nhiá»‡t.")
        return
    chart_df["_weight"] = pd.to_numeric(chart_df["Trá»ng sá»‘ %"], errors="coerce")
    chart_df["_score"] = pd.to_numeric(chart_df["Äiá»ƒm Ä‘áº¡t"], errors="coerce")
    chart_df["Äiá»ƒm nhiá»‡t"] = (chart_df["_score"] / chart_df["_weight"].replace(0, pd.NA) * 100).clip(0, 100).fillna(0)
    theta = chart_df["NhÃ³m Porter/Moat"].astype(str).tolist()
    r = chart_df["Äiá»ƒm nhiá»‡t"].astype(float).tolist()
    custom = []
    for _, row in chart_df.iterrows():
        custom.append(
            f"{row.get('NhÃ³m Porter/Moat','')}<br>Äiá»ƒm nhiá»‡t: {row.get('Äiá»ƒm nhiá»‡t',0):.1f}/100"
            f"<br>Äiá»ƒm Ä‘áº¡t: {row.get('Äiá»ƒm Ä‘áº¡t','N/A')}/{row.get('Trá»ng sá»‘ %','N/A')}"
            f"<br>TÃ­n hiá»‡u: {row.get('TÃ­n hiá»‡u','N/A')}"
        )
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=r + r[:1],
        theta=theta + theta[:1],
        fill="toself",
        name=str(getattr(company, "ticker", "Porter Moat") or "Porter Moat"),
        text=custom + custom[:1],
        hovertemplate="%{text}<extra></extra>",
    ))
    fig.update_layout(
        title="Biá»ƒu Ä‘á»“ mÃ ng nhá»‡n Ä‘iá»ƒm nhiá»‡t Porter Moat Score",
        polar=dict(radialaxis=dict(visible=True, range=[0, 100], tickfont=dict(size=10))),
        showlegend=False,
        height=520,
        margin=dict(l=50, r=50, t=70, b=45),
    )
    st.plotly_chart(fig, use_container_width=True)
    st.caption("Äiá»ƒm nhiá»‡t = Äiá»ƒm Ä‘áº¡t / Trá»ng sá»‘ cá»§a tá»«ng nhÃ³m Porter, quy Ä‘á»•i vá» thang 0-100 Ä‘á»ƒ so sÃ¡nh trá»±c quan.")


def _simplize_peer_display_df(df: pd.DataFrame, current_ticker: str = "") -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    src = _normalize_peer_universe(df)
    out = pd.DataFrame({
        "MÃ£ cá»• phiáº¿u": src.get("ticker", pd.Series(dtype=str)),
        "TÃªn doanh nghiá»‡p": src.get("company_name", pd.Series(dtype=str)),
        "GiÃ¡ hiá»‡n táº¡i": src.get("current_price", pd.Series(dtype=float)),
        "Vá»‘n hÃ³a (tá»· Ä‘á»“ng)": src.get("market_cap_bil", pd.Series(dtype=float)),
        "Biáº¿n Ä‘á»™ng giÃ¡ %": src.get("price_change_pct", pd.Series(dtype=float)),
        "7 ngÃ y %": src.get("change_7d_pct", pd.Series(dtype=float)),
        "1 nÄƒm %": src.get("change_1y_pct", pd.Series(dtype=float)),
        "P/E": src.get("pe", pd.Series(dtype=float)),
        "P/B": src.get("pb", pd.Series(dtype=float)),
        "ROE %": src.get("roe_pct", pd.Series(dtype=float)),
        "T.trÆ°á»Ÿng LNST 3Y dá»± phÃ³ng %": src.get("forecast_profit_growth_3y_pct", pd.Series(dtype=float)),
        "Tá»· suáº¥t cá»• tá»©c %": src.get("dividend_yield_pct", pd.Series(dtype=float)),
        "SÃ n": src.get("exchange", pd.Series(dtype=str)),
    })
    current = _safe_ticker(current_ticker)
    out["_active"] = out["MÃ£ cá»• phiáº¿u"].astype(str).map(_safe_ticker).eq(current)
    # Chá»‰ hiá»ƒn thá»‹ tá»‘i Ä‘a 100 mÃ£ nhÆ° yÃªu cáº§u; Æ°u tiÃªn mÃ£ Ä‘ang phÃ¢n tÃ­ch Ä‘á»©ng Ä‘áº§u, sau Ä‘Ã³ theo vá»‘n hÃ³a giáº£m dáº§n.
    cap = pd.to_numeric(out.get("Vá»‘n hÃ³a (tá»· Ä‘á»“ng)"), errors="coerce")
    out = out.assign(_cap_sort=cap.fillna(-1)).sort_values(["_active", "_cap_sort", "MÃ£ cá»• phiáº¿u"], ascending=[False, False, True]).drop(columns=["_cap_sort"]).head(100).reset_index(drop=True)
    return out


def _render_simplize_peer_sortable_table(df: pd.DataFrame, current_ticker: str = "") -> pd.DataFrame:
    """Render a sortable/selectable peer list and return edited rows."""
    display_df = _simplize_peer_display_df(df, current_ticker)
    if display_df.empty:
        st.info("ChÆ°a cÃ³ danh sÃ¡ch cá»• phiáº¿u cÃ¹ng ngÃ nh.")
        return pd.DataFrame()
    current = _safe_ticker(current_ticker)
    active_mask = display_df.get("_active", pd.Series(False, index=display_df.index)).fillna(False).astype(bool)
    display_df.insert(0, "Chá»n", active_mask)
    if "_active" in display_df.columns:
        # Äáº£m báº£o dÃ²ng mÃ£ gá»‘c luÃ´n hiá»‡n Ä‘á»§ ticker/tÃªn, khÃ´ng cÃ²n dÃ²ng trá»‘ng chá»‰ cÃ³ icon.
        display_df.loc[active_mask, "MÃ£ cá»• phiáº¿u"] = display_df.loc[active_mask, "MÃ£ cá»• phiáº¿u"].replace("", current)
        display_df.loc[active_mask & display_df["TÃªn doanh nghiá»‡p"].astype(str).str.strip().isin(["", "nan", "None"]), "TÃªn doanh nghiá»‡p"] = current + " - mÃ£ Ä‘ang phÃ¢n tÃ­ch"
        display_df.loc[active_mask, "MÃ£ cá»• phiáº¿u"] = display_df.loc[active_mask, "MÃ£ cá»• phiáº¿u"].astype(str).str.replace("ðŸŽ¯", "", regex=False).str.strip().replace("", current)
        display_df.loc[active_mask, "TÃªn doanh nghiá»‡p"] = display_df.loc[active_mask, "TÃªn doanh nghiá»‡p"].astype(str).map(lambda x: ("ðŸŽ¯ " + x.replace("ðŸŽ¯", "").strip()) if x.replace("ðŸŽ¯", "").strip() else "ðŸŽ¯ " + current + " - mÃ£ Ä‘ang phÃ¢n tÃ­ch")
    editor_df = display_df.drop(columns=["_active"], errors="ignore")
    st.caption("Báº£ng danh sÃ¡ch cÃ¹ng ngÃ nh, tá»‘i Ä‘a 100 mÃ£. DÃ²ng cÃ³ kÃ½ hiá»‡u ðŸŽ¯ lÃ  mÃ£ Ä‘ang phÃ¢n tÃ­ch, Ä‘Æ°á»£c tick máº·c Ä‘á»‹nh vÃ  tÃ´ vÃ ng thÆ°Æ¡ng hiá»‡u; cÃ³ thá»ƒ báº¥m tiÃªu Ä‘á» cá»™t Ä‘á»ƒ sort tÄƒng/giáº£m trong báº£ng.")
    def _active_peer_row_style(row: pd.Series) -> list[str]:
        code = _safe_ticker(str(row.get("MÃ£ cá»• phiáº¿u", "")).replace("ðŸŽ¯", ""))
        if code == current:
            return ["background-color:#FFF4C7; color:#3B2600; font-weight:900; border-top:2px solid #F5B21B; border-bottom:2px solid #F5B21B;" for _ in row]
        return ["" for _ in row]
    styled_editor_df = editor_df.style.apply(_active_peer_row_style, axis=1)
    edited = st.data_editor(
        styled_editor_df,
        use_container_width=True,
        height=620,
        hide_index=True,
        disabled=[c for c in editor_df.columns if c != "Chá»n"],
        key=f"simplize_peer_selector_{current}_{len(editor_df)}",
        column_config={
            "Chá»n": st.column_config.CheckboxColumn("Chá»n", help="Tick Ä‘á»ƒ Ä‘Æ°a mÃ£ vÃ o danh sÃ¡ch so sÃ¡nh", default=False),
            "GiÃ¡ hiá»‡n táº¡i": st.column_config.NumberColumn("GiÃ¡ hiá»‡n táº¡i", format="%d"),
            "Biáº¿n Ä‘á»™ng giÃ¡ %": st.column_config.NumberColumn("Biáº¿n Ä‘á»™ng giÃ¡ %", format="%.2f%%"),
            "7 ngÃ y %": st.column_config.NumberColumn("7 ngÃ y %", format="%.2f%%"),
            "1 nÄƒm %": st.column_config.NumberColumn("1 nÄƒm %", format="%.2f%%"),
            "P/E": st.column_config.NumberColumn("P/E", format="%.2f"),
            "P/B": st.column_config.NumberColumn("P/B", format="%.2f"),
            "ROE %": st.column_config.NumberColumn("ROE %", format="%.2f%%"),
            "T.trÆ°á»Ÿng LNST 3Y dá»± phÃ³ng %": st.column_config.NumberColumn("T.trÆ°á»Ÿng LNST 3Y dá»± phÃ³ng %", format="%.2f%%"),
            "Tá»· suáº¥t cá»• tá»©c %": st.column_config.NumberColumn("Tá»· suáº¥t cá»• tá»©c %", format="%.2f%%"),
            "Vá»‘n hÃ³a (tá»· Ä‘á»“ng)": st.column_config.NumberColumn("Vá»‘n hÃ³a (tá»· Ä‘á»“ng)", format="%d"),
        },
    )
    return pd.DataFrame(edited)

def _render_peer_universe_and_comparison(company, source: str, assumptions: dict, target_mos_pct: float, available_tickers: list[str], *, auto_simplize: bool = False, simplize_industry_url: str = "") -> None:
    st.subheader("So sÃ¡nh doanh nghiá»‡p cÃ¹ng ngÃ nh")
    st.caption("App tá»± láº¥y mÃ£ Ä‘ang phÃ¢n tÃ­ch tá»« Tá»•ng quan doanh nghiá»‡p, cáº­p nháº­t danh sÃ¡ch cá»• phiáº¿u cÃ¹ng ngÃ nh, rá»“i cho chá»n tá»‘i Ä‘a 10 doanh nghiá»‡p Ä‘á»ƒ so sÃ¡nh.")
    universe = _ensure_current_peer_row(_merge_peer_rows(_load_peer_universe(), [_company_to_peer_row(company, "MÃ£ Ä‘ang phÃ¢n tÃ­ch", "Tá»± thÃªm vÃ o peer universe")]), company)
    current_ticker = _safe_ticker(str(getattr(company, "ticker", "")))
    auto_peer_group = None
    if auto_simplize and current_ticker:
        with st.spinner(f"Äang láº¥y danh sÃ¡ch cÃ¹ng ngÃ nh cho {current_ticker}..."):
            simplize_peers, peer_note, raw_peer_path = _load_simplize_peer_rows(current_ticker, simplize_industry_url)
        if not simplize_peers.empty:
            universe = _ensure_current_peer_row(_merge_peer_rows(universe, simplize_peers.to_dict("records")), company)
            auto_peer_group = str(simplize_peers["peer_group"].dropna().astype(str).iloc[0]) if "peer_group" in simplize_peers.columns and not simplize_peers.empty else None
            _save_peer_universe(universe)
            st.success(_public_text(peer_note))
        else:
            st.warning(_public_text(peer_note))
        if raw_peer_path:
            pass
    col_fetch, col_clear = st.columns([0.55, 0.45])
    with col_fetch:
        if st.button("ðŸ”„ Láº¥y danh sÃ¡ch cÃ¹ng ngÃ nh", use_container_width=True):
            _simplize_industry_peers_cached.clear()
            simplize_peers, peer_note, raw_peer_path = _load_simplize_peer_rows(current_ticker, simplize_industry_url)
            if not simplize_peers.empty:
                universe = _ensure_current_peer_row(_merge_peer_rows(universe, simplize_peers.to_dict("records")), company)
                auto_peer_group = str(simplize_peers["peer_group"].dropna().astype(str).iloc[0]) if "peer_group" in simplize_peers.columns and not simplize_peers.empty else None
                _save_peer_universe(universe)
                st.success(_public_text(peer_note))
            else:
                st.warning(_public_text(peer_note))
            if raw_peer_path:
                pass
    with col_clear:
        if st.button("ðŸ§¹ XÃ³a káº¿t quáº£ so sÃ¡nh táº¡m", use_container_width=True):
            st.session_state["peer_compare_result"] = pd.DataFrame()
            st.info("ÄÃ£ xÃ³a káº¿t quáº£ so sÃ¡nh táº¡m trong phiÃªn hiá»‡n táº¡i.")

    # V23.29: bá» import CSV vÃ  pháº§n thÃªm nhanh mÃ£ ngÃ nh. MÃ£ thá»§ cÃ´ng chá»‰ nháº­p táº¡i pháº§n ra lá»‡nh so sÃ¡nh.

    universe = _ensure_current_peer_row(universe, company)
    current_rows_for_group = universe[universe["ticker"].astype(str).map(_safe_ticker).eq(current_ticker)] if "ticker" in universe.columns else pd.DataFrame()
    current_group_from_row = str(current_rows_for_group.iloc[0].get("peer_group", "")).strip() if not current_rows_for_group.empty else ""
    current_group = auto_peer_group or current_group_from_row or _company_peer_group(company)
    groups = ["Táº¥t cáº£"] + sorted([g for g in universe["peer_group"].dropna().astype(str).unique().tolist() if g.strip()])
    default_idx = groups.index(current_group) if current_group in groups else 0
    selected_group = st.selectbox("Lá»c nhÃ³m ngÃ nh", groups, index=default_idx)
    view_df = universe if selected_group == "Táº¥t cáº£" else universe[universe["peer_group"].astype(str) == selected_group]
    if current_ticker and "ticker" in universe.columns and not view_df["ticker"].astype(str).map(_safe_ticker).eq(current_ticker).any():
        cur_row = universe[universe["ticker"].astype(str).map(_safe_ticker).eq(current_ticker)]
        if not cur_row.empty:
            view_df = pd.concat([cur_row, view_df], ignore_index=True)

    st.subheader("Báº£ng danh sÃ¡ch cá»• phiáº¿u cÃ¹ng ngÃ nh")
    selection_table = _render_simplize_peer_sortable_table(view_df, current_ticker)

    st.divider()
    st.subheader("Chá»n mÃ£ vÃ  ra lá»‡nh so sÃ¡nh")
    selected_from_ticks = []
    if isinstance(selection_table, pd.DataFrame) and not selection_table.empty and "Chá»n" in selection_table.columns:
        mask = selection_table["Chá»n"].fillna(False).astype(bool)
        selected_from_ticks = [x for x in selection_table.loc[mask, "MÃ£ cá»• phiáº¿u"].astype(str).str.replace("ðŸŽ¯", "", regex=False).map(_safe_ticker).tolist() if _is_probable_peer_ticker(x)]
    manual_codes = st.text_input("Nháº­p thÃªm mÃ£ thá»§ cÃ´ng Ä‘á»ƒ so sÃ¡nh, cÃ¡ch nhau báº±ng dáº¥u ','", value="", placeholder="VÃ­ dá»¥: CII, HAH, GMD, VSC")
    selected_manual = [_safe_ticker(x) for x in re.split(r"[,;\s]+", manual_codes.upper()) if _is_probable_peer_ticker(_safe_ticker(x))]
    selected_preview = list(dict.fromkeys([current_ticker] + [x for x in selected_from_ticks + selected_manual if x and x != current_ticker]))[:10]
    st.caption(f"Danh sÃ¡ch sáº½ so sÃ¡nh: {', '.join(selected_preview) if selected_preview else 'chÆ°a chá»n'}")
    peer_source_options = list(PEER_SOURCE_DISPLAY_TO_INTERNAL.keys())
    peer_source_display = st.selectbox("Cháº¿ Ä‘á»™ dá»¯ liá»‡u khi cháº¡y so sÃ¡nh", peer_source_options, index=0, key="peer_compare_source")
    source_for_peer = _to_internal_peer_source(peer_source_display, source)
    total_requested = 1 + len([x for x in list(dict.fromkeys(selected_from_ticks + selected_manual)) if x and x != current_ticker])
    if total_requested > 10:
        st.warning("Anh Ä‘ang chá»n/nháº­p quÃ¡ 10 mÃ£ tÃ­nh cáº£ mÃ£ Ä‘ang phÃ¢n tÃ­ch; app sáº½ láº¥y mÃ£ gá»‘c + 9 mÃ£ Ä‘áº§u tiÃªn khi báº¥m so sÃ¡nh.")
    if st.button("âš–ï¸ So sÃ¡nh doanh nghiá»‡p", use_container_width=True):
        selected = selected_preview
        st.session_state["module3_base_ticker"] = current_ticker
        if len(selected) < 2:
            st.error("Cáº§n chá»n hoáº·c nháº­p tá»‘i thiá»ƒu 1 mÃ£ so sÃ¡nh ngoÃ i mÃ£ Ä‘ang phÃ¢n tÃ­ch.")
        else:
            rows, peer_rows = [], []
            preserve_keys = ["active_ticker", "module1_ticker", "module2_ticker", "shared_ticker", "active_overview_csv", "active_year_csv", "active_quarter_csv", "active_source_label", "module_sync_status"]
            preserved_state = {k: st.session_state.get(k) for k in preserve_keys if k in st.session_state}
            progress = st.progress(0, text="Äang táº£i dá»¯ liá»‡u vÃ  tÃ­nh Ä‘iá»ƒm peer...")
            try:
                for i, code in enumerate(selected, start=1):
                    row, peer_row = _peer_snapshot(code, source_for_peer, assumptions, float(target_mos_pct))
                    row["MÃ£ Ä‘ang phÃ¢n tÃ­ch"] = (_safe_ticker(code) == current_ticker)
                    rows.append(row)
                    if peer_row:
                        peer_rows.append(peer_row)
                    progress.progress(i / len(selected), text=f"ÄÃ£ xá»­ lÃ½ {i}/{len(selected)}: {code}")
            finally:
                progress.empty()
                # So sÃ¡nh peer khÃ´ng Ä‘Æ°á»£c lÃ m Ä‘á»•i mÃ£/bá»™ dá»¯ liá»‡u chÃ­nh Ä‘ang hiá»ƒn thá»‹ trÃªn dashboard.
                for k in preserve_keys:
                    if k in preserved_state:
                        st.session_state[k] = preserved_state[k]
            result = _rank_peer_comparison(pd.DataFrame(rows))
            st.session_state["peer_compare_result"] = result
            if peer_rows:
                _save_peer_universe(_merge_peer_rows(universe, peer_rows))
            st.success("ÄÃ£ hoÃ n táº¥t so sÃ¡nh peer.")

    result = st.session_state.get("peer_compare_result", pd.DataFrame())
    if isinstance(result, pd.DataFrame) and not result.empty:
        _render_important_red("Nháº­n Ä‘á»‹nh so sÃ¡nh cÃ¹ng ngÃ nh", _peer_comparison_summary(result, float(target_mos_pct)))
        display_result = result.drop(columns=[c for c in ["MÃ£ Ä‘ang phÃ¢n tÃ­ch", "Nguá»“n dá»¯ liá»‡u", "source", "Source", "NgÃ nh", "PhÃ¢n ngÃ nh"] if c in result.columns], errors="ignore")
        _render_explainable_table(display_result, "peer_compare", height=520)
        export_result = display_result
        st.download_button("Táº£i káº¿t quáº£ so sÃ¡nh peer", export_result.to_csv(index=False, encoding="utf-8-sig"), file_name=f"peer_compare_{_safe_ticker(str(getattr(company, 'ticker', '')))}.csv", mime="text/csv", use_container_width=True)


def _render_tre_sidebar_nav() -> None:
    """Manual branded navigation so Streamlit never exposes the technical root page name 'app'."""
    st.markdown("### Äiá»u hÆ°á»›ng")
    st.page_link("app.py", label="Tá»•ng quan doanh nghiá»‡p", icon="ðŸ“Š")
    st.page_link("pages/02_Dinh_gia_Porter_Moat.py", label="Äá»‹nh giÃ¡ chuyÃªn sÃ¢u", icon="ðŸ§ ")
    st.page_link("pages/03_So_sanh_doanh_nghiep.py", label="So sÃ¡nh doanh nghiá»‡p", icon="âš–ï¸")
    st.page_link("pages/04_Bao_cao_tong_hop.py", label="BÃ¡o cÃ¡o tá»•ng há»£p toÃ n bá»™ ná»™i dung", icon="ðŸ“„")
    st.divider()


def render_dashboard() -> None:
    _inject_runtime_ui_css()
    _render_brand_page_header(
        f"ðŸ§  {APP_NAME}",
        "Trecapital valuation dashboard | Tá»± Ä‘á»“ng bá»™ BCTC Tá»•ng quan doanh nghiá»‡p â‡„ Äá»‹nh giÃ¡ chuyÃªn sÃ¢u, tÃ¬m evidence internet, Ä‘á»‹nh giÃ¡ nhiá»u lá»›p vÃ  cháº¥m lá»£i tháº¿ cáº¡nh tranh theo Porter.",
    )

    available_tickers = _available_financial_tickers_cached(str(BUNDLED_XLSM)) if BUNDLED_XLSM.exists() else []

    with st.sidebar:
        _render_tre_sidebar_nav()
        st.header("Thiáº¿t láº­p phÃ¢n tÃ­ch")
        source_display = st.selectbox(
            "Cháº¿ Ä‘á»™ dá»¯ liá»‡u tÃ i chÃ­nh",
            list(DATA_SOURCE_DISPLAY_TO_INTERNAL.keys()),
            index=0,
        )
        source = _to_internal_source(source_display)
        ticker = st.text_input("MÃ£ cá»• phiáº¿u", value=st.session_state.get("module2_ticker", st.session_state.get("last_query_ticker", "DGC")), max_chars=12).upper()
        mos_canonical = _prepare_mos_widget("module2_mos_widget")
        target_mos_pct = st.selectbox(
            "Má»©c MOS yÃªu cáº§u (%)",
            MOS_OPTIONS_GLOBAL,
            index=MOS_OPTIONS_GLOBAL.index(mos_canonical),
            key="module2_mos_widget",
            on_change=_commit_mos_widget,
            args=("module2_mos_widget",),
            help="MOS dÃ¹ng chung toÃ n app: chá»n á»Ÿ Äá»‹nh giÃ¡ chuyÃªn sÃ¢u sáº½ tá»± Ä‘á»“ng bá»™ sang Tá»•ng quan doanh nghiá»‡p vÃ  má»i cÃ´ng thá»©c giÃ¡ mua/káº¿t luáº­n sáº½ nháº£y theo.",
        )
        target_mos_pct = float(st.session_state.get("target_mos_pct", target_mos_pct))
        if st.session_state.get("mos_sync_status"):
            st.caption(st.session_state["mos_sync_status"])
        if source == "Financial tÃ­ch há»£p" and available_tickers:
            chosen = st.selectbox("MÃ£ cÃ³ Ä‘á»§ BCTC trong dá»¯ liá»‡u tÃ­ch há»£p", ["-- Giá»¯ mÃ£ Ä‘ang nháº­p --"] + available_tickers, index=0)
            if chosen != "-- Giá»¯ mÃ£ Ä‘ang nháº­p --":
                ticker = chosen
        st.caption("Máº·c Ä‘á»‹nh dÃ¹ng cháº¿ Ä‘á»™ Tá»± Ä‘á»™ng: nháº­p mÃ£ á»Ÿ Tá»•ng quan doanh nghiá»‡p hoáº·c Äá»‹nh giÃ¡ chuyÃªn sÃ¢u Ä‘á»u Ä‘á»“ng bá»™ cÃ¹ng má»™t bá»™ BCTC/cache Ä‘á»ƒ Ä‘á»‹nh giÃ¡.")
        st.caption("WACC Ä‘Æ°á»£c tá»± tÃ­nh theo doanh nghiá»‡p tá»« cÆ¡ cáº¥u vá»‘n, chi phÃ­ ná»£, thuáº¿ vÃ  cost of equity proxy; khÃ´ng cÃ²n WACC tham chiáº¿u nháº­p tay.")
        run_all = st.button("ðŸ”Ž TÃ¬m kiáº¿m/cáº­p nháº­t táº¥t cáº£", use_container_width=True, help="Má»™t nÃºt duy nháº¥t: Ä‘á»“ng bá»™ dá»¯ liá»‡u tÃ i chÃ­nh tá»« Tá»•ng quan doanh nghiá»‡p sang Äá»‹nh giÃ¡ chuyÃªn sÃ¢u vÃ  tÃ¬m báº±ng chá»©ng Ä‘á»‹nh tÃ­nh cho lá»£i tháº¿/rá»§i ro/BCTC.")
        if run_all:
            _export_module1_crawler_cached.clear()
            _export_bundled_financial_cached.clear()
            _load_overview_cached.clear()
            _load_timeseries_cached.clear()
            st.session_state["module2_run_all_requested"] = True
            st.info("Äang cáº­p nháº­t toÃ n bá»™: dá»¯ liá»‡u tÃ i chÃ­nh Tá»•ng quan doanh nghiá»‡p â†’ Äá»‹nh giÃ¡ chuyÃªn sÃ¢u + báº±ng chá»©ng Ä‘á»‹nh tÃ­nh.")
        st.divider()
        terminal_growth = st.slider("Terminal growth (%)", 0.0, 6.0, 3.0, 0.5)
        target_pe = st.slider("P/E má»¥c tiÃªu thÆ°á»ng", 5.0, 20.0, 10.0, 0.5)
        st.session_state["module2_ticker"] = _safe_ticker(ticker) or "DGC"

    ticker = st.session_state["module2_ticker"]
    run_all_requested = bool(st.session_state.pop("module2_run_all_requested", False))
    effective_source = "FireAnt + Vietstock" if run_all_requested and source == "Tá»± Ä‘á»™ng tá»« dá»¯ liá»‡u tá»•ng quan" else source
    load_error = None
    try:
        company, annual_df, quarterly_df, source_label, paths = _load_data(ticker, effective_source)
    except Exception as exc:
        load_error = str(exc)
        _render_no_data(ticker, effective_source, available_tickers, load_error)
        st.stop()

    if not _has_real_financial_data(annual_df):
        _render_no_data(ticker, effective_source, available_tickers)
        st.info("Chi tiáº¿t ká»¹ thuáº­t Ä‘Ã£ Ä‘Æ°á»£c ghi trong nháº­t kÃ½ ná»™i bá»™.")
        st.stop()

    if run_all_requested:
        with st.spinner("Äang tÃ¬m báº±ng chá»©ng Ä‘á»‹nh tÃ­nh cho Äá»‹nh giÃ¡ chuyÃªn sÃ¢u..."):
            _update_module2_web_evidence(company)

    assumptions = load_assumptions(ASSUMPTIONS_PATH)
    assumptions["required_return"] = assumptions.get("required_return_pct", 13.0) / 100
    assumptions["terminal_growth"] = terminal_growth / 100
    assumptions["target_pe_normal"] = target_pe
    assumptions["target_mos_pct"] = float(target_mos_pct)

    valuation_df = build_module2_valuation_table(company, annual_df, assumptions)
    cls = classify_company(company, annual_df)
    current_price = getattr(company, "current_price", None)
    value_range = build_valuation_range(valuation_df, current_price, float(target_mos_pct))
    moat_df = build_porter_moat_scorecard(company, annual_df)
    value_chain_df = build_value_chain_table(company, annual_df)
    scenario_df = build_risk_scenario_table(company, annual_df, value_range)
    beneish_df = build_beneish_mscore_table(company, annual_df)
    accrual_quality_df = build_accrual_quality_table(company, annual_df)
    modified_jones_df = build_modified_jones_kothari_table(company, annual_df)
    rem_df = build_real_earnings_management_table(company, annual_df)
    summary = build_module2_summary(company, annual_df, valuation_df, moat_df)

    st.session_state["module2_note_context"] = {
        "company": company,
        "annual_df": annual_df,
        "quarterly_df": quarterly_df,
        "valuation_df": valuation_df,
        "value_range": value_range,
        "moat_df": moat_df,
        "value_chain_df": value_chain_df,
        "scenario_df": scenario_df,
        "beneish_df": beneish_df,
        "classification": cls,
        "assumptions": assumptions,
        "source_label": source_label,
        "target_mos_pct": float(target_mos_pct),
    }

    updated_text = html.escape(_public_text(getattr(company, "updated_at", "N/A") or "N/A"))
    current_price_text = f"{current_price:,.0f}" if current_price else "N/A"
    st.markdown(
        f"""
        <div class='ticker-title-card'>
            <div class='ticker-title-main'><span class='ticker-title-code'>{html.escape(str(company.ticker))}</span> - <span class='ticker-title-name'>{html.escape(str(company.company_name))}</span></div>
            <div class='current-price-inline-card'>
                <div class='price-label'>GiÃ¡ hiá»‡n táº¡i</div>
                <div class='price-value'>{html.escape(current_price_text)}</div>
                <div class='price-note'>Cáº­p nháº­t: {updated_text}</div>
            </div>
            <div class='ticker-title-meta'><b>PhÃ¢n loáº¡i sÆ¡ bá»™:</b> {html.escape(str(cls.company_type))} &nbsp; | &nbsp; <b>Äá»™ tin cáº­y:</b> {cls.confidence:.0f}/100</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    cols = st.columns(5)
    cols[0].metric("GiÃ¡ trá»‹ weighted", f"{value_range.weighted_vnd:,.0f}" if value_range.weighted_vnd else "N/A")
    cols[1].metric("MOS hiá»‡n táº¡i", f"{value_range.mos_to_weighted_pct:,.1f}%" if value_range.mos_to_weighted_pct is not None else "N/A")
    cols[2].metric("Moat score", f"{moat_df.attrs.get('total_score', 0):,.1f}/100")
    cols[3].metric("Moat level", str(moat_df.attrs.get("level", "N/A")))
    latest_cards = latest_metric_cards(annual_df) if annual_df is not None and not annual_df.empty else {}
    cols[4].metric("Owner Earnings", latest_cards.get("Owner Earnings", "N/A"))

    _render_important_red("TÃ³m táº¯t tá»± Ä‘á»™ng", summary)

    # V23.39: Ä‘Ã£ bá» nÃºt/khung xuáº¥t bÃ¡o cÃ¡o trong tá»«ng pháº§n; chá»‰ cÃ²n trang BÃ¡o cÃ¡o tá»•ng há»£p toÃ n bá»™ ná»™i dung á»Ÿ sidebar.

    tab_val, tab_moat, tab_chain, tab_scenario, tab_beneish, tab_web, tab_data, tab_docs = st.tabs([
        "Äá»‹nh giÃ¡ chuyÃªn sÃ¢u", "Porter Moat Score", "Chuá»—i giÃ¡ trá»‹", "Ká»‹ch báº£n & rá»§i ro", "Thao tÃºng tÃ i chÃ­nh", "Báº±ng chá»©ng Ä‘á»‹nh tÃ­nh", "Dá»¯ liá»‡u", "CÃ´ng thá»©c & giáº£ Ä‘á»‹nh"
    ])

    with tab_val:
        st.markdown("<div class='valuation-tab-compact'>", unsafe_allow_html=True)
        _render_company_type_summary_callout(cls)
        st.subheader("ÄÃ¡nh giÃ¡ trá»ng yáº¿u theo dá»¯ liá»‡u doanh nghiá»‡p")
        assessment_df = _build_strategic_assessment_table(company, annual_df, cls, value_range, moat_df, float(target_mos_pct))
        _render_explainable_table(assessment_df, "strategic_assessment", height=240)

        _render_big_recommendation(value_range.recommendation)
        st.subheader("Dáº£i giÃ¡ trá»‹ ná»™i táº¡i")
        range_df = pd.DataFrame([
            {"Chá»‰ tiÃªu": "Low", "GiÃ¡ trá»‹/cp": value_range.low_vnd},
            {"Chá»‰ tiÃªu": "Base/Median", "GiÃ¡ trá»‹/cp": value_range.base_vnd},
            {"Chá»‰ tiÃªu": "High", "GiÃ¡ trá»‹/cp": value_range.high_vnd},
            {"Chá»‰ tiÃªu": "Weighted", "GiÃ¡ trá»‹/cp": value_range.weighted_vnd},
            {"Chá»‰ tiÃªu": f"GiÃ¡ mua theo MOS chá»n {float(target_mos_pct):.0f}%", "GiÃ¡ trá»‹/cp": value_range.weighted_vnd * (1 - float(target_mos_pct) / 100) if value_range.weighted_vnd else None},
            {"Chá»‰ tiÃªu": "MOS hiá»‡n táº¡i %", "GiÃ¡ trá»‹/cp": value_range.mos_to_weighted_pct},
            {"Chá»‰ tiÃªu": "MOS yÃªu cáº§u %", "GiÃ¡ trá»‹/cp": float(target_mos_pct)},
        ])
        _render_explainable_table(range_df, "valuation_range", height=312)
        st.subheader("Báº£ng Ä‘á»‹nh giÃ¡ theo tá»«ng phÆ°Æ¡ng phÃ¡p")
        _render_explainable_table(valuation_df, "valuation_methods", height=432)
        st.caption("KhÃ´ng dÃ¹ng má»™t fair value duy nháº¥t. Há»‡ thá»‘ng chá»n trá»ng sá»‘ theo loáº¡i doanh nghiá»‡p vÃ  dá»¯ liá»‡u sáºµn cÃ³; cÃ¡c tham sá»‘ cÃ³ thá»ƒ chá»‰nh táº¡i sidebar hoáº·c cáº¥u hÃ¬nh giáº£ Ä‘á»‹nh ná»™i bá»™.")
        st.markdown("</div>", unsafe_allow_html=True)

    with tab_moat:
        st.subheader("Báº£ng Ä‘iá»ƒm lá»£i tháº¿ cáº¡nh tranh theo Porter")
        _render_important_red("Tá»•ng Ä‘iá»ƒm Porter Moat", f"{moat_df.attrs.get('total_score', 0):,.1f}/100 - {moat_df.attrs.get('level', 'N/A')}")
        _render_moat_spider_chart(moat_df, company)
        _render_explainable_table(moat_df, "moat_score", height=430)

    with tab_chain:
        st.subheader("Báº£n Ä‘á»“ chuá»—i giÃ¡ trá»‹ theo Porter")
        st.caption("Má»—i hoáº¡t Ä‘á»™ng Ä‘Æ°á»£c liÃªn káº¿t vá»›i tÃ­n hiá»‡u Ä‘á»‹nh lÆ°á»£ng hiá»‡n cÃ³ vÃ  báº±ng chá»©ng Ä‘á»‹nh tÃ­nh cáº§n tÃ¬m thÃªm trong bÃ¡o cÃ¡o/tin doanh nghiá»‡p.")
        _render_value_chain_yellow_assessment_card(value_chain_df)
        _render_value_chain_spider_chart(value_chain_df, company)
        _render_explainable_table(value_chain_df, "value_chain", height=480)

    with tab_scenario:
        st.subheader("Ká»‹ch báº£n Ä‘á»‹nh giÃ¡ vÃ  rá»§i ro")
        _render_explainable_table(scenario_df, "scenario", height=300)
        st.subheader("TÃ­n hiá»‡u ká»³ gáº§n nháº¥t")
        cards = latest_metric_cards(annual_df)
        card_df = pd.DataFrame([{"Chá»‰ tiÃªu": k, "GiÃ¡ trá»‹": v} for k, v in cards.items()])
        _render_explainable_table(card_df, "latest_cards", height=360)

    with tab_beneish:
        st.subheader("Thao tÃºng tÃ i chÃ­nh - 4 lá»›p cáº£nh bÃ¡o Ä‘á»‹nh lÆ°á»£ng")
        st.markdown(
            """
            <div class='note-card'>
            <b>NguyÃªn táº¯c:</b> app khÃ´ng káº¿t luáº­n doanh nghiá»‡p gian láº­n. Bá»‘n lá»›p dÆ°á»›i Ä‘Ã¢y chá»‰ táº¡o cá» Ä‘á» Ä‘á»‹nh lÆ°á»£ng Ä‘á»ƒ kiá»ƒm tra sÃ¢u cháº¥t lÆ°á»£ng BCTC: lá»£i nhuáº­n cÃ³ Ä‘i kÃ¨m tiá»n tháº­t khÃ´ng, accruals cÃ³ báº¥t thÆ°á»ng khÃ´ng, doanh thu/pháº£i thu/tÃ i sáº£n cÃ³ bá»‹ lÃ m Ä‘áº¹p khÃ´ng vÃ  cÃ³ dáº¥u hiá»‡u quáº£n trá»‹ lá»£i nhuáº­n qua hoáº¡t Ä‘á»™ng tháº­t khÃ´ng.<br><br>
            <b>CÃ¡ch dÃ¹ng trong app:</b> khi má»™t hoáº·c nhiá»u lá»›p cáº£nh bÃ¡o cao, hÃ£y giáº£m Ä‘á»™ tin cáº­y cá»§a lá»£i nhuáº­n káº¿ toÃ¡n khi Ä‘á»‹nh giÃ¡; Ä‘á»c ká»¹ thuyáº¿t minh doanh thu, pháº£i thu, tá»“n kho, kháº¥u hao, chi phÃ­ vá»‘n hÃ³a, giao dá»‹ch bÃªn liÃªn quan, Ã½ kiáº¿n kiá»ƒm toÃ¡n vÃ  Ä‘á»‘i chiáº¿u CFO/LNST/FCF. Vá»›i doanh nghiá»‡p tÃ i chÃ­nh/ngÃ¢n hÃ ng, cÃ¡c mÃ´ hÃ¬nh cÃ´ng nghiá»‡p nhÆ° Beneish/Jones/REM chá»‰ dÃ¹ng tham kháº£o náº¿u dá»¯ liá»‡u Ä‘á»§.
            </div>
            """,
            unsafe_allow_html=True,
        )
        financial_manipulation_summary_df = _build_financial_manipulation_summary_df(
            beneish_df,
            accrual_quality_df,
            modified_jones_df,
            rem_df,
        )
        st.subheader("Tá»•ng há»£p thao tÃºng tÃ i chÃ­nh 4 lá»›p")
        # V23.61: báº£ng tá»•ng há»£p náº±m sÃ¡t nhÃ³m phÃ¢n tÃ­ch tá»«ng lá»›p; váº«n dÃ¹ng format chuáº©n,
        # chá»‰ bá» note/click cá»§a riÃªng báº£ng tá»•ng há»£p vÃ  in Ä‘áº­m cá»™t Lá»›p.
        _render_static_html_table(financial_manipulation_summary_df, "financial_manipulation_summary", height=404)

        layer_beneish, layer_accrual, layer_jones, layer_rem = st.tabs([
            "1. Beneish M-Score",
            "2. Accrual Quality / Sloan",
            "3. Modified Jones / Kothari",
            "4. REM - hoáº¡t Ä‘á»™ng tháº­t",
        ])

        with layer_beneish:
            st.subheader("Lá»›p 1 - Beneish M-Score: cáº£nh bÃ¡o thao tÃºng lá»£i nhuáº­n báº±ng 8 biáº¿n")
            latest_mscore = beneish_df.attrs.get("latest_score") if isinstance(beneish_df, pd.DataFrame) else None
            latest_risk = beneish_df.attrs.get("latest_risk", "N/A") if isinstance(beneish_df, pd.DataFrame) else "N/A"
            latest_period = beneish_df.attrs.get("latest_period", "N/A") if isinstance(beneish_df, pd.DataFrame) else "N/A"
            latest_note = beneish_df.attrs.get("latest_note", "N/A") if isinstance(beneish_df, pd.DataFrame) else "N/A"
            mscore_text = _format_note_value(latest_mscore) if latest_mscore is not None else "N/A"
            latest_missing = ""
            if isinstance(beneish_df, pd.DataFrame) and not beneish_df.empty and "Biáº¿n thiáº¿u/cáº§n kiá»ƒm tra" in beneish_df.columns:
                latest_missing = str(beneish_df.iloc[-1].get("Biáº¿n thiáº¿u/cáº§n kiá»ƒm tra") or "")
            if latest_mscore is None:
                _render_warning_card(
                    "TÃ­n hiá»‡u Beneish M-Score",
                    f"Ká»³ má»›i nháº¥t {latest_period}: chÆ°a Ä‘á»§ 8 biáº¿n Ä‘á»ƒ tÃ­nh M-Score chÃ­nh thá»©c. Biáº¿n thiáº¿u/cáº§n kiá»ƒm tra: {latest_missing or 'N/A'}. {latest_note}"
                )
            else:
                _render_important_red(
                    "TÃ­n hiá»‡u Beneish M-Score",
                    f"Ká»³ má»›i nháº¥t {latest_period}: M-Score {mscore_text}; má»©c cáº£nh bÃ¡o: {latest_risk}. {latest_note}"
                )
            st.caption("CÃ´ng thá»©c: M = -4.84 + 0.920Ã—DSRI + 0.528Ã—GMI + 0.404Ã—AQI + 0.892Ã—SGI + 0.115Ã—DEPI - 0.172Ã—SGAI + 4.679Ã—TATA - 0.327Ã—LVGI. NgÆ°á»¡ng: M-Score > -2.22 lÃ  vÃ¹ng cáº£nh bÃ¡o. AQI Æ°u tiÃªn TSCÄ/PPE tháº­t; náº¿u nguá»“n chá»‰ cÃ³ TÃ i sáº£n dÃ i háº¡n, app dÃ¹ng AQI proxy vÃ  ghi chÃº rÃµ.")
            _render_explainable_table(beneish_df, "beneish_mscore", height=470)

        with layer_accrual:
            st.subheader("Lá»›p 2 - Accrual Quality/Sloan: lá»£i nhuáº­n cÃ³ Ä‘i kÃ¨m dÃ²ng tiá»n tháº­t khÃ´ng")
            latest_risk = accrual_quality_df.attrs.get("latest_risk", "N/A") if isinstance(accrual_quality_df, pd.DataFrame) else "N/A"
            latest_score = accrual_quality_df.attrs.get("latest_score") if isinstance(accrual_quality_df, pd.DataFrame) else None
            latest_note = accrual_quality_df.attrs.get("latest_note", "N/A") if isinstance(accrual_quality_df, pd.DataFrame) else "N/A"
            _render_warning_card("TÃ­n hiá»‡u Accrual Quality/Sloan", f"Ká»³ má»›i nháº¥t: Sloan accrual ratio {_format_note_value(latest_score)}; má»©c cáº£nh bÃ¡o: {latest_risk}. {latest_note}")
            st.caption("CÃ´ng thá»©c chÃ­nh: Sloan accrual ratio = (LNST - CFO) / Tá»•ng tÃ i sáº£n bÃ¬nh quÃ¢n. App kiá»ƒm tra thÃªm CFO/LNST, FCF/LNST vÃ  Balance-sheet accruals = Î”CA - Î”Cash - Î”CL + Î”STD - Kháº¥u hao.")
            _render_explainable_table(accrual_quality_df, "accrual_quality", height=470)

        with layer_jones:
            st.subheader("Lá»›p 3 - Modified Jones/Kothari: discretionary accruals")
            latest_risk = modified_jones_df.attrs.get("latest_risk", "N/A") if isinstance(modified_jones_df, pd.DataFrame) else "N/A"
            latest_score = modified_jones_df.attrs.get("latest_score") if isinstance(modified_jones_df, pd.DataFrame) else None
            latest_note = modified_jones_df.attrs.get("latest_note", "N/A") if isinstance(modified_jones_df, pd.DataFrame) else "N/A"
            _render_warning_card("TÃ­n hiá»‡u Modified Jones/Kothari", f"Ká»³ má»›i nháº¥t: DA Modified Jones {_format_note_value(latest_score)}; má»©c cáº£nh bÃ¡o: {latest_risk}. {latest_note}")
            st.caption("Modified Jones: TA/A(t-1)=Î±0+Î±1(1/A(t-1))+Î±2((Î”REV-Î”REC)/A(t-1))+Î±3(PPE/A(t-1))+Îµ. Kothari thÃªm ROA Ä‘á»ƒ kiá»ƒm soÃ¡t hiá»‡u quáº£ hoáº¡t Ä‘á»™ng. DA=residual Îµ.")
            _render_explainable_table(modified_jones_df, "modified_jones", height=490)

        with layer_rem:
            st.subheader("Lá»›p 4 - Real Earnings Management: quáº£n trá»‹ lá»£i nhuáº­n qua hoáº¡t Ä‘á»™ng tháº­t")
            latest_risk = rem_df.attrs.get("latest_risk", "N/A") if isinstance(rem_df, pd.DataFrame) else "N/A"
            latest_score = rem_df.attrs.get("latest_score") if isinstance(rem_df, pd.DataFrame) else None
            latest_note = rem_df.attrs.get("latest_note", "N/A") if isinstance(rem_df, pd.DataFrame) else "N/A"
            _render_warning_card("TÃ­n hiá»‡u REM", f"Ká»³ má»›i nháº¥t: REM Score {_format_note_value(latest_score)}; má»©c cáº£nh bÃ¡o: {latest_risk}. {latest_note}")
            st.caption("REM gá»“m 3 residual báº¥t thÆ°á»ng: Abnormal CFO, Abnormal PROD = COGS + Î”Inventory, vÃ  Abnormal DISEXP. CFO Ã¢m báº¥t thÆ°á»ng, sáº£n xuáº¥t dÆ°/tá»“n kho cao báº¥t thÆ°á»ng hoáº·c cáº¯t chi phÃ­ tÃ¹y Ã½ lÃ  cÃ¡c cá» Ä‘á» cáº§n kiá»ƒm tra.")
            _render_explainable_table(rem_df, "real_earnings_management", height=490)


    with tab_web:
        st.subheader("Báº±ng chá»©ng Ä‘á»‹nh tÃ­nh cho lá»£i tháº¿/rá»§i ro/BCTC")
        st.markdown(
            """
            <div class='note-card'>
            <b>Tab nÃ y lÃ  kho báº±ng chá»©ng Ä‘á»‹nh tÃ­nh phá»¥c vá»¥ kiá»ƒm tra nháº­n Ä‘á»‹nh.</b><br>
            App dÃ¹ng nÃºt <b>TÃ¬m kiáº¿m/cáº­p nháº­t táº¥t cáº£</b> Ä‘á»ƒ cáº­p nháº­t thÃ´ng tin liÃªn quan Ä‘áº¿n: BCTC, bÃ¡o cÃ¡o thÆ°á»ng niÃªn, tin cÃ´ng bá»‘, moat/lá»£i tháº¿ cáº¡nh tranh, rá»§i ro ngÃ nh, thá»‹ pháº§n, quáº£n trá»‹ vÃ  sá»± kiá»‡n báº¥t thÆ°á»ng.<br><br>
            <b>CÃ¡ch dÃ¹ng:</b> dá»¯ liá»‡u BCTC sá»‘ há»c váº«n Æ°u tiÃªn bá»™ dá»¯ liá»‡u tÃ i chÃ­nh Ä‘Ã£ chuáº©n hÃ³a; báº±ng chá»©ng Ä‘á»‹nh tÃ­nh chá»§ yáº¿u dÃ¹ng Ä‘á»ƒ <b>Ä‘á»‘i chiáº¿u, kiá»ƒm chá»©ng vÃ  giáº£i thÃ­ch</b> cÃ¡c nháº­n Ä‘á»‹nh vá» lá»£i tháº¿/rá»§i ro/BCTC. KhÃ´ng nÃªn xem Ä‘Ã¢y lÃ  cÄƒn cá»© duy nháº¥t Ä‘á»ƒ ra quyáº¿t Ä‘á»‹nh náº¿u chÆ°a Ä‘á»‘i chiáº¿u vá»›i tÃ i liá»‡u gá»‘c.
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.caption("Báº±ng chá»©ng Ä‘á»‹nh tÃ­nh Ä‘Æ°á»£c cáº­p nháº­t báº±ng nÃºt duy nháº¥t á»Ÿ sidebar: 'TÃ¬m kiáº¿m/cáº­p nháº­t táº¥t cáº£'. Náº¿u anh tÃ¬m mÃ£ á»Ÿ Tá»•ng quan doanh nghiá»‡p, dá»¯ liá»‡u nÃ y cÅ©ng Ä‘Æ°á»£c tá»± cáº­p nháº­t sáºµn cho Äá»‹nh giÃ¡ chuyÃªn sÃ¢u.")
        if st.session_state.get("module2_auto_update_status"):
            st.info(_public_text(st.session_state["module2_auto_update_status"]))
        if st.session_state.get("module2_web_note"):
            st.success(_public_text(st.session_state["module2_web_note"]))
        web_ticker = st.session_state.get("module2_web_ticker")
        if web_ticker and _safe_ticker(str(web_ticker)) != _safe_ticker(company.ticker):
            st.warning(f"Báº±ng chá»©ng Ä‘á»‹nh tÃ­nh hiá»‡n Ä‘ang thuá»™c mÃ£ {web_ticker}; báº¥m 'TÃ¬m kiáº¿m/cáº­p nháº­t táº¥t cáº£' Ä‘á»ƒ cáº­p nháº­t láº¡i cho {company.ticker}.")
        _show_table(st.session_state.get("module2_web_table", pd.DataFrame()), height=520)

    with tab_data:
        st.subheader("Dá»¯ liá»‡u nÄƒm + TTM dÃ¹ng cho Äá»‹nh giÃ¡ chuyÃªn sÃ¢u")
        _show_table(format_table_for_display(annual_df), height=480)
        st.download_button("Táº£i dá»¯ liá»‡u nÄƒm + TTM", annual_df.to_csv(index=False, encoding="utf-8-sig"), file_name=f"{company.ticker}_dinh_gia_year_ttm.csv", mime="text/csv")
        st.subheader("Dá»¯ liá»‡u quÃ½")
        _show_table(format_table_for_display(quarterly_df), height=420)
        st.download_button(
            "Táº£i dá»¯ liá»‡u quÃ½",
            quarterly_df.to_csv(index=False, encoding="utf-8-sig"),
            file_name=f"{company.ticker}_dinh_gia_quarter.csv",
            mime="text/csv",
            key=f"download_quarterly_data_{company.ticker}",
        )

    with tab_docs:
        st.subheader("CÃ´ng thá»©c vÃ  giáº£ Ä‘á»‹nh")
        st.markdown(
            """
            <div class='note-card'>
            <b>Khu vá»±c nÃ y hiá»ƒn thá»‹ cÃ´ng thá»©c, giáº£ Ä‘á»‹nh vÃ  nguyÃªn táº¯c Ä‘Ã¡nh giÃ¡.</b><br>
            CÃ¡c Ä‘Æ°á»ng dáº«n ká»¹ thuáº­t, tÃªn nhÃ  cung cáº¥p dá»¯ liá»‡u vÃ  nháº­t kÃ½ ná»™i bá»™ khÃ´ng hiá»ƒn thá»‹ trÃªn giao diá»‡n Ä‘á»ƒ Ä‘áº£m báº£o báº£o máº­t váº­n hÃ nh.
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("### TÃ³m táº¯t cÃ´ng thá»©c chÃ­nh")
        formula_df = pd.DataFrame([
            {"NhÃ³m": "GiÃ¡ trá»‹ ná»™i táº¡i", "CÃ´ng thá»©c/logic": "Dáº£i giÃ¡ trá»‹ Low - Base - High tá»« cÃ¡c phÆ°Æ¡ng phÃ¡p há»£p lá»‡; Weighted lÃ  trung bÃ¬nh trá»ng sá»‘ theo má»©c phÃ¹ há»£p dá»¯ liá»‡u vÃ  loáº¡i doanh nghiá»‡p."},
            {"NhÃ³m": "MOS", "CÃ´ng thá»©c/logic": "MOS hiá»‡n táº¡i = (GiÃ¡ trá»‹ ná»™i táº¡i - GiÃ¡ thá»‹ trÆ°á»ng) / GiÃ¡ trá»‹ ná»™i táº¡i. GiÃ¡ mua MOS = GiÃ¡ trá»‹ ná»™i táº¡i x (1 - MOS yÃªu cáº§u)."},
            {"NhÃ³m": "Owner Earnings", "CÃ´ng thá»©c/logic": "LNST + kháº¥u hao vÃ  chi phÃ­ phi tiá»n máº·t - capex duy trÃ¬ Â± thay Ä‘á»•i vá»‘n lÆ°u Ä‘á»™ng váº­n hÃ nh cáº§n thiáº¿t."},
            {"NhÃ³m": "FCF", "CÃ´ng thá»©c/logic": "DÃ²ng tiá»n tá»± do = CFO - Capex. Náº¿u Capex trong dá»¯ liá»‡u lÃ  sá»‘ Ã¢m thÃ¬ quy Ä‘á»•i Ä‘Ãºng chiá»u dÃ²ng tiá»n trÆ°á»›c khi tÃ­nh."},
            {"NhÃ³m": "ROIC/ROCE", "CÃ´ng thá»©c/logic": "ROIC chÃ­nh = NOPAT / vá»‘n Ä‘áº§u tÆ° bÃ¬nh quÃ¢n; ROCE = EBIT / capital employed khi Ä‘á»§ dá»¯ liá»‡u."},
            {"NhÃ³m": "DuPont", "CÃ´ng thá»©c/logic": "ROE Ä‘Æ°á»£c tÃ¡ch thÃ nh biÃªn lá»£i nhuáº­n, vÃ²ng quay tÃ i sáº£n vÃ  Ä‘Ã²n báº©y tÃ i chÃ­nh Ä‘á»ƒ nháº­n diá»‡n nguá»“n táº¡o ROE."},
            {"NhÃ³m": "Porter Moat", "CÃ´ng thá»©c/logic": "Äiá»ƒm moat lÃ  tá»•ng trá»ng sá»‘ cÃ¡c tiÃªu chÃ­ hiá»‡u quáº£ vá»‘n, lá»£i tháº¿ chi phÃ­/khÃ¡c biá»‡t hÃ³a, cáº¥u trÃºc ngÃ nh, runway, cháº¥t lÆ°á»£ng tÃ i chÃ­nh vÃ  rá»§i ro."},
            {"NhÃ³m": "Beneish M-Score", "CÃ´ng thá»©c/logic": "M = -4.84 + 0.920Ã—DSRI + 0.528Ã—GMI + 0.404Ã—AQI + 0.892Ã—SGI + 0.115Ã—DEPI - 0.172Ã—SGAI + 4.679Ã—TATA - 0.327Ã—LVGI. M > -2.22 lÃ  vÃ¹ng cáº£nh bÃ¡o thao tÃºng lá»£i nhuáº­n."},
            {"NhÃ³m": "Accrual Quality/Sloan", "CÃ´ng thá»©c/logic": "Sloan accrual ratio = (LNST - CFO) / Tá»•ng tÃ i sáº£n bÃ¬nh quÃ¢n. Kiá»ƒm tra thÃªm CFO/LNST, FCF/LNST vÃ  Balance-sheet accruals = Î”CA - Î”Cash - Î”CL + Î”STD - Kháº¥u hao."},
            {"NhÃ³m": "Modified Jones/Kothari", "CÃ´ng thá»©c/logic": "TA/A(t-1)=Î±0+Î±1(1/A(t-1))+Î±2((Î”REV-Î”REC)/A(t-1))+Î±3(PPE/A(t-1))+Îµ. Kothari thÃªm ROA. DA = residual Îµ; DA dÆ°Æ¡ng cao lÃ  accruals lÃ m tÄƒng lá»£i nhuáº­n."},
            {"NhÃ³m": "Real Earnings Management", "CÃ´ng thá»©c/logic": "REM kiá»ƒm tra Abnormal CFO, Abnormal PROD vÃ  Abnormal DISEXP. PROD = COGS + Î”Inventory; DISEXP dÃ¹ng chi phÃ­ bÃ¡n hÃ ng + quáº£n lÃ½ hoáº·c proxy SG&A khi thiáº¿u chi tiáº¿t."},
        ])
        formula_headers = "".join(
            f"<th>{html.escape(str(col))}</th>" for col in formula_df.columns
        )
        formula_rows = []
        for _, formula_row in formula_df.iterrows():
            formula_rows.append(
                "<tr>"
                f"<td class='formula-group'>{html.escape(str(formula_row.get('NhÃ³m', '')))}</td>"
                f"<td>{html.escape(str(formula_row.get('CÃ´ng thá»©c/logic', '')))}</td>"
                "</tr>"
            )
        st.markdown(
            f"""
            <div class='formula-table-wrap'>
              <table class='formula-table'>
                <thead><tr>{formula_headers}</tr></thead>
                <tbody>{''.join(formula_rows)}</tbody>
              </table>
            </div>
            <style>
              .formula-table-wrap {{border:1px solid #E2E8F0; border-radius:14px; overflow:hidden; margin: 8px 0 16px 0;}}
              .formula-table {{border-collapse:collapse; width:100%; font-size:13.5px; background:#FFFFFF;}}
              .formula-table th {{background:#EAF7F1; color:#123D3A; font-weight:950; text-align:left; padding:10px 12px; border-bottom:1px solid #D5E6DD;}}
              .formula-table td {{padding:9px 12px; border-bottom:1px solid #EDF2F7; color:#123D3A; vertical-align:top;}}
              .formula-table .formula-group {{font-weight:950; color:#064E47; white-space:nowrap; background:#F7FBF8;}}
            </style>
            """,
            unsafe_allow_html=True,
        )

        st.markdown("### Giáº£ Ä‘á»‹nh Ä‘ang dÃ¹ng")
        try:
            assumption_items = load_assumptions(ASSUMPTIONS_PATH).__dict__
        except Exception:
            assumption_items = {}
        if assumption_items:
            assumption_label_map = {
                "required_return_pct": "Tá»· suáº¥t sinh lá»i yÃªu cáº§u (%)",
                "terminal_growth_pct": "TÄƒng trÆ°á»Ÿng dÃ i háº¡n (%)",
                "conservative_growth_pct": "TÄƒng trÆ°á»Ÿng tháº­n trá»ng (%)",
                "base_growth_cap_pct": "Tráº§n tÄƒng trÆ°á»Ÿng cÆ¡ sá»Ÿ (%)",
                "high_growth_cap_pct": "Tráº§n tÄƒng trÆ°á»Ÿng cao (%)",
                "mos_conservative_pct": "MOS tháº­n trá»ng (%)",
                "mos_base_pct": "MOS cÆ¡ sá»Ÿ (%)",
                "target_pe_default": "P/E má»¥c tiÃªu máº·c Ä‘á»‹nh",
                "target_pe_quality": "P/E má»¥c tiÃªu doanh nghiá»‡p cháº¥t lÆ°á»£ng",
                "target_pb_bank": "P/B má»¥c tiÃªu ngÃ¢n hÃ ng/báº£o hiá»ƒm",
                "asset_haircut_cash_pct": "Haircut tiá»n (%)",
                "asset_haircut_receivables_pct": "Haircut pháº£i thu (%)",
                "asset_haircut_inventory_pct": "Haircut tá»“n kho (%)",
                "asset_haircut_fixed_assets_pct": "Haircut tÃ i sáº£n cá»‘ Ä‘á»‹nh (%)",
                "min_required_years_for_high_confidence": "Sá»‘ nÄƒm tá»‘i thiá»ƒu Ä‘á»ƒ tÄƒng Ä‘á»™ tin cáº­y",
            }
            assumption_df = pd.DataFrame([
                {"Giáº£ Ä‘á»‹nh": assumption_label_map.get(k, k), "GiÃ¡ trá»‹": v}
                for k, v in assumption_items.items()
                if not str(k).startswith("_")
            ])
            _show_table(assumption_df, height=300)
        else:
            st.warning("ChÆ°a Ä‘á»c Ä‘Æ°á»£c bá»™ giáº£ Ä‘á»‹nh Ä‘á»‹nh giÃ¡; app váº«n cháº¡y vá»›i giÃ¡ trá»‹ máº·c Ä‘á»‹nh trong engine.")

        _render_company_type_guidance(getattr(cls, "company_type", "Normal Business"))
        st.divider()
        _render_glossary_panel()
        st.divider()
        if st.button("ðŸ“„ Xuáº¥t bÃ¡o cÃ¡o Markdown Ä‘á»‹nh giÃ¡", use_container_width=True):
            out_path = REPORT_DIR / f"{company.ticker}_Valuation_Porter_Report.md"
            export_module2_report_markdown(company, valuation_df, moat_df, value_chain_df, scenario_df, out_path, annual_df)
            st.success("ÄÃ£ xuáº¥t bÃ¡o cÃ¡o Markdown vÃ o thÆ° má»¥c bÃ¡o cÃ¡o ná»™i bá»™.")
        docs = [
            APP_DIR / "docs" / "FORMULA_EXPLANATION_MODULE2.md",
            APP_DIR / "docs" / "PORTER_MOAT_SCORING_GUIDE.md",
            APP_DIR / "docs" / "BENEISH_M_SCORE_GUIDE.md",
        ]
        for doc in docs:
            if doc.exists():
                doc_label = (
                    doc.name
                    .replace("MODULE2", "DINH_GIA")
                    .replace("MODULE1", "TONG_QUAN")
                    .replace("Module", "Phan")
                    .replace("module", "phan")
                )
                with st.expander(doc_label, expanded=False):
                    doc_text = doc.read_text(encoding="utf-8").replace("Module", "Pháº§n").replace("module", "pháº§n")
                    st.markdown(doc_text)


def _module3_default_ticker() -> str:
    for key in ["module1_ticker", "active_ticker", "shared_ticker", "module2_ticker", "last_query_ticker"]:
        val = _safe_ticker(str(st.session_state.get(key, "")))
        if val:
            return val
    return "DGC"


def _minimal_company_for_ticker(ticker: str) -> CompanyOverview:
    ticker = _safe_ticker(ticker) or "DGC"
    return CompanyOverview(
        ticker=ticker,
        company_name=ticker,
        exchange="",
        industry="",
        sub_industry="",
        market_cap_bil=None,
        shares_outstanding_mil=None,
        current_price=None,
        eps=None,
        pe=None,
        pb=None,
        ps=None,
        roe=None,
        roa=None,
        roic=None,
        updated_at=pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def render_module3_comparison_dashboard() -> None:
    """Standalone So sÃ¡nh doanh nghiá»‡p page: peer universe + 10-company comparison."""
    _inject_runtime_ui_css()
    _render_brand_page_header(
        "âš–ï¸ So sÃ¡nh doanh nghiá»‡p",
        "Tá»± láº¥y mÃ£ tá»« Tá»•ng quan doanh nghiá»‡p, cáº­p nháº­t danh sÃ¡ch doanh nghiá»‡p cÃ¹ng ngÃ nh, chá»n tá»‘i Ä‘a 10 mÃ£ vÃ  cháº¥m Ä‘iá»ƒm so sÃ¡nh theo Ä‘á»‹nh giÃ¡, cháº¥t lÆ°á»£ng, dÃ²ng tiá»n vÃ  Porter Moat.",
    )
    if st.session_state.get("module_sync_status"):
        st.info(st.session_state.get("module_sync_status"))

    with st.sidebar:
        _render_tre_sidebar_nav()
        st.markdown("### Tham sá»‘ so sÃ¡nh")
        default_ticker = _module3_default_ticker()
        ticker = st.text_input("MÃ£ láº¥y tá»« Tá»•ng quan doanh nghiá»‡p / mÃ£ gá»‘c Ä‘á»ƒ crawl cÃ¹ng ngÃ nh", value=default_ticker, max_chars=12).upper()
        source_options = ["Dá»¯ liá»‡u Æ°u tiÃªn", "Dá»¯ liá»‡u tÃ­ch há»£p", "Dá»¯ liá»‡u máº«u"]
        source_display = st.selectbox("Cháº¿ Ä‘á»™ dá»¯ liá»‡u Ä‘á»ƒ Ä‘á»‹nh giÃ¡ cÃ¡c mÃ£ peer", source_options, index=0)
        source = {"Dá»¯ liá»‡u Æ°u tiÃªn": "FireAnt", "Dá»¯ liá»‡u tÃ­ch há»£p": "Financial tÃ­ch há»£p", "Dá»¯ liá»‡u máº«u": "CSV máº«u tÃ­ch há»£p"}.get(source_display, "FireAnt")
        simplize_default_url = str(st.session_state.get("module3_simplize_industry_url", ""))
        simplize_industry_url = st.text_input(
            "URL nhÃ³m ngÃ nh náº¿u muá»‘n cá»‘ Ä‘á»‹nh",
            value=simplize_default_url,
            placeholder="DÃ¡n URL nhÃ³m ngÃ nh náº¿u cáº§n",
            help="CÃ³ thá»ƒ Ä‘á»ƒ trá»‘ng; app sáº½ tá»± tÃ¬m link ngÃ nh. Khi website Ä‘á»•i bá»‘ cá»¥c, dÃ¡n trá»±c tiáº¿p URL nhÃ³m ngÃ nh táº¡i Ä‘Ã¢y Ä‘á»ƒ láº¥y nhanh hÆ¡n."
        ).strip()
        st.session_state["module3_simplize_industry_url"] = simplize_industry_url
        assumptions = load_assumptions(ASSUMPTIONS_PATH)
        assumptions["required_return"] = assumptions.get("required_return_pct", 13.0) / 100
        target_mos_pct = st.selectbox("MOS yÃªu cáº§u khi lá»c peer (%)", MOS_OPTIONS_GLOBAL, index=MOS_OPTIONS_GLOBAL.index(_normalize_mos_value(st.session_state.get("target_mos_pct", 30), 30)))
        st.session_state["target_mos_pct"] = target_mos_pct

    ticker = _safe_ticker(ticker) or default_ticker
    st.session_state["module3_base_ticker"] = ticker
    company = None
    annual_df = pd.DataFrame()
    quarterly_df = pd.DataFrame()
    source_label = source
    paths = []
    available_tickers = _available_financial_tickers_cached(str(BUNDLED_XLSM)) if BUNDLED_XLSM.exists() else []
    try:
        company, annual_df, quarterly_df, source_label, paths = _load_data(ticker, source)
        if not _has_real_financial_data(annual_df):
            raise RuntimeError("ChÆ°a cÃ³ BCTC nhiá»u ká»³ cho mÃ£ gá»‘c; váº«n cÃ³ thá»ƒ crawl danh sÃ¡ch cÃ¹ng ngÃ nh vÃ  so sÃ¡nh cÃ¡c mÃ£ khÃ¡c.")
        st.markdown(
            f"""
            <div class='ticker-title-card'>
                <div class='ticker-title-main'><span class='ticker-title-code'>{html.escape(str(company.ticker))}</span> - <span class='ticker-title-name'>{html.escape(str(company.company_name))}</span></div>
                <div class='ticker-title-meta'><b>Cháº¿ Ä‘á»™ dá»¯ liá»‡u mÃ£ gá»‘c:</b> {html.escape(source_display)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    except Exception as exc:
        company = _minimal_company_for_ticker(ticker)
        st.warning(f"KhÃ´ng táº£i Ä‘Æ°á»£c Ä‘áº§y Ä‘á»§ BCTC mÃ£ gá»‘c {ticker}: {_public_text(exc)}. So sÃ¡nh doanh nghiá»‡p váº«n cáº­p nháº­t danh sÃ¡ch cÃ¹ng ngÃ nh vÃ  cho chá»n mÃ£ Ä‘á»ƒ so sÃ¡nh.")

    _render_peer_universe_and_comparison(company, source, assumptions, float(target_mos_pct), available_tickers, auto_simplize=True, simplize_industry_url=simplize_industry_url)

    # V23.39: Ä‘Ã£ bá» nÃºt/khung xuáº¥t bÃ¡o cÃ¡o trong tá»«ng pháº§n; chá»‰ cÃ²n trang BÃ¡o cÃ¡o tá»•ng há»£p toÃ n bá»™ ná»™i dung á»Ÿ sidebar.


if __name__ == "__main__":
    render_dashboard()

