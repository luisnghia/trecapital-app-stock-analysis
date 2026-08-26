"""Read-only views and exports for immutable Fisher Top-Down macro snapshots."""

from __future__ import annotations

from collections.abc import Iterable, Mapping
from io import BytesIO
import json
import math
from pathlib import Path
import re
from typing import Any
from xml.sax.saxutils import escape

import pandas as pd


SOURCE_LABELS = {
    "analyst_override": "Analyst tự chấm",
    "automatic_suggestion": "Gợi ý tự động",
    "default": "Mặc định / chưa có dữ liệu tự động",
}


def _payload(snapshot: Mapping[str, Any]) -> dict[str, Any]:
    value = snapshot.get("payload", snapshot)
    return dict(value) if isinstance(value, Mapping) else {}


def _json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return str(value)


def _finite_or_none(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def snapshot_option_label(snapshot: Mapping[str, Any]) -> str:
    """Compact, stable label for Streamlit snapshot selectors."""
    return (
        f"#{snapshot.get('version_no', '?')} · {snapshot.get('as_of_date', '—')} · "
        f"{snapshot.get('snapshot_label', 'Snapshot chưa đặt tên')}"
    )


def snapshot_metadata_frame(
    snapshot: Mapping[str, Any],
    *,
    cycle_labels: Mapping[str, str] | None = None,
) -> pd.DataFrame:
    payload = _payload(snapshot)
    cycle_id = str(payload.get("cycle_phase", ""))
    benchmark = payload.get("benchmark") if isinstance(payload.get("benchmark"), Mapping) else {}
    rows = [
        ("Phiên bản", snapshot.get("version_no")),
        ("Ngày đánh giá", snapshot.get("as_of_date")),
        ("Tên snapshot", snapshot.get("snapshot_label")),
        ("Lý do lưu", snapshot.get("save_reason")),
        ("Người lưu", snapshot.get("created_by")),
        ("Thời điểm tạo", snapshot.get("created_at")),
        ("Phiên bản phương pháp", snapshot.get("methodology_version") or payload.get("methodology_version")),
        ("Pha chu kỳ", (cycle_labels or {}).get(cycle_id, cycle_id or "—")),
        ("Benchmark", benchmark.get("name") or benchmark.get("id") or "—"),
        ("Nguồn benchmark", benchmark.get("source_name") or "—"),
        ("URL benchmark", benchmark.get("source_url") or "—"),
        ("Ngày hiệu lực benchmark", benchmark.get("source_as_of_date") or "—"),
        ("Trạng thái benchmark", "Cần cập nhật" if benchmark.get("requires_update") else "Đã kiểm chứng"),
        ("Thời điểm tính toán", payload.get("generated_at")),
        ("Schema payload", payload.get("schema")),
        ("Payload hash", snapshot.get("payload_hash")),
        ("Source Registry hash", snapshot.get("source_registry_hash")),
        ("Source Mapping SHA-256", payload.get("source_mapping_sha256")),
    ]
    return pd.DataFrame(rows, columns=["Thuộc tính", "Giá trị"])


def snapshot_driver_frame(
    snapshot: Mapping[str, Any],
    *,
    driver_catalog: Iterable[Mapping[str, Any]] | None = None,
) -> pd.DataFrame:
    payload = _payload(snapshot)
    parameters = payload.get("parameters") if isinstance(payload.get("parameters"), Mapping) else {}
    scores = dict(parameters.get("driver_outlook") or {})
    sources = dict(parameters.get("driver_score_sources") or {})
    automatic = dict(parameters.get("automatic_driver_scores") or {})
    catalog = {str(row.get("id")): row for row in (driver_catalog or []) if row.get("id")}
    ordered_ids = list(catalog)
    ordered_ids.extend(sorted(set(scores) - set(ordered_ids)))
    rows: list[dict[str, Any]] = []
    for driver_id in ordered_ids:
        meta = catalog.get(driver_id, {})
        source_code = str(sources.get(driver_id, "default"))
        rows.append(
            {
                "Nhóm": meta.get("nhom", "—"),
                "Driver": meta.get("ten_vi", driver_id),
                "Driver ID": driver_id,
                "Điểm hiệu lực": _finite_or_none(scores.get(driver_id)),
                "Nguồn điểm": SOURCE_LABELS.get(source_code, source_code),
                "Baseline tự động": _finite_or_none(automatic.get(driver_id)),
            }
        )
    return pd.DataFrame(
        rows,
        columns=["Nhóm", "Driver", "Driver ID", "Điểm hiệu lực", "Nguồn điểm", "Baseline tự động"],
    )


def snapshot_macro_frame(snapshot: Mapping[str, Any]) -> pd.DataFrame:
    payload = _payload(snapshot)
    latest = payload.get("latest_macro_update")
    latest = latest if isinstance(latest, Mapping) else {}
    suggestions = {
        str(row.get("driver_id")): row
        for row in latest.get("suggestions", [])
        if isinstance(row, Mapping) and row.get("driver_id")
    }
    rows: list[dict[str, Any]] = []
    for observation in latest.get("observations", []):
        if not isinstance(observation, Mapping):
            continue
        driver_id = str(observation.get("driver_id", ""))
        suggestion = suggestions.get(driver_id, {})
        rows.append(
            {
                "Driver": observation.get("driver_name") or suggestion.get("driver_name") or driver_id,
                "Driver ID": driver_id,
                "Nguồn": observation.get("source_code"),
                "Series": observation.get("series_code"),
                "Kỳ": observation.get("period_label"),
                "Giá trị": _finite_or_none(observation.get("value_numeric")),
                "Kỳ trước": _finite_or_none(observation.get("previous_value_numeric")),
                "Chênh lệch": _finite_or_none(observation.get("delta_numeric")),
                "Đơn vị": observation.get("unit"),
                "Độ mới": observation.get("freshness_status"),
                "Điểm gợi ý": _finite_or_none(suggestion.get("suggested_score")),
                "Độ tin cậy": suggestion.get("confidence"),
                "Lập luận": suggestion.get("rationale"),
                "URL nguồn": observation.get("source_url"),
            }
        )
    return pd.DataFrame(
        rows,
        columns=[
            "Driver", "Driver ID", "Nguồn", "Series", "Kỳ", "Giá trị", "Kỳ trước",
            "Chênh lệch", "Đơn vị", "Độ mới", "Điểm gợi ý", "Độ tin cậy", "Lập luận", "URL nguồn",
        ],
    )


def snapshot_ranking_frame(snapshot: Mapping[str, Any]) -> pd.DataFrame:
    rows = []
    for row in _payload(snapshot).get("ranking", []):
        if not isinstance(row, Mapping):
            continue
        rows.append(
            {
                "Xếp hạng": row.get("rank"),
                "Mã ngành": row.get("sector_code"),
                "Ngành": row.get("sector_name"),
                "Tính chất": row.get("character"),
                "Điểm kinh tế": row.get("economy_score"),
                "Điểm chính trị": row.get("politics_score"),
                "Điểm tâm lý": row.get("sentiment_score"),
                "Điểm chu kỳ": row.get("cycle_score"),
                "Điểm tổng hợp": row.get("score"),
            }
        )
    return pd.DataFrame(rows)


def snapshot_weights_frame(snapshot: Mapping[str, Any]) -> pd.DataFrame:
    rows = []
    for row in _payload(snapshot).get("weights", []):
        if not isinstance(row, Mapping):
            continue
        rows.append(
            {
                "Mã ngành": row.get("sector_code"),
                "Ngành": row.get("sector_name"),
                "Điểm tổng hợp": row.get("sector_score"),
                "Khuyến nghị": row.get("signal"),
                "Hệ số tilt": row.get("tilt_factor"),
                "Tỷ trọng benchmark %": row.get("benchmark_weight_pct"),
                "Tỷ trọng đề xuất %": row.get("proposed_weight_pct"),
                "Độ lệch điểm %": row.get("tilt_pct"),
            }
        )
    return pd.DataFrame(rows)


def snapshot_checks_frame(snapshot: Mapping[str, Any]) -> pd.DataFrame:
    checks = _payload(snapshot).get("sync_checks", [])
    return pd.DataFrame([dict(row) for row in checks if isinstance(row, Mapping)])


def snapshot_detail_frames(
    snapshot: Mapping[str, Any],
    *,
    driver_catalog: Iterable[Mapping[str, Any]] | None = None,
    cycle_labels: Mapping[str, str] | None = None,
) -> dict[str, pd.DataFrame]:
    """Return all read-only snapshot sections in a stable export/display order."""
    return {
        "Tổng quan": snapshot_metadata_frame(snapshot, cycle_labels=cycle_labels),
        "Portfolio Drivers": snapshot_driver_frame(snapshot, driver_catalog=driver_catalog),
        "Dữ liệu vĩ mô": snapshot_macro_frame(snapshot),
        "Xếp hạng ngành": snapshot_ranking_frame(snapshot),
        "Tỷ trọng đề xuất": snapshot_weights_frame(snapshot),
        "Kiểm tra dữ liệu": snapshot_checks_frame(snapshot),
    }


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, (dict, list, tuple)):
        return _json_text(value)
    return value


def build_snapshot_excel_bytes(
    snapshot: Mapping[str, Any],
    *,
    driver_catalog: Iterable[Mapping[str, Any]] | None = None,
    cycle_labels: Mapping[str, str] | None = None,
) -> bytes:
    """Create a formatted, auditable Excel workbook for exactly one immutable snapshot."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    frames = snapshot_detail_frames(
        snapshot,
        driver_catalog=driver_catalog,
        cycle_labels=cycle_labels,
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = f"Fisher Top-Down Snapshot #{snapshot.get('version_no', '')}"
    workbook.properties.subject = "Immutable macro assessment snapshot"
    workbook.properties.creator = str(snapshot.get("created_by") or "Trecapital")
    thin = Side(style="thin", color="D9E5E1")
    header_fill = PatternFill("solid", fgColor="0B7F75")
    title_fill = PatternFill("solid", fgColor="EAF7F1")
    for sheet_name, frame in frames.items():
        safe_name = sheet_name[:31]
        sheet = workbook.create_sheet(safe_name)
        column_count = max(1, len(frame.columns))
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
        title_cell = sheet.cell(1, 1, f"Fisher Top-Down - {sheet_name} - Snapshot #{snapshot.get('version_no', '?')}")
        title_cell.font = Font(name="Arial", size=14, bold=True, color="123D3A")
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 26
        for col_index, column in enumerate(frame.columns, start=1):
            cell = sheet.cell(3, col_index, str(column))
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=4):
            for col_index, value in enumerate(row, start=1):
                cell = sheet.cell(row_index, col_index, _excel_value(value))
                cell.font = Font(name="Arial", size=9, color="123D3A")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = "#,##0.0"
                    if float(cell.value) < 0:
                        cell.font = Font(name="Arial", size=9, color="C00000")
        sheet.freeze_panes = "A4"
        if not frame.empty:
            sheet.auto_filter.ref = f"A3:{get_column_letter(column_count)}{3 + len(frame)}"
        for col_index, column in enumerate(frame.columns, start=1):
            values = [str(column)] + [_json_text(value) for value in frame[column].head(120)]
            longest = max((len(value) for value in values), default=10)
            sheet.column_dimensions[get_column_letter(col_index)].width = min(45, max(11, longest + 2))
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_title_rows = "3:3"
        sheet.sheet_properties.outlinePr.summaryBelow = True
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.45
        sheet.page_margins.bottom = 0.45
        sheet.auto_filter.ref = sheet.auto_filter.ref if not frame.empty else None

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_font_paths() -> tuple[Path, Path]:
    candidates: list[tuple[Path, Path]] = []
    try:
        from importlib.util import find_spec

        matplotlib_spec = find_spec("matplotlib")
        if matplotlib_spec is not None:
            package_locations = list(matplotlib_spec.submodule_search_locations or [])
            matplotlib_root = None
            if package_locations:
                matplotlib_root = Path(package_locations[0])
            elif matplotlib_spec.origin:
                matplotlib_root = Path(matplotlib_spec.origin).resolve().parent
            if matplotlib_root is not None:
                matplotlib_fonts = matplotlib_root / "mpl-data" / "fonts" / "ttf"
                candidates.append(
                    (
                        matplotlib_fonts / "DejaVuSans.ttf",
                        matplotlib_fonts / "DejaVuSans-Bold.ttf",
                    )
                )
    except (ImportError, AttributeError, TypeError, ValueError):
        pass

    candidates.extend(
        [
            (
                Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
                Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
            ),
            (
                Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
                Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
            ),
        ]
    )
    for regular, bold in candidates:
        if regular.is_file() and bold.is_file():
            return regular, bold
    raise RuntimeError(
        "Không tìm thấy font Unicode để xuất PDF tiếng Việt. "
        "Hãy cài dependency matplotlib hoặc bộ font DejaVu/Liberation Sans."
    )


def _pdf_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    if isinstance(value, float):
        return f"{value:,.2f}".rstrip("0").rstrip(".")
    return _json_text(value)


def build_snapshot_pdf_bytes(
    snapshot: Mapping[str, Any],
    *,
    driver_catalog: Iterable[Mapping[str, Any]] | None = None,
    cycle_labels: Mapping[str, str] | None = None,
) -> bytes:
    """Create a paginated Vietnamese PDF report for exactly one immutable snapshot."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle

    regular_path, bold_path = _pdf_font_paths()
    regular_name = "TreSnapshotSans"
    bold_name = "TreSnapshotSansBold"
    if regular_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(regular_name, str(regular_path)))
        pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))

    frames = snapshot_detail_frames(
        snapshot,
        driver_catalog=driver_catalog,
        cycle_labels=cycle_labels,
    )
    buffer = BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"Fisher Top-Down Snapshot #{snapshot.get('version_no', '')}",
        author=str(snapshot.get("created_by") or "Trecapital"),
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TreTitle", parent=styles["Title"], fontName=bold_name, fontSize=18,
        leading=22, textColor=colors.HexColor("#123D3A"), alignment=TA_CENTER, spaceAfter=8,
    )
    section_style = ParagraphStyle(
        "TreSection", parent=styles["Heading2"], fontName=bold_name, fontSize=13,
        leading=16, textColor=colors.HexColor("#0B7F75"), spaceAfter=7,
    )
    cell_style = ParagraphStyle(
        "TreCell", parent=styles["BodyText"], fontName=regular_name, fontSize=6.5,
        leading=8.2, textColor=colors.HexColor("#123D3A"),
    )
    header_style = ParagraphStyle(
        "TreHeader", parent=cell_style, fontName=bold_name, textColor=colors.white, alignment=TA_CENTER,
    )
    story: list[Any] = [
        Paragraph("Fisher Top-Down - Lịch sử đánh giá vĩ mô", title_style),
        Paragraph(
            escape(snapshot_option_label(snapshot)) + "<br/>Chế độ chỉ đọc - snapshot bất biến",
            ParagraphStyle("TreSubtitle", parent=cell_style, fontSize=9, leading=12, alignment=TA_CENTER),
        ),
        Spacer(1, 8),
    ]
    available_width = page_size[0] - doc.leftMargin - doc.rightMargin
    for section_index, (section_name, frame) in enumerate(frames.items()):
        if section_index:
            story.append(PageBreak())
        story.append(Paragraph(escape(section_name), section_style))
        if frame.empty:
            story.append(Paragraph("Không có dữ liệu được lưu trong snapshot này.", cell_style))
            continue
        raw_widths = [min(130.0, max(45.0, 5.0 * len(str(column)))) for column in frame.columns]
        scale = available_width / sum(raw_widths)
        col_widths = [width * scale for width in raw_widths]
        table_data = [[Paragraph(escape(str(column)), header_style) for column in frame.columns]]
        for row in frame.itertuples(index=False, name=None):
            table_data.append([Paragraph(escape(_pdf_text(value)), cell_style) for value in row])
        table = LongTable(table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        commands: list[tuple] = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B7F75")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for row_index in range(1, len(table_data)):
            if row_index % 2 == 0:
                commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F7FBF8")))
        table.setStyle(TableStyle(commands))
        story.append(table)

    def _footer(canvas, _doc):
        canvas.saveState()
        canvas.setFont(regular_name, 7)
        canvas.setFillColor(colors.HexColor("#61736F"))
        canvas.drawString(doc.leftMargin, 7 * mm, "Trecapital - Fisher Top-Down Snapshot")
        canvas.drawRightString(page_size[0] - doc.rightMargin, 7 * mm, f"Trang {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def snapshot_export_filename(snapshot: Mapping[str, Any], extension: str) -> str:
    version = re.sub(r"[^0-9A-Za-z_-]+", "_", str(snapshot.get("version_no", "unknown")))
    as_of = re.sub(r"[^0-9A-Za-z_-]+", "_", str(snapshot.get("as_of_date", "unknown")))
    suffix = extension.lower().lstrip(".")
    return f"fisher_topdown_snapshot_v{version}_{as_of}.{suffix}"


__all__ = [
    "SOURCE_LABELS",
    "build_snapshot_excel_bytes",
    "build_snapshot_pdf_bytes",
    "snapshot_checks_frame",
    "snapshot_detail_frames",
    "snapshot_driver_frame",
    "snapshot_export_filename",
    "snapshot_macro_frame",
    "snapshot_metadata_frame",
    "snapshot_option_label",
    "snapshot_ranking_frame",
    "snapshot_weights_frame",
]
