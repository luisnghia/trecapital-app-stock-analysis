from __future__ import annotations

"""Read Financial/xFINN-style .xlsm workbooks and normalize to Tổng quan doanh nghiệp schema.

This adapter is built from the structure of the user's file `Financial-v1.3.0.xlsm`:
- BÁO CÁO TÀI CHÍNH: controls and peer tickers
- CÂN ĐỐI KẾ TOÁN / CÂN ĐỐI KẾ TOÁN - QUÝ
- KẾT QUẢ KINH DOANH / KẾT QUẢ KINH DOANH - QUÝ
- LƯU CHUYỂN TIỀN TỆ GT / LƯU CHUYỂN TIỀN TỆ GT - QUÝ
- CHỈ SỐ TÀI CHÍNH / CHỈ SỐ TÀI CHÍNH - QUÝ
- TỔNG QUAN / DOANH NGHIỆP for profile and price overview

No API token is needed because this provider reads the workbook directly. It also avoids the common bug
where chart periods become duplicated or show as decimals by parsing periods strictly and de-duplicating
on the normalized year/quarter key before sorting ascending.
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
import math
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook

from .base import ProviderResult, normalize_columns, MODULE1_OVERVIEW_COLUMNS, MODULE1_TIMESERIES_COLUMNS

VND_TO_BIL = 1_000_000_000


@dataclass(frozen=True)
class ParsedQuarter:
    year: int
    quarter: int

    @property
    def label(self) -> str:
        return f"Q{self.quarter}/{self.year}"


def _norm_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("đ", "d").replace("Đ", "D")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value) if isinstance(value, float) else False:
            return None
        return float(value)
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan", "null", "-"}:
        return None
    # Keep sign and decimal symbols; tolerate Vietnamese formatted strings like "363,470 tỷ" or "17.7x".
    multiplier = 1.0
    if "nghìn tỷ" in text.lower() or "ngan ty" in _norm_text(text):
        multiplier = 1_000_000.0
    elif "tỷ" in text.lower() or "ty" in _norm_text(text):
        multiplier = 1_000.0 if re.search(r"\d[,\.]\d{3}\b", text) else 1.0
    elif "tr" in text.lower() or "triệu" in text.lower() or "trieu" in _norm_text(text):
        multiplier = 1.0
    cleaned = re.sub(r"[^0-9,\.\-]", "", text)
    if "," in cleaned and "." not in cleaned:
        parts = cleaned.split(",")
        if len(parts) == 2 and len(parts[1]) != 3:
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    try:
        return float(cleaned) * multiplier
    except Exception:
        return None


def _parse_year(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value if 1900 <= value <= 2200 else None
    if isinstance(value, float) and value.is_integer():
        y = int(value)
        return y if 1900 <= y <= 2200 else None
    text = str(value).strip()
    if re.fullmatch(r"\d{4}(?:\.0+)?", text):
        y = int(float(text))
        return y if 1900 <= y <= 2200 else None
    m = re.search(r"(19\d{2}|20\d{2}|21\d{2}|22\d{2})", text)
    if m and "q" not in text.lower():
        return int(m.group(1))
    return None


def _parse_quarter(value: Any) -> ParsedQuarter | None:
    if value is None:
        return None
    text = str(value).strip().upper().replace(" ", "")
    m = re.fullmatch(r"Q([1-4])[/\-.](\d{4})", text)
    if m:
        return ParsedQuarter(year=int(m.group(2)), quarter=int(m.group(1)))
    # Accept 2024Q1 or 2024-Q1 if the web source returns this style.
    m = re.fullmatch(r"(\d{4})[/\-.]?Q([1-4])", text)
    if m:
        return ParsedQuarter(year=int(m.group(1)), quarter=int(m.group(2)))
    return None


def _period_key(value: Any, kind: str) -> tuple[int, int, str] | None:
    if kind.upper() == "Q":
        q = _parse_quarter(value)
        return (q.year, q.quarter, q.label) if q else None
    y = _parse_year(value)
    return (y, 0, str(y)) if y else None


def _find_ticker_row(ws, ticker: str) -> int | None:
    ticker = ticker.upper().strip()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=2).value
        if isinstance(value, str) and value.strip().upper() == ticker:
            return row
    return None


def _read_header_periods(ws, ticker_row: int, kind: str, start_col: int = 4) -> list[tuple[int, str, int, int]]:
    """Return [(col, label, year, quarter)] with duplicates removed by normalized period.

    If the workbook contains repeated periods, the first occurrence is retained. In Financial-v1.3.0,
    the first occurrence is the left-most/current data column.
    """
    seen: set[tuple[int, int]] = set()
    out: list[tuple[int, str, int, int]] = []
    for col in range(start_col, ws.max_column + 1):
        raw = ws.cell(row=ticker_row, column=col).value
        key = _period_key(raw, kind)
        if not key:
            continue
        year, quarter, label = key
        dedup_key = (year, quarter)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)
        out.append((col, label, year, quarter))
    return out


def _block_end(ws, ticker_row: int) -> int:
    # Financial sheets store one ticker block after another. A new block begins with an alphanumeric code in col B
    # followed by recognizable first financial statement row. Use a safe scan window if the next ticker is not obvious.
    for row in range(ticker_row + 1, min(ws.max_row, ticker_row + 260) + 1):
        value = ws.cell(row=row, column=2).value
        next_value = ws.cell(row=row + 1, column=2).value if row + 1 <= ws.max_row else None
        if isinstance(value, str) and re.fullmatch(r"[A-Z0-9]{2,8}", value.strip().upper()):
            nv = _norm_text(next_value)
            if any(x in nv for x in ["tai san", "tong doanh thu", "luu chuyen", "dilutedeps", "basiceps"]):
                return row - 1
    return min(ws.max_row, ticker_row + 240)


def _find_label_row(ws, ticker_row: int, patterns: Iterable[str]) -> int | None:
    normalized_patterns = [_norm_text(p) for p in patterns]
    end = _block_end(ws, ticker_row)
    for row in range(ticker_row + 1, end + 1):
        label = _norm_text(ws.cell(row=row, column=2).value)
        if not label:
            continue
        if any(p in label for p in normalized_patterns):
            return row
    return None


def _extract_series(ws, ticker: str, patterns: Iterable[str], kind: str, div: float = 1.0) -> dict[str, float | None]:
    ticker_row = _find_ticker_row(ws, ticker)
    if ticker_row is None:
        return {}
    value_row = _find_label_row(ws, ticker_row, patterns)
    if value_row is None:
        return {}
    periods = _read_header_periods(ws, ticker_row, kind)
    out: dict[str, float | None] = {}
    for col, label, _year, _quarter in periods:
        v = _to_float(ws.cell(row=value_row, column=col).value)
        out[label] = (v / div) if v is not None else None
    return out


def _sum_series(*series: dict[str, float | None]) -> dict[str, float | None]:
    keys = sorted(set().union(*(s.keys() for s in series)))
    out: dict[str, float | None] = {}
    for k in keys:
        vals = [s.get(k) for s in series if s.get(k) is not None]
        out[k] = sum(vals) if vals else None
    return out


def _ratio(num: float | None, den: float | None) -> float | None:
    if num is None or den in (None, 0):
        return None
    try:
        return num / den
    except Exception:
        return None


def _pct_ratio(num: float | None, den: float | None) -> float | None:
    r = _ratio(num, den)
    return r * 100 if r is not None else None


def _latest_non_null(values: dict[str, float | None]) -> float | None:
    parsed = []
    for k, v in values.items():
        y = _parse_year(k)
        q = _parse_quarter(k)
        if v is not None:
            parsed.append((q.year if q else (y or 0), q.quarter if q else 0, v))
    if not parsed:
        return None
    parsed.sort()
    return parsed[-1][2]


def _maybe_pct(v: float | None) -> float | None:
    if v is None:
        return None
    # Financial ratios in the workbook are usually decimals, e.g. 0.206 = 20.6%.
    if abs(v) <= 3:
        return v * 100
    return v


class ExcelFinancialProvider:
    def __init__(self, workbook_path: str | Path, raw_dir: str | Path | None = None):
        self.workbook_path = Path(workbook_path)
        self.raw_dir = Path(raw_dir) if raw_dir else None

    def fetch(self, ticker: str) -> ProviderResult:
        ticker = ticker.upper().strip()
        wb = load_workbook(self.workbook_path, data_only=True, read_only=False)
        overview = self._build_overview(wb, ticker)
        annual = self._build_timeseries(wb, ticker, kind="Y")
        quarterly = self._build_timeseries(wb, ticker, kind="Q")
        return ProviderResult(
            overview=normalize_columns(overview, MODULE1_OVERVIEW_COLUMNS),
            annual=normalize_columns(annual, MODULE1_TIMESERIES_COLUMNS),
            quarterly=normalize_columns(quarterly, MODULE1_TIMESERIES_COLUMNS),
            raw_path=self.workbook_path,
            note=f"Loaded from Excel workbook: {self.workbook_path.name}",
        )

    def export_csv(self, ticker: str, out_dir: str | Path) -> ProviderResult:
        out_dir = Path(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        result = self.fetch(ticker)
        result.overview.to_csv(out_dir / "company_overview_sample.csv", index=False, encoding="utf-8-sig")
        result.annual.to_csv(out_dir / "financial_timeseries_year.csv", index=False, encoding="utf-8-sig")
        result.quarterly.to_csv(out_dir / "financial_timeseries_quarter.csv", index=False, encoding="utf-8-sig")
        return result

    def _get_sheet(self, wb, *names: str):
        for name in names:
            if name in wb.sheetnames:
                return wb[name]
        return None

    def _build_overview(self, wb, ticker: str) -> pd.DataFrame:
        stock_list = self._get_sheet(wb, "BÁO CÁO TÀI CHÍNH", "DANH SÁCH MÃ")
        overview_ws = self._get_sheet(wb, "TỔNG QUAN")
        business_ws = self._get_sheet(wb, "DOANH NGHIỆP")
        ratio_ws = self._get_sheet(wb, "CHỈ SỐ TÀI CHÍNH")
        annual = self._build_timeseries(wb, ticker, kind="Y")

        workbook_business_ticker = ""
        if business_ws is not None:
            workbook_business_ticker = str(business_ws["C3"].value or "").strip().upper()
        overview_title = str(overview_ws["C5"].value or "").strip().upper() if overview_ws is not None else ""
        business_matches_ticker = bool(workbook_business_ticker and workbook_business_ticker == ticker)
        overview_matches_ticker = bool(overview_title.startswith(ticker + " ") or overview_title.startswith(ticker + "-") or overview_title == ticker)

        company_name = ""
        exchange = ""
        industry = ""
        sub_industry = ""
        if stock_list is not None:
            # Prefer the configured peer list in BÁO CÁO TÀI CHÍNH; fallback to DANH SÁCH MÃ layout.
            for row in range(1, stock_list.max_row + 1):
                code = stock_list.cell(row=row, column=3).value if stock_list.title == "BÁO CÁO TÀI CHÍNH" else stock_list.cell(row=row, column=2).value
                if isinstance(code, str) and code.strip().upper() == ticker:
                    if stock_list.title == "BÁO CÁO TÀI CHÍNH":
                        company_name = str(stock_list.cell(row=row, column=4).value or "")
                        exchange = str(stock_list.cell(row=row, column=5).value or "")
                        sub_industry = str(stock_list.cell(row=row, column=7).value or "")
                    else:
                        company_name = str(stock_list.cell(row=row, column=3).value or "")
                        exchange = str(stock_list.cell(row=row, column=4).value or "")
                        sub_industry = str(stock_list.cell(row=row, column=6).value or "")
                    break
        if business_ws is not None and business_matches_ticker:
            if not company_name:
                company_name = str(business_ws["C7"].value or business_ws["C4"].value or company_name)
            exchange = exchange or str(business_ws["E5"].value or "")
            industry = str(business_ws["C6"].value or "") or industry
            sub_industry = str(business_ws["C5"].value or "") or sub_industry
        price = shares = market_cap = eps = pe = pb = ps = roe = roa = roic = None
        updated_at = "Đọc từ Financial-v1.3.0.xlsm; nếu mã không có trong workbook thì chỉ dùng làm fallback."
        if overview_ws is not None and overview_matches_ticker:
            price = _to_float(overview_ws["C6"].value)
            market_cap = _to_float(overview_ws["C24"].value)
            shares = _to_float(overview_ws["G24"].value)
            pe = _to_float(overview_ws["C27"].value)
            pb = _to_float(overview_ws["D27"].value)
            roe = _maybe_pct(_to_float(overview_ws["E27"].value))
            roa = _maybe_pct(_to_float(overview_ws["F27"].value))
            eps = _to_float(overview_ws["H27"].value)
            if overview_ws["C7"].value:
                updated_at = str(overview_ws["C7"].value)
        if annual is not None and not annual.empty:
            latest = annual.tail(1).iloc[0]
            eps = eps if eps is not None else _to_float(latest.get("eps_vnd"))
            roe = roe if roe is not None else _to_float(latest.get("roe_pct"))
            roa = roa if roa is not None else _to_float(latest.get("roa_pct"))
            roic = roic if roic is not None else _to_float(latest.get("roic_pct"))
            if price and eps and not pe:
                pe = price * 1000 / eps if price < 1000 else price / eps
            if price and latest.get("revenue_bil") and shares and not ps:
                market_cap_bil = market_cap if market_cap is not None else (price * shares / 1000 if price < 1000 else price * shares / 1_000_000)
                ps = market_cap_bil / latest.get("revenue_bil") if latest.get("revenue_bil") else None
        return pd.DataFrame([{
            "ticker": ticker,
            "company_name": company_name or ticker,
            "exchange": exchange,
            "industry": industry,
            "sub_industry": sub_industry,
            "market_cap_bil": market_cap,
            "shares_outstanding_mil": shares,
            "current_price": price * 1000 if price is not None and price < 1000 else price,
            "eps": eps,
            "pe": pe,
            "pb": pb,
            "ps": ps,
            "roe": roe,
            "roa": roa,
            "roic": roic,
            "updated_at": updated_at,
        }])

    def _series_sources(self, wb, ticker: str, kind: str) -> dict[str, dict[str, float | None]]:
        suffix = " - QUÝ" if kind.upper() == "Q" else ""
        income = self._get_sheet(wb, f"KẾT QUẢ KINH DOANH{suffix}", "KQKD_Q" if kind.upper() == "Q" else "KQKD")
        cashflow = self._get_sheet(wb, f"LƯU CHUYỂN TIỀN TỆ GT{suffix}", "LCTT GT_Q" if kind.upper() == "Q" else "LCTT GT", "LƯU CHUYỂN TIỀN TỆ GT - QUÝ" if kind.upper() == "Q" else "LƯU CHUYỂN TIỀN TỆ GT")
        balance = self._get_sheet(wb, f"CÂN ĐỐI KẾ TOÁN{suffix}", "CDKT_Q" if kind.upper() == "Q" else "CDKT")
        ratios = self._get_sheet(wb, f"CHỈ SỐ TÀI CHÍNH{suffix}", "CHI SO TC_Q" if kind.upper() == "Q" else "CHI SO TC")
        src: dict[str, dict[str, float | None]] = {}
        if income is not None:
            src["gross_revenue_bil"] = _extract_series(income, ticker, ["tong doanh thu hoat dong kinh doanh", "tong doanh thu"], kind, VND_TO_BIL)
            src["revenue_bil"] = _extract_series(income, ticker, ["doanh thu thuan"], kind, VND_TO_BIL) or src["gross_revenue_bil"]
            src["gross_profit_bil"] = _extract_series(income, ticker, ["loi nhuan gop"], kind, VND_TO_BIL)
            src["pretax_profit_bil"] = _extract_series(income, ticker, ["tong loi nhuan ke toan truoc thue", "loi nhuan truoc thue"], kind, VND_TO_BIL)
            src["net_profit_bil"] = _extract_series(income, ticker, ["loi nhuan sau thue cua co dong cua cong ty me", "loi nhuan sau thue thu nhap doanh nghiep"], kind, VND_TO_BIL)
        if cashflow is not None:
            src["cfo_bil"] = _extract_series(cashflow, ticker, ["luu chuyen tien thuan tu hoat dong kinh doanh"], kind, VND_TO_BIL)
            src["cfi_bil"] = _extract_series(cashflow, ticker, ["luu chuyen tien thuan tu hoat dong dau tu"], kind, VND_TO_BIL)
            src["cff_bil"] = _extract_series(cashflow, ticker, ["luu chuyen tien thuan tu hoat dong tai chinh"], kind, VND_TO_BIL)
            src["capex_bil"] = _extract_series(cashflow, ticker, ["tien chi de mua sam", "mua sam, xay dung tscd"], kind, VND_TO_BIL)
            src["cash_dividend_bil"] = _extract_series(cashflow, ticker, ["co tuc, loi nhuan da tra cho chu so huu"], kind, VND_TO_BIL)
            src["noncash_adjustments_bil"] = _extract_series(cashflow, ticker, ["dieu chinh cho cac khoan"], kind, VND_TO_BIL)
            src["depreciation_bil"] = _extract_series(cashflow, ticker, ["khau hao tscd"], kind, VND_TO_BIL)
            src["operating_cash_before_wc_bil"] = _extract_series(cashflow, ticker, ["loi nhuan tu hoat dong kinh doanh truoc thay doi von luu dong"], kind, VND_TO_BIL)
            src["receivables_change_bil"] = _extract_series(cashflow, ticker, ["tang, giam cac khoan phai thu"], kind, VND_TO_BIL)
            src["inventory_change_bil"] = _extract_series(cashflow, ticker, ["tang, giam hang ton kho"], kind, VND_TO_BIL)
            src["payables_change_bil"] = _extract_series(cashflow, ticker, ["tang, giam cac khoan phai tra"], kind, VND_TO_BIL)
            src["prepaid_change_bil"] = _extract_series(cashflow, ticker, ["tang giam chi phi tra truoc"], kind, VND_TO_BIL)
            src["other_current_assets_change_bil"] = _extract_series(cashflow, ticker, ["tang giam tai san ngan han khac"], kind, VND_TO_BIL)
            src["interest_paid_bil"] = _extract_series(cashflow, ticker, ["tien lai vay da tra", "tien lai vay phai tra"], kind, VND_TO_BIL)
            src["tax_paid_bil"] = _extract_series(cashflow, ticker, ["thue thu nhap doanh nghiep da nop"], kind, VND_TO_BIL)
            src["other_operating_cash_in_bil"] = _extract_series(cashflow, ticker, ["tien thu khac tu hoat dong kinh doanh"], kind, VND_TO_BIL)
            src["other_operating_cash_out_bil"] = _extract_series(cashflow, ticker, ["tien chi khac tu hoat dong kinh doanh"], kind, VND_TO_BIL)
            src["equity_issued_bil"] = _extract_series(cashflow, ticker, ["tien thu tu phat hanh co phieu", "nhan von gop cua chu so huu"], kind, VND_TO_BIL)
            src["buyback_bil"] = _extract_series(cashflow, ticker, ["mua lai co phieu", "tien chi tra von gop cho cac chu so huu"], kind, VND_TO_BIL)
            src["debt_raised_bil"] = _extract_series(cashflow, ticker, ["tien vay ngan han, dai han nhan duoc"], kind, VND_TO_BIL)
            src["debt_repaid_bil"] = _extract_series(cashflow, ticker, ["tien chi tra no goc vay"], kind, VND_TO_BIL)
            src["investment_subsidiary_bil"] = _sum_series(
                _extract_series(cashflow, ticker, ["dau tu gop von vao cong ty lien doanh lien ket"], kind, VND_TO_BIL),
                _extract_series(cashflow, ticker, ["tien chi dau tu gop von vao don vi khac"], kind, VND_TO_BIL),
            )
            wc_series = [
                src["receivables_change_bil"], src["inventory_change_bil"], src["payables_change_bil"], src["prepaid_change_bil"], src["other_current_assets_change_bil"],
                src["interest_paid_bil"], src["tax_paid_bil"], src["other_operating_cash_in_bil"], src["other_operating_cash_out_bil"],
            ]
            src["working_capital_change_bil"] = _sum_series(*wc_series)
            src["net_debt_cashflow_bil"] = _sum_series(src["debt_raised_bil"], src["debt_repaid_bil"])
        if balance is not None:
            src["current_assets_bil"] = _extract_series(balance, ticker, ["tai san luu dong va dau tu ngan han", "tai san ngan han"], kind, VND_TO_BIL)
            src["cash_equivalents_bil"] = _extract_series(balance, ticker, ["tien va cac khoan tuong duong tien"], kind, VND_TO_BIL)
            src["short_term_investments_bil"] = _extract_series(balance, ticker, ["cac khoan dau tu tai chinh ngan han"], kind, VND_TO_BIL)
            src["accounts_receivable_bil"] = _extract_series(balance, ticker, ["cac khoan phai thu ngan han", "phai thu ngan han", "phai thu khach hang"], kind, VND_TO_BIL)
            src["inventory_bil"] = _extract_series(balance, ticker, ["hang ton kho"], kind, VND_TO_BIL)
            src["fixed_assets_bil"] = _extract_series(balance, ticker, ["tai san dai han"], kind, VND_TO_BIL)
            src["current_liabilities_bil"] = _extract_series(balance, ticker, ["no ngan han"], kind, VND_TO_BIL)
            src["accounts_payable_bil"] = _extract_series(balance, ticker, ["phai tra nguoi ban ngan han", "phai tra nguoi ban"], kind, VND_TO_BIL)
            # V23.16: lấy đúng nợ vay chịu lãi từ Bảng cân đối kế toán để WACC thay đổi theo từng năm.
            # Trước đây adapter chưa map các dòng vay/nợ thuê tài chính nên WACC thường rơi về 100% vốn chủ,
            # làm các năm bị giống nhau nếu beta proxy không đổi.
            src["short_term_debt_bil"] = _extract_series(balance, ticker, [
                "vay va no thue tai chinh ngan han"
            ], kind, VND_TO_BIL)
            src["current_portion_long_term_debt_bil"] = _extract_series(balance, ticker, [
                "vay va no dai han den han phai tra", "no dai han den han phai tra", "vay dai han den han phai tra"
            ], kind, VND_TO_BIL)
            src["long_term_debt_bil"] = _extract_series(balance, ticker, [
                "vay va no thue tai chinh dai han"
            ], kind, VND_TO_BIL)
            src["bonds_payable_bil"] = _extract_series(balance, ticker, [
                "trai phieu chuyen doi", "trai phieu phat hanh", "trai phieu"
            ], kind, VND_TO_BIL)
            # Không dùng pattern quá rộng "nợ thuê tài chính" ở đây vì sẽ bắt trùng dòng vay ngắn hạn/dài hạn.
            src["lease_liabilities_bil"] = {}
            src["total_assets_bil"] = _extract_series(balance, ticker, ["tong cong tai san"], kind, VND_TO_BIL)
            src["equity_bil"] = _extract_series(balance, ticker, ["nguon von chu so huu", "von chu so huu"], kind, VND_TO_BIL)
            src["liabilities_bil"] = _extract_series(balance, ticker, ["no phai tra"], kind, VND_TO_BIL)
            src["cash_and_short_investments_bil"] = _sum_series(src["cash_equivalents_bil"], src["short_term_investments_bil"])
        if ratios is not None:
            # Financial-v1.3.0 stores ratios as decimals. Convert to percent where applicable.
            ratio_map = {
                "eps_vnd": ["BasicEPS", "BasicEPS_TTM"],
                "roe_pct": ["ROE", "ROE_MRQ"],
                "roa_pct": ["ROA", "ROA_MRQ"],
                "roic_pct": ["ROIC", "ROIC_MRQ"],
                "gross_margin_pct": ["GrossMargin", "GrossMargin_MRQ"],
                "net_margin_pct": ["ProfitMargin", "NetProfitMargin", "ProfitMargin_MRQ"],
                "asset_turnover": ["AssetsTurnover", "AssetsTurnover_MRQ"],
                "equity_multiplier": ["EquityMultiplier", "FinancialLeverage"],
            }
            for field, labels in ratio_map.items():
                values = _extract_series(ratios, ticker, labels, kind, 1.0)
                if field.endswith("_pct"):
                    values = {k: _maybe_pct(v) for k, v in values.items()}
                src[field] = values
        return src

    def _build_timeseries(self, wb, ticker: str, kind: str) -> pd.DataFrame:
        src = self._series_sources(wb, ticker, kind)
        period_keys: set[str] = set()
        for values in src.values():
            period_keys.update(values.keys())
        parsed = []
        for p in period_keys:
            key = _period_key(p, kind)
            if key:
                parsed.append((*key, p))
        parsed.sort(key=lambda x: (x[0], x[1]))
        limit = 20 if kind.upper() == "Q" else 10
        parsed = parsed[-limit:]
        rows = []
        # Owner earnings uses an average maintenance capex assumption. Use last 5 normalized periods available.
        capex_map = src.get("capex_bil", {})
        for year, quarter, label, raw_label in parsed:
            row: dict[str, Any] = {
                "ticker": ticker,
                "period_type": kind.upper(),
                "period": label,
                "year": year,
                "quarter": quarter if kind.upper() == "Q" else "",
            }
            for field, values in src.items():
                row[field] = values.get(raw_label) if raw_label in values else values.get(label)
            # Derive core metrics if missing.
            row["free_cash_flow_bil"] = _sum_series({label: row.get("cfo_bil")}, {label: row.get("capex_bil")}).get(label)
            if row.get("working_capital_change_bil") is None:
                row["working_capital_change_bil"] = _sum_series(
                    {label: row.get("receivables_change_bil")}, {label: row.get("inventory_change_bil")}, {label: row.get("payables_change_bil")},
                    {label: row.get("prepaid_change_bil")}, {label: row.get("other_current_assets_change_bil")}, {label: row.get("interest_paid_bil")},
                    {label: row.get("tax_paid_bil")}, {label: row.get("other_operating_cash_in_bil")}, {label: row.get("other_operating_cash_out_bil")},
                ).get(label)
            if row.get("noncash_adjustments_bil") is None and row.get("operating_cash_before_wc_bil") is not None and row.get("pretax_profit_bil") is not None:
                row["noncash_adjustments_bil"] = row.get("operating_cash_before_wc_bil") - row.get("pretax_profit_bil")
            if row.get("noncash_adjustments_bil") is None:
                row["noncash_adjustments_bil"] = row.get("depreciation_bil")
            row["net_debt_cashflow_bil"] = _sum_series({label: row.get("debt_raised_bil")}, {label: row.get("debt_repaid_bil")}).get(label)
            row["cash_and_short_investments_bil"] = _sum_series({label: row.get("cash_equivalents_bil")}, {label: row.get("short_term_investments_bil")}).get(label)
            row["fcf_to_pretax"] = _ratio(row.get("free_cash_flow_bil"), row.get("pretax_profit_bil"))
            row["nibt_to_fcf"] = _ratio(row.get("pretax_profit_bil"), row.get("free_cash_flow_bil"))
            row["noncash_to_pretax"] = _ratio(row.get("noncash_adjustments_bil"), row.get("pretax_profit_bil"))
            row["wc_to_pretax"] = _ratio(row.get("working_capital_change_bil"), row.get("pretax_profit_bil"))
            row["capex_to_pretax"] = _ratio(row.get("capex_bil"), row.get("pretax_profit_bil"))
            row["gross_margin_pct"] = row.get("gross_margin_pct") if row.get("gross_margin_pct") is not None else _pct_ratio(row.get("gross_profit_bil"), row.get("revenue_bil"))
            row["net_margin_pct"] = row.get("net_margin_pct") if row.get("net_margin_pct") is not None else _pct_ratio(row.get("net_profit_bil"), row.get("revenue_bil"))
            row["asset_turnover"] = row.get("asset_turnover") if row.get("asset_turnover") is not None else _ratio(row.get("revenue_bil"), row.get("total_assets_bil"))
            row["equity_multiplier"] = row.get("equity_multiplier") if row.get("equity_multiplier") is not None else _ratio(row.get("total_assets_bil"), row.get("equity_bil"))
            row["roe_pct"] = row.get("roe_pct") if row.get("roe_pct") is not None else _pct_ratio(row.get("net_profit_bil"), row.get("equity_bil"))
            row["roa_pct"] = row.get("roa_pct") if row.get("roa_pct") is not None else _pct_ratio(row.get("net_profit_bil"), row.get("total_assets_bil"))
            # Simple ROIC fallback when source lacks ROIC: use pre-tax profit / (equity + liabilities) = pre-tax / assets.
            row["roic_pct"] = row.get("roic_pct") if row.get("roic_pct") is not None else _pct_ratio(row.get("pretax_profit_bil"), row.get("total_assets_bil"))
            row["roe_dupont_pct"] = (
                row["net_margin_pct"] / 100 * row["asset_turnover"] * row["equity_multiplier"] * 100
                if row.get("net_margin_pct") is not None and row.get("asset_turnover") is not None and row.get("equity_multiplier") is not None else None
            )
            row["cfo_to_net_profit"] = _ratio(row.get("cfo_bil"), row.get("net_profit_bil"))
            row["fcf_to_net_profit"] = _ratio(row.get("free_cash_flow_bil"), row.get("net_profit_bil"))
            rows.append(row)
        # Maintenance capex: rolling average of capex over data already sorted ascending. This follows the Owner Earnings idea:
        # earnings + depreciation + working-capital change + average maintenance capex (capex is usually negative).
        df = pd.DataFrame(rows)
        if not df.empty:
            capex = pd.to_numeric(df.get("capex_bil"), errors="coerce")
            maintenance_capex = capex.rolling(5, min_periods=1).mean()
            df["owner_earnings_bil"] = (
                pd.to_numeric(df.get("net_profit_bil"), errors="coerce")
                + pd.to_numeric(df.get("depreciation_bil"), errors="coerce").fillna(0)
                + pd.to_numeric(df.get("working_capital_change_bil"), errors="coerce").fillna(0)
                + maintenance_capex.fillna(0)
            )
            shares = None
            try:
                # derive EPS if not available using latest shares from overview sheet if possible.
                ov_ws = self._get_sheet(wb, "TỔNG QUAN")
                shares = _to_float(ov_ws["G24"].value) if ov_ws is not None else None
            except Exception:
                shares = None
            if shares:
                if "eps_vnd" not in df.columns:
                    df["eps_vnd"] = None
                eps_calc = pd.to_numeric(df.get("net_profit_bil"), errors="coerce") * 1_000_000_000 / (shares * 1_000_000)
                df["eps_vnd"] = pd.to_numeric(df["eps_vnd"], errors="coerce").fillna(eps_calc)
                df["oeps_vnd"] = pd.to_numeric(df["owner_earnings_bil"], errors="coerce") * 1_000_000_000 / (shares * 1_000_000)
        return df.reset_index(drop=True)
