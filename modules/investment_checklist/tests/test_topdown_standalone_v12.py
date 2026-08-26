from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
import sqlite3

import httpx
from openpyxl import load_workbook
import pandas as pd
from pypdf import PdfReader
import pytest

from adapters.base import ProviderResult
import module_topdown_engine as engine
from module_topdown_macro_update import resolve_effective_driver_scores, run_macro_update
from module_topdown_screening_data import screening_row_from_provider
from module_topdown_snapshot_export import (
    build_snapshot_excel_bytes,
    build_snapshot_pdf_bytes,
    snapshot_detail_frames,
    snapshot_export_filename,
)
from module_topdown_snapshot_store import TopDownMacroSnapshotStore, compare_snapshots


NOW = datetime(2026, 8, 24, 8, 0, tzinfo=timezone.utc)


def test_default_benchmark_is_official_current_and_becomes_stale_deterministically():
    inp = engine.default_input()
    assert inp.benchmark_id == "msci_vietnam_2026_07"
    assert sum(inp.benchmark_weights.values()) == pytest.approx(100.01)
    meta = next(
        item for item in engine.benchmark_config()["benchmarks"]
        if item["id"] == inp.benchmark_id
    )
    assert meta["source_name"] == "MSCI Vietnam Index Factsheet"
    assert meta["source_url"].startswith("https://www.msci.com/")
    assert engine.benchmark_reliability(meta, today=date(2026, 8, 24))[0] is True
    stale, stale_note = engine.benchmark_reliability(meta, today=date(2026, 9, 16))
    assert stale is False
    assert "đã cũ" in stale_note


def test_pdf_export_uses_runtime_unicode_fonts_without_repository_binaries():
    regular, bold = __import__("module_topdown_snapshot_export")._pdf_font_paths()
    assert regular.name in {"DejaVuSans.ttf", "LiberationSans-Regular.ttf"}
    assert bold.name in {"DejaVuSans-Bold.ttf", "LiberationSans-Bold.ttf"}
    assert regular.is_file() and bold.is_file()
    assert not Path("assets/fonts/DejaVuSans.ttf").exists()
    assert not Path("assets/fonts/DejaVuSans-Bold.ttf").exists()


def test_screening_derives_trecapital_metrics_without_manual_entry():
    overview = pd.DataFrame(
        [
            {
                "ticker": "DCM",
                "company_name": "Đạm Cà Mau",
                "industry": "Hóa chất cơ bản / Phân bón",
                "sub_industry": "Phân bón",
                "market_cap_bil": 1_000.0,
                "pe": 10.0,
                "pb": 2.0,
                "ps": 1.5,
            }
        ]
    )
    quarterly = pd.DataFrame(
        [
            {
                "year": 2025 + (index // 4),
                "quarter": index % 4 + 1,
                "cfo_bil": 10.0,
                "interest_bearing_debt_bil": 50.0,
                "equity_bil": 100.0,
            }
            for index in range(4)
        ]
    )
    result = ProviderResult(overview, pd.DataFrame(), quarterly, note="fixture")
    row = screening_row_from_provider("DCM", result, source="Trecapital fixture").row

    assert row["Mã ngành"] == "MAT"
    assert row["Vốn hóa (tỷ đồng)"] == 1_000.0
    assert row["P/CF (lần)"] == 25.0
    assert row["Nợ vay/Vốn chủ (lần)"] == 0.5
    assert row["GTGD bình quân 20 phiên (tỷ đồng)"] is None
    assert "GTGD bình quân 20 phiên" in row["Ghi chú dữ liệu"]


def test_screening_missing_metric_can_never_pass():
    row = engine.mau_bang_sang_loc(1).iloc[0].to_dict()
    row.update(
        {
            "Mã CK": "DCM",
            "Tên doanh nghiệp": "Đạm Cà Mau",
            "Mã ngành": "MAT",
            "Vốn hóa (tỷ đồng)": 10_000.0,
            "P/E (lần)": 8.0,
            "P/B (lần)": 1.5,
            "P/CF (lần)": 7.0,
            "P/S (lần)": 1.0,
            "Nợ vay/Vốn chủ (lần)": 0.2,
            # Liquidity intentionally missing.
        }
    )
    criteria = engine.scoring_config()["sang_loc_dinh_luong_mac_dinh"]["rong"]
    result = engine.chay_sang_loc(pd.DataFrame([row]), criteria).iloc[0]
    assert result["Kết quả"] == "Thiếu dữ liệu"
    assert "GTGD bình quân 20 phiên" in result["Lý do loại"]


def test_standalone_macro_update_uses_world_bank_fallback_and_never_writes_driver():
    calls: list[str] = []

    def handler(request: httpx.Request) -> httpx.Response:
        calls.append(str(request.url))
        if "NGDP_RPCH" in request.url.path:
            return httpx.Response(403, json={"error": "blocked"}, request=request)
        if "NY.GDP.MKTP.KD.ZG" in request.url.path:
            return httpx.Response(
                200,
                json=[
                    {"page": 1},
                    [
                        {"date": "2025", "value": 7.1, "obs_status": ""},
                        {"date": "2024", "value": 6.1, "obs_status": ""},
                    ],
                ],
                request=request,
            )
        return httpx.Response(404, json={"error": "fixture"}, request=request)

    client = httpx.Client(transport=httpx.MockTransport(handler))
    result = run_macro_update(["gdp_growth"], http_client=client, api_keys={}, now=NOW)
    client.close()

    assert len(calls) == 2
    assert result["status"] == "completed"
    assert result["trigger_type"] == "manual_click"
    assert result["guardrail"].endswith("no company/checklist write")
    observation = result["observations"][0]
    assert observation["source_code"] == "world_bank_wdi"
    assert observation["series_code"] == "NY.GDP.MKTP.KD.ZG"
    assert observation["fallback_from"] == "imf_datamapper:NGDP_RPCH"
    assert result["suggestions"][0]["suggested_score"] == 2


def test_automatic_driver_suggestion_is_baseline_but_analyst_override_wins():
    resolved = resolve_effective_driver_scores(
        {"gdp_growth": 0.0, "inflation": 1.0, "employment": -1.0},
        {
            "gdp_growth": "default",
            "inflation": "analyst_override",
            "employment": "automatic_suggestion",
        },
        [
            {"driver_id": "gdp_growth", "suggested_score": 2},
            {"driver_id": "inflation", "suggested_score": -2},
            {"driver_id": "employment", "suggested_score": None},
        ],
    )

    assert resolved["effective_scores"]["gdp_growth"] == 2.0
    assert resolved["score_sources"]["gdp_growth"] == "automatic_suggestion"
    assert resolved["effective_scores"]["inflation"] == 1.0
    assert resolved["score_sources"]["inflation"] == "analyst_override"
    assert resolved["applied_driver_ids"] == ["gdp_growth"]
    assert resolved["analyst_override_ids"] == ["inflation"]
    assert resolved["research_gap_ids"] == ["employment"]


def _snapshot_payload(score: float, rank: int) -> dict:
    return {
        "schema": "fisher-topdown-macro-snapshot-v1",
        "parameters": {"driver_outlook": {"gdp_growth": score}},
        "ranking": [
            {
                "sector_code": "MAT",
                "sector_name": "Nguyên vật liệu",
                "rank": rank,
                "score": 60.0 + score,
            }
        ],
    }


def test_macro_snapshot_store_is_append_only_and_comparable(tmp_path):
    store = TopDownMacroSnapshotStore(tmp_path / "macro.db")
    store.initialize()
    first = store.save(
        _snapshot_payload(0.0, 2),
        as_of_date="2026-08-01",
        snapshot_label="Baseline",
        save_reason="Start proxy",
        created_by="analyst",
        methodology_version="V1",
    )
    second = store.save(
        _snapshot_payload(1.0, 1),
        as_of_date="2026-08-24",
        snapshot_label="After update",
        save_reason="Macro changed",
        created_by="analyst",
        methodology_version="V1",
    )
    assert first["version_no"] == 1 and second["version_no"] == 2
    rows = store.list()
    assert [row["version_no"] for row in rows] == [2, 1]
    delta = compare_snapshots(rows[0], rows[1])
    assert delta["drivers"][0]["Thay đổi"] == 1.0
    assert delta["sectors"][0]["Hạng cũ"] == 2
    assert delta["sectors"][0]["Hạng mới"] == 1

    with sqlite3.connect(tmp_path / "macro.db") as conn, pytest.raises(
        sqlite3.IntegrityError, match="append-only"
    ):
        conn.execute("UPDATE topdown_macro_snapshots SET snapshot_label='mutated' WHERE id=1")


def test_snapshot_detail_and_excel_pdf_exports_use_the_immutable_payload(tmp_path):
    payload = {
        "schema": "fisher-topdown-macro-snapshot-v1",
        "methodology_version": "V1.4",
        "generated_at": "2026-08-24T09:12:57+00:00",
        "cycle_phase": "mid",
        "benchmark": {"id": "vnindex", "name": "VN-Index"},
        "parameters": {
            "driver_outlook": {"gdp_growth": 1.0, "inflation": -1.0},
            "driver_score_sources": {
                "gdp_growth": "automatic_suggestion",
                "inflation": "analyst_override",
            },
            "automatic_driver_scores": {"gdp_growth": 1.0, "inflation": 0.0},
        },
        "latest_macro_update": {
            "retrieved_at": "2026-08-24T09:12:57+00:00",
            "observations": [
                {
                    "driver_id": "gdp_growth",
                    "driver_name": "Tăng trưởng GDP thực",
                    "source_code": "world_bank_wdi",
                    "series_code": "NY.GDP.MKTP.KD.ZG",
                    "period_label": "2025",
                    "value_numeric": 8.02,
                    "previous_value_numeric": 7.04,
                    "delta_numeric": 0.98,
                    "unit": "%",
                    "freshness_status": "current",
                    "source_url": "https://data.worldbank.org/",
                }
            ],
            "suggestions": [
                {
                    "driver_id": "gdp_growth",
                    "driver_name": "Tăng trưởng GDP thực",
                    "suggested_score": 1,
                    "confidence": "high",
                    "rationale": "Tăng nhanh hơn kỳ trước.",
                }
            ],
        },
        "ranking": [
            {
                "rank": 1,
                "sector_code": "IND",
                "sector_name": "Công nghiệp",
                "character": "Chu kỳ",
                "economy_score": 3.0,
                "politics_score": 0.0,
                "sentiment_score": 0.0,
                "cycle_score": 3.0,
                "score": 60.0,
            }
        ],
        "weights": [
            {
                "sector_code": "IND",
                "sector_name": "Công nghiệp",
                "sector_score": 60.0,
                "signal": "Tăng tỷ trọng",
                "tilt_factor": 1.1,
                "benchmark_weight_pct": 10.0,
                "proposed_weight_pct": 11.0,
                "tilt_pct": 1.0,
            }
        ],
        "sync_checks": [{"Hạng mục kiểm tra": "Tổng tỷ trọng", "Tình trạng": "Đạt"}],
    }
    store = TopDownMacroSnapshotStore(tmp_path / "export.db")
    store.initialize()
    snapshot = store.save(
        payload,
        as_of_date="2026-08-24",
        snapshot_label="Baseline vĩ mô",
        save_reason="Lưu proxy tháng 8",
        created_by="analyst",
        methodology_version="V1.4",
        source_registry_hash="registry-hash",
    )
    catalog = [
        {"id": "gdp_growth", "nhom": "Kinh tế", "ten_vi": "Tăng trưởng GDP thực"},
        {"id": "inflation", "nhom": "Kinh tế", "ten_vi": "Lạm phát"},
    ]
    frames = snapshot_detail_frames(
        snapshot,
        driver_catalog=catalog,
        cycle_labels={"mid": "Giữa chu kỳ"},
    )
    assert list(frames) == [
        "Tổng quan",
        "Portfolio Drivers",
        "Dữ liệu vĩ mô",
        "Xếp hạng ngành",
        "Tỷ trọng đề xuất",
        "Kiểm tra dữ liệu",
    ]
    drivers = frames["Portfolio Drivers"].set_index("Driver ID")
    assert drivers.loc["gdp_growth", "Nguồn điểm"] == "Gợi ý tự động"
    assert drivers.loc["inflation", "Nguồn điểm"] == "Analyst tự chấm"
    assert frames["Dữ liệu vĩ mô"].iloc[0]["Giá trị"] == pytest.approx(8.02)

    excel_bytes = build_snapshot_excel_bytes(
        snapshot,
        driver_catalog=catalog,
        cycle_labels={"mid": "Giữa chu kỳ"},
    )
    assert excel_bytes.startswith(b"PK")
    workbook = load_workbook(BytesIO(excel_bytes), data_only=False)
    assert workbook.sheetnames == [
        "Tổng quan",
        "Portfolio Drivers",
        "Dữ liệu vĩ mô",
        "Xếp hạng ngành",
        "Tỷ trọng đề xuất",
        "Kiểm tra dữ liệu",
    ]
    assert workbook["Portfolio Drivers"]["D5"].value == -1.0
    assert workbook["Portfolio Drivers"]["D5"].font.color.rgb == "00C00000"

    pdf_bytes = build_snapshot_pdf_bytes(
        snapshot,
        driver_catalog=catalog,
        cycle_labels={"mid": "Giữa chu kỳ"},
    )
    assert pdf_bytes.startswith(b"%PDF")
    pdf = PdfReader(BytesIO(pdf_bytes))
    assert len(pdf.pages) >= 6
    assert "Fisher Top-Down" in (pdf.pages[0].extract_text() or "")
    assert snapshot_export_filename(snapshot, "xlsx") == "fisher_topdown_snapshot_v1_2026-08-24.xlsx"


def test_snapshot_history_ui_is_read_only_and_supports_arbitrary_comparison_and_exports():
    dashboard = Path("module_topdown_dashboard.py").read_text(encoding="utf-8")
    assert '"Chọn snapshot để xem chi tiết"' in dashboard
    assert '"Bản mới"' in dashboard and '"Bản gốc"' in dashboard
    assert '"⬇️ Tải snapshot Excel"' in dashboard
    assert '"⬇️ Tải snapshot PDF"' in dashboard
    assert "Snapshot cũ không được nạp ngược" in dashboard
    assert "store.update(" not in dashboard
    assert "store.delete(" not in dashboard


def test_fisher_page_and_checklist_routes_are_decoupled():
    dashboard = Path("module_topdown_dashboard.py").read_text(encoding="utf-8")
    legacy = Path("modules/investment_checklist/ui/integration_preview.py").read_text(encoding="utf-8")
    fast = Path("modules/investment_checklist/ui/integration_preview_v3.py").read_text(encoding="utf-8")

    assert '"💾 Snapshot vĩ mô"' in dashboard
    assert '"🔄 Cập nhật dữ liệu vĩ mô mới nhất"' in dashboard
    assert '"🔄 Lấy dữ liệu & sàng lọc"' in dashboard
    assert '"🏢 Phân tích cổ phiếu 5 bước",' not in dashboard
    assert "_bridge_to_other_modules(" not in dashboard
    assert "topdown_phase9_applied" not in dashboard
    assert '"🔄 Latest Data Update"' not in legacy + fast


def test_topdown_tables_wrap_responsively_and_autosize_component_frames():
    dashboard = Path("module_topdown_dashboard.py").read_text(encoding="utf-8")

    assert "table-layout:fixed" in dashboard
    assert "white-space:normal!important" in dashboard
    assert "overflow-wrap:anywhere" in dashboard
    assert "data-label=" in dashboard
    assert "@media (max-width:760px)" in dashboard
    assert "streamlit:setFrameHeight" in dashboard
    assert "scrolling=False" in dashboard
    assert "st.dataframe(" not in dashboard
